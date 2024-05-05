'''
See COPYRIGHT.md for copyright information.
'''
import os
from datetime import timedelta
from collections import OrderedDict
from copy import deepcopy
from lxml import etree
from arelle import ViewFile
from arelle.Aspect import Aspect, aspectModels, aspectRuleAspects, aspectModelAspect, aspectStr
from arelle.ModelInstanceObject import ModelDimensionValue
from arelle.PrototypeInstanceObject import FactPrototype
from arelle.PythonUtil import OrderedSet
from arelle.ModelRenderingObject import (StrctMdlBreakdown, StrctMdlStructuralNode,
                                         DefnMdlClosedDefinitionNode, DefnMdlRuleDefinitionNode, DefnMdlAspectNode,
                                         OPEN_ASPECT_ENTRY_SURROGATE, ROLLUP_SPECIFIES_MEMBER, ROLLUP_FOR_DIMENSION_RELATIONSHIP_NODE,
                                         aspectStrctNodes)
from arelle.rendering.RenderingLayout import layoutTable
from arelle.rendering.RenderingResolution import RENDER_UNITS_PER_CHAR
from arelle.ModelValue import QName
from arelle.ModelXbrl import DEFAULT
from arelle.ViewFile import HTML, XLSX

# deferred opening of openpyxl so it's not needed in site-packages unless it is used
Workbook = cell = utils = Font = PatternFill = Border = Alignment = Color = fills = Side = Comment = None


# change tableModel for namespace needed for consistency suite
'''
from arelle.XbrlConst import (tableModelMMDD as tableModelNamespace,
                              tableModelMMDDQName as tableModelQName)
'''
from arelle import XbrlConst
from arelle.XmlUtil import innerTextList, child, elementFragmentIdentifier, addQnameValue
from collections import defaultdict

emptySet = set()
emptyList = []

headerOmittedRollupAspects  = {
    Aspect.CONCEPT,
    Aspect.COMPLETE_SEGMENT, Aspect.COMPLETE_SCENARIO, Aspect.NON_XDT_SEGMENT, Aspect.NON_XDT_SCENARIO,
    Aspect.VALUE, Aspect.SCHEME,
    Aspect.PERIOD_TYPE, Aspect.START, Aspect.END, Aspect.INSTANT,
    Aspect.UNIT_MEASURES, Aspect.MULTIPLY_BY, Aspect.DIVIDE_BY}


def viewRenderedGrid(modelXbrl, outfile, lang=None, viewTblELR=None, sourceView=None, diffToFile=False, cssExtras="", table=None):
    modelXbrl.modelManager.showStatus(_("saving rendered tagle"))
    view = ViewRenderedGrid(modelXbrl, outfile, lang, cssExtras)

    if sourceView is not None:
        if sourceView.lytMdlTblMdl:
            lytMdlTblMdl = sourceView.lytMdlTblMdl
        else:
            lytMdlTblMdl = layoutTable(sourceView)
    else:
        layoutTable(view, table)
        lytMdlTblMdl = view.lytMdlTblMdl
    if view.type == HTML and view.tblElt is not None: # may be None if there is no table
        view.viewHTML(lytMdlTblMdl)
    elif view.type == XLSX:
        view.viewXLSX(lytMdlTblMdl)
    view.close()
    modelXbrl.modelManager.showStatus(_("rendered table saved to {0}").format(outfile), clearAfter=5000)

class ViewRenderedGrid(ViewFile.View):
    def __init__(self, modelXbrl, outfile, lang, cssExtras):
        global Workbook, cell, utils, Font, PatternFill, Border, Alignment, Color, fills, Side, Comment
        # find table model namespace based on table namespace
        self.tableModelNamespace = XbrlConst.tableModel
        for xsdNs in modelXbrl.namespaceDocs.keys():
            if xsdNs in (XbrlConst.tableMMDD, XbrlConst.table):
                self.tableModelNamespace = xsdNs + "/model"
                break
        super(ViewRenderedGrid, self).__init__(modelXbrl, outfile,
                                               'tableModel',
                                               lang,
                                               style="rendering",
                                               cssExtras=cssExtras)
        if self.type == XLSX:
            if Workbook is None:
                from openpyxl import Workbook, cell, utils
                from openpyxl.styles import Font, PatternFill, Border, Alignment, Color, fills, Side
                from openpyxl.comments import Comment

        class nonTkBooleanVar():
            def __init__(self, value=True):
                self.value = value
            def set(self, value):
                self.value = value
            def get(self):
                return self.value
        # context menu boolean vars (non-tkinter boolean
        self.ignoreDimValidity = nonTkBooleanVar(value=True)
        self.openBreakdownLines = 0 # layout model conformance suite requires no open entry lines
    def viewReloadDueToMenuAction(self, *args):
        self.view()

    def view(self, lytMdlTblMdl):
        if self.type == HTML:
            self.viewHTML(lytMdlTblMdl)
        elif self.type == XLSX:
            self.viewXLSX(lytMdlTblMdl)

    def viewHTML(self, lytMdlTblMdl):
        for lytMdlTableSet in lytMdlTblMdl.lytMdlTableSets:
            self.tblElt.append(etree.Comment(f"TableSet linkbase file: {lytMdlTableSet.srcFile}, line {lytMdlTableSet.srcLine}"))
            self.tblElt.append(etree.Comment(f"TableSet linkrole: {lytMdlTableSet.srcLinkrole}"))
            tblSetHdr = etree.SubElement(self.tblElt, "{http://www.w3.org/1999/xhtml}tr")
            etree.SubElement(tblSetHdr, "{http://www.w3.org/1999/xhtml}td").text = lytMdlTableSet.label
            for lytMdlTable in lytMdlTableSet.lytMdlTables:
                if lytMdlTable.strctMdlTable.tblParamValues:
                    # show any parameters
                    pTableRow = etree.SubElement(self.tblElt, "{http://www.w3.org/1999/xhtml}tr")
                    pRowCell = etree.SubElement(pTableRow, "{http://www.w3.org/1999/xhtml}td")
                    paramTable = etree.SubElement(pRowCell, "{http://www.w3.org/1999/xhtml}table",
                                                  attrib={"border":"1", "cellspacing":"0", "cellpadding":"4", "style":"font-size:8pt;"})
                    pHdrRow = etree.SubElement(paramTable, "{http://www.w3.org/1999/xhtml}tr")
                    etree.SubElement(pHdrRow, "{http://www.w3.org/1999/xhtml}th").text = "parameter"
                    etree.SubElement(pHdrRow, "{http://www.w3.org/1999/xhtml}th").text = "value"
                    for name, value in lytMdlTable.strctMdlTable.tblParamValues.items():
                        pTableRow = etree.SubElement(paramTable, "{http://www.w3.org/1999/xhtml}tr")
                        etree.SubElement(pTableRow, "{http://www.w3.org/1999/xhtml}td").text = str(name)
                        etree.SubElement(pTableRow, "{http://www.w3.org/1999/xhtml}td").text = str(value)
                # each Z is a separate table in the outer table
                lytMdlZHdrs = lytMdlTable.lytMdlAxisHeaders("z")
                if lytMdlZHdrs is not None:
                    lytMdlZHdrGroups = lytMdlZHdrs.lytMdlGroups
                    numZtbls = lytMdlTable.numBodyCells("z") or 1 # must have at least 1 z entry
                    zHdrElts = [OrderedDict() for i in range(numZtbls)]
                    for lytMdlZGrp in lytMdlZHdrs.lytMdlGroups:
                        for lytMdlZHdr in lytMdlZGrp.lytMdlHeaders:
                            zRow = 0
                            if all(lytMdlCell.isOpenAspectEntrySurrogate for lytMdlCell in lytMdlZHdr.lytMdlCells):
                                continue # skip header with only open aspect entry surrogate
                            for lytMdlZCell in lytMdlZHdr.lytMdlCells:
                                zConstraint = tuple(c.aspect for c in lytMdlZCell.lytMdlConstraints)
                                for iSpan in range(lytMdlZCell.span):
                                    if not lytMdlZCell.rollup:
                                        zHdrElts[zRow][zConstraint] = tuple(lbl[0] for lbl in lytMdlZCell.labels)
                                    zRow += 1
                else:
                    zHdrElts = [{}]
                    numZtbls = 1
                zTbl = 0
                lytMdlZBodyCell = lytMdlTable.lytMdlBodyChildren[0] # examples only show one z cell despite number of tables
                for lytMdlYBodyCell in lytMdlZBodyCell.lytMdlBodyChildren:
                    zTableRow = etree.SubElement(self.tblElt, "{http://www.w3.org/1999/xhtml}tr")
                    zRowCell = etree.SubElement(zTableRow, "{http://www.w3.org/1999/xhtml}td")
                    zCellTable = etree.SubElement(zRowCell, "{http://www.w3.org/1999/xhtml}table",
                                                  attrib={"border":"1", "cellspacing":"0", "cellpadding":"4", "style":"font-size:8pt;"})
                    lytMdlXHdrs = lytMdlTable.lytMdlAxisHeaders("x")
                    lytMdlYHdrs = lytMdlTable.lytMdlAxisHeaders("y")
                    nbrXcolHdrs = lytMdlTable.headerDepth("x")
                    nbrYrowHdrs = lytMdlTable.headerDepth("y")
                    # build y row headers
                    numYrows = lytMdlTable.numBodyCells("y")
                    yRowHdrs = [[] for i in range(numYrows)] # list of list of row header elements for each row
                    for lytMdlYGrp in lytMdlYHdrs.lytMdlGroups:
                        for lytMdlYHdr in lytMdlYGrp.lytMdlHeaders:
                            yRow = 0
                            if all(lytMdlCell.isOpenAspectEntrySurrogate for lytMdlCell in lytMdlYHdr.lytMdlCells):
                                continue # skip header with only open aspect entry surrogate
                            for lytMdlYCell in lytMdlYHdr.lytMdlCells:
                                for iLabel in range(lytMdlYHdr.maxNumLabels):
                                    if lytMdlYCell.isOpenAspectEntrySurrogate:
                                        continue # strip all open aspect entry surrogates from layout model file
                                    attrib = {"style":"max-width:100em;text-align:left;"}
                                    if lytMdlYCell.rollup:
                                        attrib["class"] = "yAxisTopSpanLeg"
                                    else:
                                        attrib["class"] = "yAxisHdr"
                                    if lytMdlYCell.span > 1:
                                        attrib["rowspan"] = str(lytMdlYCell.span)
                                    rowHdrElt = etree.Element("{http://www.w3.org/1999/xhtml}th", attrib=attrib)
                                    rowHdrElt.text = lytMdlYCell.labelXmlText(iLabel,"\u00a0")
                                    yRowHdrs[yRow].append(rowHdrElt)
                                yRow += lytMdlYCell.span
                    firstColHdr = True
                    for lytMdlGroup in lytMdlXHdrs.lytMdlGroups:
                        for lytMdlHeader in lytMdlGroup.lytMdlHeaders:
                            if all(lytMdlCell.isOpenAspectEntrySurrogate for lytMdlCell in lytMdlHeader.lytMdlCells):
                                continue # skip header with only open aspect entry surrogate
                            for iLabel in range(lytMdlHeader.maxNumLabels):
                                rowElt = etree.SubElement(zCellTable, "{http://www.w3.org/1999/xhtml}tr")
                                if firstColHdr:
                                    zHdrElt = etree.SubElement(rowElt, "{http://www.w3.org/1999/xhtml}th",
                                                               attrib={"class":"tableHdr",
                                                                       "style":"max-width:100em;",
                                                                       "colspan": str(nbrYrowHdrs),
                                                                       "rowspan": str(nbrXcolHdrs)})
                                    if zHdrElts[zTbl]:
                                        zHdrTblElt = etree.SubElement(zHdrElt,"{http://www.w3.org/1999/xhtml}table",
                                                                      attrib={"style":"border-top:none;border-left:none;border-right:none;border-bottom:none;"})
                                        for zHdrLblRow in zHdrElts[zTbl].values():
                                            zHdrRowElt = etree.SubElement(zHdrTblElt, "{http://www.w3.org/1999/xhtml}tr")
                                            for lbl in zHdrLblRow:
                                                lblElt = etree.SubElement(zHdrRowElt, "{http://www.w3.org/1999/xhtml}th",
                                                                     attrib={"class":"tableHdr","style":"max-width:100em;font-size:8pt;text-align:left;border-top:none;border-left:none;border-right:none;border-bottom:none;"})
                                                if lbl != OPEN_ASPECT_ENTRY_SURROGATE:
                                                    lblElt.text = lbl
                                    else:
                                        zHdrElt.text = '\u00a0'
                                    firstColHdr = False
                                for lytMdlCell in lytMdlHeader.lytMdlCells:
                                    if lytMdlCell.isOpenAspectEntrySurrogate:
                                        continue # strip all open aspect entry surrogates from layout model file
                                    attrib = {"style":"max-width:100em;"}
                                    if lytMdlCell.rollup:
                                        attrib["class"] = "xAxisSpanLeg"
                                    else:
                                        attrib["class"] = "xAxisHdr"
                                    if lytMdlCell.span > 1:
                                        attrib["colspan"] = str(lytMdlCell.span)
                                    etree.SubElement(rowElt, "{http://www.w3.org/1999/xhtml}th", attrib=attrib
                                                     ).text = lytMdlCell.labelXmlText(iLabel,"\u00a0")
                    yRowNum = 0
                    for lytMdlXBodyCell in lytMdlYBodyCell.lytMdlBodyChildren:
                        if True: # not any(lytMdlCell.isOpenAspectEntrySurrogate for lytMdlCell in lytMdlXBodyCell.lytMdlBodyChildren):
                            rowElt = etree.SubElement(zCellTable, "{http://www.w3.org/1999/xhtml}tr")
                            if yRowNum < len(yRowHdrs):
                                for rowHdrElt in yRowHdrs[yRowNum]:
                                    rowElt.append(rowHdrElt)
                            for lytMdlCell in lytMdlXBodyCell.lytMdlBodyChildren:
                                if lytMdlCell.isOpenAspectEntrySurrogate:
                                    continue
                                justify = "left"
                                for f, v, justify in lytMdlCell.facts:
                                    break;
                                colElt = etree.SubElement(rowElt, "{http://www.w3.org/1999/xhtml}td",
                                                          attrib={"class":"cell",
                                                                  "style":f"text-align:{justify};width:8em;"}
                                                 ).text = "\n".join(v for f, v, justify in lytMdlCell.facts)
                            yRowNum += 1
                    if zTbl < len(lytMdlZBodyCell.lytMdlBodyChildren) - 1:
                        zTbl += 1

    def viewXLSX(self, lytMdlTblMdl):
        self.setAutoFilter(False) # filtering not comfortable with grid of tables
        thinBorder = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        xColHdrBorder = Border(left=Side(style='thin'), top=Side(style='thin'), right=Side(style='thin'))
        yRowHdrBorder = Border(left=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        for lytMdlTableSet in lytMdlTblMdl.lytMdlTableSets:
            numCols = 1
            titleXlsRow = self.xlsxRow + 1
            titleCell = self.xlsxWs.cell(row=self.xlsxRow+1, column=1)
            titleCell.comment = Comment(
                f"TableSet linkbase file: {lytMdlTableSet.srcFile}, line {lytMdlTableSet.srcLine} \n"
                f"TableSet linkrole: {lytMdlTableSet.srcLinkrole}",
                "Arelle")
            titleCell.value = lytMdlTableSet.label
            titleCell.alignment = Alignment(wrap_text=True)
            titleCell.border = thinBorder
            # left align and merger to width of table
            self.xlsxRow += 1
            for lytMdlTable in lytMdlTableSet.lytMdlTables:
                if lytMdlTable.strctMdlTable.tblParamValues:
                    # show any parameters
                    params = ["parameter = value"]
                    for name, value in lytMdlTable.strctMdlTable.tblParamValues.items():
                        params.append(f"{name} = {value}")
                    self.xlsxWs.cell(row=self.xlsxRow+1, column=iCol+1).value = "\n".join(params)
                    self.xlsxRow += 1
                # each Z is a separate table in the outer table
                lytMdlZHdrs = lytMdlTable.lytMdlAxisHeaders("z")
                if lytMdlZHdrs is not None:
                    lytMdlZHdrGroups = lytMdlZHdrs.lytMdlGroups
                    numZtbls = lytMdlTable.numBodyCells("z") or 1 # must have at least 1 z entry
                    zHdrLbls = [[] for i in range(numZtbls)]
                    for lytMdlZGrp in lytMdlZHdrs.lytMdlGroups:
                        for lytMdlZHdr in lytMdlZGrp.lytMdlHeaders:
                            zRow = 0
                            if all(lytMdlCell.isOpenAspectEntrySurrogate for lytMdlCell in lytMdlZHdr.lytMdlCells):
                                continue # skip header with only open aspect entry surrogate
                            for lytMdlZCell in lytMdlZHdr.lytMdlCells:
                                for iSpan in range(lytMdlZCell.span):
                                    zHdrLbls[zRow].append([lbl[0] for lbl in lytMdlZCell.labels])
                                    zRow += 1
                else:
                    zHdrLbls = [[]]
                    numZtbls = 1
                zTbl = 0
                lytMdlZBodyCell = lytMdlTable.lytMdlBodyChildren[0] # examples only show one z cell despite number of tables
                for lytMdlYBodyCell in lytMdlZBodyCell.lytMdlBodyChildren:
                    lytMdlXHdrs = lytMdlTable.lytMdlAxisHeaders("x")
                    lytMdlYHdrs = lytMdlTable.lytMdlAxisHeaders("y")
                    nbrXcolHdrs = lytMdlTable.headerDepth("x")
                    nbrYrowHdrs = lytMdlTable.headerDepth("y")
                    # build y row headers
                    numYrows = lytMdlTable.numBodyCells("y")
                    yRowHdrs = [[] for i in range(numYrows)] # list of list of row header elements for each row
                    for lytMdlYGrp in lytMdlYHdrs.lytMdlGroups:
                        for lytMdlYHdr in lytMdlYGrp.lytMdlHeaders:
                            yRow = 0
                            if all(lytMdlCell.isOpenAspectEntrySurrogate for lytMdlCell in lytMdlYHdr.lytMdlCells):
                                continue # skip header with only open aspect entry surrogate
                            for lytMdlYCell in lytMdlYHdr.lytMdlCells:
                                for iLabel in range(lytMdlYHdr.maxNumLabels):
                                    if lytMdlYCell.isOpenAspectEntrySurrogate:
                                        continue # strip all open aspect entry surrogates from layout model file
                                    rowHdrElt = {"align": "left"}
                                    if lytMdlYCell.rollup:
                                        rowHdrElt["class"] = "yAxisTopSpanLeg"
                                    else:
                                        rowHdrElt["class"] = "yAxisHdr"
                                    if lytMdlYCell.span > 1:
                                        rowHdrElt["rowspan"] = lytMdlYCell.span
                                    rowHdrElt["text"] = lytMdlYCell.labelXmlText(iLabel,"")
                                    yRowHdrs[yRow].append(rowHdrElt)
                                yRow += lytMdlYCell.span
                    yHdrCols = max(len(yRowHdrs[y]) for y in range(len(yRowHdrs)))
                    yFirstHdrRow = self.xlsxRow + 1
                    # upper left row/col hdr col, contains z axis labels, if any
                    zHdrCell = self.xlsxWs.cell(row=yFirstHdrRow, column=1)
                    zHdrCell.fill = PatternFill(patternType=fills.FILL_SOLID, fgColor=Color("EEEEEE"))
                    zHdrCell.border = thinBorder
                    v = "\n".join(((" ".join(lbl for lbl in zHdrRowLbls))
                                   for zHdrRowLbls in zHdrLbls[zTbl]
                                   if zHdrRowLbls))
                    if v:
                        zHdrCell.value = v
                        zHdrCell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    for lytMdlGroup in lytMdlXHdrs.lytMdlGroups:
                        for lytMdlHeader in lytMdlGroup.lytMdlHeaders:
                            if all(lytMdlCell.isOpenAspectEntrySurrogate for lytMdlCell in lytMdlHeader.lytMdlCells):
                                continue # skip header with only open aspect entry surrogate
                            for iLabel in range(lytMdlHeader.maxNumLabels):
                                xlsxCol = yHdrCols + 1
                                self.xlsxRow += 1
                                for i, lytMdlCell in enumerate(lytMdlHeader.lytMdlCells):
                                    if lytMdlCell.isOpenAspectEntrySurrogate:
                                        continue # strip all open aspect entry surrogates from layout model file
                                    c = self.xlsxWs.cell(row=self.xlsxRow, column=xlsxCol)
                                    c.value = lytMdlCell.labelXmlText(iLabel, None)
                                    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                                    c.fill = PatternFill(patternType=fills.FILL_SOLID, fgColor=Color("EEEEEE"))
                                    c.border = xColHdrBorder
                                    if lytMdlCell.span > 1:
                                        self.xlsxWs.merge_cells(range_string='%s%s:%s%s' % (utils.get_column_letter(xlsxCol), self.xlsxRow, utils.get_column_letter(xlsxCol + lytMdlCell.span - 1), self.xlsxRow))
                                        xlsxCol += lytMdlCell.span
                                    else:
                                        xlsxCol += 1
                    # upper left row/col hdr col
                    if yHdrCols > 1:
                        self.xlsxWs.merge_cells(range_string='%s%s:%s%s' % (utils.get_column_letter(1), yFirstHdrRow, utils.get_column_letter(yHdrCols), self.xlsxRow))
                    yRowNum = 0
                    rowspans = len(yRowHdrs[yRowNum]) * [0] # remaining rowspan per row col
                    for lytMdlXBodyCell in lytMdlYBodyCell.lytMdlBodyChildren:
                        if True: # not any(lytMdlCell.isOpenAspectEntrySurrogate for lytMdlCell in lytMdlXBodyCell.lytMdlBodyChildren):
                            self.xlsxRow += 1
                            xlsxCol = 1
                            if yRowNum < len(yRowHdrs):
                                for i, rowHdrElt in enumerate(yRowHdrs[yRowNum]):
                                    while xlsxCol-1 < len(rowspans) and rowspans[xlsxCol-1] > 0:
                                        rowspans[xlsxCol-1] -= 1
                                        xlsxCol += 1
                                    rowspan = rowHdrElt.get("rowspan", 1)
                                    c = self.xlsxWs.cell(row=self.xlsxRow, column=xlsxCol)
                                    v = rowHdrElt.get("text")
                                    c.value = v
                                    c.alignment = Alignment(horizontal=rowHdrElt.get("align", "left" if xlsxCol == 1 or not v.isnumeric() else "center"),
                                                            vertical="center", wrap_text=True)
                                    c.fill = PatternFill(patternType=fills.FILL_SOLID, fgColor=Color("EEEEEE"))
                                    c.border = yRowHdrBorder
                                    if rowspan > 1:
                                        rowspans[xlsxCol-1] = rowspan - 1
                                        self.xlsxWs.merge_cells(range_string='%s%s:%s%s' % (utils.get_column_letter(xlsxCol), self.xlsxRow, utils.get_column_letter(xlsxCol), self.xlsxRow + rowspan - 1))
                                    xlsxCol += 1
                            for lytMdlCell in lytMdlXBodyCell.lytMdlBodyChildren:
                                if lytMdlCell.isOpenAspectEntrySurrogate:
                                    xlsxCol += 1
                                    continue
                                justify = "left"
                                for f, v, justify in lytMdlCell.facts:
                                    break; # sets justify to first fact
                                if len(lytMdlCell.facts) == 0:
                                    v = None
                                elif len(lytMdlCell.facts) == 1:
                                    v = lytMdlCell.facts[0][1]
                                else:
                                    v = "\n".join(v for f, v, justify in lytMdlCell.facts)
                                c = self.xlsxWs.cell(row=self.xlsxRow, column=xlsxCol)
                                if v is not None:
                                    c.value = v
                                    c.alignment = Alignment(horizontal=justify, vertical="top", wrap_text=True)
                                c.border = thinBorder
                                xlsxCol += 1
                            if xlsxCol > numCols:
                                numCols = xlsxCol
                        yRowNum += 1
                    if zTbl < len(lytMdlZBodyCell.lytMdlBodyChildren) - 1:
                        zTbl += 1
                        self.xlsxRow += 1 # add blank row between z tables
                self.xlsxRow += 1 # add blank row between tables in tableset
            self.setColWidths((numCols) * [12])
            self.xlsxWs.merge_cells(range_string='%s%s:%s%s' % (utils.get_column_letter(1), titleXlsRow, utils.get_column_letter(numCols - 1), titleXlsRow))
