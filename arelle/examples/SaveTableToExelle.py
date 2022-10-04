'''
This module is an example to convert Html Tables into Xlsx (Excel) tables
Preconfigured here to use SEC Edgar Rendering R files as input

See COPYRIGHT.md for copyright information.
'''
import os, sys, re
from lxml import etree, html
from openpyxl.workbook import Workbook
from openpyxl.worksheet import ColumnDimension
from openpyxl.cell import get_column_letter
from openpyxl.style import Alignment

class Report():
    def __init__(self, longName, shortName, htmlFileName):
        self.longName = longName
        self.shortName = shortName
        self.htmlFileName = htmlFileName
    def __repr__(self):
        return ("report(longName='{}', shortName='{}', htmlFileName='{}')"
                .format(self.longName, self.shortName, self.htmlFileName))

def intCol(elt, attrName, default=None):
    try:
        return int(elt.get(attrName, default))
    except (TypeError, ValueError):
        return default

numberPattern = re.compile(r"\s*([$]\s*)?[(]?\s*[+-]?[0-9,]+([.][0-9]*)?[)-]?\s*$")
displayNonePattern = re.compile(r"\s*display:\s*none;")

def saveTableToExelle(rFilesDir):

    # get reports from FilingSummary
    reports = []
    try:
        fsdoc = etree.parse(os.path.join(rFilesDir, "FilingSummary.xml"))
        for rElt in fsdoc.iter(tag="Report"):
            reports.append(Report(rElt.findtext("LongName"),
                                  rElt.findtext("ShortName"),
                                  rElt.findtext("HtmlFileName")))
    except (EnvironmentError,
            etree.LxmlError) as err:
        print("FilingSummary.xml: directory {0} error: {1}".format(rFilesDir, err))

    wb = Workbook(encoding='utf-8')
    # remove predefined sheets
    for sheetName in wb.get_sheet_names():
        ws = wb.get_sheet_by_name(sheetName)
        if ws is not None:
            wb.remove_sheet(ws)

    sheetNames = set() # prevent duplicates

    for reportNum, report in enumerate(reports):
        sheetName = report.shortName[:31]  # max length 31 for excel title
        if sheetName in sheetNames:
            sheetName = sheetName[:31-len(str(reportNum))] + str(reportNum)
        sheetNames.add(sheetName)
        ws = wb.create_sheet(title=sheetName)

        try:
            # doesn't detect utf-8 encoding the normal way, pass it a string
            #htmlSource = ''
            #with open(os.path.join(rFilesDir, report.htmlFileName), 'rt', encoding='utf-8') as fh:
            #    htmlSource = fh.read()
            #rdoc = html.document_fromstring(htmlSource)
            rdoc = html.parse(os.path.join(rFilesDir, report.htmlFileName))
            row = -1
            mergedAreas = {}  # colNumber: (colspan,lastrow)
            for tableElt in rdoc.iter(tag="table"):
                # skip pop up tables
                if tableElt.get("class") ==  "authRefData":
                    continue
                if tableElt.getparent().tag == "div":
                    style = tableElt.getparent().get("style")
                    if style and displayNonePattern.match(style):
                        continue
                colWidths = {}
                for rowNum, trElt in enumerate(tableElt.iter(tag="tr")):
                    # remove passed mergedAreas
                    for mergeCol in [col
                                     for col, mergedArea in mergedAreas.items()
                                     if mergedArea[1] > rowNum]:
                        del mergedAreas[mergeCol]
                    col = 0
                    for coltag in ("th", "td"):
                        for cellElt in trElt.iter(tag=coltag):
                            if col == 0:
                                row += 1 # new row
                            if col in mergedAreas:
                                col += mergedAreas[col][0] - 1
                            text = cellElt.text_content()
                            colspan = intCol(cellElt, "colspan", 1)
                            rowspan = intCol(cellElt, "rowspan", 1)
                            #if col not in colWidths:
                            #    colWidths[col] = 10.0 # some kind of default width
                            for elt in cellElt.iter():
                                style = elt.get("style")
                                if style and "width:" in style:
                                    try:
                                        kw, sep, width = style.partition("width:")
                                        if "px" in width:
                                            width, sep, kw = width.partition("px")
                                            width = float(width) * 0.67777777
                                        else:
                                            width = float(width)
                                        colWidths[col] = width
                                    except ValueError:
                                        pass
                            if rowspan > 1:
                                mergedAreas[col] = (colspan, row + rowspan - 1)
                            cell = ws.cell(row=row,column=col)
                            if text:
                                cell.value = text
                                if numberPattern.match(text):
                                    cell.style.alignment.horizontal = Alignment.HORIZONTAL_RIGHT
                                else:
                                    cell.style.alignment.wrap_text = True
                            if colspan > 1 or rowspan > 1:
                                ws.merge_cells(start_row=row, end_row=row+rowspan-1, start_column=col, end_column=col+colspan-1)
                            cell.style.alignment.vertical = Alignment.VERTICAL_TOP
                            if coltag == "th":
                                cell.style.alignment.horizontal = Alignment.HORIZONTAL_CENTER
                                cell.style.font.bold = True
                            cell.style.font.size = 9  # some kind of default size
                            col += colspan
                for col, width in colWidths.items():
                    ws.column_dimensions[get_column_letter(col+1)].width = width
        except (EnvironmentError,
                etree.LxmlError) as err:
            print("{0}: directory {1} error: {2}".format(report.htmlFileName, rFilesDir, err))

    wb.save(os.path.join(rFilesDir, "exelleOut.xlsx"))

if __name__ == "__main__":

    # test directory
    saveTableToExelle(r"C:\Users\Herm Fischer\Documents\mvsl\projects\SEC\14.1\R-files\wpoRfiles")
