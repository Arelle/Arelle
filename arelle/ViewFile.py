'''
See COPYRIGHT.md for copyright information.
'''
import csv, io, json, sys
import regex as re
from lxml import etree
from decimal import Decimal
from arelle.FileSource import FileNamedStringIO

NoneType = type(None) # for isinstance testing

# deferred opening of openpyxl so it's not needed in site-packages unless it is used
Workbook = cell = utils = Font = PatternFill = Border = Alignment = Color = fills = Side = None

NOOUT = 0
CSV   = 1
XLSX  = 2
HTML  = 3
XML   = 4
JSON  = 5
TYPENAMES = ["NOOUT", "CSV", "XLSX", "HTML", "XML", "JSON"] # null means no output
nonNameCharPattern =  re.compile(r"[^\w\-\.:]")

class View:
    # note that cssExtras override any css entries provided by this module if they have the same name
    def __init__(self, modelXbrl, outfile, rootElementName, lang=None, style="table", cssExtras=""):
        global Workbook, cell, utils, Font, PatternFill, Border, Alignment, Color, fills, Side
        self.modelXbrl = modelXbrl
        self.lang = lang
        if lang and lang[:2] in {'ar', 'he'}:
            self.langDir = 'rtl'
            self.langAlign = 'right'
        else:
            self.langDir = 'ltr'
            self.langAlign = 'left'
        if outfile is None:
            self.type = NOOUT
        elif isinstance(outfile, FileNamedStringIO):
            if outfile.fileName in ("html", "xhtml"):
                self.type = HTML
            elif outfile.fileName == "csv":
                self.type = CSV
            elif outfile.fileName == "json":
                self.type = JSON
            else:
                self.type = XML
        elif outfile.endswith(".html") or outfile.endswith(".htm") or outfile.endswith(".xhtml"):
            self.type = HTML
        elif outfile.endswith(".xml"):
            self.type = XML
        elif outfile.endswith(".json"):
            self.type = JSON
        elif outfile.endswith(".xlsx"):
            self.type = XLSX
            if Workbook is None:
                from openpyxl import Workbook, cell, utils
                from openpyxl.styles import Font, PatternFill, Border, Alignment, Color, fills, Side
        else:
            self.type = CSV
        self.outfile = outfile
        if style == "rendering": # for rendering, preserve root element name
            self.rootElementName = rootElementName
        else: # root element is formed from words in title or description
            self.rootElementName = rootElementName[0].lower() + nonNameCharPattern.sub("", rootElementName.title())[1:]
        self.numHdrCols = 0
        self.treeCols = 0  # set to number of tree columns for auto-tree-columns
        if modelXbrl:
            if not lang:
                self.lang = modelXbrl.modelManager.defaultLang
        if self.type == NOOUT:
            self.xmlDoc = None
            self.tblElt = None
        elif self.type == CSV:
            if isinstance(self.outfile, FileNamedStringIO):
                self.csvFile = self.outfile
            else:
                # note: BOM signature required for Excel to open properly with characters > 0x7f
                self.csvFile = open(outfile, 'w', newline='', encoding='utf-8-sig')
            self.csvWriter = csv.writer(self.csvFile, dialect="excel")
        elif self.type == XLSX:
            self.xlsxWb = Workbook()
            # remove pre-existing worksheets
            while len(self.xlsxWb.worksheets)>0:
                self.xlsxWb.remove(self.xlsxWb.worksheets[0])
            self.xlsxWs = self.xlsxWb.create_sheet(title=rootElementName)
            self.xlsxRow = 0
            self.xlsxColWrapText = [] # bool true if col is always wrap text
        elif self.type == HTML:
            if style == "rendering":
                html = io.StringIO(
'''
<html xmlns="http://www.w3.org/1999/xhtml" dir="''' + self.langDir + '''">
    <head>
        <meta http-equiv="content-type" content="text/html;charset=utf-8" />
        <STYLE type="text/css">
            table {font-family:Arial,sans-serif;vertical-align:middle;white-space:normal;}
            th {background:#eee;}
            td {}
            .tableHdr{border-top:.5pt solid windowtext;border-right:none;border-bottom:none;border-left:.5pt solid windowtext;}
            .zAxisHdr{border-top:.5pt solid windowtext;border-right:.5pt solid windowtext;border-bottom:none;border-left:.5pt solid windowtext;}
            .xAxisSpanLeg,.yAxisSpanLeg,.yAxisSpanArm{border-top:none;border-right:none;border-bottom:none;border-left:.5pt solid windowtext;}
            .xAxisHdrValue{border-top:.5pt solid windowtext;border-right:none;border-bottom:none;border-left:1.0pt solid windowtext;}
            .xAxisHdr{border-top:.5pt solid windowtext;border-right:none;border-bottom:none;border-left:.5pt solid windowtext;}
            .yAxisHdrWithLeg{vertical-align:middle;border-top:.5pt solid windowtext;border-right:none;border-bottom:none;border-left:.5pt solid windowtext;}
            .yAxisHdrWithChildrenFirst{border-top:none;border-right:none;border-bottom:.5pt solid windowtext;border-left:.5pt solid windowtext;}
            .yAxisHdrAbstract{border-top:.5pt solid windowtext;border-right:none;border-bottom:none;border-left:.5pt solid windowtext;}
            .yAxisHdrAbstractChildrenFirst{border-top:none;border-right:none;border-bottom:.5pt solid windowtext;border-left:.5pt solid windowtext;}
            .yAxisHdr{border-top:.5pt solid windowtext;border-right:none;border-bottom:none;border-left:.5pt solid windowtext;}
            .cell{border-top:1.0pt solid windowtext;border-right:.5pt solid windowtext;border-bottom:.5pt solid windowtext;border-left:.5pt solid windowtext;}
            .abstractCell{border-top:1.0pt solid windowtext;border-right:.5pt solid windowtext;border-bottom:.5pt solid windowtext;border-left:.5pt solid windowtext;background:#e8e8e8;}
            .blockedCell{border-top:1.0pt solid windowtext;border-right:.5pt solid windowtext;border-bottom:.5pt solid windowtext;border-left:.5pt solid windowtext;background:#eee;}
            .tblCell{border-top:.5pt solid windowtext;border-right:.5pt solid windowtext;border-bottom:.5pt solid windowtext;border-left:.5pt solid windowtext;}
            ''' + cssExtras + '''
        </STYLE>
    </head>
    <body>
        <table border="1" cellspacing="0" cellpadding="4" style="font-size:8pt;">
        </table>
    </body>
</html>
'''
                )
            else:
                html = io.StringIO(
'''
<html xmlns="http://www.w3.org/1999/xhtml" dir="''' + self.langDir + '''">
    <head>
        <meta http-equiv="content-type" content="text/html;charset=utf-8" />
        <STYLE type="text/css">
            table {font-family:Arial,sans-serif;vertical-align:middle;white-space:normal;
                    border-top:.5pt solid windowtext;border-right:1.5pt solid windowtext;border-bottom:1.5pt solid windowtext;border-left:.5pt solid windowtext;}
            th {background:#eee;}
            td {}
            .tableHdr{border-top:.5pt solid windowtext;border-right:none;border-bottom:.5pt solid windowtext;border-left:.5pt solid windowtext;}
            .rowSpanLeg{width:1.0em;border-top:none;border-right:none;border-bottom:none;border-left:.5pt solid windowtext;}
            .tableCell{border-top:.5pt solid windowtext;border-right:none;border-bottom:none;border-left:.5pt solid windowtext;}
            .tblCell{border-top:.5pt solid windowtext;border-right:none;border-bottom:.5pt solid windowtext;border-left:.5pt solid windowtext;}
        </STYLE>
    </head>
    <body>
        <table cellspacing="0" cellpadding="4" style="font-size:8pt;">
        </table>
    </body>
</html>
'''
                )
            self.xmlDoc = etree.parse(html)
            html.close()
            self.tblElt = None
            for self.tblElt in self.xmlDoc.iter(tag="{http://www.w3.org/1999/xhtml}table"):
                break
        elif self.type == XML:
            html = io.StringIO("<{0}/>".format(self.rootElementName))
            self.xmlDoc = etree.parse(html)
            html.close()
            self.docEltLevels = [self.xmlDoc.getroot()]
            self.tblElt = self.docEltLevels[0]
        elif self.type == JSON:
            self.entries = []
            self.entryLevels = [self.entries]
            self.jsonObject = {self.rootElementName: self.entries}

    def setColWidths(self, colWidths):
        # widths in monospace character counts (as with xlsx files)
        if self.type == XLSX:
            for iCol, colWidth in enumerate(colWidths):
                colLetter = utils.get_column_letter(iCol+1)
                self.xlsxWs.column_dimensions[colLetter].width = colWidth

    def setColWrapText(self, colColWrapText):
        # list with True for columns to be word wrapped in every row including heading
        if self.type == XLSX:
            self.xlsxColWrapText = colColWrapText

    def addRow(self, cols, asHeader=False, treeIndent=0, colSpan=1, xmlRowElementName=None, xmlRowEltAttr=None, xmlRowText=None, xmlCol0skipElt=False, xmlColElementNames=None, lastColSpan=None, arcRole=None):
        if asHeader and len(cols) > self.numHdrCols:
            self.numHdrCols = len(cols)
        if self.type == CSV:
            self.csvWriter.writerow(cols if not self.treeCols else
                                    ([None for i in range(treeIndent)] +
                                     cols[0:1] +
                                     [None for i in range(treeIndent, self.treeCols - 1)] +
                                     cols[1:]))
        elif self.type == XLSX:
            cell = None
            for iCol, col in enumerate(cols):
                cell = self.xlsxWs.cell(row=self.xlsxRow+1, column=iCol+1)
                if asHeader:
                    cell.value = col.replace('\u00AD','') # remove soft-breaks
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.fill = PatternFill(patternType=fills.FILL_SOLID, fgColor=Color("00FFBF5F"))
                else:
                    cell.value = col if isinstance(col,(str,int,float,Decimal,NoneType)) else str(col)
                    if iCol == 0 and self.treeCols and treeIndent > 0:
                        cell.alignment = Alignment(indent=treeIndent)
                if self.xlsxColWrapText and iCol < len(self.xlsxColWrapText) and self.xlsxColWrapText[iCol]:
                    cell.alignment = Alignment(wrap_text=True)
            if lastColSpan and cell is not None:
                self.xlsxWs.merge_cells(range_string='%s%s:%s%s' % (utils.get_column_letter(iCol+1),
                                                                    self.xlsxRow+1,
                                                                    utils.get_column_letter(iCol+lastColSpan),
                                                                    self.xlsxRow+1))
            if asHeader and self.xlsxRow == 0:
                self.xlsxWs.freeze_panes = self.xlsxWs["A2"] # freezes row 1 and no columns
            self.xlsxRow += 1
        elif self.type == HTML:
            tr = etree.SubElement(self.tblElt, "{http://www.w3.org/1999/xhtml}tr")
            td = None
            for i, col in enumerate(cols + [None for emptyCol in range(self.numHdrCols - colSpan + 1 - len(cols))]):
                attrib = {}
                if asHeader:
                    attrib["class"] = "tableHdr"
                    colEltTag = "{http://www.w3.org/1999/xhtml}th"
                else:
                    colEltTag = "{http://www.w3.org/1999/xhtml}td"
                    attrib["class"] = "tableCell"
                if i == 0:
                    if self.treeCols - 1 > treeIndent:
                        attrib["colspan"] = str(self.treeCols - treeIndent + colSpan - 1)
                    elif colSpan > 1:
                        attrib["colspan"] = str(colSpan)
                if i == 0 and self.treeCols:
                    for indent in range(treeIndent):
                        etree.SubElement(tr, colEltTag,
                                         attrib={"class":"rowSpanLeg"},
                                         ).text = '\u00A0'  # produces &nbsp;
                td = etree.SubElement(tr, colEltTag, attrib=attrib)
                td.text = str(col) if col else '\u00A0'  # produces &nbsp;
            if lastColSpan and td is not None:
                td.set("colspan", str(lastColSpan))
        elif self.type == XML:
            if asHeader:
                # save column element names
                self.xmlRowElementName = xmlRowElementName or "row"
                self.columnEltNames = [col[0].lower() + nonNameCharPattern.sub('', col[1:])
                                       for col in cols]
            else:
                if treeIndent < len(self.docEltLevels) and self.docEltLevels[treeIndent] is not None:
                    parentElt = self.docEltLevels[treeIndent]
                else:
                    # problem, error message? unexpected indent
                    parentElt = self.docEltLevels[0]
                # escape attributes content
                escapedRowEltAttr = dict(((k, v.replace("&","&amp;").replace("<","&lt;"))
                                          for k,v in xmlRowEltAttr.items())
                                         if xmlRowEltAttr else ())
                rowElt = etree.SubElement(parentElt, xmlRowElementName or self.xmlRowElementName, attrib=escapedRowEltAttr)
                if treeIndent + 1 >= len(self.docEltLevels): # extend levels as needed
                    for extraColIndex in range(len(self.docEltLevels) - 1, treeIndent + 1):
                        self.docEltLevels.append(None)
                self.docEltLevels[treeIndent + 1] = rowElt
                if not xmlColElementNames: xmlColElementNames = self.columnEltNames
                if len(cols) == 1 and not xmlCol0skipElt:
                    rowElt.text = xmlRowText if xmlRowText else cols[0]
                else:
                    isDimensionName = isDimensionValue = False
                    elementName = "element" # need a default
                    for i, col in enumerate(cols):
                        if (i != 0 or not xmlCol0skipElt) and col:
                            if i < len(xmlColElementNames):
                                elementName = xmlColElementNames[i]
                                if elementName == "dimensions":
                                    elementName = "dimension" # one element per dimension
                                    isDimensionName = True
                            if isDimensionName:
                                isDimensionValue = True
                                isDimensionName = False
                                dimensionName = str(col)
                            else:
                                elt = etree.SubElement(rowElt, elementName)
                                elt.text = str(col).replace("&","&amp;").replace("<","&lt;")
                                if isDimensionValue:
                                    elt.set("name", dimensionName)
                                    isDimensionName = True
                                    isDimensionValue = False
        elif self.type == JSON:
            if asHeader:
                # save column element names
                self.xmlRowElementName = xmlRowElementName
                self.columnEltNames = [col[0].lower() + nonNameCharPattern.sub('', col[1:])
                                       for col in cols]
            else:
                if treeIndent < len(self.entryLevels) and self.entryLevels[treeIndent] is not None:
                    entries = self.entryLevels[treeIndent]
                else:
                    # problem, error message? unexpected indent
                    entries = self.entryLevels[0]
                entry = []
                if xmlRowElementName:
                    entry.append(xmlRowElementName)
                elif self.xmlRowElementName:
                    entry.append(self.xmlRowElementName)
                if xmlRowEltAttr:
                    entry.append(xmlRowEltAttr)
                else:
                    entry.append({})
                entries.append(entry)
                if treeIndent + 1 >= len(self.entryLevels): # extend levels as needed
                    for extraColIndex in range(len(self.entryLevels) - 1, treeIndent + 1):
                        self.entryLevels.append(None)
                self.entryLevels[treeIndent + 1] = entry
                if not xmlColElementNames: xmlColElementNames = self.columnEltNames
                if len(cols) == 1 and not xmlCol0skipElt:
                    entry.append(xmlRowText if xmlRowText else cols[0])
                else:
                    content = {}
                    entry.append(content)
                    for i, col in enumerate(cols):
                        if (i != 0 or not xmlCol0skipElt) and col and i < len(xmlColElementNames):
                                elementName = xmlColElementNames[i]
                                if elementName == "dimensions":
                                    value = dict((str(cols[i]),str(cols[i+1])) for i in range(i, len(cols), 2))
                                else:
                                    value = str(col)
                                content[elementName] = value
        if asHeader and lastColSpan:
            self.numHdrCols += lastColSpan - 1

    def close(self, noWrite=False):
        if self.type == CSV:
            if not isinstance(self.outfile, FileNamedStringIO):
                self.csvFile.close()
        elif self.type == XLSX:
            # add filtering
            self.xlsxWs.auto_filter.ref = 'A1:{}{}'.format(utils.get_column_letter(self.xlsxWs.max_column), len(self.xlsxWs['A']))
            self.xlsxWb.save(self.outfile)
        elif self.type != NOOUT and not noWrite:
            fileType = TYPENAMES[self.type]
            try:
                from arelle import XmlUtil
                if isinstance(self.outfile, FileNamedStringIO):
                    fh = self.outfile
                else:
                    fh = open(self.outfile, "w", encoding="utf-8")
                if self.type == JSON:
                    fh.write(json.dumps(self.jsonObject, ensure_ascii=False))
                else:
                    XmlUtil.writexml(fh, self.xmlDoc, encoding="utf-8",
                                     xmlcharrefreplace= (self.type == HTML) )
                if not isinstance(self.outfile, FileNamedStringIO):
                    fh.close()
                self.modelXbrl.info("info", _("Saved output %(type)s to %(file)s"), file=self.outfile, type=fileType)
            except (IOError, EnvironmentError) as err:
                self.modelXbrl.exception("arelle:htmlIOError", _("Failed to save output %(type)s to %(file)s: %(error)s"), file=self.outfile, type=fileType, error=err)
        self.modelXbrl = None
        if self.type == HTML:
            self.tblElt = None
        elif self.type == XML:
            self.docEltLevels = None

        self.__dict__.clear() # dereference everything after closing document
