# -*- coding: utf-8 -*-
'''
loadFromExcel.py is an example of a plug-in that will load an extension taxonomy from Excel
input and optionally save an (extension) DTS.

See COPYRIGHT.md for copyright information.
'''
import os, io, time
import regex as re
from arelle import XbrlConst
from arelle.Version import authorLabel, copyrightLabel
from collections import defaultdict

headerWidths = {
        "label": 40,
        "prefix": 20,
        "name": 36,
        "type": 24,
        "abstract": 7,
        "nillable": 6,
        "substitutionGroup": 30,
        "periodType": 9,
        "balance": 8,
        "depth": 5,
        "preferredLabel": 16,
        "calculationParent": 20,
        "calculationWeight": 12
        }

headersStyles = (
    # ELR match string, font (size: 20 * pt size)
    (r"http://[^/]+\.edinet-fsa\.go\.jp", {"name":"MS UI Gothic", "size":12.0}, (
        # (display string, content, label role, lang [,indented])
        ("標準ラベル（日本語）", "label", XbrlConst.standardLabel, "ja", "indented"),
        ("冗長ラベル（日本語）", "label", XbrlConst.verboseLabel, "ja"),
        ("標準ラベル（英語）", "label", XbrlConst.standardLabel, "en"),
        ("冗長ラベル（英語）", "label", XbrlConst.verboseLabel, "en"),
        ("名前空間プレフィックス", "prefix"),
        ("要素名", "name"),
        ("type", "type"),
        ("substitutionGroup", "substitutionGroup"),
        ("periodType", "periodType"),
        ("balance", "balance"),
        ("abstract", "abstract"),
        ("nillable", "nillable"),
        ("depth", "depth"),
        ("preferred label", "preferredLabel"),
        ("calculation parent", "calculationParent"), # qname
        ("calculation weight", "calculationWeight")
        )),
    (r"http://[^/]+/us-gaap/", {"name":"Calibri", "size":10.0}, (
        ("label", "label", XbrlConst.standardLabel, "en-US", "indented"),
        ("label, standard", "label", XbrlConst.standardLabel, "en-US", "overridePreferred"),
        ("label, terse", "label", XbrlConst.terseLabel, "en-US"),
        ("label, verbose", "label", XbrlConst.verboseLabel, "en-US"),
        ("prefix", "prefix"),
        ("name", "name"),
        ("type", "type"),
        ("substitutionGroup", "substitutionGroup"),
        ("periodType", "periodType"),
        ("balance", "balance"),
        ("abstract", "abstract"),
        ("nillable", "nillable"),
        ("depth", "depth"),
        ("preferred label", "preferredLabel"),
        ("calculation parent", "calculationParent"), # qname
        ("calculation weight", "calculationWeight"),
        )),
    # generic pattern taxonomy
    (r"http://[^/]+/us-gaap/", {"name":"Calibri", "size":10.0}, (
        ("label", "label", XbrlConst.standardLabel, "en", "indented"),
        ("prefix", "prefix"),
        ("name", "name"),
        ("type", "type"),
        ("substitutionGroup", "substitutionGroup"),
        ("periodType", "periodType"),
        ("balance", "balance"),
        ("abstract", "abstract"),
        ("nillable", "nillable"),
        ("depth", "depth"),
        ("preferred label", "preferredLabel"),
        ("calculation parent", "calculationParent"), # qname
        ("calculation weight", "calculationWeight"),
        )),
    )

MAXINDENT = 10

dtsWsHeaders = (
    ("specification", 14),
    ("file type", 12),
    ("prefix (schema)\ntype (linkbase)\nargument (other)", 20),
    ("file, href or role definition", 60),
    ("namespace URI", 60),
    )

def saveLoadableExcel(dts, excelFile):
    from arelle import ModelDocument, XmlUtil
    from openpyxl import Workbook, cell
    from openpyxl.styles import Font, PatternFill, Border, Alignment, Color, fills, Side
    from openpyxl.worksheet.dimensions import ColumnDimension
    from openpyxl.utils import get_column_letter

    workbook = Workbook()
    # remove pre-existing worksheets
    while len(workbook.worksheets)>0:
        workbook.remove_sheet(workbook.worksheets[0])
    conceptsWs = workbook.create_sheet(title="Concepts")
    dtsWs = workbook.create_sheet(title="DTS")

    # identify type of taxonomy
    conceptsWsHeaders = None
    cellFontArgs = None
    for doc in dts.urlDocs.values():
        if doc.type == ModelDocument.Type.SCHEMA and doc.inDTS:
            for i in range(len(headersStyles)):
                if re.match(headersStyles[i][0], doc.targetNamespace):
                    cellFontArgs = headersStyles[i][1] # use as arguments to Font()
                    conceptsWsHeaders = headersStyles[i][2]
                    break
    if conceptsWsHeaders is None:
        dts.info("error:saveLoadableExcel",
         _("Referenced taxonomy style not identified, assuming general pattern."),
         modelObject=dts)
        cellFontArgs = headersStyles[-1][1] # use as arguments to Font()
        conceptsWsHeaders = headersStyles[-1][2]


    hdrCellFont = Font(**cellFontArgs)
    hdrCellFill = PatternFill(patternType=fills.FILL_SOLID,
                              fgColor=Color("00FFBF5F")) # Excel's light orange fill color = 00FF990
    cellFont = Font(**cellFontArgs)

    def writeCell(ws,row,col,value,fontBold=False,borders=True,indent=0,hAlign=None,vAlign=None,hdr=False):
        cell = ws.cell(row=row,column=col)
        cell.value = value
        if hdr:
            cell.font = hdrCellFont
            cell.fill = hdrCellFill
            if not hAlign: hAlign = "center"
            if not vAlign: vAlign = "center"
        else:
            cell.font = cellFont
            if not hAlign: hAlign = "left"
            if not vAlign: vAlign = "top"
        if borders:
            cell.border = Border(top=Side(border_style="thin"),
                                 left=Side(border_style="thin"),
                                 right=Side(border_style="thin"),
                                 bottom=Side(border_style="thin"))
        cell.alignment = Alignment(horizontal=hAlign, vertical=vAlign, wrap_text=True, indent=indent)

    # sheet 1 col widths
    for i, hdr in enumerate(conceptsWsHeaders):
        colLetter = get_column_letter(i+1)
        conceptsWs.column_dimensions[colLetter] = ColumnDimension(conceptsWs, customWidth=True)
        conceptsWs.column_dimensions[colLetter].width = headerWidths.get(hdr[1], 40)

    # sheet 2 headers
    for i, hdr in enumerate(dtsWsHeaders):
        colLetter = get_column_letter(i+1)
        dtsWs.column_dimensions[colLetter] = ColumnDimension(conceptsWs, customWidth=True)
        dtsWs.column_dimensions[colLetter].width = hdr[1]
        writeCell(dtsWs, 1, i+1, hdr[0], hdr=True)

    # referenced taxonomies
    conceptsRow = 1
    dtsRow = 3
    # identify extension schema
    extensionSchemaDoc = None
    if dts.modelDocument.type == ModelDocument.Type.SCHEMA:
        extensionSchemaDoc = dts.modelDocument
    elif dts.modelDocument.type == ModelDocument.Type.INSTANCE:
        for doc, docReference in dts.modelDocument.referencesDocument.items():
            if "href" in docReference.referenceTypes:
                extensionSchemaDoc = doc
                break
    if extensionSchemaDoc is None:
        dts.info("error:saveLoadableExcel",
         _("Unable to identify extension taxonomy."),
         modelObject=dts)
        return

    for doc, docReference in extensionSchemaDoc.referencesDocument.items():
        if "import" in docReference.referenceTypes and doc.targetNamespace != XbrlConst.xbrli:
            writeCell(dtsWs, dtsRow, 1, "import")
            writeCell(dtsWs, dtsRow, 2, "schema")
            writeCell(dtsWs, dtsRow, 3, XmlUtil.xmlnsprefix(doc.xmlRootElement, doc.targetNamespace))
            writeCell(dtsWs, dtsRow, 4, doc.uri)
            writeCell(dtsWs, dtsRow, 5, doc.targetNamespace)
            dtsRow += 1

    dtsRow += 1

    doc = extensionSchemaDoc
    writeCell(dtsWs, dtsRow, 1, "extension")
    writeCell(dtsWs, dtsRow, 2, "schema")
    writeCell(dtsWs, dtsRow, 3, XmlUtil.xmlnsprefix(doc.xmlRootElement, doc.targetNamespace))
    writeCell(dtsWs, dtsRow, 4, os.path.basename(doc.uri))
    writeCell(dtsWs, dtsRow, 5, doc.targetNamespace)
    dtsRow += 1

    for doc, docReference in extensionSchemaDoc.referencesDocument.items():
        if "href" in docReference.referenceTypes and doc.type == ModelDocument.Type.LINKBASE:
            linkbaseType = ""
            role = docReference.referringModelObject.get("{http://www.w3.org/1999/xlink}role") or ""
            if role.startswith("http://www.xbrl.org/2003/role/") and role.endswith("LinkbaseRef"):
                linkbaseType = os.path.basename(role)[0:-11]
            writeCell(dtsWs, dtsRow, 1, "extension")
            writeCell(dtsWs, dtsRow, 2, "linkbase")
            writeCell(dtsWs, dtsRow, 3, linkbaseType)
            writeCell(dtsWs, dtsRow, 4, os.path.basename(doc.uri))
            writeCell(dtsWs, dtsRow, 5, "")
            dtsRow += 1

    dtsRow += 1

    # extended link roles defined in this document
    for roleURI, roleTypes in sorted(dts.roleTypes.items(),
            # sort on definition if any else URI
            key=lambda item: (item[1][0].definition if len(item[1]) and item[1][0].definition else item[0])):
        for roleType in roleTypes:
            if roleType.modelDocument == extensionSchemaDoc:
                writeCell(dtsWs, dtsRow, 1, "extension")
                writeCell(dtsWs, dtsRow, 2, "role")
                writeCell(dtsWs, dtsRow, 3, "")
                writeCell(dtsWs, dtsRow, 4, roleType.definition)
                writeCell(dtsWs, dtsRow, 5, roleURI)
                dtsRow += 1

    # tree walk recursive function
    def treeWalk(row, depth, concept, preferredLabel, arcrole, preRelSet, visited):
        if concept is not None:
            # calc parents
            calcRelSet = dts.relationshipSet(XbrlConst.summationItem, preRelSet.linkrole)
            calcRel = None
            for modelRel in calcRelSet.toModelObject(concept):
                calcRel = modelRel
                break
            for i, hdr in enumerate(conceptsWsHeaders):
                colType = hdr[1]
                value = ""
                if colType == "name":
                    value = str(concept.name)
                elif colType == "prefix" and concept.qname is not None:
                    value = concept.qname.prefix
                elif colType == "type" and concept.type is not None:
                    value = str(concept.type.qname)
                elif colType == "substitutionGroup":
                    value = str(concept.substitutionGroupQname)
                elif colType == "abstract":
                    value = "true" if concept.isAbstract else "false"
                elif colType == "nillable":
                    if concept.isNillable:
                        value = "true"
                elif colType == "periodType":
                    value = concept.periodType
                elif colType == "balance":
                    value = concept.balance
                elif colType == "label":
                    role = hdr[2]
                    lang = hdr[3]
                    if role == XbrlConst.standardLabel:
                        if "indented" in hdr:
                            roleUri = preferredLabel
                        elif "overridePreferred" in hdr:
                            if preferredLabel and preferredLabel != XbrlConst.standardLabel:
                                roleUri = role
                            else:
                                roleUri = "**no value**" # skip putting a value in this column
                        else:
                            roleUri = role
                    else:
                        roleUri = role
                    if roleUri != "**no value**":
                        value = concept.label(roleUri,
                                              linkroleHint=preRelSet.linkrole,
                                              lang=lang,
                                              fallbackToQname=(role == XbrlConst.standardLabel))
                elif colType == "preferredLabel" and preferredLabel:
                    if preferredLabel.startswith("http://www.xbrl.org/2003/role/"):
                        value = os.path.basename(preferredLabel)
                    else:
                        value = preferredLabel
                elif colType == "calculationParent" and calcRel is not None:
                    calcParent = calcRel.fromModelObject
                    if calcParent is not None:
                        value = str(calcParent.qname)
                elif colType == "calculationWeight" and calcRel is not None:
                    value = calcRel.weight
                elif colType == "depth":
                    value = depth
                if "indented" in hdr:
                    indent = min(depth, MAXINDENT)
                else:
                    indent = 0
                writeCell(conceptsWs, row, i+1, value, indent=indent)
            row += 1
            if concept not in visited:
                visited.add(concept)
                for modelRel in preRelSet.fromModelObject(concept):
                    if modelRel.toModelObject is not None:
                        row = treeWalk(row, depth + 1, modelRel.toModelObject, modelRel.preferredLabel, arcrole, preRelSet, visited)
                visited.remove(concept)
        return row

    # use presentation relationships for conceptsWs
    arcrole = XbrlConst.parentChild
    # sort URIs by definition
    linkroleUris = []
    relationshipSet = dts.relationshipSet(arcrole)
    if relationshipSet:
        for linkroleUri in relationshipSet.linkRoleUris:
            modelRoleTypes = dts.roleTypes.get(linkroleUri)
            if modelRoleTypes:
                roledefinition = (modelRoleTypes[0].genLabel(strip=True) or modelRoleTypes[0].definition or linkroleUri)
            else:
                roledefinition = linkroleUri
            linkroleUris.append((roledefinition, linkroleUri))
        linkroleUris.sort()

        # for each URI in definition order
        for roledefinition, linkroleUri in linkroleUris:
            # write linkrole
            writeCell(conceptsWs, conceptsRow, 1, (roledefinition or linkroleUri), borders=False)  # ELR has no boarders, just font specified
            conceptsRow += 1
            # write header row
            for i, hdr in enumerate(conceptsWsHeaders):
                writeCell(conceptsWs, conceptsRow, i+1, hdr[0], hdr=True)
            conceptsRow += 1
            # elr relationships for tree walk
            linkRelationshipSet = dts.relationshipSet(arcrole, linkroleUri)
            for rootConcept in linkRelationshipSet.rootConcepts:
                conceptsRow = treeWalk(conceptsRow, 0, rootConcept, None, arcrole, linkRelationshipSet, set())
            conceptsRow += 1 # double space rows between tables
    else:
        # write header row
        for i, hdr in enumerate(conceptsWsHeaders):
            writeCell(conceptsWs, conceptsRow, i, hdr[0], hdr=True)
        conceptsRow += 1
        # get lang
        lang = None
        for i, hdr in enumerate(conceptsWsHeaders):
            colType = hdr[1]
            if colType == "label":
                lang = hdr[3]
                if colType == "label":
                    role = hdr[2]
                    lang = hdr[3]
        lbls = defaultdict(list)
        for concept in set(dts.qnameConcepts.values()): # may be twice if unqualified, with and without namespace
            lbls[concept.label(role,lang=lang)].append(concept.objectId())
        srtLbls = sorted(lbls.keys())
        excludedNamespaces = XbrlConst.ixbrlAll.union(
            (XbrlConst.xbrli, XbrlConst.link, XbrlConst.xlink, XbrlConst.xl,
             XbrlConst.xbrldt,
             XbrlConst.xhtml))
        for label in srtLbls:
            for objectId in lbls[label]:
                concept = dts.modelObject(objectId)
                if concept.modelDocument.targetNamespace not in excludedNamespaces:
                    for i, hdr in enumerate(conceptsWsHeaders):
                        colType = hdr[1]
                        value = ""
                        if colType == "name":
                            value = str(concept.qname.localName)
                        elif colType == "prefix":
                            value = concept.qname.prefix
                        elif colType == "type":
                            value = str(concept.type.qname)
                        elif colType == "substitutionGroup":
                            value = str(concept.substitutionGroupQname)
                        elif colType == "abstract":
                            value = "true" if concept.isAbstract else "false"
                        elif colType == "periodType":
                            value = concept.periodType
                        elif colType == "balance":
                            value = concept.balance
                        elif colType == "label":
                            role = hdr[2]
                            lang = hdr[3]
                            value = concept.label(role, lang=lang)
                        elif colType == "depth":
                            value = 0
                        if "indented" in hdr:
                            indent = min(0, MAXINDENT)
                        else:
                            indent = 0
                        writeCell(conceptsWs, conceptsRow, i, value, indent=indent)
                    conceptsRow += 1

    try:
        excelFilename = excelFile if excelFile.lower().endswith(".xlsx") else excelFile + ".xlsx"
        workbook.save(excelFilename)
        dts.info("info:saveLoadableExcel",
            _("Saved Excel file: %(excelFile)s"),
            excelFile=os.path.basename(excelFilename),
            modelXbrl=dts)
    except Exception as ex:
        dts.error("exception:saveLoadableExcel",
            _("File saving exception: %(error)s"), error=ex,
            modelXbrl=dts)

def saveLoadableExcelMenuEntender(cntlr, menu, *args, **kwargs):
    # Extend menu with an item for the savedts plugin
    menu.add_command(label="Save Loadable Excel",
                     underline=0,
                     command=lambda: saveLoadableExcelMenuCommand(cntlr) )

def saveLoadableExcelMenuCommand(cntlr):
    # save DTS menu item has been invoked
    if cntlr.modelManager is None or cntlr.modelManager.modelXbrl is None:
        cntlr.addToLog("No taxonomy loaded.")
        return
        # get file name into which to save log file while in foreground thread
    excelFile = cntlr.uiFileDialog("save",
            title=_("arelle - Save Loadable Excel file"),
            initialdir=cntlr.config.setdefault("loadableExcelFileDir","."),
            filetypes=[(_("Excel file .xlsx"), "*.xlsx")],
            defaultextension=".xlsx")
    if not excelFile:
        return False
    import os
    cntlr.config["loadableExcelFileDir"] = os.path.dirname(excelFile)
    cntlr.saveConfig()

    import threading
    thread = threading.Thread(target=lambda
                                  _dts=cntlr.modelManager.modelXbrl,
                                  _excelFile=excelFile:
                                        saveLoadableExcel(_dts, _excelFile))
    thread.daemon = True
    thread.start()

def saveLoadableExcelCommandLineOptionExtender(parser, *args, **kwargs):
    # extend command line options with a save DTS option
    parser.add_option("--save-loadable-excel",
                      dest="saveLoadableExcel",
                      help=_("Save Loadable Excel file"))

def saveLoadableExcelCommandLineXbrlRun(cntlr, options, modelXbrl, *args, **kwargs):
    # extend XBRL-loaded run processing for this option
    excelFile = getattr(options, "saveLoadableExcel", None)
    if excelFile:
        if cntlr.modelManager is None or cntlr.modelManager.modelXbrl is None:
            cntlr.addToLog("No taxonomy loaded.")
            return
        saveLoadableExcel(cntlr.modelManager.modelXbrl, excelFile)

__pluginInfo__ = {
    'name': 'Save Loadable Excel',
    'version': '0.9',
    'description': "This plug-in saves XBRL in Excel that can be loaded as an extension DTS.",
    'license': 'Apache-2',
    'author': authorLabel,
    'copyright': copyrightLabel,
    # classes of mount points (required)
    'CntlrWinMain.Menu.Tools': saveLoadableExcelMenuEntender,
    'CntlrCmdLine.Options': saveLoadableExcelCommandLineOptionExtender,
    'CntlrCmdLine.Xbrl.Run': saveLoadableExcelCommandLineXbrlRun,
}
