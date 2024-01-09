"""
See COPYRIGHT.md for copyright information.

## Overview

The Load From OIM plugin is designed to load reports in Arelle from JSON and CSV that adhere to the Open Information
Model (OIM) XBRL Specification. It also offers the option to save a loaded report as an xBRL-XML instance. It is
designed to work seamlessly with the Save Loadable OIM plugin, allowing for efficient data handling in Arelle.

## Key Features

- **Multiple Formats**: Enables loading data from JSON and CSV OIM formats as well as XLSX.
- **Seamless Integration**: Compatible with the Save Loadable OIM plugin for saving and loading reports.
- **GUI and CLI Compatibility**: Available for use in both GUI and CLI modes.
- **Save xBRL-XML Instance**: Optionally save the data as an xBRL-XML instance.

## Usage Instructions

### Command Line Usage

- **Load OIM Report**:
  To load an OIM report, specify the file path to the JSON, CSV, or XLSX file:
  ```bash
  python arelleCmdLine.py --plugins loadFromOIM --file filing-document.json
  ```

- **Save xBRL-XML Instance**:
  Use the `--saveOIMinstance` argument to save an xBRL-XML instance from an OIM report:
  ```bash
  python arelleCmdLine.py --plugins loadFromOIM --file filing-document.json --saveOIMinstance example.xbrl
  ```

### GUI Usage

* **Load OIM Report**:
  1. Using the normal `File` menu `Open File...` dialog, select the CSV, JSON, or XLSX file.
  2. Provide a name for the XBRL-XML instance to save.
"""
import os, sys, io, time, traceback, json, csv, logging, zipfile, datetime, isodate
from math import isnan, log10
from regex import compile as re_compile, match as re_match, sub as re_sub, DOTALL as re_DOTALL
from lxml import etree
from collections import defaultdict, OrderedDict
from arelle.ModelDocument import Type, create as createModelDocument
from arelle.ModelDtsObject import ModelResource
from arelle import XbrlConst, ModelDocument, ModelXbrl, PackageManager, ValidateXbrlDimensions
from arelle.ModelObject import ModelObject
from arelle.PluginManager import pluginClassMethods
from arelle.ModelValue import qname, dateTime, DateTime, DATETIME, yearMonthDuration, dayTimeDuration
from arelle.PrototypeInstanceObject import DimValuePrototype
from arelle.PythonUtil import attrdict, flattenToSet, strTruncate
from arelle.UrlUtil import isHttpUrl, isAbsolute as isAbsoluteUri, isValidUriReference
from arelle.ValidateDuplicateFacts import DuplicateTypeArg, getDuplicateFactSetsWithType
from arelle.Version import authorLabel, copyrightLabel
from arelle.XbrlConst import (xbrli, qnLinkLabel, standardLabelRoles, qnLinkReference, standardReferenceRoles,
                              qnLinkPart, gen, link, defaultLinkRole, footnote, factFootnote, isStandardRole,
                              conceptLabel, elementLabel, conceptReference, all as hc_all, notAll as hc_notAll,
                              xhtml, qnXbrliDateItemType,
                              dtrPrefixedContentItemTypes, dtrPrefixedContentTypes, dtrSQNameNamesItemTypes, dtrSQNameNamesTypes,
                              lrrRoleHrefs, lrrArcroleHrefs)
from arelle.XmlUtil import addChild, addQnameValue, copyIxFootnoteHtml, setXmlns
from arelle.XmlValidateConst import VALID
from arelle.XmlValidate import integerPattern, languagePattern, NCNamePattern, QNamePattern, validate as xmlValidate
from arelle.ValidateXbrlCalcs import inferredDecimals, rangeValue

nsOims = ("https://xbrl.org/2021",
          "http://www.xbrl.org/WGWD/YYYY-MM-DD",
          "https://www.xbrl.org/WGWD/YYYY-MM-DD",
          "http://www.xbrl.org/((~status_date_uri~))",
          "https://xbrl.org/((~status_date_uri~))"
         )
nsOimCes = ("https://xbrl.org/2021/oim-common/error",
            "http://www.xbrl.org/WGWD/YYYY-MM-DD/oim-common/error",
            "http://www.xbrl.org/CR/2020-05-06/oim-common/error",
            "http://www.xbrl.org/((~status_date_uri~))/oim-common/error",
            "https://xbrl.org/((~status_date_uri~))/oim-common/error"
    )
jsonDocumentTypes = (
        "https://xbrl.org/2021/xbrl-json",
        "http://www.xbrl.org/WGWD/YYYY-MM-DD/xbrl-json",
        "http://www.xbrl.org/YYYY-MM-DD/xbrl-json",
        "https://xbrl.org/((~status_date_uri~))/xbrl-json" # allows loading of XII "template" test cases without CI production
    )
csvDocumentTypes = (
        "https://xbrl.org/2021/xbrl-csv",
        "http://www.xbrl.org/WGWD/YYYY-MM-DD/xbrl-csv",
        "http://xbrl.org/YYYY/xbrl-csv",
        "https://xbrl.org/((~status_date_uri~))/xbrl-csv" # allows loading of XII "template" test cases without CI production
    )
csvDocinfoObjects = {"documentType", "namespaces", "taxonomy", "extends", "final", "linkTypes", "linkGroups"}
csvExtensibleObjects = {"namespaces", "linkTypes", "linkGroups", "features", "final", "tableTemplates", "tables", "dimensions", "parameters"}


reservedLinkTypesAndGroups = {
        "footnote":         "http://www.xbrl.org/2003/arcrole/fact-footnote",
        "explanatoryFact":  "http://www.xbrl.org/2009/arcrole/fact-explanatoryFact",
        "_":     "http://www.xbrl.org/2003/role/link"
    }
reservedLinkTypeAndGroupAliases = {
        "http://www.xbrl.org/2003/arcrole/fact-footnote": "footnote",
        "http://www.xbrl.org/2009/arcrole/fact-explanatoryFact": "explanatoryFact",
        "http://www.xbrl.org/2003/role/link": "_"
    }


XLINKTYPE = "{http://www.w3.org/1999/xlink}type"
XLINKLABEL = "{http://www.w3.org/1999/xlink}label"
XLINKARCROLE = "{http://www.w3.org/1999/xlink}arcrole"
XLINKFROM = "{http://www.w3.org/1999/xlink}from"
XLINKTO = "{http://www.w3.org/1999/xlink}to"
XLINKHREF = "{http://www.w3.org/1999/xlink}href"
XMLLANG = "{http://www.w3.org/XML/1998/namespace}lang"

JSONdocumentType = "http://www.xbrl.org/WGWD/YYYY-MM-DD/xbrl-json"

NSReservedAliasURIs = {
    "xbrl": nsOims,
    "xs": (XbrlConst.xsd,),
    "enum2": XbrlConst.enum2s,
    # "oimce": nsOimCes,
    "xbrli": (XbrlConst.xbrli,),
    "xs": (XbrlConst.xsd,),
    "utr": (XbrlConst.utr,),
    "iso4217": (XbrlConst.iso4217,),
    #"xbrle":  [ns + "/error" for ns in nsOims],
    #"xbrlxe": [ns + "/xbrl-xml/error" for ns in nsOims]
    }
JSONNSReservedAliasURIs = {
    # xbrlje no longer reserved, issue #381
    # "xbrlje": [ns + "/xbrl-json/error" for ns in nsOims],
    }
CSVNSReservedAliasURIs = {
    "xbrlce": [ns + "/xbrl-csv/error" for ns in nsOims],
    }
JSONNSReservedURIAliases = {} # #381 no longer reserved - dict((ns + "/xbrl-json/error", "xbrlje") for ns in nsOims)
CSVNSReservedURIAliases = dict((ns + "/xbrl-csv/error", "xbrlce") for ns in nsOims)
NSReservedAliasURIPrefixes = { # for starts-with checking
    # "dtr-type": "http://www.xbrl.org/dtr/type/",
    }
NSReservedURIAlias = {}

OIMDefaultContextElement = "scenario"
OIMReservedAliasURIs = {
    # "namespaces": NSReservedAliasURIs,  -- generated at load time
    "linkTypes": reservedLinkTypesAndGroups,
    "linkGroups": reservedLinkTypesAndGroups
    }
OIMReservedURIAlias = {
    #"namespaces": NSReservedURIAlias, -- generated at load time
    "linkTypes": reservedLinkTypeAndGroupAliases,
    "linkGroups": reservedLinkTypeAndGroupAliases
    }

EMPTY_DICT = {}
EMPTY_LIST = []

DUPJSONKEY = "!@%duplicateKeys%@!"
DUPJSONVALUE = "!@%duplicateValues%@!"

UTF_7_16_Pattern = re_compile(r"(?P<utf16>(^([\x00][^\x00])+$)|(^([^\x00][\x00])+$))|(?P<utf7>^\s*\+AHs-)")
UTF_7_16_Bytes_Pattern = re_compile(br"(?P<utf16>(^([\x00][^\x00])+$)|(^([^\x00][\x00])+$))|(?P<utf7>^\s*\+AHs-)")
EBCDIC_Bytes_Pattern = re_compile(b"^[\x40\x4a-\x4f\x50\x5a-\x5f\x60-\x61\x6a-\x6f\x79-\x7f\x81-\x89\x8f\x91-\x99\xa1-\xa9\xb0\xba-\xbb\xc1-\xc9\xd1-\xd9\xe0\xe2-\xe9\xf0-\xf9\xff\x0a\x0d]+$")
NEVER_EBCDIC_Bytes_Pattern = re_compile(b"[\x30-\x31\x3e\x41-\x49\x51-\x59\x62-\x69\x70-\x78\x80\x8a-\x8e\x90\x9a-\x9f\xa0\xaa-\xaf\xb1-\xb9\xbc-\xbf\xca-\xcf\xda-\xdf\xe1\xea-\xef\xfa-\xfe]")
JSONmetadataPattern = re_compile(r"\s*\{.*\"documentInfo\"\s*:.*\}", re_DOTALL)
NoCanonicalPattern = attrdict(match=lambda s: True)
CanonicalFloatPattern = re_compile(r"^-?[0-9]\.[0-9]([0-9]*[1-9])?E-?([1-9][0-9]*|0)$|^-?INF$|^NaN$")
CanonicalIntegerPattern = re_compile(r"^-?([1-9][0-9]*)?[0-9]$")
CanonicalXmlTypePattern = {
    "boolean": re_compile("^true$|^false$"),
    "date": re_compile(r"-?[0-9]{4}-[0-9]{2}-[0-9]{2}Z?$"),
    "dateTime": re_compile(r"-?[0-9]{4}-[0-9]{2}-[0-9]{2}T([01][0-9]|20|21|22|23):[0-9]{2}:[0-9]{2}(\.[0-9]([0-9]*[1-9])?)?Z?$"),
    "XBRLI_DATEUNION": re_compile(r"-?[0-9]{4}-[0-9]{2}-[0-9]{2}Z?$|-?[0-9]{4}-[0-9]{2}-[0-9]{2}T([01][0-9]|20|21|22|23):[0-9]{2}:[0-9]{2}(\.[0-9]([0-9]*[1-9])?)?Z?$"),
    "time": re_compile(r"-?([01][0-9]|20|21|22|23):[0-9]{2}:[0-9]{2}(\.[0-9]([0-9]*[1-9])?)?Z?$"),
    "decimal": re_compile(r"^[-]?([1-9][0-9]*)?[0-9]\.[0-9]([0-9]*[1-9])?$"),
    "float": CanonicalFloatPattern,
    "double": CanonicalFloatPattern,
    "hexBinary": re_compile(r"^([0-9A-F][0-9A-F])*$"),
    "integer": CanonicalIntegerPattern,
    "language": re_compile(r"[a-z]{1,8}(-[a-z0-9]{1,8})*$"),
    "nonPositiveInteger": CanonicalIntegerPattern,
    "negativeInteger": CanonicalIntegerPattern,
    "long": CanonicalIntegerPattern,
    "int": CanonicalIntegerPattern,
    "short": CanonicalIntegerPattern,
    "byte": CanonicalIntegerPattern,
    "nonNegativeInteger": CanonicalIntegerPattern,
    "unsignedLong": CanonicalIntegerPattern,
    "unsignedInt": CanonicalIntegerPattern,
    "unsignedShort": CanonicalIntegerPattern,
    "unsignedByte": CanonicalIntegerPattern,
    "positiveInteger": CanonicalIntegerPattern,
    }
IdentifierPattern = re_compile(
    "^[_A-Za-z\xC0-\xD6\xD8-\xF6\xF8-\xFF\u0100-\u02FF\u0370-\u037D\u037F-\u1FFF\u200C-\u200D\u2070-\u218F\u2C00-\u2FEF\u3001-\uD7FF\uF900-\uFDCF\uFDF0-\uFFFD]"
     r"[_\-"
     "\xB7A-Za-z0-9\xC0-\xD6\xD8-\xF6\xF8-\xFF\u0100-\u02FF\u0370-\u037D\u037F-\u1FFF\u200C-\u200D\u2070-\u218F\u2C00-\u2FEF\u3001-\uD7FF\uF900-\uFDCF\uFDF0-\uFFFD\u0300-\u036F\u203F-\u2040]*$")
RowIdentifierPattern = re_compile(
     r"[_\-"
     "\xB7A-Za-z0-9\xC0-\xD6\xD8-\xF6\xF8-\xFF\u0100-\u02FF\u0370-\u037D\u037F-\u1FFF\u200C-\u200D\u2070-\u218F\u2C00-\u2FEF\u3001-\uD7FF\uF900-\uFDCF\uFDF0-\uFFFD\u0300-\u036F\u203F-\u2040]*$")
PeriodPattern = re_compile(
    r"^-?[0-9]{4}-[0-9]{2}-[0-9]{2}T([01][0-9]|20|21|22|23):[0-9]{2}:[0-9]{2}(\.[0-9]([0-9]*[1-9])?)?Z?"
    r"(/-?[0-9]{4}-[0-9]{2}-[0-9]{2}T([01][0-9]|20|21|22|23):[0-9]{2}:[0-9]{2}(\.[0-9]([0-9]*[1-9])?)?Z?)?$"
    )
PrefixedQName = re_compile(
    "[_A-Za-z\xC0-\xD6\xD8-\xF6\xF8-\xFF\u0100-\u02FF\u0370-\u037D\u037F-\u1FFF\u200C-\u200D\u2070-\u218F\u2C00-\u2FEF\u3001-\uD7FF\uF900-\uFDCF\uFDF0-\uFFFD]"
     r"[_\-\."
     "\xB7A-Za-z0-9\xC0-\xD6\xD8-\xF6\xF8-\xFF\u0100-\u02FF\u0370-\u037D\u037F-\u1FFF\u200C-\u200D\u2070-\u218F\u2C00-\u2FEF\u3001-\uD7FF\uF900-\uFDCF\uFDF0-\uFFFD\u0300-\u036F\u203F-\u2040]*:"
    "[_A-Za-z\xC0-\xD6\xD8-\xF6\xF8-\xFF\u0100-\u02FF\u0370-\u037D\u037F-\u1FFF\u200C-\u200D\u2070-\u218F\u2C00-\u2FEF\u3001-\uD7FF\uF900-\uFDCF\uFDF0-\uFFFD]"
     r"[_\-\."
     "\xB7A-Za-z0-9\xC0-\xD6\xD8-\xF6\xF8-\xFF\u0100-\u02FF\u0370-\u037D\u037F-\u1FFF\u200C-\u200D\u2070-\u218F\u2C00-\u2FEF\u3001-\uD7FF\uF900-\uFDCF\uFDF0-\uFFFD\u0300-\u036F\u203F-\u2040]*")
SpecialValuePattern = re_compile("##|#empty$|#nil$|#none$")
SQNamePattern = re_compile(
    "[_A-Za-z\xC0-\xD6\xD8-\xF6\xF8-\xFF\u0100-\u02FF\u0370-\u037D\u037F-\u1FFF\u200C-\u200D\u2070-\u218F\u2C00-\u2FEF\u3001-\uD7FF\uF900-\uFDCF\uFDF0-\uFFFD]"
     r"[_\-\."
     "\xB7A-Za-z0-9\xC0-\xD6\xD8-\xF6\xF8-\xFF\u0100-\u02FF\u0370-\u037D\u037F-\u1FFF\u200C-\u200D\u2070-\u218F\u2C00-\u2FEF\u3001-\uD7FF\uF900-\uFDCF\uFDF0-\uFFFD\u0300-\u036F\u203F-\u2040]*:"
    r"\S+")
UnitPrefixedQNameSubstitutionChar = "\x07" # replaces PrefixedQName in unit pattern
UnitPattern = re_compile(
    # QNames are replaced by \x07 in these expressions
    # numerator only (no parentheses)
    "(^\x07$)|(^\x07([*]\x07)+$)|"
    # numerator and optional denominator, with parentheses if more than one term in either
    "(^((\x07)|([(]\x07([*]\x07)+[)]))([/]((\x07)|([(]\x07([*]\x07)+[)])))?$)"
    )
UrlInvalidPattern = re_compile(
    r"^[ \t\n\r]+[^ \t\n\r]*|.*[^ \t\n\r][ \t\n\r]+$|" # leading or trailing whitespace
    r".*[^ \t\n\r]([\t\n\r]+|[ \t\n\r]{2,})[^ \t\n\r]|" # embedded uncollapsed whitespace
    r".*%[^0-9a-fA-F]|.*%[0-9a-fA-F][^0-9a-fA-F]|.*#.*#" # invalid %nn or two ##s
    )
WhitespacePattern = re_compile(r"[ \t\n\r]")
WhitespaceUntrimmedPattern = re_compile(r"^[ \t\n\r]|.*[ \t\n\r]$")

xlUnicodePattern = re_compile("_x([0-9A-F]{4})_")

precisionZeroPattern = re_compile(r"^\s*0+\s*$")
decimalsSuffixPattern = re_compile(r".*[0-9.][\r\n\t ]*d[\r\n\t ]*(0|-?[1-9][0-9]*|INF)[\r\n\t ]*$") # test starting 1 position before the d

RegexPatternType = type(decimalsSuffixPattern)

htmlBodyTemplate = "<body xmlns='http://www.w3.org/1999/xhtml'>\n{0}\n</body>\n"
xhtmlTagPrefix = "{http://www.w3.org/1999/xhtml}"
DimensionsKeyPattern = re_compile(r"^(concept|entity|period|unit|language|(\w+:\w+))$")

nonDiscoveringXmlInstanceElements = {qname(link, "roleRef"), qname(link, "arcroleRef")}

UNSUPPORTED_DATA_TYPES = dtrPrefixedContentItemTypes + (
    qname(xbrli,"fractionItemType"), )

# CSV Files
CSV_PARAMETER_FILE = 1
CSV_FACTS_FILE = 2
CSV_HAS_HEADER_ROW = True

# allowed duplicates settings
NONE = 1
COMPLETE = 2
CONSISTENT = 3
ALL = 4
AllowedDuplicatesFeatureValues = {"none": NONE, "complete": COMPLETE, "consistent": CONSISTENT, "all": ALL}
DisallowedDescription = {NONE: "Disallowed", COMPLETE: "Non-complete", CONSISTENT: "Inconsistent", ALL: "Allowed"}
DuplicateTypeArgMap = {NONE: DuplicateTypeArg.ALL, COMPLETE: DuplicateTypeArg.INCOMPLETE, CONSISTENT: DuplicateTypeArg.INCONSISTENT, ALL: DuplicateTypeArg.NONE}

class SQNameType:
    pass # fake class for detecting SQName type in JSON structure check

class QNameType:
    pass # fake class for detecting QName type in JSON structure check

class LangType:
    pass

class URIType:
    pass

class IdentifierType:
    pass

class NoRecursionCheck:
    pass

class CheckPrefix:
    pass

class KeyIsNcName:
    pass


UnrecognizedDocMemberTypes = {
    "/documentInfo": dict,
    "/documentInfo/documentType": str,
    }
UnrecognizedDocRequiredMembers = {
    "/": {"documentInfo"},
    "/documentInfo/": {"documentType","taxonomy"},
    }

JsonMemberTypes = {
    # keys are json pointer with * meaning any id,  and *:* meaning any SQName or QName, for array no index is used
    # report
    "/documentInfo": dict,
    "/facts": dict,
    "/*:*": (int,bool,str,dict,list,type(None),NoRecursionCheck,CheckPrefix), # custom extensions
    # documentInfo
    "/documentInfo/baseURL": URIType,
    "/documentInfo/documentType": str,
    "/documentInfo/features": dict,
    "/documentInfo/features/*:*": (int,float,bool,str,type(None)),
    "/documentInfo/namespaces": dict,
    "/documentInfo/namespaces/*": URIType,
    "/documentInfo/linkTypes": dict,
    "/documentInfo/linkTypes/*": str,
    "/documentInfo/linkGroups": dict,
    "/documentInfo/linkGroups/*": str,
    "/documentInfo/taxonomy": list,
    "/documentInfo/taxonomy/": str,
    "/documentInfo/*:*": (int,float,bool,str,dict,list,type(None),NoRecursionCheck,CheckPrefix), # custom extensions
    # facts
    "/facts/*": dict,
    "/facts/*/value": (str,type(None)),
    "/facts/*/decimals": int,
    "/facts/*/dimensions": dict,
    "/facts/*/links": dict,
    "/facts/*/links/*": dict,
    "/facts/*/links/*/*": list,
    "/facts/*/links/*/*/": str,
    # dimensions
    "/facts/*/dimensions/concept": QNameType,
    "/facts/*/dimensions/entity": SQNameType,
    "/facts/*/dimensions/period": str,
    "/facts/*/dimensions/unit": str,
    "/facts/*/dimensions/language": LangType,
    "/facts/*/dimensions/noteId": str,
    "/facts/*/dimensions/*:*": (str,type(None)),
    # custom properties on fact are unchecked
    "/facts/*/*:*": (int,float,bool,str,dict,list,type(None),NoRecursionCheck,CheckPrefix), # custom extensions
    }
JsonRequiredMembers = {
    "/": {"documentInfo"},
    "/documentInfo/": {"documentType","taxonomy"},
    "/facts/*/": {"value","dimensions"},
    "/facts/*/dimensions/": {"concept"}
    }

CsvMemberTypes = {
    # report
    "/documentInfo": dict,
    "/tableTemplates": dict,
    "/tables": dict,
    "/parameters": dict,
    "/parameters/*": str,
    "/parameterURL": str,
    "/dimensions": dict,
    "/decimals": (int,str),
    "/links": dict,
    "/*:*": (int,float,bool,str,dict,list,type(None),NoRecursionCheck,CheckPrefix), # custom extensions
    # documentInfo
    "/documentInfo/baseURL": URIType,
    "/documentInfo/documentType": str,
    "/documentInfo/features": dict,
    "/documentInfo/features/*:*": (int,float,bool,str,type(None)),
    "/documentInfo/final": dict,
    "/documentInfo/namespaces": dict,
    "/documentInfo/namespaces/*": URIType,
    "/documentInfo/linkTypes": dict,
    "/documentInfo/linkTypes/*": str,
    "/documentInfo/linkGroups": dict,
    "/documentInfo/linkGroups/*": str,
    "/documentInfo/taxonomy": list,
    "/documentInfo/taxonomy/": str,
    "/documentInfo/extends": list,
    "/documentInfo/extends/": URIType,
    "/documentInfo/*:*": (int,float,bool,str,dict,list,type(None),NoRecursionCheck,CheckPrefix), # custom extensions
    # documentInfo/final
    "/documentInfo/final/namespaces": bool,
    "/documentInfo/final/taxonomy": bool,
    "/documentInfo/final/linkTypes": bool,
    "/documentInfo/final/linkGroups": bool,
    "/documentInfo/final/features": bool,
    "/documentInfo/final/tableTemplates": bool,
    "/documentInfo/final/tables": bool,
    "/documentInfo/final/dimensions": bool,
    "/documentInfo/final/final": bool,
    "/documentInfo/final/parameters": bool,
    "/documentInfo/final/parameterURL": bool,
    # table templates
    "/tableTemplates/*": dict,
    "/tableTemplates/*/rowIdColumn": str,
    "/tableTemplates/*/columns": dict,
    "/tableTemplates/*/decimals": (int,str),
    "/tableTemplates/*/dimensions": dict,
    "/tableTemplates/*:*": (int,float,bool,str,dict,list,type(None),NoRecursionCheck,CheckPrefix), # custom extensions
    "/tableTemplates/*/dimensions/concept": str,
    "/tableTemplates/*/dimensions/entity": str,
    "/tableTemplates/*/dimensions/period": str,
    "/tableTemplates/*/dimensions/unit": str,
    "/tableTemplates/*/dimensions/language": str,
    "/tableTemplates/*/dimensions/*:*": str,
    "/tableTemplates/*/dimensions/$*": str,
    #"/tableTemplates/*/transposed": bool,
    # columns
    "/tableTemplates/*/columns/*": dict,
    "/tableTemplates/*/columns/*/comment": bool,
    "/tableTemplates/*/columns/*/decimals": (int,str),
    "/tableTemplates/*/columns/*/dimensions": dict,
    "/tableTemplates/*/columns/*/*:*": (int,float,bool,str,dict,list,type(None),NoRecursionCheck,CheckPrefix), # custom extensions
    # dimensions (column)
    "/tableTemplates/*/columns/*/dimensions/concept": str,
    "/tableTemplates/*/columns/*/dimensions/entity": str,
    "/tableTemplates/*/columns/*/dimensions/period": str,
    "/tableTemplates/*/columns/*/dimensions/unit": str,
    "/tableTemplates/*/columns/*/dimensions/language": str,
    "/tableTemplates/*/columns/*/dimensions/*:*": str,
    "/tableTemplates/*/columns/*/dimensions/$*": str,
    # property groups (column)
    "/tableTemplates/*/columns/*/propertiesFrom": list,
    "/tableTemplates/*/columns/*/propertiesFrom/": str,
    "/tableTemplates/*/columns/*/propertyGroups": dict,
    "/tableTemplates/*/columns/*/propertyGroups/*": dict,
    "/tableTemplates/*/columns/*/propertyGroups/*/decimals": (int,str),
    "/tableTemplates/*/columns/*/propertyGroups/*/dimensions": dict,
    "/tableTemplates/*/columns/*/propertyGroups/*/dimensions/concept": str,
    "/tableTemplates/*/columns/*/propertyGroups/*/dimensions/entity": str,
    "/tableTemplates/*/columns/*/propertyGroups/*/dimensions/period": str,
    "/tableTemplates/*/columns/*/propertyGroups/*/dimensions/unit": str,
    "/tableTemplates/*/columns/*/propertyGroups/*/dimensions/language": str,
    "/tableTemplates/*/columns/*/propertyGroups/*/dimensions/*:*": str,
    "/tableTemplates/*/columns/*/propertyGroups/*/dimensions/$*": str,
    # dimensions (top level)
    "/dimensions/concept": str,
    "/dimensions/entity": str,
    "/dimensions/period": str,
    "/dimensions/unit": str,
    "/dimensions/language": str,
    "/dimensions/noteId": str,
    "/dimensions/*:*": str,
    "/dimensions/$*": str,
    # tables
    "/tables/*": dict,
    "/tables/*/url": str,
    "/tables/*/template": str,
    "/tables/*/optional": bool,
    "/tables/*/parameters": dict,
    "/tables/*/parameters/*": str,
    "/tables/*/*:*": (int,float,bool,str,dict,list,type(None),NoRecursionCheck,CheckPrefix), # custom extensions
    # links
    "/links/*": (dict,KeyIsNcName),
    # link group
    "/links/*/*": (dict,KeyIsNcName),
    # fact links
    "/links/*/*/*": list,
    # fact IDs
    "/links/*/*/*/*": str,
    }
CsvRequiredMembers = {
    "/": {"documentInfo"},
    "/documentInfo/": {"documentType"},
    "/tableTemplates/*/": {"columns"},
    "/tables/*/": {"url"}
    }
EMPTY_SET = set()

def jsonGet(tbl, key, default=None):
    if isinstance(tbl, dict):
        return tbl.get(key, default)
    return default

# singleton special values
class Singleton(str):
    def __init__(self, value):
        self.value = value
    def __str__(self):
        return self.value

EMPTY_CELL = Singleton("")
NONE_CELL = Singleton("")
INVALID_REFERENCE_TARGET = Singleton("")

def csvCellValue(cellValue):
    # CSV table in Appendix A
    if cellValue == "#nil": # nil value
        return None
    elif cellValue == "": # empty cell
        return EMPTY_CELL
    elif cellValue == "#none":
        return NONE_CELL
    elif cellValue == "#empty": # empty string
        return ""
    elif isinstance(cellValue, str) and cellValue.startswith("#"):
        if cellValue.startswith("##"):
            return cellValue[1:]
        else:
            raise OIMException("xbrlce:unknownSpecialValue",
                               _("Unknown special value %(specialValue)s"),
                               specialValue=cellValue)
    else:
        return cellValue

def xlUnicodeChar(match):
    return chr(int(match.group(1), 16))

def xlValue(cell): # excel values may have encoded unicode, such as _0000D_
    v = cell.value
    if isinstance(v, str):
        v = xlUnicodePattern.sub(xlUnicodeChar, v).replace('\r\n','\n').replace('\r','\n')
    elif v is None:
        v = ""
    elif isinstance(v, float):
        return str(round(v, 14)) # Deal with general numbers which may be imprecise
    else:
        v = str(v)
    return csvCellValue(v)

def parseMetadataCellValues(metadataTable):
    for dimName in metadataTable.keys():
        dimValue = metadataTable[dimName]
        # CSV table in Appendix A (similar to "cellValue"
        if dimValue is None or dimValue == "#nil":
            metadataTable[dimName] = None
        elif dimValue == "" and dimName != "period": # empty cell except for period
            metadataTable[dimName] = EMPTY_CELL
        elif dimValue == "#none":
            metadataTable[dimName] = NONE_CELL
        elif isinstance(dimValue, str) and dimValue.startswith("##"):
            metadataTable[dimName] = dimValue[1:]

def xlTrimHeaderRow(row):
    numEmptyCellsAtEndOfRow = 0
    for i in range(len(row)-1, -1, -1):
        if row[i] in (None, ""):
            numEmptyCellsAtEndOfRow += 1
        else:
            break
    if numEmptyCellsAtEndOfRow:
        return row[:-numEmptyCellsAtEndOfRow]
    return row

class OIMException(Exception):
    def __init__(self, code=None, message=None, **kwargs):
        self.code = code
        self.message = message
        self.msgArgs = kwargs
        self.args = ( self.__repr__(), )
    def __repr__(self):
        if self.code and self.message:
            return _('[{0}] exception {1}').format(self.code, self.message % self.msgArgs)
        else:
            return "Errors noted in log"

class NotOIMException(Exception):
    def __init__(self, **kwargs):
        self.args = ( self.__repr__(), )
    def __repr__(self):
        return _('[NotOIM] not an OIM document')

class FactProduced():
    def clear(self):
        self.modelFact = None
        self.dimensionsUsed = set()
        self.invalidReferenceTarget = None

PER_ISO = 0
PER_INCLUSIVE_DATES = 1
PER_SINGLE_DAY = 2
PER_MONTH = 3
PER_YEAR = 4
PER_QTR = 5
PER_HALF = 6
PER_WEEK = 7
ONE_DAY = dayTimeDuration("P1D")
ONE_MONTH = yearMonthDuration("P1M")
ONE_YEAR = yearMonthDuration("P1Y")
ONE_QTR = yearMonthDuration("P3M")
ONE_HALF = yearMonthDuration("P6M")

periodForms = ((PER_ISO, re_compile("([0-9]{4}-[0-9]{2}-[0-9]{2}T[0-9]{2}:[0-9]{2}:[0-9]{2}(Z|[+-][0-2][0-9]([:]?)[0-5][0-9]+)?(/[0-9]{4}-[0-9]{2}-[0-9]{2}T[0-9]{2}:[0-9]{2}:[0-9]{2})?(Z|[+-][0-2][0-9]([:]?)[0-5][0-9]+)?)$")),
               (PER_INCLUSIVE_DATES, re_compile("([0-9]{4}-[0-9]{2}-[0-9]{2})[.][.]([0-9]{4}-[0-9]{2}-[0-9]{2})$")),
               (PER_SINGLE_DAY, re_compile("([0-9]{4}-[0-9]{2}-[0-9]{2})(@(start|end))?$")),
               (PER_MONTH,  re_compile("([0-9]{4}-[0-9]{2})(@(start|end))?$")),
               (PER_YEAR, re_compile("([0-9]{4})(@(start|end))?$")),
               (PER_QTR, re_compile("([0-9]{4})Q([1-4])(@(start|end))?$")),
               (PER_HALF, re_compile("([0-9]{4})H([1-2])(@(start|end))?$")),
               (PER_WEEK, re_compile("([0-9]{4}W[1-5]?[0-9])(@(start|end))?$")))

def csvPeriod(cellValue, startOrEnd=None):
    if cellValue is EMPTY_CELL or cellValue is NONE_CELL:
        return NONE_CELL # Forever period (absent in xBRL-JSON)
    if cellValue is None: # #nil is not valid for date
        return cellValue  # stays None
    isoDuration = None
    for perType, perFormMatch in periodForms:
        m = perFormMatch.match(cellValue)
        if m:
            try:
                if perType == PER_ISO:
                    if not m.group(4) and startOrEnd: # instant date
                        return "referenceTargetNotDuration"
                    isoDuration = cellValue
                    startendSuffixGroup = 0
                elif perType == PER_INCLUSIVE_DATES:
                    isoDuration = "{}/{}".format(dateTime(m.group(1)), dateTime(m.group(2)) + ONE_DAY)
                    startendSuffixGroup = 0
                elif perType == PER_SINGLE_DAY:
                    isoDuration = "{}/{}".format(dateTime(m.group(1)), dateTime(m.group(1)) + ONE_DAY)
                    startendSuffixGroup = 3
                elif perType == PER_MONTH:
                    moStart = dateTime(m.group(1) + "-01")
                    isoDuration = "{}/{}".format(moStart, moStart + ONE_MONTH)
                    startendSuffixGroup = 3
                elif perType == PER_YEAR:
                    yrStart = dateTime(m.group(1) + "-01-01")
                    isoDuration = "{}/{}".format(yrStart, yrStart + ONE_YEAR)
                    startendSuffixGroup = 3
                elif perType == PER_QTR:
                    qtrStart = dateTime(m.group(1) + "-{:02}-01".format(int(m.group(2))*3 - 2))
                    isoDuration = "{}/{}".format(qtrStart, qtrStart + ONE_QTR)
                    startendSuffixGroup = 4
                elif perType == PER_HALF:
                    qtrStart = dateTime(m.group(1) + "-{:02}-01".format(int(m.group(2))*6 - 5))
                    isoDuration = "{}/{}".format(qtrStart, qtrStart + ONE_HALF)
                    startendSuffixGroup = 4
                elif perType == PER_WEEK:
                    weekStart = dateTime(isodate.parse_date(m.group(1)))
                    isoDuration = "{}T00:00:00/{}T00:00:00".format(weekStart, weekStart + datetime.timedelta(7))
                    startendSuffixGroup = 3
                if startendSuffixGroup and m.group(startendSuffixGroup):
                    if startOrEnd:
                        # period specifier is being applied to an instant date
                        return "referenceTargetNotDuration"
                    startOrEnd = m.group(startendSuffixGroup)
            except ValueError:
                return None
        if isoDuration:
            if startOrEnd == "start":
                return isoDuration.partition("/")[0]
            elif startOrEnd == "end":
                return isoDuration.partition("/")[2]
            return isoDuration
    return None

# no longer used because transpose is not supported
def transposer(rowIterator, default=""):
    cells = [row for row in rowIterator]
    if cells:
        colsCount = max(len(row) for row in cells)
        rowsCount = len(cells)
        for colIndex in range(colsCount):
            yield [(cells[rowIndex][colIndex] if colIndex < len(cells[colIndex]) else default)
                   for rowIndex in range(rowsCount)]

def idDeduped(modelXbrl, id):
    for i in range(99999):
        if i == 0:
            candidateId = id
        else:
            candidateId = "{}.{}".format(id, i)
        if candidateId not in modelXbrl.modelDocument.idObjects:
            return candidateId
    return None

def oimEquivalentFacts(f1, f2):
        if f1.context is None or f1.concept is None:
            return False # need valid context and concept for v-Equality of nonTuple
        if f1.isNil:
            return f2.isNil
        if f2.isNil:
            return False
        if not f1.context.isEqualTo(f2.context):
            return False
        elif type(f1.xValue) == type(f2.xValue):
            if f1.concept.isLanguage and f2.concept.isLanguage and f1.xValue is not None and f2.xValue is not None:
                return f1.xValue.lower() == f2.xValue.lower() # required to handle case insensitivity
            if isinstance(f1.xValue, DateTime): # with/without time makes values unequal
                return f1.xValue.dateOnly == f2.xValue.dateOnly and f1.xValue == f2.xValue
            return f1.xValue == f2.xValue # required to handle date/time with 24 hrs.
        return f1.value == f2.value


def checkForDuplicates(modelXbrl, allowedDups, footnoteIDs):
    duplicateTypeArg = DuplicateTypeArgMap[allowedDups]
    for duplicateFactSet in getDuplicateFactSetsWithType(modelXbrl.facts, duplicateTypeArg.duplicateType()):
        fList = duplicateFactSet.facts
        f0 = fList[0]
        modelXbrl.error("oime:disallowedDuplicateFacts",
                        "%(disallowance)s duplicate fact values %(element)s: %(values)s, %(contextIDs)s.",
                        modelObject=fList, disallowance=DisallowedDescription[allowedDups], element=f0.qname,
                        contextIDs=", ".join(sorted(set(f.contextID for f in fList))),
                        values=", ".join(strTruncate(f.value,64) for f in fList))


def getTaxonomyContextElement(modelXbrl):
    # https://www.xbrl.org/Specification/xbrl-xml/REC-2021-10-13/xbrl-xml-REC-2021-10-13.html#sec-dimensions
    # The spec states that if in the DTS:
    # 1. neither segment nor scenario is present, scenario is used.
    # 2. segment is present and scenario is not, segment is used.
    # 3. scenario is present and segment is not, scenario is used.
    # 4. segment and scenario are present and facts are valid against both of them, scenario is used.
    # 5. segment and scenario are present and facts are only valid against scenario, scenario is used.
    # 6. segment and scenario are present and facts are only valid against segment, segment is used.
    # 7. segment and scenario are present and facts are invalid against both, the choice is made arbitrarily.
    # We don't yet inspect dimensional validity and therefore incorrectly use scenario in case 6.
    taxonomyContextRefTypes = {
        modelRelationship.contextElement
        for hasHypercubeRelationship in (XbrlConst.all, XbrlConst.notAll)
        for modelRelationship in modelXbrl.relationshipSet(hasHypercubeRelationship).modelRelationships
    }
    return taxonomyContextRefTypes.pop() if len(taxonomyContextRefTypes) == 1 else OIMDefaultContextElement

def loadFromOIM(cntlr, error, warning, modelXbrl, oimFile, mappedUri):
    from openpyxl import load_workbook
    from openpyxl.cell import Cell
    from arelle import ModelDocument, ModelXbrl, XmlUtil
    from arelle.ModelDocument import ModelDocumentReference
    from arelle.ModelValue import qname

    _return = None # modelDocument or an exception

    try:
        currentAction = "initializing"
        startingErrorCount = len(modelXbrl.errors) if modelXbrl else 0
        startedAt = time.time()

        currentAction = "determining file type"
        isJSON = False
        # isCSV means metadata loaded from separate JSON file (but instance data can be in excel or CSV)
        isCSV = False # oimFile.endswith(".csv") # this option is not currently supported
        instanceFileName = os.path.splitext(oimFile)[0] + ".xbrl"

        currentAction = "loading and parsing OIM file"
        loadDictErrors = []
        def openCsvReader(csvFilePath, fileType):
            _file = modelXbrl.fileSource.file(csvFilePath, binary=True)[0]
            bytes = _file.read(16) # test encoding
            try:
                m = EBCDIC_Bytes_Pattern.match(bytes)
                if m and not NEVER_EBCDIC_Bytes_Pattern.findall(bytes):
                    raise OIMException("xbrlce:invalidCSVFileFormat",
                          _("CSV file MUST use utf-8 encoding: %(file)s, appears to be EBCDIC"),
                          file=csvFilePath)
                m = UTF_7_16_Bytes_Pattern.match(bytes)
                if m:
                    raise OIMException("xbrlce:invalidCSVFileFormat",
                          _("CSV file MUST use utf-8 encoding: %(file)s, appears to be %(encoding)s"),
                          file=csvFilePath, encoding=m.lastgroup)
                _file.close()
            except UnicodeDecodeError as ex:
                raise OIMException("xbrlce:invalidCSVFileFormat",
                      _("CSV file MUST use utf-8 encoding: %(file)s, appears to be %(encoding)s"),
                      file=csvFilePath, encoding=m.lastgroup)
            _file = modelXbrl.fileSource.file(csvFilePath, encoding='utf-8-sig')[0]
            if CSV_HAS_HEADER_ROW:
                try:
                    chars = _file.read(1024)
                    _dialect = csv.Sniffer().sniff(chars, delimiters=[',', '\t', ';', '|']) # also check for disallowed potential separators
                    if _dialect.lineterminator not in ("\r", "\n", "\r\n"):
                        raise OIMException("xbrlce:invalidCSVFileFormat",
                                           _("CSV line ending is not CR, LF or CR LF, file %(file)s"),
                                          file=csvFilePath)
                    if _dialect.delimiter not in (","):
                        raise OIMException({CSV_PARAMETER_FILE: "xbrlce:invalidParameterCSVFile",
                                            CSV_FACTS_FILE: "xbrlce:invalidHeaderValue"}[fileType],
                                           _("CSV deliminator %(deliminator)s is not comma: file %(file)s"),
                                          file=csvFilePath, deliminator=repr(_dialect.delimiter))
                except csv.Error as ex:
                    # possibly can't br sniffed because there's only one column in the rows
                    _dialect = None
                    for char in chars:
                        if char in  (",", "\n", "\r"):
                            _dialect = "excel"
                            break
                        elif char == "\t":
                            _dialect = "excel-tab"
                            break
                    if not _dialect:
                        raise OIMException("xbrlce:invalidCSVFileFormat",
                                           _("CSV file %(file)s: %(error)s"),
                                          file=csvFilePath, error=str(ex))
                except UnicodeDecodeError as ex:
                    raise OIMException("xbrlce:invalidCSVFileFormat",
                                       _("CSV file must use utf-8 encoding %(file)s: %(error)s"),
                                      file=csvFilePath, error=str(ex))
                _file.seek(0)
            else:
                # check for comma or tab in first line
                _dialect = "excel" # fallback if no first line tab is determinable
                for char in _file.read(1024):
                    if char in (",", "\n", "\r", ";", "|"): # ;, | force invalid parameter file detection
                        _dialect = "excel"
                        break
                    elif char == "\t": # only way to sniff first row deliminator if value contains SQName semicolon
                        _dialect = "excel-tab"
                        break
                _file.seek(0)
            return csv.reader(_file, _dialect)

        def ldError(msgCode, msgText, **kwargs):
            loadDictErrors.append((msgCode, msgText, kwargs))
        def loadDict(keyValuePairs):
            _dict = OrderedDict() # preserve fact order in resulting instance
            _valueKeyDict = {}
            for key, value in keyValuePairs:
                if isinstance(value, dict):
                    if key in ("namespaces", "linkTypes", "linkGroups"):
                        normalizedDict = OrderedDict()
                        normalizedValueKeyDict = {}
                        if DUPJSONKEY in value:
                            normalizedDict[DUPJSONKEY] = value[DUPJSONKEY]
                        if DUPJSONVALUE in value:
                            normalizedDict[DUPJSONVALUE] = value[DUPJSONVALUE]
                        for _key, _value in value.items():
                            if not isinstance(_value, str):
                                continue # skip dup key/value entries
                            # _key = _key.strip() # per !178 keys have only normalized values, don't normalize key
                            # _value = _value.strip()
                            if _key in normalizedDict: # don't put the duplicate in the dictionary but report it as error
                                if DUPJSONKEY not in normalizedDict:
                                    normalizedDict[DUPJSONKEY] = []
                                normalizedDict[DUPJSONKEY].append((_key, _value, normalizedDict[_key]))
                            else: # do put into dictionary, only report if it's a map object
                                normalizedDict[_key] = _value
                                if _value in normalizedValueKeyDict:
                                    if DUPJSONVALUE not in normalizedDict:
                                        normalizedDict[DUPJSONVALUE] = []
                                    normalizedDict[DUPJSONVALUE].append((_value, _key, normalizedValueKeyDict[_value]))
                                else:
                                    normalizedValueKeyDict[_value] = _key
                            if not NCNamePattern.match(_key):
                                ldError("{}:invalidJSONStructure",
                                      _("The %(map)s alias \"%(alias)s\" must be a canonical NCName value"),
                                      modelObject=modelXbrl, map=key, alias=_key)
                            if UrlInvalidPattern.match(_value):
                                ldError("{}:invalidJSONStructure",
                                      _("The %(map)s alias \"%(alias)s\" URI must be a canonical URI value: \"%(URI)s\"."),
                                      modelObject=modelXbrl, map=key, alias=_key, URI=_value)
                            elif not (_value and isAbsoluteUri(_value)) or UrlInvalidPattern.match(_value):
                                ldError("oimce:invalidURI",
                                        _("The %(map)s \"%(alias)s\" URI is invalid: \"%(URI)s\"."),
                                        modelObject=modelXbrl, map=key, alias=_key, URI=_value)
                        value.clear() # replace with normalized values
                        for _key, _value in normalizedDict.items():
                            value[_key] = _value
                    if DUPJSONKEY in value:
                        for _errKey, _errValue, _otherValue in value[DUPJSONKEY]:
                            if key in ("namespaces", "linkTypes", "linkGroups"):
                                ldError("{}:invalidJSON", # {} expanded when loadDictErrors are processed
                                                _("The %(map)s alias \"%(prefix)s\" is used on uri \"%(uri1)s\" and uri \"\"%(uri2)s."),
                                                modelObject=modelXbrl, map=key, prefix=_errKey, uri1=_errValue, uri2=_otherValue)
                            else:
                                ldError("{}:invalidJSON", # {} expanded when loadDictErrors are processed
                                                _("The %(obj)s key \"%(key)s\" is used on multiple objects."),
                                                modelObject=modelXbrl, obj=key, key=_errKey)
                        del value[DUPJSONKEY]
                    if DUPJSONVALUE in value:
                        if key in ("namespaces", "linkTypes", "linkGroups"):
                            for _errValue, _errKey, _otherKey in value[DUPJSONVALUE]:
                                ldError("oimce:multipleAliasesForURI",
                                                _("The \"%(map)s\" value \"%(uri)s\" is used on alias \"%(alias1)s\" and alias \"%(alias2)s\"."),
                                                modelObject=modelXbrl, map=key, uri=_errValue, alias1=_errKey, alias2=_otherKey)
                        del value[DUPJSONVALUE]
                if key in _dict: # don't put the duplicate in the dictionary but report it as error
                    if DUPJSONKEY not in _dict:
                        _dict[DUPJSONKEY] = []
                    _dict[DUPJSONKEY].append((key, value, _dict[key]))
                else: # do put into dictionary, only report if it's a map object
                    _dict[key] = value
                    if isinstance(value, str):
                        if value in _valueKeyDict:
                            if DUPJSONVALUE not in _dict:
                                _dict[DUPJSONVALUE] = []
                            _dict[DUPJSONVALUE].append((value, key, _valueKeyDict[value]))
                        else:
                            _valueKeyDict[value] = key
            return _dict

        primaryOimFile = oimFile
        extensionProperties = OrderedDict() # key is property QName, value is property path

        def loadOimObject(oimFile, extendingFile, visitedFiles, extensionChain, primaryReportParameters=None): # returns oimObject, oimWb
            # isXL means metadata loaded from Excel (but instance data can be in excel or CSV)
            isXL = oimFile.endswith(".xlsx") or oimFile.endswith(".xls")
            # same logic as modelDocument.load
            normalizedUrl = modelXbrl.modelManager.cntlr.webCache.normalizeUrl(oimFile, extendingFile)
            if modelXbrl.fileSource.isMappedUrl(normalizedUrl):
                mappedUrl = modelXbrl.fileSource.mappedUrl(normalizedUrl)
            elif PackageManager.isMappedUrl(normalizedUrl):
                mappedUrl = PackageManager.mappedUrl(normalizedUrl)
            else:
                mappedUrl = modelXbrl.modelManager.disclosureSystem.mappedUrl(normalizedUrl)
            if modelXbrl.fileSource.isInArchive(mappedUrl):
                filepath = mappedUrl
            else:
                filepath = modelXbrl.modelManager.cntlr.webCache.getfilename(mappedUrl) # , reload=reloadCache, checkModifiedTime=kwargs.get("checkModifiedTime",False))
                if filepath:
                    url = modelXbrl.modelManager.cntlr.webCache.normalizeUrl(filepath)
            if filepath and filepath.endswith(".csv") or ("metadata" in filepath and filepath.endswith(".json")):
                errPrefix = "xbrlce"
            else:
                errPrefix = "xbrlje"
            # prevent recursion
            if filepath in extensionChain:
                raise OIMException("{}:cycleInExtensionChain".format(errPrefix),
                      _("File MUST NOT extend itself: %(file)s cycles to %(extendingFile)s"),
                      file=filepath, extendingFile=extendingFile)
            elif filepath in visitedFiles:
                return None
            visitedFiles.add(filepath)
            extensionChain.add(filepath)
            if not isXL:
                try:
                    _file = modelXbrl.fileSource.file(filepath, encoding="utf-8-sig")[0]
                    with _file as f:
                        chars = f.read(16) # test encoding
                        m = UTF_7_16_Pattern.match(chars)
                        if m:
                            raise OIMException("{}:invalidJSON".format(errPrefix),
                                  _("File MUST use utf-8 encoding: %(file)s, appears to be %(encoding)s"),
                                  file=filepath, encoding=m.lastgroup)
                        else:
                            f.seek(0)
                            oimObject = json.load(f, object_pairs_hook=loadDict)
                except UnicodeDecodeError as ex:
                    raise OIMException("{}:invalidJSON".format(errPrefix),
                          _("File MUST use utf-8 encoding: %(file)s, error %(error)s"),
                          file=filepath, error=str(ex))
                except json.JSONDecodeError as ex:
                    raise OIMException("{}:invalidJSON".format(errPrefix),
                            "JSON error while %(action)s, %(file)s, error %(error)s",
                            file=filepath, action=currentAction, error=ex)
                # check for top-level key duplicates
                if isinstance(oimObject, dict) and DUPJSONKEY in oimObject:
                    for _errKey, _errValue, _otherValue in oimObject[DUPJSONKEY]:
                        error("{}:invalidJSON".format(errPrefix),
                              _("The key %(key)s is used on multiple objects"),
                              modelObject=modelXbrl, key=_errKey)
                    del oimObject[DUPJSONKEY]
                oimWb = None
            elif isXL:
                _file = modelXbrl.fileSource.file(filepath, binary=True)[0]
                with _file as f:
                    oimWb = load_workbook(f, data_only=True)
                if "metadata" not in oimWb:
                    raise OIMException("xbrlwe:missingWorkbookWorksheets",
                                       _("Unable to identify worksheet tabs for metadata"))
                _foundMatch = False
                for row in range(1,10): # allow metadata to be indented or surrounded by column and row title columns
                    for col in range(1,10):
                        _metadata = xlValue(oimWb["metadata"].cell(row=row,column=col))
                        if _metadata and JSONmetadataPattern.match(_metadata): # find JSON metadata cell
                            _foundMatch = True
                            break
                    if _foundMatch:
                        break
                try:
                    oimObject = json.loads(_metadata, object_pairs_hook=loadDict)
                except UnicodeDecodeError as ex:
                    raise OIMException("{}:invalidJSON".format(errPrefix),
                          _("File MUST use utf-8 encoding: %(file)s \"metadata\" worksheet, error %(error)s"),
                          file=filepath, error=str(ex))
                except json.JSONDecodeError as ex:
                    raise OIMException("{}:invalidJSON".format(errPrefix),
                            "JSON error while %(action)s, %(file)s \"metadata\" worksheet, error %(error)s",
                            file=filepath, action=currentAction, error=ex)
            # allow report setup or extension objects processing
            for pluginXbrlMethod in pluginClassMethods("LoadFromOim.DocumentSetup"):
                pluginXbrlMethod(modelXbrl, oimObject, oimFile)
            # identify document type (JSON or CSV)
            documentInfo = jsonGet(oimObject, "documentInfo", {})
            documentType = jsonGet(documentInfo, "documentType")
            documentBase = jsonGet(documentInfo, "baseURL")
            if documentType in jsonDocumentTypes:
                isCSV = False
                isJSON = True
                errPrefix = "xbrlje"
                oimMemberTypes = JsonMemberTypes
                oimRequiredMembers = JsonRequiredMembers
            elif documentType in csvDocumentTypes:
                isJSON = False
                isCSV = not isXL
                errPrefix = "xbrlce"
                oimMemberTypes = CsvMemberTypes
                oimRequiredMembers = CsvRequiredMembers
            else: # if wrong type defer to type checking
                isCSV = False
                isJSON = False
                #errPrefix was set earlier based on file name
                oimMemberTypes = UnrecognizedDocMemberTypes
                oimRequiredMembers = UnrecognizedDocRequiredMembers
            isCSVorXL = isCSV or isXL

            # report loadDict errors
            for msgCode, msgText, kwargs in loadDictErrors:
                error(msgCode.format(errPrefix), msgText, href=filepath, **kwargs)
            del loadDictErrors[:]

            invalidMemberTypes = []
            invalidSQNames = []
            missingRequiredMembers = []
            unexpectedMembers = []
            def showPathObj(parts, obj): # this can be replaced with jsonPath syntax if appropriate
                try:
                    shortObjStr = json.dumps(obj)
                except TypeError:
                    shortObjStr = str(obj)
                if len(shortObjStr) > 34:
                    shortObjStr = "{:.32}...".format(shortObjStr)
                return "/{}={}".format("/".join(str(p) for p in parts), shortObjStr)
            def checkMemberTypes(obj, path, pathParts):
                if (isinstance(obj,dict)):
                    for missingMbr in oimRequiredMembers.get(path,EMPTY_SET) - obj.keys():
                        missingRequiredMembers.append(path + missingMbr)
                    for mbrName, mbrObj in obj.items():
                        mbrPath = path + mbrName
                        pathParts.append(mbrName)
                        # print("mbrName {} mbrObj {}".format(mbrName, mbrObj))
                        if mbrPath in oimMemberTypes:
                            mbrTypes = oimMemberTypes[mbrPath]
                            if (mbrTypes is SQNameType or (isinstance(mbrTypes,tuple) and SQNameType in mbrTypes)):
                                if not isinstance(mbrObj, str) or not SQNamePattern.match(mbrObj):
                                    invalidSQNames.append(showPathObj(pathParts, mbrObj))
                            elif (not ((mbrTypes is QNameType or (isinstance(mbrTypes,tuple) and QNameType in mbrTypes)) and isinstance(mbrObj, str) and QNamePattern.match(mbrObj)) and
                                not ((mbrTypes is LangType or (isinstance(mbrTypes,tuple) and LangType in mbrTypes)) and isinstance(mbrObj, str) and languagePattern.match(mbrObj)) and
                                not ((mbrTypes is URIType or (isinstance(mbrTypes,tuple) and URIType in mbrTypes)) and isinstance(mbrObj, str) and isValidUriReference(mbrObj) and not WhitespaceUntrimmedPattern.match(mbrObj)) and
                                #not (mbrTypes is IdentifierType and isinstance(mbrObj, str) and isinstance(mbrObj, str) and IdentifierPattern.match(mbrObj)) and
                                not ((mbrTypes is int or (isinstance(mbrTypes,tuple) and int in mbrTypes)) and isinstance(mbrObj, str) and CanonicalIntegerPattern.match(mbrObj)) and
                                not isinstance(mbrObj, mbrTypes)):
                                invalidMemberTypes.append(showPathObj(pathParts, mbrObj))
                        elif ":" in mbrName and path + "*:*" in oimMemberTypes:
                            _mbrTypes = oimMemberTypes[path + "*:*"]
                            if not (QNamePattern.match(mbrName) and isinstance(mbrObj, _mbrTypes)):
                                invalidMemberTypes.append(showPathObj(pathParts, mbrObj))
                            elif isinstance(_mbrTypes,tuple):
                                if CheckPrefix in _mbrTypes:
                                    extensionProperties[mbrName] = showPathObj(pathParts, mbrObj)
                                if NoRecursionCheck in _mbrTypes:
                                    continue # custom types, block recursive check
                            mbrPath = path + "*:*" # for recursion
                        elif path + "*" in oimMemberTypes:
                            mbrTypes = oimMemberTypes[path + "*"]
                            if (not ((mbrTypes is URIType or (isinstance(mbrTypes,tuple) and isinstance(mbrObj, str) and URIType in mbrTypes)) and isValidUriReference(mbrObj)) and
                                not isinstance(mbrObj, mbrTypes)):
                                invalidMemberTypes.append(showPathObj(pathParts, mbrObj))
                            if isinstance(mbrTypes,tuple) and KeyIsNcName in mbrTypes and not NCNamePattern.match(mbrName):
                                invalidMemberTypes.append(showPathObj(pathParts, mbrObj))
                            mbrPath = path + "*" # for recursion
                        else:
                            unexpectedMembers.append(showPathObj(pathParts, mbrObj))
                        if isinstance(mbrObj, (dict,list)):
                            checkMemberTypes(mbrObj, mbrPath + "/", pathParts)
                        pathParts.pop() # remove mbrName
                if (isinstance(obj,list)):
                    mbrNdx = 1
                    for mbrObj in obj:
                        mbrPath = path # list entry just uses path ending in /
                        pathParts.append(mbrNdx)
                        if mbrPath in oimMemberTypes:
                            mbrTypes = oimMemberTypes[mbrPath]
                            if (not (mbrTypes is IdentifierType and isinstance(mbrObj, str) and isinstance(mbrObj, str) and IdentifierPattern.match(mbrObj)) and
                                not ((mbrTypes is URIType or (isinstance(mbrTypes,tuple) and URIType in mbrTypes)) and isinstance(mbrObj, str) and isValidUriReference(mbrObj) and not WhitespaceUntrimmedPattern.match(mbrObj)) and
                                not isinstance(mbrObj, mbrTypes)):
                                invalidMemberTypes.append(showPathObj(pathParts, mbrObj))
                        if isinstance(mbrObj, (dict,list)):
                            checkMemberTypes(mbrObj, mbrPath + "/", pathParts)
                        pathParts.pop() # remove mbrNdx
                        mbrNdx += 1
            checkMemberTypes(oimObject, "/", [])
            numErrorsBeforeJsonCheck = len(modelXbrl.errors)
            if not isJSON and not isCSV and not isXL:
                error("oimce:unsupportedDocumentType",
                      _("Unrecognized /documentInfo/docType: %(documentType)s"),
                      documentType=documentType)
                extensionChain.discard(filepath)
                return {}
            if missingRequiredMembers or unexpectedMembers:
                msg = []
                if missingRequiredMembers:
                    msg.append(_("Required element(s) are missing from metadata: %(missing)s"))
                if unexpectedMembers:
                    msg.append(_("Unexpected element(s) in metadata: %(unexpected)s"))
                error("{}:invalidJSONStructure".format(errPrefix),
                      "\n ".join(msg), documentType=documentType,
                      sourceFileLine=oimFile, missing=", ".join(missingRequiredMembers), unexpected=", ".join(unexpectedMembers))
            if invalidMemberTypes:
                error("{}:invalidJSONStructure".format(errPrefix),
                      _("Invalid JSON structure member types in metadata: %(members)s"),
                      sourceFileLine=oimFile, members=", ".join(invalidMemberTypes))
            if invalidSQNames:
                error("oimce:invalidSQName".format(errPrefix),
                      _("Invalid SQNames in metadata: %(members)s"),
                      sourceFileLine=oimFile, members=", ".join(invalidMemberTypes))

            if isCSV and not primaryReportParameters:
                primaryReportParameters = oimObject.setdefault("parameters", {})

            # read reportParameters if in a CSV file relative to parent metadata file
            if isinstance(oimObject.get("parameterURL"), str):
                parameterURL = oimObject["parameterURL"]
                parameterFilePath = os.path.join(os.path.dirname(primaryOimFile), parameterURL)
                if modelXbrl.fileSource.exists(parameterFilePath):
                    problems = []
                    badIdentifiers = []
                    identifiersInThisFile = set()
                    for i, row in enumerate(openCsvReader(parameterFilePath, CSV_PARAMETER_FILE)):
                        if i == 0:
                            if row != ["name", "value"]:
                                problems.append(_("The first row must only consist of \"name\" and \"value\" but contains: {}").format(",".join(row)))
                        elif len(row) > 0 and row[0]:
                            name = row[0]
                            if not IdentifierPattern.match(name):
                                badIdentifiers.append(_("Row {} column 1 is not a valid identifier: {}").format(i+1, name))
                            elif len(row) < 2 or not row[1]:
                                problems.append(_("Row {} value column 2 missing").format(i+1))
                            elif any(cell for cell in row[2:]):
                                problems.append(_("Row {} columns 3 - {} must be empty").format(i+1, len(row)))
                            # no longer illegal to override primary report parameters... but between csv files is it illegal?
                            #elif row[0] in primaryReportParameters:
                            #    if primaryReportParameters[row[0]] != row[1]:
                            #        error("xbrlce:illegalReportParameterRedefinition",
                            #             _("Report parameter %(name)s redefined in file %(file)s, report value %(value1)s, csv value %(value2)s"),
                            #             file=parameterURL, name=row[0], value1=primaryReportParameters[row[0]], value2=row[1])
                            elif name in identifiersInThisFile:
                                problems.append(_("Row {} column 1 is has a repeated identifier: {}").format(i+1, name))
                            else:
                                identifiersInThisFile.add(name)
                                primaryReportParameters[name] = row[1]
                        elif any(cell for cell in row):
                            problems.append(_("Row {} has no identifier, all columns must be empty").format(i+1))
                    if badIdentifiers:
                        error("xbrlce:invalidIdentifier",
                              _("Report parameter file %(file)s:\n %(issues)s"),
                              file=parameterURL, issues=", \n".join(badIdentifiers))
                    if problems:
                        error("xbrlce:invalidParameterCSVFile",
                              _("Report parameter file %(file)s issues:\n %(issues)s"),
                              file=parameterURL, issues=", \n".join(problems))
                else:
                    error("xbrlce:missingParametersFile",
                          _("Report parameter file is missing: %(file)s"),
                          file=parameterURL)

            if isCSVorXL: # normalize relative taxonomy URLs to primary document or nearest absolute parent
                t = documentInfo.get("taxonomy",())
                for i, tUrl in enumerate(t):
                    t[i] = modelXbrl.modelManager.cntlr.webCache.normalizeUrl(tUrl, normalizedUrl)

            if isCSVorXL and "extends" in documentInfo:
                # process extension
                for extendedFile in documentInfo["extends"]:
                    try:
                        extendedOimObject = loadOimObject(extendedFile, mappedUrl, visitedFiles, extensionChain)
                    except IOError:
                        error("{}:unresolvableBaseMetadataFile".format(errPrefix),
                              _("Extending document file not found: %(extendingFile)s, referenced from %(extendedFile)s"),
                              extendingFile=extendedFile, extendedFile=oimFile)
                        raise OIMException()
                    if extendedOimObject is None:
                        continue # None returned when directed cycle blocks reloading same file
                    # extended must be CSV
                    extendedDocumentInfo = extendedOimObject.get("documentInfo", EMPTY_DICT)
                    extendedDocumentType = extendedDocumentInfo.get("documentType")
                    extendedFinal = extendedDocumentInfo.get("final", EMPTY_DICT)
                    if extendedDocumentType != documentType:
                        error("{}:multipleDocumentTypesInExtensionChain".format(errPrefix),
                              _("Extended documentType %(extendedDocumentType)s must same as extending documentType %(documentType)s in file %(extendedFile)s"),
                              extendedFile=extendedFile, extendedDocumentType=extendedDocumentType, documentType=documentType)
                        raise OIMException()
                    oimParameters = oimObject.setdefault("parameters", {})
                    for paramName, paramValue in extendedOimObject.get("parameters",{}).items():
                        if paramName in oimParameters and oimParameters[paramName] != paramValue:
                            error("xbrlce:illegalReportParameterRedefinition",
                                  _("Report parameter %(name)s redefined in file %(file)s, extended value %(value1)s, extending value %(value2)s"),
                                  file=extendedFile, name=paramName, value1=oimParameters[paramName], value2=paramValue)
                        else:
                            oimParameters[paramName] = paramValue
                    for parent, extendedParent, excludedObjectNames in (
                        (documentInfo, extendedDocumentInfo, {"documentType", "extends"}),
                        (oimObject, extendedOimObject, {"documentInfo"})):
                        for objectName in extendedFinal:
                            if objectName not in excludedObjectNames and objectName not in extendedParent and objectName in parent:
                                error("xbrlce:illegalExtensionOfFinalProperty",
                                      _("Extended file %(extendedFile)s redefines final object %(finalObjectName)s"),
                                      extendedFile=extendedFile, finalObjectName=objectName)
                        for objectName in extendedParent.keys() - excludedObjectNames:
                            if objectName in csvExtensibleObjects:
                                for parProp, parPropValue in parent.get(objectName,EMPTY_DICT).items():
                                    if extendedFinal.get(objectName, False) and parProp not in extendedParent.get(objectName,EMPTY_DICT):
                                        error("xbrlce:illegalExtensionOfFinalProperty",
                                              _("Extended file %(extendedFile)s specifies final object %(objectName)s property %(property)s"),
                                              extendedFile=oimFile, objectName=objectName, property=parProp)
                                for extProp, extPropValue in extendedParent.get(objectName,EMPTY_DICT).items():
                                    if extProp in parent.get(objectName,EMPTY_DICT):
                                        if json.dumps(extPropValue,sort_keys=True) != json.dumps(parent[objectName][extProp],sort_keys=True): # ordered dicts, especially nested are not comparable
                                            error("xbrlce:conflictingMetadataValue" if extendedFinal.get(objectName, False)
                                                  else "xbrlce:conflictingMetadataValue",
                                                  _("Extended file %(extendedFile)s redefines object %(objectName)s property %(property)s"),
                                                  extendedFile=extendedFile, objectName=objectName, property=extProp)
                                    else:
                                        if objectName not in parent:
                                            parent[objectName] = {}
                                        parent[objectName][extProp] = extPropValue
                            elif objectName in parent:
                                if objectName == "taxonomy":
                                    for extPropValue in extendedParent["taxonomy"]:
                                        if extPropValue not in parent["taxonomy"]:
                                            parent["taxonomy"].append(extPropValue)
                                elif extendedParent[objectName] != parent[objectName]:
                                    error("xbrlce:illegalRedefinitionOfNonExtensibleProperty",
                                          _("Extended file %(extendedFile)s redefines object %(objectName)s"),
                                          extendedFile=extendedFile, objectName=objectName)
                            else:
                                parent[objectName] = extendedParent[objectName]

            if extendingFile is None: # entry oimFile
                if ("taxonomy" in documentInfo or isCSV) and not documentInfo.get("taxonomy",()):
                    error("oime:noTaxonomy",
                          _("The list of taxonomies MUST NOT be empty."))
                if  len(modelXbrl.errors) > numErrorsBeforeJsonCheck:
                    raise OIMException()

                oimObject["=entryParameters"] = (isJSON, isCSV, isXL, isCSVorXL, oimWb, documentInfo, documentType, documentBase)

            extensionChain.discard(filepath)
            return oimObject

        errorIndexBeforeLoadOim = len(modelXbrl.errors)
        oimObject = loadOimObject(oimFile, None, set(), set())
        try:
            isJSON, isCSV, isXL, isCSVorXL, oimWb, oimDocumentInfo, documentType, documentBase = oimObject["=entryParameters"]
        except KeyError:
            raise OIMException() # no document
        del oimObject["=entryParameters"]

        currentAction = "identifying Metadata objects"
        taxonomyRefs = oimDocumentInfo.get("taxonomy", EMPTY_LIST)
        namespaces = oimDocumentInfo.get("namespaces", EMPTY_DICT)
        linkTypes = oimDocumentInfo.get("linkTypes", EMPTY_DICT)
        linkGroups = oimDocumentInfo.get("linkGroups", EMPTY_DICT)
        featuresDict = oimDocumentInfo.get("features", EMPTY_DICT)
        documentInfoProperties = {"documentType", "features", "namespaces", "linkTypes", "linkGroups", "taxonomy", "baseURL"}
        oimObjectProperties = {}
        factProperties = {"decimals", "dimensions", "links", "value"}
        canonicalValuesFeature = False
        if isJSON:
            errPrefix = "xbrlje"
            valErrPrefix = "xbrlje"
            OIMReservedAliasURIs["namespaces"] = NSReservedAliasURIs.copy()
            OIMReservedAliasURIs["namespaces"].update(JSONNSReservedAliasURIs)
            OIMReservedURIAlias["namespaces"] = NSReservedURIAlias.copy()
            OIMReservedURIAlias["namespaces"].update(JSONNSReservedURIAliases)
            factItems = oimObject.get("facts",{}).items()
            footnotes = oimObject.get("facts",{}).values() # shares this object
            canonicalValuesFeature = featuresDict.get("xbrl:canonicalValues") in (True, "true")
        else: # isCSVorXL
            errPrefix = "xbrlce"
            valErrPrefix = "xbrlce"
            OIMReservedAliasURIs["namespaces"] = NSReservedAliasURIs.copy()
            OIMReservedAliasURIs["namespaces"].update(CSVNSReservedAliasURIs)
            OIMReservedURIAlias["namespaces"] = NSReservedURIAlias.copy()
            OIMReservedURIAlias["namespaces"].update(CSVNSReservedURIAliases)
            reportDimensions = oimObject.get("dimensions", EMPTY_DICT)
            reportDecimals = oimObject.get("decimals", None)
            reportParameters = oimObject.get("parameters", {}) # fresh empty dict because csv-loaded parameters get added
            tableTemplates = oimObject.get("tableTemplates", EMPTY_DICT)
            tables = oimObject.get("tables", EMPTY_DICT)
            footnotes = (oimObject.get("links", {}), )
            final = oimObject.get("final", EMPTY_DICT)
            documentInfoProperties.add("extends")
            documentInfoProperties.add("final")
            reportProperties = {"documentInfo", "tableTemplates", "tables", "parameters", "parameterURL", "dimensions", "decimals", "links"}
            columnProperties = {"comment", "decimals", "dimensions", "propertyGroups", "parameterURL", "propertiesFrom"}

        entityNaQName = qname(re_sub("/xbrl-(json|csv)$","/entities",documentType), "NA")
        allowedDuplicatesFeature = ALL
        v = featuresDict.get("xbrl:allowedDuplicates")
        if v is not None:
            if v in AllowedDuplicatesFeatureValues:
                allowedDuplicatesFeature = AllowedDuplicatesFeatureValues[v]
            else:
                error("{}:invalidJSONStructure".format(errPrefix),
                      _("The xbbrl:allowedDuplicates feature has an invalid value: %(value)s"),
                      value=v)

        # check extension properties (where metadata specifies CheckPrefix)
        for extPropSQName, extPropertyPath in extensionProperties.items():
            extPropPrefix = extPropSQName.partition(":")[0]
            if extPropPrefix not in namespaces:
                error("oimce:unboundPrefix",
                      _("The extension property QName prefix was not defined in namespaces: %(extensionProperty)s."),
                      modelObject=modelXbrl, extensionProperty=extPropertyPath)

        # check features
        for featureSQName, isActive in featuresDict.items():
            featurePrefix = featureSQName.partition(":")[0]
            if featurePrefix not in namespaces:
                error("oimce:unboundPrefix",
                      _("The feature QName prefix was not defined in namespaces: %(feature)s."),
                      modelObject=modelXbrl, feature=featureSQName)

        # check maps
        for alias, uris in NSReservedAliasURIs.items():
            for uri in uris:
                NSReservedURIAlias[uri] = alias

        for map in ("namespaces", "linkTypes", "linkGroups"):
            for key, value in oimDocumentInfo.get(map, EMPTY_DICT).items():
                if key in OIMReservedAliasURIs[map] and value not in OIMReservedAliasURIs[map][key]:
                    error("oimce:invalidURIForReservedAlias",
                          _("The %(map)s URI \"%(uri)s\" is used on standard alias \"%(alias)s\" which requires URI \"%(standardUri)s\"."),
                          modelObject=modelXbrl, map=map, alias=key, uri=value, standardUri=OIMReservedAliasURIs[map][key][0])
                elif value in OIMReservedURIAlias[map] and key != OIMReservedURIAlias[map][value]:
                    error("oimce:invalidAliasForReservedURI",
                          _("The %(map)s URI \"%(uri)s\" is bound to alias \"%(key)s\" instead of standard alias \"%(alias)s\"."),
                          modelObject=modelXbrl, map=key, key=key, uri=value, alias=OIMReservedURIAlias[map][value])

        # check baseURL
        if documentBase and not isAbsoluteUri(documentBase):
            error("oime:invalidBaseURL",
                  _("The base-url must be absolute: \"%(url)s\"."),
                  modelObject=modelXbrl, url=documentBase)

        factProduced = FactProduced() # pass back fact info to csv Fact producer

        if isCSVorXL:
            currentAction = "loading CSV facts tables"
            _dir = os.path.dirname(oimFile)

            def csvFacts():
                parseMetadataCellValues(reportDimensions)
                for tableId, table in tables.items():
                    _file = tablePath = None
                    try: # note that decoder errors may occur late during streaming of rows
                        tableTemplateId = table.get("template", tableId)
                        tableTemplate = tableTemplates[tableTemplateId]
                        # tableIsTransposed = tableTemplate.get("transposed", False)
                        tableDecimals = tableTemplate.get("decimals")
                        tableDimensions = tableTemplate.get("dimensions", EMPTY_DICT)
                        parseMetadataCellValues(tableDimensions)
                        tableIsOptional = table.get("optional", False)
                        tableParameters = table.get("parameters", EMPTY_DICT)
                        rowIdColName = tableTemplate.get("rowIdColumn")
                        tableUrl = table["url"]
                        tableParameterColNames = set()
                        hasHeaderError = False # set to true blocks handling file beyond header row

                        # compile column dependencies
                        factDimensions = {} # keys are column, values are dimensions object
                        factDecimals = {} # keys are column
                        propertyGroups = {}
                        propertiesFrom = {}
                        dimensionsColumns = set()
                        commentColumns = set()
                        extensionColumnProperties = defaultdict(dict)
                        for colId, colProperties in tableTemplate["columns"].items():
                            isCommentColumn = colProperties.get("comment") == True
                            if isCommentColumn:
                                commentColumns.add(colId)
                            else:
                                factDimensions[colId] = colProperties.get("dimensions")
                                factDecimals[colId] = colProperties.get("decimals")
                            isFactColumn = "dimensions" in colProperties
                            if "propertiesFrom" in colProperties:
                                isFactColumn = True
                                propertiesFrom[colId] = colProperties["propertiesFrom"]
                            if not isFactColumn and not isCommentColumn:
                                dimensionsColumns.add(colId) # neither comment nor fact column
                            isPropertyGroupColumn = "propertyGroups" in colProperties
                            if isPropertyGroupColumn:
                                propertyGroups[colId] = colProperties["propertyGroups"]
                            for extPropSQName, prop in colProperties.items():
                                if extPropSQName not in columnProperties:
                                    extensionColumnProperties[colId][extPropSQName] = prop
                        # check table parameters
                        tableParameterReferenceNames = set()
                        def checkParamRef(paramValue, factColName=None, dimName=None):
                            if isinstance(paramValue, str) and paramValue.startswith("$") and not paramValue.startswith("$$"):
                                paramName = paramValue[1:].partition("@")[0]
                                tableParameterReferenceNames.add(paramName)
                        unitDims = set()
                        for factColName, colDims in factDimensions.items():
                            if colDims is not None:
                                factDims = set()
                                for inheritedDims in (colDims, tableDimensions, reportDimensions):
                                    for dimName, dimValue in inheritedDims.items():
                                        checkParamRef(dimValue, factColName, dimName)
                                        factDims.add(dimName)
                                parseMetadataCellValues(colDims)
                                for _factDecimals in (factDecimals.get(factColName), tableDecimals, reportDecimals):
                                    if "decimals" not in factDims:
                                        checkParamRef(_factDecimals, factColName, "decimals")

                        if hasHeaderError:
                            return
                        # determine whether table is a CSV file or an Excel range.
                        # Local range can be sheetname! or !rangename
                        # url to workbook with range must be url#sheet! or url#!range or url!range (unencoded !)
                        tableWb = None
                        _file = None
                        _rowIterator = None
                        _cellValue = None
                        if isXL and not ("#" in tableUrl or ".xlsx" in tableUrl or ".csv" in tableUrl):
                            # local Workbook range
                            tableWb = oimWb
                            _cellValue = xlValue
                            xlSheetName, _sep, xlNamedRange = tableUrl.partition('!')
                        else:
                            # check if there's a reference to an Excel workbook file
                            if "#" in tableUrl:
                                tableUrl, _sep, sheetAndRange = tableUrl.partition("#")
                                xlSheetName, _sep, xlNamedRange = sheetAndRange.partition('!')
                            tablePath = os.path.join(_dir, tableUrl)
                            # Remove unnecessary relative segments within path. Effected paths are handled fine
                            # when loading from directories, but this fails when loading from ZIP archives.
                            # OIM conformance suites expect this to be supported:
                            # oim-conf-2021-10-13.zip/300-csv-conformant-processor/V-11,
                            #  "/300-csv-conformant-processor/./helloWorld-value-date-table2-facts.csv"
                            # oim-conf-2021-10-13.zip/300-csv-conformant-processor/V-12
                            #  "/300-csv-conformant-processor/./helloWorld-SQNameSpecial-facts.csv"
                            tablePath = os.path.normpath(tablePath)
                            if not modelXbrl.fileSource.exists(tablePath):
                                if not tableIsOptional:
                                    error("xbrlce:missingRequiredCSVFile",
                                          _("Table %(table)s missing, url: %(url)s"),
                                          table=tableId, url=tableUrl)
                                continue
                            if tableUrl.endswith(".xlsx"):
                                _file = modelXbrl.fileSource.file(tablePath, binary=True)[0]
                                tableWb = load_workbook(_file, data_only=True)
                                _cellValue = xlValue
                            else:
                                # must be CSV
                                _rowIterator = openCsvReader(tablePath, CSV_FACTS_FILE)
                                _cellValue = csvCellValue
                                # if tableIsTransposed:
                                #    _rowIterator = transposer(_rowIterator)
                        if tableWb is not None:
                            hasSheetname = xlSheetName and xlSheetName in tableWb
                            hasNamedRange = xlNamedRange and xlNamedRange in tableWb.defined_names
                            if xlSheetName and not hasSheetname:
                                if tableIsOptional:
                                    continue
                                raise OIMException("xbrlwe:missingTable",
                                                   _("Referenced table tab(s): %(missing)s"),
                                                   missing=tableUrl)
                            if xlNamedRange and not hasNamedRange:
                                if tableIsOptional:
                                    continue
                                raise OIMException("xbrlwe:missingTableNamedRange",
                                                   _("Referenced named ranges tab(s): %(missing)s"),
                                                   missing=tableRangeName)
                            if hasNamedRange: # check type of range
                                defn = tableWb.defined_names[xlNamedRange]
                                if defn.type != "RANGE":
                                    raise OIMException("xbrlwe:unusableRange",
                                                       _("Referenced range does not refer to a range: %(tableRange)s"),
                                                       tableRange=tableRangeName)
                            _rowIterator = []
                            if hasNamedRange:
                                for _tableName, _xlCellsRange in tableWb.defined_names[xlNamedRange].destinations:
                                    rows = tableWb[_tableName][_xlCellsRange]
                                    if isinstance(rows, Cell):
                                        _rowIterator.append((rows, ))
                                    else:
                                        _rowIterator.extend(rows)
                            else: # use whole table
                                _rowIterator = tableWb[xlSheetName]
                            # if tableIsTransposed:
                            #     _rowIterator = transposer(_rowIterator)

                        rowIds = set()
                        paramRefColNames = set()
                        potentialInvalidReferenceTargets = {} # dimName: referenceTarget
                        for rowIndex, row in enumerate(_rowIterator):
                            if rowIndex == 0:
                                header = [_cellValue(cell) for cell in row]
                                emptyHeaderCols = set()
                                if isXL: # trim empty cells
                                    header = xlTrimHeaderRow(header)
                                colNameIndex = dict((name, colIndex) for colIndex, name in enumerate(header))
                                idColIndex = colNameIndex.get(rowIdColName)
                                for colIndex, colName in enumerate(header):
                                    if colName == "":
                                        emptyHeaderCols.add(colIndex)
                                    elif not IdentifierPattern.match(colName):
                                        hasHeaderError = True
                                        error("xbrlce:invalidHeaderValue",
                                              _("Table %(table)s CSV file header column %(column)s is not a valid identifier: %(identifier)s, url: %(url)s"),
                                              table=tableId, column=colIndex+1, identifier=colName, url=tableUrl)
                                    elif colName not in factDimensions and colName not in commentColumns:
                                        hasHeaderError = True
                                        error("xbrlce:unknownColumn",
                                              _("Table %(table)s CSV file header column %(column)s is not in table template definition: %(identifier)s, url: %(url)s"),
                                              table=tableId, column=colIndex+1, identifier=colName, url=tableUrl)
                                    elif colNameIndex[colName] != colIndex:
                                        error("xbrlce:repeatedColumnIdentifier",
                                              _("Table %(table)s CSV file header columns %(column)s and %(column2)s repeat identifier: %(identifier)s, url: %(url)s"),
                                              table=tableId, column=colIndex+1, column2=colNameIndex[colName]+1, identifier=colName, url=tableUrl)
                                    if colName in tableParameterReferenceNames and colName not in commentColumns:
                                        paramRefColNames.add(colName)
                                #missingPropFromCols = flattenToSet(propertiesFrom.values()) - colNameIndex.keys()
                                #if missingPropFromCols:
                                #    raise OIMException("xbrlce:invalidPropertyGroupColumnReference",
                                #                  _("Table %(table)s propertyFrom %(propFromColumns)s column missing, url: %(url)s"),
                                #                  table=tableId, propFromColumns=", ".join(sorted(missingPropFromCols)), url=tableUrl)
                                # check parameter references
                                checkedDims = set()
                                checkedParams = set()
                                def dimChecks():
                                    for colName, colDims in factDimensions.items():
                                        if colDims:
                                            yield colDims, "column {} dimension".format(colName)
                                    # no way to check parameterGroup dimensions at header-row processing time
                                    for dims, source in ((tableDimensions, "table dimension"),
                                                         (reportDimensions, "report dimension"),
                                                         ):
                                        yield dims, source
                                    for colName, dec in factDecimals.items():
                                        yield {"decimals": dec}, "column {} decimals".format(colName)
                                    for dec, source in ((tableDecimals, "table decimals"),
                                                        (reportDecimals, "report decimals")):
                                        if source:
                                            yield {"decimals": dec}, source
                                for inheritedDims, dimSource in dimChecks():
                                    for dimName, dimValue in inheritedDims.items():
                                        if dimName not in checkedDims:
                                            dimValue = inheritedDims[dimName]
                                            # resolve column-relative dimensions
                                            if isinstance(dimValue, str):
                                                if dimValue.startswith("$"):
                                                    dimValue = dimValue[1:]
                                                    if not dimValue.startswith("$"):
                                                        dimValue, _sep, dimAttr = dimValue.partition("@")
                                                        if _sep and dimAttr not in ("start", "end"):
                                                            hasHeaderError = True
                                                            error("xbrlce:invalidPeriodSpecifier",
                                                                  _("Table %(table)s %(source)s %(dimension)s period-specifier invalid: %(target)s, url: %(url)s"),
                                                                  table=tableId, source=dimSource, dimension=dimName, target=dimAttr, url=tableUrl)
                                                        if dimValue not in checkedParams:
                                                            checkedParams.add(dimValue)
                                                            if dimValue in ("rowNumber", ) or (dimValue in header and dimValue not in commentColumns) or dimValue in tableParameters or dimValue in reportParameters:
                                                                checkedDims.add(dimValue)
                                                            else:
                                                                potentialInvalidReferenceTargets[dimName] = dimValue
                                                elif ":" in dimName and ":" in dimValue:
                                                    dimConcept = modelXbrl.qnameConcepts.get(qname(dimName, namespaces))
                                                    if dimConcept is not None and dimConcept.isExplicitDimension:
                                                        memConcept = modelXbrl.qnameConcepts.get(qname(dimValue, namespaces))
                                                        if memConcept is not None and modelXbrl.dimensionDefaultConcepts.get(dimConcept) == memConcept:
                                                            error("xbrlce:invalidDimensionValue",
                                                                  _("Table %(table)s %(source)s %(dimension)s value must not be the default member %(member)s, url: %(url)s"),
                                                                  table=tableId, source=dimSource, dimension=dimName, member=dimValue, url=tableUrl)
                                for commentCol in commentColumns:
                                    colNameIndex.pop(commentCol,None) # remove comment columns from col name index
                                unreportedFactDimensionColumns = factDimensions.keys() - set(header)
                                reportedDimensionsColumns = dimensionsColumns & set(header)
                                if hasHeaderError:
                                    break # stop processing table
                            else:
                                rowId = None
                                paramColsWithValue = set()
                                paramColsUsed = set()
                                emptyCols = set()
                                emptyHeaderColsWithValue = []
                                if isXL and all(cell.value in (None, "") for cell in row): # skip empty excel rows
                                    continue
                                rowPropGroups = {} # colName, propGroupObject for property groups in this row
                                rowPropGroupsUsed = set() # colNames used by propertiesFrom of fact col producing a fact
                                hasRowError = False
                                rowPropGrpParamRefs = set()
                                for propGrpName, propGrpObjects in propertyGroups.items():
                                    propGrpColIndex = colNameIndex.get(propGrpName, 999999999)
                                    if propGrpColIndex < len(row):
                                        propGrpColValue = _cellValue(row[propGrpColIndex])
                                        if propGrpColValue is NONE_CELL:
                                            error("xbrlce:illegalUseOfNone",
                                                  _("Table %(table)s row %(row)s column %(column)s must not have #none, from %(source)s, url: %(url)s"),
                                                  table=tableId, row=rowIndex+1, column=colName, url=tableUrl, source=dimSource)
                                            hasRowError = True
                                        elif propGrpColValue in propGrpObjects:
                                            rowPropGroups[propGrpName] = propGrpObjects[propGrpColValue]
                                        else:
                                            error("xbrlce:unknownPropertyGroup",
                                                  _("Table %(table)s unknown property group row %(row)s column %(column)s group %(propertyGroup)s, url: %(url)s"),
                                                  table=tableId, row=rowIndex+1, column=rowIdColName, url=tableUrl, propertyGroup=propGrpName)
                                            hasRowError = True
                                if hasRowError:
                                    continue
                                for colIndex, colValue in enumerate(row):
                                    if colIndex >= len(header):
                                        if _cellValue(colValue) != EMPTY_CELL:
                                            emptyHeaderColsWithValue.append(colIndex)
                                        continue
                                    cellPropGroup = {}
                                    propGroupDimSource = {}
                                    colName = header[colIndex]
                                    if colName == "":
                                        if _cellValue(colValue) != EMPTY_CELL:
                                            emptyHeaderColsWithValue.append(colIndex)
                                        continue
                                    if colName in commentColumns:
                                        continue
                                    propFromColNames = propertiesFrom.get(colName,EMPTY_LIST)
                                    for propFromColName in propFromColNames:
                                        if propFromColName in rowPropGroups:
                                            for prop, val in rowPropGroups[propFromColName].items():
                                                if isinstance(val, dict):
                                                    _valDict = cellPropGroup.setdefault(prop, {})
                                                    for dim, _val in val.items():
                                                        _valDict[dim] = _val
                                                        propGroupDimSource[dim] = propFromColName
                                                        if _val.startswith("$") and not _val.startswith("$$"):
                                                            rowPropGrpParamRefs.add(_val.partition("@")[0][1:])
                                                else:
                                                    cellPropGroup[prop] = val
                                                    propGroupDimSource[prop] = propFromColName
                                    if factDimensions[colName] is None:
                                        if colName in paramRefColNames:
                                            value = _cellValue(row[colNameIndex[colName]])
                                            if value:
                                                paramColsWithValue.add(colName)
                                            elif value is EMPTY_CELL or value is NONE_CELL:
                                                emptyCols.add(colName)
                                        if not cellPropGroup:
                                            continue # not a fact column
                                    for rowPropGrpParamRef in rowPropGrpParamRefs:
                                        value = _cellValue(row[colNameIndex[rowPropGrpParamRef]])
                                        if value is EMPTY_CELL or value is NONE_CELL:
                                            emptyCols.add(rowPropGrpParamRef)
                                    # assemble row and fact Ids
                                    if idColIndex is not None and not rowId:
                                        if idColIndex < len(row):
                                            rowId = _cellValue(row[idColIndex])
                                        if not rowId:
                                            error("xbrlce:missingRowIdentifier",
                                                  _("Table %(table)s missing row %(row)s column %(column)s row identifier, url: %(url)s"),
                                                  table=tableId, row=rowIndex+1, column=rowIdColName, url=tableUrl)
                                        elif not RowIdentifierPattern.match(rowId):
                                            error("xbrlce:invalidRowIdentifier",
                                                  _("Table %(table)s row %(row)s column %(column)s is not valid as a row identifier: %(identifier)s, url: %(url)s"),
                                                  table=tableId, row=rowIndex+1, column=rowIdColName, identifier=rowId, url=tableUrl)
                                        elif rowId in rowIds:
                                            error("xbrlce:repeatedRowIdentifier",
                                                  _("Table %(table)s row %(row)s column %(column)s is a duplicate: %(identifier)s, url: %(url)s"),
                                                  table=tableId, row=rowIndex+1, column=rowIdColName, identifier=rowId, url=tableUrl)
                                        else:
                                            rowIds.add(rowId)
                                            paramColsUsed.add(rowIdColName)
                                    factId = "{}.r_{}.{}".format(tableId, rowId or rowIndex, colName) # pre-pend r_ to rowId col value or row number if no rowId col value
                                    fact = {}
                                    # if this is an id column
                                    cellValue = _cellValue(colValue) # nil facts return None, #empty string is ""
                                    if cellValue is EMPTY_CELL: # no fact produced
                                        continue
                                    if cellValue is NONE_CELL:
                                        error("xbrlce:illegalUseOfNone",
                                              _("Table %(table)s row %(row)s column %(column)s must not have #none, from %(source)s, url: %(url)s"),
                                              table=tableId, row=rowIndex+1, column=colName, url=tableUrl, source=dimSource)
                                        continue
                                    if cellPropGroup:
                                        for propFromColName in propFromColNames:
                                            rowPropGroupsUsed.add(propFromColName)
                                    if colName in extensionColumnProperties: # merge extension properties to fact
                                        fact.update(extensionColumnProperties[colName])
                                    fact["value"] = cellValue
                                    fact["dimensions"] = colFactDims = {}
                                    noValueDimNames = set()
                                    factDimensionSourceCol = {} # track consumption of column value dynamically
                                    factDimensionPropGrpCol = {}
                                    for inheritedDims, dimSource in ((factDimensions[colName], "column dimension"),
                                                                     (cellPropGroup.get("dimensions",EMPTY_DICT), "propertyGroup {}".format(propFromColNames)),
                                                                     (tableDimensions, "table dimension"),
                                                                     (reportDimensions, "report dimension")):
                                        for dimName, dimValue in inheritedDims.items():
                                            if dimName not in colFactDims and dimName not in noValueDimNames:
                                                dimValue = inheritedDims[dimName]
                                                dimAttr = None
                                                # resolve column-relative dimensions
                                                if isinstance(dimValue, str) and dimValue.startswith("$"):
                                                    dimValue = dimValue[1:]
                                                    if not dimValue.startswith("$"):
                                                        paramName, _sep, dimAttr = dimValue.partition("@")
                                                        if paramName == "rowNumber":
                                                            dimValue = str(rowIndex)
                                                        elif paramName in colNameIndex:
                                                            dimValue = _cellValue(row[colNameIndex[paramName]])
                                                            if dimValue is EMPTY_CELL or dimValue is NONE_CELL: # csv file empty cell or  none
                                                                dimValue = NONE_CELL
                                                            else:
                                                                factDimensionSourceCol[dimName] = paramName
                                                        elif paramName in tableParameters:
                                                            dimValue = tableParameters[paramName]
                                                        elif paramName in reportParameters:
                                                            dimValue = reportParameters[paramName]
                                                        elif paramName in unreportedFactDimensionColumns:
                                                            dimValue = NONE_CELL
                                                        else:
                                                            dimValue = INVALID_REFERENCE_TARGET
                                                # else if in parameters?
                                                if dimName == "period" and dimValue is not INVALID_REFERENCE_TARGET:
                                                    _dimValue = csvPeriod(dimValue, dimAttr)
                                                    if _dimValue == "referenceTargetNotDuration":
                                                        error("xbrlce:referenceTargetNotDuration",
                                                              _("Table %(table)s row %(row)s column %(column)s has instant date with period reference \"%(date)s\", from %(source)s, url: %(url)s"),
                                                              table=tableId, row=rowIndex+1, column=colName, date=dimValue, url=tableUrl, source=dimSource)
                                                        dimValue = NONE_CELL
                                                    elif _dimValue is None: # bad format, raised value error
                                                        error("xbrlce:invalidPeriodRepresentation",
                                                              _("Table %(table)s row %(row)s column %(column)s has lexical syntax issue with date \"%(date)s\", from %(source)s, url: %(url)s"),
                                                              table=tableId, row=rowIndex+1, column=colName, date=dimValue, url=tableUrl, source=dimSource)
                                                        dimValue = NONE_CELL
                                                    else:
                                                        dimValue = _dimValue
                                                if dimValue is NONE_CELL:
                                                    noValueDimNames.add(dimName)
                                                else:
                                                    colFactDims[dimName] = dimValue
                                                if dimSource.startswith("propertyGroup"):
                                                    factDimensionPropGrpCol[dimName] = propGroupDimSource[dimName]
                                    if factDecimals.get(colName) is not None:
                                        dimValue = factDecimals[colName]
                                        dimSource = "column decimals"
                                    elif "decimals" in cellPropGroup:
                                        dimValue = cellPropGroup["decimals"]
                                        dimSource = "propertyGroup " + propFromColName
                                        factDimensionPropGrpCol["decimals"] = propGroupDimSource[dimName]
                                    elif tableDecimals is not None:
                                        dimValue = tableDecimals
                                        dimSource = "table decimals"
                                    elif reportDecimals is not None:
                                        dimValue = reportDecimals
                                        dimSource = "report decimals"
                                    else:
                                        dimValue = None
                                        dimSource = "absent"
                                    if dimValue is not None:
                                        validCsvCell = False
                                        if isinstance(dimValue, str) and dimValue.startswith("$"):
                                            paramName = dimValue[1:].partition("@")[0]
                                            if paramName in colNameIndex:
                                                dimSource += " from CSV column " + paramName
                                                dimValue = _cellValue(row[colNameIndex[paramName]])
                                                validCsvCell = integerPattern.match(dimValue or "") is not None # is None if is_XL
                                                if dimValue is not NONE_CELL and dimValue != "" and dimValue != "#none":
                                                    factDimensionSourceCol["decimals"] = paramName
                                            elif paramName in tableParameters:
                                                dimSource += " from table parameter " + paramName
                                                dimValue = tableParameters[paramName]
                                                if dimValue != "" and dimValue != "#none" and integerPattern.match(dimValue):
                                                    dimValue = int(dimValue)
                                            elif paramName in reportParameters:
                                                dimSource += " from report parameter " + paramName
                                                dimValue = reportParameters[paramName]
                                                if dimValue != "" and dimValue != "#none" and integerPattern.match(dimValue):
                                                    dimValue = int(dimValue)
                                            else:
                                                dimValue = INVALID_REFERENCE_TARGET
                                                validCsvCell = True # must wait to see if it's used later
                                        if dimValue is INVALID_REFERENCE_TARGET:
                                            fact["decimals"] = dimValue # allow referencing if not overridden by decimals suffix
                                        elif dimValue is not NONE_CELL and dimValue != "" and dimValue != "#none":
                                            if isinstance(dimValue, int) or validCsvCell:
                                                fact["decimals"] = dimValue
                                            else:
                                                error("xbrlce:invalidDecimalsValue",
                                                      _("Table %(table)s row %(row)s column %(column)s has invalid decimals \"%(decimals)s\", from %(source)s, url: %(url)s"),
                                                      table=tableId, row=rowIndex+1, column=colName, decimals=dimValue, url=tableUrl, source=dimSource)
                                    yield (factId, fact)
                                    if factProduced.invalidReferenceTarget:
                                        error("xbrlce:invalidReferenceTarget",
                                              _("Table %(table)s %(dimension)s target not in table columns, parameters or report parameters: %(target)s, url: %(url)s"),
                                              table=tableId, dimension=factProduced.invalidReferenceTarget, target=potentialInvalidReferenceTargets.get(factProduced.invalidReferenceTarget), url=tableUrl)
                                        break # stop processing table
                                    for dimName, dimSource in factDimensionSourceCol.items():
                                        if dimName in factProduced.dimensionsUsed:
                                            paramColsUsed.add(dimSource)
                                    for dimName in factProduced.dimensionsUsed:
                                        if dimName in factDimensionPropGrpCol:
                                            paramColsUsed.add(factDimensionPropGrpCol[dimName])

                                unmappedParamCols = (paramColsWithValue | rowPropGrpParamRefs | reportedDimensionsColumns) - paramColsUsed - emptyCols
                                if unmappedParamCols:
                                    error("xbrlce:unmappedCellValue",
                                          _("Table %(table)s row %(row)s unmapped parameter columns %(columns)s, url: %(url)s"),
                                          table=tableId, row=rowIndex+1, columns=", ".join(sorted(unmappedParamCols)), url=tableUrl)
                                unmappedPropGrps = rowPropGroups.keys() - rowPropGroupsUsed
                                if unmappedPropGrps:
                                    error("xbrlce:unmappedCellValue",
                                          _("Table %(table)s row %(row)s unmapped property group columns %(columns)s, url: %(url)s"),
                                          table=tableId, row=rowIndex+1, columns=", ".join(sorted(unmappedPropGrps)), url=tableUrl)
                                if emptyHeaderColsWithValue:
                                    error("xbrlce:unmappedCellValue",
                                          _("Table %(table)s row %(row)s empty-header columns with unmapped values in columns %(columns)s, url: %(url)s"),
                                          table=tableId, row=rowIndex+1, columns=", ".join(str(c) for c in emptyHeaderColsWithValue), url=tableUrl)

                    except UnicodeDecodeError as ex:
                        raise OIMException("{}:invalidJSON".format(errPrefix),
                              _("File MUST use utf-8 encoding: %(file)s, error %(error)s"),
                              file=tablePath, error=str(ex))
                    tableWb = None # dereference
                    _rowIterator = None # dereference
                    if _file is not None:
                        _file.close()

            factItems = csvFacts()

        currentAction = "identifying default dimensions"
        if modelXbrl is not None:
            ValidateXbrlDimensions.loadDimensionDefaults(modelXbrl) # needs dimension defaults

        currentAction = "validating OIM"

        # create the instance document
        currentAction = "creating instance document"
        # relativize taxonomyRefs to base where feasible
        txBase = os.path.dirname(documentBase or (modelXbrl.entryLoadingUrl if modelXbrl else ""))
        for i, tUrl in enumerate(taxonomyRefs or ()):
            if not isAbsoluteUri(tUrl) and os.path.isabs(tUrl) and not isAbsoluteUri(txBase) and os.path.isabs(txBase):
                taxonomyRefs[i] = os.path.relpath(tUrl, txBase)
        prevErrLen = len(modelXbrl.errors) # track any xbrl validation errors
        if modelXbrl: # pull loader implementation
            modelXbrl.blockDpmDBrecursion = True
            modelXbrl.modelDocument = _return = createModelDocument(
                  modelXbrl,
                  Type.INSTANCE,
                  instanceFileName,
                  schemaRefs=taxonomyRefs,
                  isEntry=True,
                  initialComment="extracted from OIM {}".format(mappedUri),
                  documentEncoding="utf-8",
                  base=documentBase or modelXbrl.entryLoadingUrl)
            modelXbrl.modelDocument.inDTS = True
        else: # API implementation
            modelXbrl = ModelXbrl.create(
                cntlr.modelManager,
                Type.INSTANCE,
                instanceFileName,
                schemaRefs=taxonomyRefs,
                isEntry=True,
                initialComment="extracted from OIM {}".format(mappedUri),
                base=documentBase)
            _return = modelXbrl.modelDocument
        if len(modelXbrl.errors) > prevErrLen:
            error("oime:invalidTaxonomy",
                  _("Unable to obtain a valid taxonomy from URLs provided"),
                  modelObject=modelXbrl)

        currentAction = "identifying default dimensions"
        if modelXbrl is not None:
            ValidateXbrlDimensions.loadDimensionDefaults(modelXbrl) # needs dimension defaults

        # validate statically defined templates
        if isCSVorXL:
            currentAction = "checking statically defined dimensions in CSV templates"
            prevErrLen = len(modelXbrl.errors) # track any xbrl validation errors
            reportParametersUsed = set()

            def checkIdentifier(identifier, *pathSegs):
                if not IdentifierPattern.match(identifier):
                    error("xbrlce:invalidIdentifier",
                          _("Invalid identifier: %(identifier)s at %(path)s"),
                          sourceFileLine=oimFile, identifier=identifier, path="/".join(pathSegs))
                    return False
                return True # identifier is ok

            def checkSQName(sqname, *pathSegs):
                if not SQNamePattern.match(sqname):
                    error("oimce:invalidSQName",
                          _("Invalid SQName: %(sqname)s"),
                          sourceFileLine=oimFile, sqname=sqname, path="/".join(pathSegs))
                    return False
                return True # SQName is ok

            def checkDim(tblTmpl, dimName, dimValue, *pathSegs):
                if dimValue is not None:
                    if isinstance(dimValue,str) and dimValue.startswith("#") and not SpecialValuePattern.match(dimValue):
                        error("xbrlce:unknownSpecialValue",
                              _("Unknown special value: %(value)s at %(path)s"),
                              modelObject=modelXbrl, value=dimValue, path="/".join(pathSegs+(dimName,)))
                    elif dimValue == "#nil" and ":" not in dimName and dimName not in ("concept", "period", "value", "entity", "unit"):
                        error("xbrlce:invalidJSONStructure",
                              _("Invalid value: %(value)s at %(path)s"),
                              modelObject=modelXbrl, value=dimValue, path="/".join(pathSegs+(dimName,)))
                    elif isinstance(dimValue,str) and dimValue.startswith("$") and not dimValue.startswith("$$"):
                        paramName, _sep, periodSpecifier = dimValue[1:].partition("@")
                        if _sep and periodSpecifier not in ("start", "end"):
                            error("xbrlce:invalidPeriodSpecifier",
                                  _("Parameter period-specifier invalid: %(periodSpecifier)s at %(path)s"),
                                  periodSpecifier=periodSpecifier, path="/".join(pathSegs+(dimName,)))
                        if not IdentifierPattern.match(paramName):
                            error("xbrlce:invalidReference",
                                  _("Parameter reference invalid: %(target)s at %(path)s"),
                                  target=paramName, path="/".join(pathSegs+(dimName,)))
                        reportParametersUsed.add(paramName)
                        if tblTmpl:
                            tblTmpl.setdefault("_parametersUsed",set()).add(paramName)
                    elif not (isinstance(dimValue,str) and dimValue.startswith("$")):
                        if dimName == "concept":
                            if dimValue != "#none":
                                if not isinstance(dimValue,str) or ":" not in dimValue or not QNamePattern.match(dimValue): # allow #nil
                                    error("xbrlce:invalidConceptQName",
                                          _("Concept does not match lexical QName pattern: %(concept)s at %(path)s"),
                                          modelObject=modelXbrl, concept=dimValue, path="/".join(pathSegs+(dimName,)))
                                else:
                                    conceptQn = qname(dimValue, namespaces)
                                    if conceptQn is None: # bad prefix
                                        error("oimce:unboundPrefix",
                                              _("The QName prefix could not be resolved with available namespaces: %(concept)s at %(path)s"),
                                              modelObject=modelXbrl, concept=dimValue, path="/".join(pathSegs+(dimName,)))
                                    elif conceptQn.localName != "note" or conceptQn.namespaceURI not in nsOims:
                                        concept = modelXbrl.qnameConcepts.get(conceptQn)
                                        if concept is None:
                                            error("oime:unknownConcept",
                                                  _("The concept QName could not be resolved with available DTS: %(concept)s at %(path)s"),
                                                  modelObject=modelXbrl, concept=dimValue, path="/".join(pathSegs+(dimName,)))
                                        elif concept.isItem and concept.isAbstract:
                                            error("oime:valueForAbstractConcept",
                                                  _("Value provided for abstract concept by %(concept)s at %(path)s"),
                                                  modelObject=modelXbrl, concept=dimValue, path="/".join(pathSegs+(dimName,)))
                                        elif ((concept.instanceOfType(UNSUPPORTED_DATA_TYPES) and not concept.instanceOfType(dtrSQNameNamesItemTypes))
                                              or concept.isTuple):
                                            error("oime:unsupportedConceptDataType",
                                                  _("Concept has unsupported data type, %(dataType)s: %(concept)s at %(path)s"),
                                                  modelObject=modelXbrl, concept=dimValue, dataType=concept.typeQname, path="/".join(pathSegs+(dimName,)))
                        elif dimName == "unit":
                            if dimValue == "xbrli:pure":
                                error("oime:illegalPureUnit",
                                      _("Unit MUST NOT have single numerator measure xbrli:pure with no denominators: %(unit)s at %(path)s"),
                                      modelObject=modelXbrl, unit=dimValue, path="/".join(pathSegs+(dimName,)))
                            elif dimValue != "#none" and not UnitPattern.match( PrefixedQName.sub(UnitPrefixedQNameSubstitutionChar, dimValue) ):
                                error("oimce:invalidUnitStringRepresentation",
                                      _("Unit string representation is lexically invalid, %(unit)s at %(path)s"),
                                      modelObject=modelXbrl, unit=dimValue, path="/".join(pathSegs+(dimName,)))
                        elif dimName == "entity":
                            if dimValue != "#none":
                                checkSQName(dimValue or "", *(pathSegs+(dimName,)) )
                                dimQname = qname(dimValue, namespaces)
                                if dimQname == entityNaQName:
                                    error("oime:invalidUseOfReservedIdentifier",
                                          _("The entity core dimension MUST NOT have a scheme of 'https://xbrl.org/.../entities' with an identifier of 'NA': %(entity)s at %(path)s"),
                                          modelObject=modelXbrl, entity=dimQname, path="/".join(pathSegs+(dimName,)))
                        elif dimName == "period":
                            if dimValue != "#none" and not PeriodPattern.match(csvPeriod(dimValue) or ""):
                                error("xbrlce:invalidPeriodRepresentation",
                                      _("The period has lexically invalid dateTime %(period)s at %(path)s"),
                                      modelObject=modelXbrl, period=dimValue, path="/".join(pathSegs+(dimName,)))
                        elif dimName == "language":
                            if dimValue != "#none" and not languagePattern.match(dimValue or ""):
                                error("xbrlce:invalidLanguageCode",
                                      _("The language is lexically invalid %(language)s at %(path)s"),
                                      modelObject=modelXbrl, language=dimValue, path="/".join(pathSegs+(dimName,)))
                        elif dimName == "decimals":
                            if dimValue != "#none" and not isinstance(dimValue,int) and not integerPattern.match(str(dimValue) or ""):
                                error("xbrlce:invalidDecimalsValue",
                                      _("Decimals is lexically invalid %(language)s at %(path)s"),
                                      modelObject=modelXbrl, language=dimValue, path="/".join(pathSegs+(dimName,)))
                        elif dimName == "xbrl:noteId":
                            error("xbrlce:invalidJSONStructure",
                                  _("NoteId dimension must not be explicitly defined at %(path)s"),
                                  modelObject=modelXbrl, qname=dimName, path="/".join(pathSegs+(dimName,)))
                        elif dimName.startswith("xbrl:"):
                            error("xbrlce:invalidJSONStructure",
                                  _("Taxonomy-defined dimension must not have xbrl prefix: %(qname)s at %(path)s"),
                                  modelObject=modelXbrl, qname=dimName, path="/".join(pathSegs+(dimName,)))
                        elif ":" in dimName: # taxonomy defined dimension
                            dimQname = qname(dimName, namespaces)
                            dimConcept = modelXbrl.qnameConcepts.get(dimQname)
                            if dimConcept is None:
                                error("oime:unknownDimension",
                                      _("Taxonomy-defined dimension QName not be resolved with available DTS: %(qname)s at %(path)s"),
                                      modelObject=modelXbrl, qname=dimQname, path="/".join(pathSegs+(dimName,)))
                            elif dimConcept.isExplicitDimension:
                                mem = qname(dimValue, namespaces)
                                if mem is None:
                                    error("{}:invalidDimensionValue".format(valErrPrefix),
                                          _("Taxonomy-defined explicit dimension value is invalid: %(memberQName)s at %(path)s"),
                                          modelObject=modelXbrl, memberQName=dimValue, path="/".join(pathSegs+(dimName,)))
                            elif dimConcept.isTypedDimension:
                                # a modelObject xml element is needed for all of the instance functions to manage the typed dim
                                _type = dimConcept.typedDomainElement.type
                                if (_type is not None and
                                    _type.qname != qnXbrliDateItemType and
                                    (_type.localName in ("complexType", "union", "list", "ENTITY", "ENTITIES", "ID", "IDREF", "IDREFS", "NMTOKEN", "NMTOKENS", "NOTATION")
                                     or _type.isDerivedFrom(dtrPrefixedContentTypes))):
                                    error("oime:unsupportedDimensionDataType",
                                          _("Taxonomy-defined typed dimension value is complex: %(memberQName)s at %(path)s"),
                                          modelObject=modelXbrl, memberQName=dimValue, path="/".join(pathSegs+(dimName,)))
                if pathSegs[-1] in ("/dimensions", "dimensions") and not DimensionsKeyPattern.match(dimName):
                    error("oimce:invalidSQName",
                          _("Invalid SQName: %(sqname)s"),
                          sourceFileLine=oimFile, sqname=dimName, path="/".join(pathSegs))

            # check reportParameterNames
            for reportParameterName in reportParameters.keys():
                checkIdentifier(reportParameterName, "/parameters")
            for dimName, dimValue in reportDimensions.items():
                checkDim(None, dimName, dimValue, "/dimensions")
            checkDim(None, "decimals", reportDecimals, "/")

            # check table template statically defined dimensions, regardless of use
            for tblTmplId, tblTmpl in tableTemplates.items():
                checkIdentifier(tblTmplId, "/tableTemplates")
                propertyGroupCols = set()
                columns = tblTmpl.get("columns",EMPTY_DICT)
                for columnId, column in columns.items():
                    checkIdentifier(columnId, "/tableTemplates", tblTmplId, "columns", columnId)
                    if "propertyGroups" in column:
                        propertyGroupCols.add(columnId)
                    isCommentColumn = column.get("comment") == True
                    isFactColumn = "dimensions" in column or "propertiesFrom" in column
                    isPropertyGroupColumn = "propertyGroups" in column
                    if (isPropertyGroupColumn and isFactColumn) or (isCommentColumn and (isPropertyGroupColumn or isFactColumn)):
                        error("xbrlce:conflictingColumnType",
                              _("Conflicting column type at %(path)s"),
                              path="/tableTemplates/{}/columns/{}".format(tblTmplId, columnId))
                    if not isFactColumn and "decimals" in column:
                        error("xbrlce:misplacedDecimalsOnNonFactColumn",
                              _("Column has decimals on a non-fact column at %(path)s"),
                              path="/tableTemplates/{}/columns/{}".format(tblTmplId, columnId))

                for dimName, dimValue in tblTmpl.get("dimensions",EMPTY_DICT).items():
                    checkDim(tblTmpl, dimName, dimValue, "/tableTemplates", tblTmplId, "dimensions")
                checkDim(tblTmpl, "decimals", tblTmpl.get("decimals",None), "/tableTemplates", tblTmplId)
                for columnId, column in columns.items():
                    for dimName, dimValue in column.get("dimensions",EMPTY_DICT).items():
                        checkDim(tblTmpl, dimName, dimValue, "/tableTemplates", tblTmplId, "columns", columnId, "dimensions")
                    checkDim(tblTmpl, "decimals", column.get("decimals",None), "/tableTemplates", tblTmplId, "columns", columnId)
                    for propGrpName, propGrp in column.get("propertyGroups",EMPTY_DICT).items():
                        checkIdentifier(propGrpName, "/tableTemplates", tblTmplId, "columns", columnId, "propertyGroups", propGrpName)
                        for dimName, dimValue in propGrp.get("dimensions",EMPTY_DICT).items():
                            checkDim(tblTmpl, dimName, dimValue, "/tableTemplates", tblTmplId, "columns", columnId, "propertyGroups", propGrpName, "dimensions")
                        checkDim(tblTmpl, "decimals", propGrp.get("decimals",None), "/tableTemplates", tblTmplId, "columns", columnId, "propertyGroups", propGrpName)
                    decPGs = set()
                    dimPGs = defaultdict(set)
                    for propertyFrom in column.get("propertiesFrom",()):
                        if propertyFrom not in propertyGroupCols:
                            error("xbrlce:invalidPropertyGroupColumnReference",
                                  _("PropertiesFrom value is not a column in table: %(propertyFrom)s at %(path)s"),
                                  modelObject=modelXbrl, propertyFrom=propertyFrom, path="/tableTemplates/{}/columns/{}/propertiesFrom".format(tblTmplId,columnId))
                        else:
                            for propGrp in columns[propertyFrom].get("propertyGroups",EMPTY_DICT).values():
                                if "decimals" in propGrp:
                                    decPGs.add(propertyFrom)
                                for dim in propGrp.get("dimensions",EMPTY_DICT).keys():
                                    dimPGs[dim].add(propertyFrom)
                    if len(decPGs) > 1:
                        error("xbrlce:repeatedPropertyGroupDecimalsProperty",
                              _("PropertiesFrom references repeat decimals property: %(propFroms)s at %(path)s."),
                              propFroms=", ".join(decPGs), path="/tableTemplates/{}/columns/{}/propertiesFrom".format(tblTmplId,columnId))
                    if any(len(dimCols) > 1 for dimCols in dimPGs.values()):
                        error("xbrlce:repeatedPropertyGroupDimension",
                              _("PropertiesFrom references repeat dimensions from: %(propFroms)s, dimension: %(dimensions)s at %(path)s."),
                              propFroms=", ".join(sorted(set(c for d,cs in dimPGs.items() if len(cs) > 1 for c in cs))),
                              dimensions=", ".join(sorted(d for d,cs in dimPGs.items() if len(cs) > 1)),
                              path="/tableTemplates/{}/columns/{}/propertiesFrom".format(tblTmplId,columnId))


                rowIdColName = tblTmpl.get("rowIdColumn")
                if rowIdColName:
                    if rowIdColName not in columns:
                        error("xbrlce:undefinedRowIdColumn",
                              _("RowIdColumn is not defined in columns: %(rowIdColumn)s at %(path)s"),
                              rowIdColumn=rowIdColName, path="/tableTemplates/{}".format(tblTmplId))
                    elif columns[rowIdColName].get("comment") == True:
                        error("xbrlce:invalidRowIdColumn",
                              _("RowIdColumn must not be a comment column: %(rowIdColumn)s at %(path)s"),
                              rowIdColumn=rowIdColName, path="/tableTemplates/{}".format(tblTmplId))

            # table static checks
            for tableId, table in tables.items():
                checkIdentifier(tableId, "/tables")
                tblTmplId = table.get("template", tableId)
                if checkIdentifier(tblTmplId, "/tables/{}/template".format(tableId)) and tblTmplId not in tableTemplates:
                    error("xbrlce:unknownTableTemplate",
                          _("Referenced template is missing: %(tableTemplateId)s at %(path)s"),
                          modelObject=modelXbrl, tableTemplateId=tblTmplId, path="/tables/{}/template".format(tableId))
                tblTmpl = tableTemplates.get(tblTmplId)
                for tblParamName in table.get("parameters", EMPTY_DICT):
                    if not IdentifierPattern.match(tblParamName):
                        error("xbrlce:invalidParameterName",
                              _("Parameter name is not a valid identifier: %(tableParameterName)s at path: %(path)s"),
                              tableParameterName=tblParamName, path="/tables/{}/parameters".format(tableId))
                    # check for table parameter usage by its template
                    if tblTmpl and tblParamName not in tblTmpl.get("_parametersUsed",EMPTY_SET):
                        error("xbrlce:unreferencedParameter",
                              _("Parameter name is not referenced: %(tableParameterName)s at path: %(path)s"),
                              tableParameterName=tblParamName, path="/tables/{}/parameters".format(tableId))

            unreferencedReportParams = reportParameters.keys() - reportParametersUsed
            if unreferencedReportParams:
                error("xbrlce:unreferencedParameter",
                      _("Report parameters not referenced: %(parameters)s"),
                      parameters=", ".join(sorted(unreferencedReportParams)))


            if len(modelXbrl.errors) > prevErrLen:
                return NotOIMException() # no point to going ahead.

        firstCntxUnitFactElt = None

        cntxTbl = {}
        unitTbl = {}
        xbrlNoteTbl = {} # fact ID: note fact
        noteFactIDsNotReferenced = set()

        currentAction = "creating facts"
        factNum = 0 # for synthetic fact number
        if isJSON:
            syntheticFactFormat = "_f{{:0{}}}".format(int(log10(len(factItems) or 1))) #want
        else:
            syntheticFactFormat = "_f{}" #want

        numFactCreationXbrlErrors = 0

        contextElement = getTaxonomyContextElement(modelXbrl)
        for id, fact in factItems:
            factProduced.clear()

            dimensions = fact.get("dimensions", EMPTY_DICT)
            if "concept" not in dimensions:
                error("oime:missingConceptDimension",
                      _("The concept core dimension MUST be present on fact: %(id)s."),
                      modelObject=modelXbrl, id=id)
                continue
            if not id:
                id = syntheticFactFormat.format(factNum)
                factNum += 1
            conceptSQName = dimensions["concept"]
            if conceptSQName is INVALID_REFERENCE_TARGET:
                factProduced.invalidReferenceTarget = "concept"
                continue
            factProduced.dimensionsUsed.add("concept")
            if isCSVorXL and (not isinstance(conceptSQName,str) or ":" not in conceptSQName or not QNamePattern.match(conceptSQName or "")): # allow #nil
                error("xbrlce:invalidConceptQName",
                      _("Concept does not match lexical QName pattern: %(concept)s."),
                      modelObject=modelXbrl, concept=conceptSQName)
                continue
            conceptPrefix = conceptSQName.partition(":")[0]
            if conceptPrefix not in namespaces:
                error("oimce:unboundPrefix",
                      _("The concept QName prefix was not defined in namespaces: %(concept)s."),
                      modelObject=modelXbrl, concept=conceptSQName)
                continue
            conceptQn = qname(conceptSQName, namespaces)
            if conceptQn.localName == "note" and conceptQn.namespaceURI in nsOims:
                xbrlNoteTbl[id] = fact
                if "language" not in dimensions:
                    error("oime:missingLanguageForNoteFact",
                          _("Missing language dimension for footnote fact %(id)s"),
                          modelObject=modelXbrl, id=id)
                else:
                    factProduced.dimensionsUsed.add("language")
                if isCSVorXL:
                    dimensions["noteId"] = id # infer this dimension
                elif "noteId" not in dimensions:
                    error("oime:missingNoteIDDimension",
                          _("Missing noteId dimension for footnote fact %(id)s"),
                          modelObject=modelXbrl, id=id)
                elif dimensions.get("noteId") != id:
                    error("oime:invalidNoteIDValue",
                          _("The noteId dimension value, %(noteId)s, must be the same as footnote fact id, %(id)s"),
                          modelObject=modelXbrl, id=id, noteId=dimensions["noteId"])
                else:
                    noteFactIDsNotReferenced.add(id)
                if dimensions.get("unit") and not isCSVorXL:
                    error("oime:misplacedUnitDimension",
                          _("The unit core dimension MUST NOT be present on footnote fact %(id)s: %(unit)s."),
                          modelObject=modelXbrl, id=id, unit=dimensions.get("unit"))
                if fact.get("decimals") and not isCSVorXL:
                    error("oime:misplacedDecimalsProperty",
                          _("The decimals property MUST NOT be present on footnote fact %(id)s: %(decimals)s"),
                          modelObject=modelXbrl, id=id, decimals=decimals)
                unexpectedDimensions = [d for d in dimensions if d in ("entity", "period") or ":" in d]
                if unexpectedDimensions:
                    error("oime:misplacedNoteFactDimension",
                          _("Unexpected dimension(s) for footnote fact %(id)s: %(dimensions)s"),
                          modelObject=modelXbrl, id=id, dimensions=", ".join(sorted(unexpectedDimensions)))
                try:
                    unacceptableTopElts = set()
                    unacceptablePrefixes = set()
                    valueXhtmlElts = etree.XML(htmlBodyTemplate.format(fact.get("value","")))
                    for elt in valueXhtmlElts.iterchildren():
                        if not elt.tag.startswith(xhtmlTagPrefix):
                            unacceptableTopElts.add(elt.tag)
                    for elt in valueXhtmlElts.iter():
                        if elt.tag.startswith(xhtmlTagPrefix) and elt.prefix:
                            unacceptablePrefixes.add(elt.prefix)
                    if unacceptableTopElts:
                        error("oime:invalidXHTMLFragment",
                              _("xbrl:note MUST have xhtml top level elements in the default xhtml namespace, fact %(id)s, elements %(elements)s"),
                              modelObject=modelXbrl, id=id, elements=", ".join(sorted(unacceptableTopElts)))
                    if unacceptablePrefixes:
                        error("oime:xhtmlElementInNonDefaultNamespace",
                              _("xbrl:note MUST have xhtml elements in the default xhtml namespace, fact %(id)s, non-default prefixes: %(prefixes)s"),
                              modelObject=modelXbrl, id=id, prefixes=", ".join(sorted(unacceptablePrefixes)))
                except (etree.XMLSyntaxError,
                        UnicodeDecodeError) as err:
                    error("oime:invalidXHTMLFragment",
                          _("Xhtml error for footnote fact %(id)s: %(error)s"),
                          modelObject=modelXbrl, id=id, error=str(err))
                continue
            elif "noteId" in dimensions:
                error("oime:misplacedNoteIDDimension",
                      _("Unexpected noteId dimension on non-footnote fact, id %(id)s"),
                      modelObject=modelXbrl, id=id, noteId=dimensions["noteId"])
            concept = modelXbrl.qnameConcepts.get(conceptQn)
            if concept is None:
                error("oime:unknownConcept",
                      _("The concept QName could not be resolved with available DTS: %(concept)s."),
                      modelObject=modelXbrl, concept=conceptQn)
                continue
            attrs = {}
            if ((concept.instanceOfType(UNSUPPORTED_DATA_TYPES) and not concept.instanceOfType(dtrSQNameNamesItemTypes))
                   or concept.isTuple):
                error("oime:unsupportedConceptDataType",
                      _("Concept has unsupported data type, %(value)s: %(concept)s."),
                      modelObject=modelXbrl, concept=conceptSQName, value=fact["value"])
                continue
            elif concept.isItem:
                if concept.isAbstract:
                    error("oime:valueForAbstractConcept",
                          _("Value provided for abstract concept by fact %(factId)s, concept %(concept)s."),
                          modelObject=modelXbrl, factId=id, concept=conceptSQName)
                    continue # skip creating fact because context would be bad
                if "language" in dimensions:
                    lang = dimensions["language"]
                    if lang is INVALID_REFERENCE_TARGET:
                        factProduced.invalidReferenceTarget = "language"
                        continue
                    if concept.type.isOimTextFactType:
                        if isJSON and not lang.islower():
                            error("xbrlje:invalidLanguageCodeCase",
                                  _("Language MUST be lower case: \"%(lang)s\", fact %(factId)s, concept %(concept)s."),
                                  modelObject=modelXbrl, factId=id, concept=conceptSQName, lang=lang)
                        factProduced.dimensionsUsed.add("language")
                        attrs["{http://www.w3.org/XML/1998/namespace}lang"] = lang
                    elif not isCSVorXL:
                        error("oime:misplacedLanguageDimension",
                              _("Language \"%(lang)s\" provided for non-text concept by fact %(factId)s, concept %(concept)s."),
                              modelObject=modelXbrl, factId=id, concept=conceptSQName, lang=lang)
                        continue # skip creating fact because language would be bad
                entityAsQn = entityNaQName
                entitySQName = dimensions.get("entity")
                if entitySQName is INVALID_REFERENCE_TARGET:
                    factProduced.invalidReferenceTarget = "entity"
                    continue
                if entitySQName is not None and entitySQName is not NONE_CELL:
                    factProduced.dimensionsUsed.add("entity")
                    if not SQNamePattern.match(entitySQName):
                        error("oimce:invalidSQName",
                              _("Entity has an invalid value: %(entity)s."),
                              modelObject=modelXbrl, entity=entitySQName)
                        continue
                    entityPrefix = entitySQName.partition(":")[0]
                    if entityPrefix not in namespaces:
                        error("oimce:unboundPrefix",
                              _("Entity QName prefix was not defined in namespaces: %(entity)s."),
                              modelObject=modelXbrl, entity=entitySQName)
                    else:
                        entityAsQn = qname(entitySQName, namespaces)
                        if entityAsQn == entityNaQName:
                            error("oime:invalidUseOfReservedIdentifier",
                                  _("The entity core dimension MUST NOT have a scheme of 'https://xbrl.org/.../entities' with an identifier of 'NA': %(entity)s."),
                                  modelObject=modelXbrl, entity=entitySQName)
                            continue
                if "period" in dimensions:
                    period = dimensions["period"]
                    if period is INVALID_REFERENCE_TARGET:
                        factProduced.invalidReferenceTarget = "period"
                        continue
                    elif period is NONE_CELL:
                        period = "forever"
                    elif period is None or not PeriodPattern.match(period):
                        error("xbrlce:invalidPeriodRepresentation" if isCSVorXL else "oimce:invalidPeriodRepresentation",
                              _("The fact %(factId)s, concept %(element)s has a lexically invalid period dateTime %(periodError)s"),
                              modelObject=modelXbrl, factId=id, element=conceptQn, periodError=period)
                        continue
                    else:
                        factProduced.dimensionsUsed.add("period")
                else:
                    period = "forever"
                cntxKey = ( # hashable context key
                    ("periodType", concept.periodType),
                    ("entity", entityAsQn),
                    ("period", period)) + tuple(sorted(
                        (dimName, dimVal["value"] if isinstance(dimVal,dict) else dimVal)
                        for dimName, dimVal in dimensions.items()
                        if ":" in dimName))
                _start, _sep, _end = period.rpartition('/')
                if period == "forever":
                    _periodType = "forever"
                elif _start == _end or not _start:
                    _periodType = "instant"
                else:
                    _periodType = "duration"
                if concept.periodType == "instant" and _periodType == "forever":
                    error("oime:missingPeriodDimension",
                          _("Missing period for %(periodType)s fact %(factId)s."),
                          modelObject=modelXbrl, factId=id, periodType=concept.periodType, period=period)
                    continue # skip creating fact because context would be bad
                elif ((concept.periodType == "duration" and (_periodType != "forever" and (not _start or _start == _end))) or
                      (concept.periodType == "instant" and _start and _start != _end)):
                    error("oime:invalidPeriodDimension",
                          _("Invalid period for %(periodType)s fact %(factId)s period %(period)s."),
                          modelObject=modelXbrl, factId=id, periodType=concept.periodType, period=period)
                    continue # skip creating fact because context would be bad
                elif cntxKey in cntxTbl:
                    _cntx = cntxTbl[cntxKey]
                    for dimName, dimVal in cntxKey[3:]:
                        factProduced.dimensionsUsed.add(dimName)
                else:
                    cntxId = 'c-{:02}'.format(len(cntxTbl) + 1)
                    qnameDims = {}
                    hasDimErr = False
                    for dimName, dimVal in dimensions.items():
                        if ":" in dimName:
                            if dimVal is INVALID_REFERENCE_TARGET:
                                factProduced.invalidReferenceTarget = dimName
                                hasDimErr = True
                                break
                            factProduced.dimensionsUsed.add(dimName)
                            dimQname = qname(dimName, namespaces)
                            if isJSON and dimQname.namespaceURI in nsOims:
                                error("xbrlje:invalidJSONStructure",
                                      _("Fact %(factId)s taxonomy-defined dimension QName must not be xbrl prefixed: %(qname)s."),
                                      modelObject=modelXbrl, factId=id, qname=dimQname)
                                continue
                            dimConcept = modelXbrl.qnameConcepts.get(dimQname)
                            if dimConcept is None:
                                error("oime:unknownDimension",
                                      _("Fact %(factId)s taxonomy-defined dimension QName not be resolved with available DTS: %(qname)s."),
                                      modelObject=modelXbrl, factId=id, qname=dimQname)
                                continue
                            if dimVal is None:
                                memberAttrs = {"{http://www.w3.org/2001/XMLSchema-instance}nil": "true"}
                            else:
                                memberAttrs = None
                            if isinstance(dimVal, dict):
                                dimVal = dimVal["value"]
                            elif dimVal is not None:
                                dimVal = str(dimVal) # may be int or boolean
                            if dimConcept.isExplicitDimension:
                                mem = qname(dimVal, namespaces)
                                if mem is None:
                                    error("{}:invalidDimensionValue".format(valErrPrefix),
                                          _("Fact %(factId)s taxonomy-defined explicit dimension value is invalid: %(memberQName)s."),
                                          modelObject=modelXbrl, factId=id, memberQName=dimVal)
                                    continue
                                memConcept = modelXbrl.qnameConcepts.get(mem)
                                if memConcept is not None and modelXbrl.dimensionDefaultConcepts.get(dimConcept) == memConcept:
                                    error("{}:invalidDimensionValue".format("oime" if valErrPrefix == "xbrlje" else valErrPrefix),
                                          _("Fact %(factId)s taxonomy-defined explicit dimension value must not be the default member: %(memberQName)s."),
                                          modelObject=modelXbrl, factId=id, memberQName=dimVal)
                                    continue
                            elif dimConcept.isTypedDimension:
                                # a modelObject xml element is needed for all of the instance functions to manage the typed dim
                                if dimConcept.typedDomainElement.baseXsdType in ("ENTITY", "ENTITIES", "ID", "IDREF", "IDREFS", "NMTOKEN", "NMTOKENS", "NOTATION") or (
                                   dimConcept.typedDomainElement.instanceOfType(dtrPrefixedContentTypes) and not dimConcept.typedDomainElement.instanceOfType(dtrSQNameNamesTypes)) or (
                                    dimConcept.typedDomainElement.type is not None and
                                    dimConcept.typedDomainElement.type.qname != XbrlConst.qnXbrliDateUnion and
                                    (dimConcept.typedDomainElement.type.localName == "complexType" or
                                     any(c.localName in ("union","list") for c in dimConcept.typedDomainElement.type.iterchildren()))):
                                    error("oime:unsupportedDimensionDataType",
                                          _("Fact %(factId)s taxonomy-defined typed dimension value is not supported: %(memberQName)s."),
                                          modelObject=modelXbrl, factId=id, memberQName=dimVal)
                                    continue
                                if (canonicalValuesFeature and dimVal is not None and
                                    not CanonicalXmlTypePattern.get(dimConcept.typedDomainElement.baseXsdType, NoCanonicalPattern).match(dimVal)):
                                    error("xbrlje:nonCanonicalValue",
                                          _("Numeric typed dimension must have canonical %(type)s value \"%(value)s\": %(concept)s."),
                                          modelObject=modelXbrl, type=dimConcept.typedDomainElement.baseXsdType, concept=dimConcept, value=dimVal)
                                mem = addChild(modelXbrl.modelDocument, dimConcept.typedDomainElement.qname, text=dimVal, attributes=memberAttrs, appendChild=False)
                            else:
                                mem = None # absent typed dimension
                            if mem is not None:
                                qnameDims[dimQname] = DimValuePrototype(modelXbrl, None, dimQname, mem, contextElement)
                    if hasDimErr:
                        continue
                    try:
                        _start, _sep, _end = period.rpartition('/')
                        if period == "forever":
                            startDateTime = endDateTime = None
                        elif _start == _end or not _start:
                            startDateTime = None
                            endDateTime = dateTime(_end, type=DATETIME)
                        else:
                            startDateTime = dateTime(_start, type=DATETIME)
                            endDateTime = dateTime(_end, type=DATETIME)
                        numFactCreationXbrlErrors -= len(modelXbrl.errors) # track any xbrl validation errors
                        prevErrLen = len(modelXbrl.errors)
                        _cntx = modelXbrl.createContext(
                                                entityAsQn.namespaceURI,
                                                entityAsQn.localName,
                                                _periodType,
                                                startDateTime,
                                                endDateTime,
                                                None, # no dimensional validity checking (like formula does)
                                                qnameDims, [], [],
                                                id=cntxId)
                        if len(modelXbrl.errors) > prevErrLen:
                            if any(err == "xmlSchema:valueError" for err in modelXbrl.errors[prevErrLen:]):
                                error("{}:invalidDimensionValue".format(valErrPrefix),
                                      _("Fact %(factId)s taxonomy-defined dimension value errors noted above."),
                                      modelObject=modelXbrl, factId=id)
                                continue
                            numFactCreationXbrlErrors += sum(err != "xmlSchema:valueError" for err in modelXbrl.errors[prevErrLen:])
                    except ValueError as err:
                        error("xbrlce:invalidPeriodRepresentation" if isCSVorXL else "oimce:invalidPeriodRepresentation",
                              _("Invalid period for fact %(factId)s period %(period)s, %(error)s."),
                              modelObject=modelXbrl, factId=id, period=period, error=err)
                        continue
                    cntxTbl[cntxKey] = _cntx
                    if firstCntxUnitFactElt is None:
                        firstCntxUnitFactElt = _cntx
                unitKey = dimensions.get("unit")
                if concept.isNumeric:
                    if unitKey is INVALID_REFERENCE_TARGET:
                        factProduced.invalidReferenceTarget = "unit"
                        continue
                    if unitKey == "xbrli:pure":
                        error("oime:illegalPureUnit",
                              _("Unit MUST NOT have single numerator measure xbrli:pure with no denominators."),
                              modelObject=modelXbrl, unit=unitKey)
                        continue
                    if unitKey: # not empty cells
                        factProduced.dimensionsUsed.add("unit")
                    if (unitKey or None) in unitTbl: # either None or EMPTY_CELL match None for default pure unit
                        _unit = unitTbl[unitKey or None]
                    else:
                        _unit = None
                        # validate unit
                        if unitKey and not UnitPattern.match( PrefixedQName.sub(UnitPrefixedQNameSubstitutionChar, unitKey) ):
                            error("oimce:invalidUnitStringRepresentation",
                                  _("Unit string representation is lexically invalid, %(unit)s"),
                                  modelObject=modelXbrl, unit=unitKey)
                            continue
                        else:
                            if not unitKey:
                                _muls = [XbrlConst.qnXbrliPure]
                                _divs = []
                                unitKey = None # use None for pure unit key (may be either no value or empty cell value)
                            else:
                                _mul, _sep, _div = unitKey.partition('/')
                                if _mul.startswith('('):
                                    _mul = _mul[1:-1]
                                _muls = [u for u in _mul.split('*') if u]
                                if _div.startswith('('):
                                    _div = _div[1:-1]
                                _divs = [u for u in _div.split('*') if u]
                                if _muls != sorted(_muls) or _divs != sorted(_divs):
                                    error("oimce:invalidUnitStringRepresentation",
                                          _("Unit string representation measures are not in alphabetical order, %(unit)s"),
                                          modelObject=modelXbrl, unit=unitKey)
                            try:
                                mulQns = [qname(u, namespaces, prefixException=OIMException("oimce:unboundPrefix",
                                                                                          _("Unit prefix is not declared: %(unit)s"),
                                                                                          unit=u))
                                          for u in _muls]
                                divQns = [qname(u, namespaces, prefixException=OIMException("oimce:unboundPrefix",
                                                                                          _("Unit prefix is not declared: %(unit)s"),
                                                                                          unit=u))
                                          for u in _divs]
                                unitId = 'u-{:02}'.format(len(unitTbl) + 1)
                                for _measures in mulQns, divQns:
                                    for _measure in _measures:
                                        addQnameValue(modelXbrl.modelDocument, _measure)
                                _unit = modelXbrl.createUnit(mulQns, divQns, id=unitId)
                                if firstCntxUnitFactElt is None:
                                    firstCntxUnitFactElt = _unit
                            except OIMException as ex:
                                error(ex.code, ex.message, modelObject=modelXbrl, **ex.msgArgs)
                        unitTbl[unitKey] = _unit
                else:
                    _unit = None
                    if unitKey is not None and not isCSVorXL:
                        error("oime:misplacedUnitDimension",
                              _("The unit core dimension MUST NOT be present on non-numeric facts: %(concept)s, unit %(unit)s."),
                              modelObject=modelXbrl, concept=conceptSQName, unit=unitKey)

                attrs["contextRef"] = _cntx.id

                if fact.get("value") is None:
                    if not concept.isNillable:
                        error("{}:invalidFactValue".format("oime" if isJSON else valErrPrefix),
                              _("Nil value applied to non-nillable concept: %(concept)s."),
                              modelObject=modelXbrl, concept=conceptSQName)
                        continue
                    attrs[XbrlConst.qnXsiNil] = "true"
                    text = None
                elif concept.isEnumeration2Item:
                    text = fact["value"]
                    if concept.instanceOfType(XbrlConst.qnEnumerationSetItemType2020):
                        if len(text):
                            qnames = text.split(" ")
                        else:
                            qnames = () # empty enumerations set
                    else: # single value may be a QName with whitespaces
                        qnames = (text.strip(),)
                    expandedNames = set()
                    if canonicalValuesFeature and not all(qnames[i] < qnames[i+1] for i in range(len(qnames)-1)):
                        error("xbrlje:nonCanonicalValue",
                              _("Enumeration item must be canonically ordered, %(value)s: %(concept)s."),
                              modelObject=modelXbrl, concept=conceptSQName, value=fact["value"])
                    isFactValid = True
                    for qn in qnames:
                        if not PrefixedQName.match(qn):
                            isFactValid = False
                        else:
                            _qname = qname(qn, namespaces)
                            if not _qname:
                                error("oimce:unboundPrefix",
                                      _("Enumeration item QName prefix was not defined in namespaces, %(qname)s: %(concept)s."),
                                      modelObject=modelXbrl, concept=conceptSQName, qname=qn)
                                continue
                            else:
                                expandedNames.add(_qname.expandedName)
                    if isFactValid:
                        text = " ".join(sorted(expandedNames))
                    else:
                        error("{}:invalidFactValue".format(valErrPrefix),
                              _("Enumeration item must be %(canonicalOrdered)slist of QNames: %(concept)s."),
                              modelObject=modelXbrl, concept=conceptSQName, canonicalOrdered="a canonical ordered " if canonicalValuesFeature else "")
                        continue
                else:
                    text = fact["value"]
                    if (canonicalValuesFeature and text is not None and
                        not CanonicalXmlTypePattern.get(concept.baseXsdType, NoCanonicalPattern).match(text)):
                        error("xbrlje:nonCanonicalValue",
                              _("Item must have canonical %(type)s value \"%(value)s\": %(concept)s."),
                              modelObject=modelXbrl, type=concept.baseXsdType, concept=conceptSQName, value=text)

                decimals = fact.get("decimals")
                if concept.isNumeric:
                    if isCSVorXL and isinstance(text, str): # don't check for suffix if not CSV/XL or None or int
                        _number, _sep, _decimals = text.partition("d")
                        if _sep:
                            if decimalsSuffixPattern.match(text):
                                decimals = _decimals.strip()
                                text = _number.strip()
                            else:
                                error("xbrlce:invalidDecimalsSuffix",
                                      _("Fact %(factId)s has invalid decimals \"%(decimals)s\""),
                                      modelObject=modelXbrl, factId=id, decimals=_sep+_decimals)
                                continue # skip processing this fact
                        elif decimals is not None:
                            if decimals is INVALID_REFERENCE_TARGET:
                                factProduced.invalidReferenceTarget = "decimals"
                                continue
                            factProduced.dimensionsUsed.add("decimals")
                    if _unit is None:
                        continue # skip creating fact because unit was invalid
                    attrs["unitRef"] = _unit.id
                    if text is not None: # no decimals for nil value
                        attrs["decimals"] = decimals if decimals is not None else "INF"
                    elif decimals is not None:
                        error("oime:misplacedDecimalsProperty",
                              _("The decimals property MUST NOT be present on nil facts: %(concept)s, decimals %(decimals)s"),
                              modelObject=modelXbrl, concept=conceptSQName, decimals=decimals)
                        continue
                elif decimals is not None and not isCSVorXL:
                    # includes nil facts for JSON (but not CSV)
                    error("oime:misplacedDecimalsProperty",
                          _("The decimals property MUST NOT be present on non-numeric facts: %(concept)s, decimals %(decimals)s"),
                          modelObject=modelXbrl, concept=conceptSQName, decimals=decimals)
            else:
                text = None #tuple

            attrs["id"] = id
            if "id" not in fact: # needed for footnote generation
                fact["id"] = id

            # is value a QName?
            if concept.baseXbrliType == "QNameItemType": # renormalize prefix of instance fact
                text = addQnameValue(modelXbrl.modelDocument, qname(text.strip(), namespaces))

            prevErrLen = len(modelXbrl.errors) # track any xbrl validation errors
            factProduced.modelFact = f = modelXbrl.createFact(conceptQn, attributes=attrs, text=text, validate=False)
            if firstCntxUnitFactElt is None:
                firstCntxUnitFactElt = f

            xmlValidate(modelXbrl, f)
            if len(modelXbrl.errors) > prevErrLen:
                numFactCreationXbrlErrors += sum(err != "xmlSchema:valueError" for err in modelXbrl.errors[prevErrLen:])
                if any(err == "xmlSchema:valueError" for err in modelXbrl.errors[prevErrLen:]):
                    error("{}:invalidFactValue".format(valErrPrefix),
                          _("Fact %(factId)s value error noted above."),
                          modelObject=modelXbrl, factId=id)

        currentAction = "creating footnotes"
        footnoteLinks = OrderedDict() # ELR elements
        factLocs = {} # index by (linkrole, factId)
        footnoteLinkNotes = defaultdict(set) # linkrole: noteIds
        # footnoteNbr = 0
        locNbr = 0
        definedInstanceRoles = set()
        undefinedFootnoteTypes = set()
        undefinedFootnoteGroups = set()
        undefinedLinkTargets = set()
        undefinedLinkSources = set() # csv only
        footnotesIdTargets = set()
        for factOrFootnote in footnotes:
            if isJSON:
                for ftGroups in factOrFootnote.get("links", {}).values():
                    for ftTgtIds in ftGroups.values():
                        for tgtId in ftTgtIds:
                            if tgtId in xbrlNoteTbl:
                                footnotesIdTargets.add(tgtId)
            elif isCSVorXL: # footnotes contains footnote objects
                for ftGroups in factOrFootnote.values():
                    for ftSrcIdTgtIds in ftGroups.values():
                        for ftTgtIds in ftSrcIdTgtIds.values():
                            for tgtId in ftTgtIds:
                                footnotesIdTargets.add(tgtId)
        for factOrFootnote in footnotes:
            if isJSON:
                factFootnotes = []
                for ftType, ftGroups in factOrFootnote.get("links", {}).items():
                    ftSrcId = factOrFootnote.get("id")
                    if ftSrcId is None:
                        ftSrcId = factOrFootnote.get("dimensions",EMPTY_DICT).get("xbrl:noteId")
                    if ftType not in linkTypes:
                        undefinedFootnoteTypes.add(ftType)
                    else:
                        for ftGroup, ftTgtIds in ftGroups.items():
                            if ftGroup not in linkGroups:
                                undefinedFootnoteGroups.add(ftGroup)
                            else:
                                footnote = {"id": ftSrcId,
                                            "footnoteGroup": linkGroups[ftGroup],
                                            "footnoteType": linkTypes[ftType]}
                                for tgtId in ftTgtIds:
                                    if tgtId in xbrlNoteTbl:
                                        footnote.setdefault("noteRefs", []).append(tgtId)
                                        noteFactIDsNotReferenced.discard(tgtId)
                                    elif tgtId in modelXbrl.modelDocument.idObjects:
                                        footnote.setdefault("factRefs", []).append(tgtId)
                                    else:
                                        undefinedLinkTargets.add(tgtId)
                                factFootnotes.append(footnote)
            elif isCSVorXL: # footnotes contains footnote objects
                factFootnotes = []
                for ftType, ftGroups in factOrFootnote.items():
                    if ftType not in linkTypes:
                        undefinedFootnoteTypes.add(ftType)
                    else:
                        for ftGroup, ftSrcIdTgtIds in ftGroups.items():
                            if ftGroup not in linkGroups:
                                undefinedFootnoteGroups.add(ftGroup)
                            else:
                                for ftSrcId, ftTgtIds in ftSrcIdTgtIds.items():
                                    footnote = {"id": ftSrcId,
                                                "footnoteGroup": linkGroups[ftGroup],
                                                "footnoteType": linkTypes[ftType]}
                                    if ftSrcId not in modelXbrl.modelDocument.idObjects:
                                        undefinedLinkSources.add(ftSrcId)
                                    for tgtId in ftTgtIds:
                                        if tgtId in xbrlNoteTbl:
                                            footnote.setdefault("noteRefs", []).append(tgtId)
                                            noteFactIDsNotReferenced.discard(tgtId)
                                        elif tgtId in modelXbrl.modelDocument.idObjects:
                                            footnote.setdefault("factRefs", []).append(tgtId)
                                        else:
                                            undefinedLinkTargets.add(tgtId)
                                    factFootnotes.append(footnote)
            for footnote in factFootnotes:
                factId = footnote["id"]
                linkrole = footnote["footnoteGroup"]
                arcrole = footnote["footnoteType"]
                skipThisFootnote = False
                if not factId or not linkrole or not arcrole or not (
                    footnote.get("factRefs") or footnote.get("footnote") is not None or footnote.get("noteRefs") is not None):
                    if not linkrole:
                        warning("xbrlje:unknownLinkGroup",
                                        _("FootnoteId has no linkrole %(footnoteId)s."),
                                        modelObject=modelXbrl, footnoteId=footnote.get("footnoteId"))
                    if not arcrole:
                        warning("xbrlje:unknownLinkType",
                                        _("FootnoteId has no arcrole %(footnoteId)s."),
                                        modelObject=modelXbrl, footnoteId=footnote.get("footnoteId"))
                    continue
                for refType, refValue, roleTypes, lrrRoles in (("role", linkrole, modelXbrl.roleTypes, lrrRoleHrefs),
                                                               ("arcrole", arcrole, modelXbrl.arcroleTypes, lrrArcroleHrefs)):
                    if not (XbrlConst.isStandardRole(refValue) or XbrlConst.isStandardArcrole(refValue)):
                        if refValue not in definedInstanceRoles:
                            if refValue in roleTypes or refValue in lrrRoles:
                                definedInstanceRoles.add(refValue)
                                if refValue in roleTypes:
                                    hrefElt = roleTypes[refValue][0]
                                    href = hrefElt.modelDocument.uri + "#" + hrefElt.id
                                else:
                                    href = lrrRoles[refValue]
                                elt = addChild(modelXbrl.modelDocument.xmlRootElement,
                                               qname(link, refType+"Ref"),
                                               attributes=(("{http://www.w3.org/1999/xlink}href", href),
                                                           ("{http://www.w3.org/1999/xlink}type", "simple")),
                                               beforeSibling=firstCntxUnitFactElt)
                                href = modelXbrl.modelDocument.discoverHref(elt)
                                if href:
                                    _elt, hrefDoc, hrefId = href
                                    _defElt = hrefDoc.idObjects.get(hrefId)
                                    if _defElt is not None:
                                        _uriAttrName = refType + "URI"
                                        _uriAttrValue = _defElt.get(_uriAttrName)
                                        if _uriAttrValue:
                                            elt.set(_uriAttrName, _uriAttrValue)
                            else:
                                error("xbrlxe:nonStandardRoleDefinitionNotInDTS",
                                      _("Footnote %(sourceId)s %(roleType)s %(role)s not defined in DTS"),
                                      modelObject=modelXbrl, sourceId=factId, roleType=refType, role=refValue)
                                skipThisFootnote = True
                if skipThisFootnote:
                    continue
                if linkrole not in footnoteLinks:
                    footnoteLinks[linkrole] = addChild(modelXbrl.modelDocument.xmlRootElement,
                                                       XbrlConst.qnLinkFootnoteLink,
                                                       attributes={"{http://www.w3.org/1999/xlink}type": "extended",
                                                                   "{http://www.w3.org/1999/xlink}role": linkrole})
                footnoteLink = footnoteLinks[linkrole]
                factIDs = (factId,)
                if factId in xbrlNoteTbl: # factId is a note, not a fact
                    fromLabel = "f_{}".format(factId)
                    factLocs[(linkrole, factIDs)] = fromLabel
                    noteFactIDsNotReferenced.discard(factId)
                elif (linkrole, factIDs) not in factLocs:
                    locNbr += 1
                    locLabel = "l_{:02}".format(locNbr)
                    factLocs[(linkrole, factIDs)] = locLabel
                    addChild(footnoteLink, XbrlConst.qnLinkLoc,
                             attributes={XLINKTYPE: "locator",
                                         XLINKHREF: "#" + factId,
                                         XLINKLABEL: locLabel})
                locFromLabel = factLocs[(linkrole, factIDs)]
                if "noteRefs" in footnote:
                    # footnoteNbr += 1
                    # footnoteToLabel = "f_{:02}".format(footnoteNbr)
                    for noteId in footnote.get("noteRefs"):
                        footnoteToLabel = "f_{}".format(noteId)
                        noteFactIDsNotReferenced.discard(noteId)
                        if noteId not in footnoteLinkNotes[linkrole]:
                            footnoteLinkNotes[linkrole].add(noteId)
                            xbrlNote = xbrlNoteTbl[noteId]
                            attrs = {XLINKTYPE: "resource",
                                     XLINKLABEL: footnoteToLabel,
                                     "id": idDeduped(modelXbrl, noteId),
                                     # "oimNoteId": noteId
                                     }
                            #if noteId in footnotesIdTargets: # footnote resource is target of another footnote loc
                            #    attrs["id"] = noteId
                            try:
                                if "dimensions" in xbrlNote:
                                    attrs[XMLLANG] = xbrlNote["dimensions"]["language"]
                                elif "aspects" in xbrlNote:
                                    attrs[XMLLANG] = xbrlNote["aspects"]["language"]
                            except KeyError:
                                pass
                            tgtElt = addChild(footnoteLink, XbrlConst.qnLinkFootnote, attributes=attrs)
                            srcElt = etree.fromstring("<footnote xmlns=\"http://www.w3.org/1999/xhtml\">{}</footnote>"
                                                      .format(xbrlNote["value"]), parser=modelXbrl.modelDocument.parser)
                            if srcElt.__len__() > 0: # has html children
                                setXmlns(modelXbrl.modelDocument, "xhtml", "http://www.w3.org/1999/xhtml")
                            copyIxFootnoteHtml(srcElt, tgtElt, withText=True, isContinChainElt=False)
                            xmlValidate(modelXbrl, tgtElt)
                        footnoteArc = addChild(footnoteLink,
                                               XbrlConst.qnLinkFootnoteArc,
                                               attributes={XLINKTYPE: "arc",
                                                           XLINKARCROLE: arcrole,
                                                           XLINKFROM: locFromLabel,
                                                           XLINKTO: footnoteToLabel})
                if "factRefs" in footnote:
                    factRef = footnote.get("factRefs")
                    fact2IDs = tuple(sorted(factRef))
                    if (linkrole, fact2IDs) not in factLocs:
                        locNbr += 1
                        locLabel = "l_{:02}".format(locNbr)
                        factLocs[(linkrole, fact2IDs)] = locLabel
                        for factId in fact2IDs:
                            addChild(footnoteLink, XbrlConst.qnLinkLoc,
                                     attributes={XLINKTYPE: "locator",
                                                 XLINKHREF: "#" + factId,
                                                 XLINKLABEL: locLabel})
                    footnoteToLabel = factLocs[(linkrole, fact2IDs)]
                    footnoteArc = addChild(footnoteLink,
                                           XbrlConst.qnLinkFootnoteArc,
                                           attributes={XLINKTYPE: "arc",
                                                       XLINKARCROLE: arcrole,
                                                       XLINKFROM: locFromLabel,
                                                       XLINKTO: footnoteToLabel})
                    if arcrole == factFootnote:
                        error("oime:illegalStandardFootnoteTarget",
                              _("Standard footnote %(sourceId)s targets must be an xbrl:note, targets %(targetIds)s."),
                              modelObject=modelXbrl, sourceId=factId, targetIds=", ".join(fact2IDs))
        if noteFactIDsNotReferenced:
            error("oime:unusedNoteFact",
                    _("Note facts MUST be referenced by at least one link group, IDs: %(noteFactIds)s."),
                    modelObject=modelXbrl, noteFactIds=", ".join(sorted(noteFactIDsNotReferenced)))
        if footnoteLinks:
            modelXbrl.modelDocument.linkbaseDiscover(footnoteLinks.values(), inInstance=True)

        if undefinedLinkSources:
            error("{}:unknownLinkSource".format(errPrefix),
                  _("These link sources are not defined in facts: %(ftTargets)s."),
                  modelObject=modelXbrl, ftTargets=", ".join(sorted(undefinedLinkSources)))
        if undefinedLinkTargets:
            error("{}:unknownLinkTarget".format(errPrefix),
                  _("These link targets are not defined in facts: %(ftTargets)s."),
                  modelObject=modelXbrl, ftTargets=", ".join(sorted(undefinedLinkTargets)))
        if undefinedFootnoteTypes:
            error("{}:unknownLinkType".format(errPrefix),
                  _("These footnote types are not defined in footnoteTypes: %(ftTypes)s."),
                  modelObject=modelXbrl, ftTypes=", ".join(sorted(undefinedFootnoteTypes)))
        if undefinedFootnoteGroups:
            error("{}:unknownLinkGroup".format(errPrefix),
                  _("These footnote groups are not defined in footnoteGroups: %(ftGroups)s."),
                  modelObject=modelXbrl, ftGroups=", ".join(sorted(undefinedFootnoteGroups)))

        currentAction = "checking for duplicates"
        checkForDuplicates(modelXbrl, allowedDuplicatesFeature, footnotesIdTargets)

        currentAction = "done loading facts and footnotes"

        if numFactCreationXbrlErrors:
            error("oime:invalidXBRL",
                  _("%(count)s XBRL errors noted above."),
                  modelObject=modelXbrl, count=numFactCreationXbrlErrors)


        #cntlr.addToLog("Completed in {0:.2} secs".format(time.time() - startedAt),
        #               messageCode="loadFromExcel:info")
    except NotOIMException as ex:
        _return = ex # not an OIM document
    except Exception as ex:
        _return = ex
        if isinstance(ex, OIMException):
            if ex.code and ex.message:
                error(ex.code, ex.message, modelObject=modelXbrl, **ex.msgArgs)
        else:
            error("arelleOIMloader:error",
                    "Error while %(action)s, error %(error)s\n traceback %(traceback)s",
                    modelObject=modelXbrl, action=currentAction, error=ex,
                    traceback=traceback.format_tb(sys.exc_info()[2]))

    global lastFilePath, lastFilePath
    lastFilePath = None
    lastFilePathIsOIM = False
    return _return

lastFilePath = None
lastFilePathIsOIM = False

def isOimLoadable(modelXbrl, mappedUri, normalizedUri, filepath, **kwargs):
    global lastFilePath, lastFilePathIsOIM
    lastFilePath = filepath
    lastFilePathIsOIM = False
    _ext = os.path.splitext(filepath)[1]
    if _ext in (".csv", ".json", ".xlsx", ".xls"):
        lastFilePathIsOIM = True
    elif isHttpUrl(normalizedUri) and '?' in _ext: # query parameters and not .json, may be JSON anyway
        with io.open(filepath, 'rt', encoding='utf-8') as f:
            _fileStart = f.read(4096)
        if _fileStart and re_match(r"\s*\{\s*\"documentType\":\s*\"http:\\+/\\+/www.xbrl.org\\+/WGWD\\+/YYYY-MM-DD\\+/xbrl-json\"", _fileStart):
            lastFilePathIsOIM = True
    return lastFilePathIsOIM

def oimLoader(modelXbrl, mappedUri, filepath, *args, **kwargs):
    if filepath != lastFilePath or not lastFilePathIsOIM:
        return None # not an OIM file

    cntlr = modelXbrl.modelManager.cntlr
    cntlr.showStatus(_("Loading OIM file: {0}").format(os.path.basename(filepath)))
    doc = loadFromOIM(cntlr, modelXbrl.error, modelXbrl.warning, modelXbrl, filepath, mappedUri)
    if doc is None:
        return None # not an OIM file
    modelXbrl.loadedFromOIM = True
    modelXbrl.loadedFromOimErrorCount = len(modelXbrl.errors)
    return doc

def guiXbrlLoaded(cntlr, modelXbrl, attach, *args, **kwargs):
    if cntlr.hasGui and getattr(modelXbrl, "loadedFromOIM", False):
        from arelle import ModelDocument
        from tkinter.filedialog import askdirectory
        instanceFile = cntlr.uiFileDialog("save",
                title=_("arelle - Save XBRL instance document"),
                initialdir=cntlr.config.setdefault("outputInstanceDir","."),
                filetypes=[(_("XBRL file .xbrl"), "*.xbrl"), (_("XBRL file .xml"), "*.xml")],
                defaultextension=".xbrl")
        if not instanceFile:
            return False
        cntlr.config["outputInstanceDir"] = os.path.dirname(instanceFile)
        cntlr.saveConfig()
        if instanceFile:
            modelXbrl.modelDocument.save(instanceFile, updateFileHistory=False)
            cntlr.showStatus(_("Saving XBRL instance: {0}").format(os.path.basename(instanceFile)))
        cntlr.showStatus(_("OIM loading completed"), 3500)

def cmdLineXbrlLoaded(cntlr, options, modelXbrl, *args, **kwargs):
    if options.saveOIMinstance and getattr(modelXbrl, "loadedFromOIM", False):
        doc = modelXbrl.modelDocument
        cntlr.showStatus(_("Saving XBRL instance: {0}").format(doc.basename))
        responseZipStream = kwargs.get("responseZipStream")
        if responseZipStream is not None:
            _zip = zipfile.ZipFile(responseZipStream, "a", zipfile.ZIP_DEFLATED, True)
        else:
            _zip = None
        doc.save(options.saveOIMinstance, _zip)
        if responseZipStream is not None:
            _zip.close()
            responseZipStream.seek(0)

def validateFinally(val, *args, **kwargs):
    modelXbrl = val.modelXbrl
    if getattr(modelXbrl, "loadedFromOIM", False):
        if modelXbrl.loadedFromOimErrorCount < len(modelXbrl.errors):
            modelXbrl.error("oime:invalidTaxonomy",
                                _("XBRL validation errors were logged for this instance."),
                                modelObject=modelXbrl)
    else:
        # validate xBRL-XML instances
        fractionFacts = []
        tupleFacts = []
        precisionZeroFacts = []
        contextsInUse = set()
        for f in modelXbrl.factsInInstance: # facts in document order (no sorting required for messages)
            concept = f.concept
            if concept is not None:
                if concept.isFraction:
                    fractionFacts.append(f)
                elif concept.isTuple:
                    tupleFacts.append(f)
                elif concept.isNumeric:
                    if f.precision is not None and precisionZeroPattern.match(f.precision):
                        precisionZeroFacts.append(f)
            context = f.context
            if context is not None:
                contextsInUse.add(context)
        if fractionFacts:
            modelXbrl.error("xbrlxe:unsupportedFraction", # this pertains only to xBRL-XML validation (JSON and CSV were checked during loading when loadedFromOIM is True)
                            _("Instance has %(count)s facts with fraction facts"),
                            modelObject=fractionFacts, count=len(fractionFacts))
        if tupleFacts:
            modelXbrl.error("xbrlxe:unsupportedTuple",
                            _("Instance has %(count)s tuple facts"),
                            modelObject=tupleFacts, count=len(tupleFacts))
        if precisionZeroFacts:
            modelXbrl.error("xbrlxe:unsupportedZeroPrecisionFact",
                            _("Instance has %(count)s precision zero facts"),
                            modelObject=precisionZeroFacts, count=len(precisionZeroFacts))
        containers = {"segment", "scenario"}
        dimContainers = set(t for c in contextsInUse for t in containers if c.dimValues(t))
        if len(dimContainers) > 1:
            modelXbrl.error("xbrlxe:inconsistentDimensionsContainer",
                            _("All hypercubes within the DTS of a report MUST be defined for use on the same container (either \"segment\" or \"scenario\")"),
                            modelObject=modelXbrl)
        contextsWithNonDimContent = set()
        contextsWithComplexTypedDimensions = set()
        for context in contextsInUse:
            if context.nonDimValues("segment"):
                contextsWithNonDimContent.add(context)
            if context.nonDimValues("scenario"):
                contextsWithNonDimContent.add(context)
            for modelDimension in context.qnameDims.values():
                if modelDimension.isTyped:
                    typedMember = modelDimension.typedMember
                    if isinstance(typedMember, ModelObject):
                        modelConcept = modelXbrl.qnameConcepts.get(typedMember.qname)
                        if modelConcept is not None and modelConcept.type is not None and modelConcept.type.localName == "complexType":
                            contextsWithComplexTypedDimensions.add(context)
        if contextsWithNonDimContent:
            modelXbrl.error("xbrlxe:nonDimensionalSegmentScenarioContent",
                            _("Contexts MUST not contain non-dimensional content: %(contexts)s"),
                            modelObject=contextsWithNonDimContent,
                            contexts=", ".join(sorted(c.id for c in contextsWithNonDimContent)))
        if contextsWithComplexTypedDimensions:
            modelXbrl.error("xbrlxe:unsupportedComplexTypedDimension",  # this pertains only to xBRL-XML validation (JSON and CSV were checked during loading when loadedFromOIM is True)
                            _("Instance has contexts with complex typed dimensions: %(contexts)s"),
                            modelObject=contextsWithComplexTypedDimensions,
                            contexts=", ".join(sorted(c.id for c in contextsWithComplexTypedDimensions)))

        footnoteRels = modelXbrl.relationshipSet("XBRL-footnotes")
        # ext group and link roles
        unsupportedExtRoleRefs = defaultdict(list) # role/arcrole and footnote relationship objects referencing it
        footnoteELRs = set()
        footnoteArcroles = set()
        roleDefiningDocs = defaultdict(set)
        def docInSchemaRefedDTS(thisdoc, roleTypeDoc, visited=None):
            if visited is None:
                visited = set()
            visited.add(thisdoc)
            for doc, docRef in thisdoc.referencesDocument.items():
                if thisdoc.type != Type.INSTANCE or docRef.referringModelObject.qname not in nonDiscoveringXmlInstanceElements:
                    if doc == roleTypeDoc or (doc not in visited and docInSchemaRefedDTS(doc, roleTypeDoc, visited)):
                        return True
            visited.remove(thisdoc)
            return False
        for rel in footnoteRels.modelRelationships:
            if not isStandardRole(rel.linkrole):
                footnoteELRs.add(rel.linkrole)
            if rel.arcrole != factFootnote:
                footnoteArcroles.add(rel.arcrole)
        for elr in footnoteELRs:
            for roleType in modelXbrl.roleTypes[elr]:
                roleDefiningDocs[elr].add(roleType.modelDocument)
        for arcrole in footnoteArcroles:
            for arcroleType in modelXbrl.arcroleTypes[arcrole]:
                roleDefiningDocs[arcrole].add(arcroleType.modelDocument)
        extRoles = set(role
                      for role, docs in roleDefiningDocs.items()
                      if not any(docInSchemaRefedDTS(modelXbrl.modelDocument, doc) for doc in docs))
        if extRoles:
            modelXbrl.error("xbrlxe:unsupportedExternalRoleRef",
                            _("Role and arcrole definitions MUST be in standard or schemaRef discoverable sources"),
                            modelObject=modelXbrl, roles=", ".join(sorted(extRoles)))

        # todo: multi-document inline instances
        for elt in modelXbrl.modelDocument.xmlRootElement.iter("{http://www.xbrl.org/2003/linkbase}footnote", "{http://www.xbrl.org/2013/inlineXBRL}footnote"):
            if isinstance(elt, ModelResource) and getattr(elt, "xValid", 0) >= VALID:
                if not footnoteRels.toModelObject(elt):
                    modelXbrl.error("xbrlxe:unlinkedFootnoteResource",
                                    _("Unlinked footnote element %(label)s: %(value)s"),
                                    modelObject=elt, label=elt.xlinkLabel, value=elt.xValue[:100])
                if elt.role not in (None, "", footnote):
                    modelXbrl.error("xbrlxe:nonStandardFootnoteResourceRole",
                                    _("Footnotes MUST have standard footnote resource role, %(role)s is disallowed, %(label)s: %(value)s"),
                                    modelObject=elt, role=elt.role, label=elt.xlinkLabel, value=elt.xValue[:100])
        # xml base on anything
        for elt in modelXbrl.modelDocument.xmlRootElement.getroottree().iterfind("//{*}*[@{http://www.w3.org/XML/1998/namespace}base]"):
            modelXbrl.error("xbrlxe:unsupportedXmlBase",
                            _("Instance MUST NOT contain xml:base attributes: element %(qname)s, xml:base %(base)s"),
                            modelObject=elt, qname=elt.qname if isinstance(elt, ModelObject) else elt.tag,
                            base=elt.get("{http://www.w3.org/XML/1998/namespace}base"))
        # todo: multi-document inline instances
        if modelXbrl.modelDocument.type in (ModelDocument.Type.INSTANCE, ModelDocument.Type.INLINEXBRL, ModelDocument.Type.INLINEXBRLDOCUMENTSET):
            for doc in modelXbrl.modelDocument.referencesDocument.keys():
                if doc.type == Type.LINKBASE:
                    val.modelXbrl.error("xbrlxe:unsupportedLinkbaseReference",
                                        _("Linkbase reference not allowed from instance document."),
                                        modelObject=(modelXbrl.modelDocument,doc))

def excelLoaderOptionExtender(parser, *args, **kwargs):
    parser.add_option("--saveOIMinstance",
                      action="store",
                      dest="saveOIMinstance",
                      help=_("Save a instance loaded from OIM into this file name."))

__pluginInfo__ = {
    'name': 'Load From OIM',
    'version': '1.2',
    'description': "This plug-in loads XBRL instance data from OIM (JSON, CSV or Excel) and saves the resulting XBRL Instance.",
    'license': 'Apache-2',
    'author': authorLabel,
    'copyright': copyrightLabel,
    # classes of mount points (required)
    'ModelDocument.IsPullLoadable': isOimLoadable,
    'ModelDocument.PullLoader': oimLoader,
    'CntlrWinMain.Xbrl.Loaded': guiXbrlLoaded,
    'CntlrCmdLine.Options': excelLoaderOptionExtender,
    'CntlrCmdLine.Xbrl.Loaded': cmdLineXbrlLoaded,
    'Validate.XBRL.Finally': validateFinally
}
