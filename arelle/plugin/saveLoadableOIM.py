"""
See COPYRIGHT.md for copyright information.

## Overview

The Save Loadable OIM plugin facilitates the saving of re-loadable instances in JSON or CSV formats, adhering to the
Open Information Model (OIM) XBRL Specification. It is designed to work seamlessly with the Load From OIM plugin,
allowing for efficient data handling in Arelle.

## Key Features

- **Multiple Formats**: Enables saving data in JSON and CSV OIM formats as well as XLSX.
- **Seamless Integration**: Compatible with the Load From OIM plugin for saving and loading reports.
- **GUI and CLI Compatibility**: Available for use in both GUI and CLI modes.
- **Test Case Augmentation**: Allows saving OIM files with a custom suffix during test suite runs.

## Usage Instructions

### Command Line Usage

- **Single-Instance Mode**:
  Save a file in single-instance mode by specifying the file path and extension:
  ```bash
  python arelleCmdLine.py --plugins saveLoadableOIM --file filing-documents.zip --saveLoadableOIM example.json
  ```

- **Test Case Operation**:
  Augment test case operations by specifying a suffix for the read-me-first file in a test suite:
  ```bash
  python arelleCmdLine.py --plugins saveLoadableOIM --file filing-documents.zip --saveTestcaseOimFileSuffix -savedOim.csv
  ```

- **Deduplicate facts**
  To save an OIM instance with duplicate fact removed use the `--deduplicateOimFacts` argument with either `complete`,
  `consistent-pairs`, or `consistent-sets` as the value.
  For details on what eaxctly consitutes a duplicate fact and why there are multiple options read the
  [Fact Deduplication][fact-deduplication] documentation.
  ```bash
  python arelleCmdLine.py --plugins saveLoadableOIM --file filing-documents.zip --saveLoadableOIM example.json --deduplicateOimFacts complete
  ```

[fact-deduplication]: project:/user_guides/fact_deduplication.md

### GUI Usage

- **Save Re-Loadable Output**:
  1. Load the desired report in Arelle.
  2. Go to `Tools` > `Save Loadable OIM`.
  3. Specify a filename and choose the desired file format (JSON, CSV, or XLSX).

## Additional Notes

If the loaded report refers to a taxonomy on the local file system, the OIM instance needs to be saved in the same
directory to maintain the validity of the relative file system path to the taxonomy.
"""

from __future__ import annotations

import csv
import io
import json
import operator
import os
import threading
import zipfile
from collections.abc import Iterable
from dataclasses import dataclass
from datetime import datetime
from decimal import Decimal
from math import isinf, isnan
from numbers import Number
from optparse import OptionParser
from pathlib import Path
from typing import TYPE_CHECKING, Any, BinaryIO, Callable, Optional, cast

import regex as re
from openpyxl import Workbook
from openpyxl.cell.cell import WriteOnlyCell
from openpyxl.styles import Alignment, Color, PatternFill, fills
from openpyxl.worksheet.dimensions import ColumnDimension

from arelle import ModelDocument, ValidateDuplicateFacts, XbrlConst
from arelle.ModelInstanceObject import ModelContext, ModelFact
from arelle.ModelRelationshipSet import ModelRelationshipSet
from arelle.ModelValue import (
    DateTime,
    DayTimeDuration,
    IsoDuration,
    QName,
    Time,
    YearMonthDuration,
    gDay,
    gMonth,
    gMonthDay,
    gYear,
    gYearMonth,
    qname,
    tzinfoStr,
)
from arelle.typing import TypeGetText
from arelle.UrlUtil import relativeUri
from arelle.utils.PluginData import PluginData
from arelle.utils.PluginHooks import PluginHooks
from arelle.ValidateXbrlCalcs import inferredDecimals
from arelle.Version import authorLabel, copyrightLabel

if TYPE_CHECKING:
    from tkinter import Menu

    from arelle.Cntlr import Cntlr
    from arelle.CntlrCmdLine import CntlrCmdLine
    from arelle.CntlrWinMain import CntlrWinMain
    from arelle.ModelXbrl import ModelXbrl
    from arelle.RuntimeOptions import RuntimeOptions

_: TypeGetText

PLUGIN_NAME = "Save Loadable OIM"

oimErrorPattern = re.compile("oime|oimce|xbrlje|xbrlce")
nsOim = "https://xbrl.org/2021"
qnOimConceptAspect = qname("concept", noPrefixIsNoNamespace=True)
qnOimLangAspect = qname("language", noPrefixIsNoNamespace=True)
qnOimPeriodAspect = qname("period", noPrefixIsNoNamespace=True)
qnOimEntityAspect = qname("entity", noPrefixIsNoNamespace=True)
qnOimUnitAspect = qname("unit", noPrefixIsNoNamespace=True)

reservedUriAliases = {
    nsOim: "xbrl",
    XbrlConst.defaultLinkRole: "_",
    XbrlConst.factExplanatoryFact: "explanatoryFact",
    XbrlConst.factFootnote: "footnote",
    XbrlConst.iso4217: "iso4217",
    XbrlConst.utr: "utr",
    XbrlConst.xbrli: "xbrli",
    XbrlConst.xsd: "xs",
}

ONE = Decimal(1)
TEN = Decimal(10)
NILVALUE = "nil"
SCHEMA_LB_REFS = {
    qname("{http://www.xbrl.org/2003/linkbase}schemaRef"),
    # qname("{http://www.xbrl.org/2003/linkbase}linkbaseRef")
}
ROLE_REFS = {
    qname("{http://www.xbrl.org/2003/linkbase}roleRef"),
    qname("{http://www.xbrl.org/2003/linkbase}arcroleRef"),
}
ENTITY_NA_QNAME = ("https://xbrl.org/entities", "NA")
csvOpenMode = "w"
csvOpenNewline = ""

OimFact = dict[str, Any]
OimReport = dict[str, Any]

class NamespacePrefixes:
    def __init__(self, prefixesByNamespace: dict[str, str] | None = None) -> None:
        self._prefixesByNamespace: dict[str, str] = prefixesByNamespace or {}
        self._usedPrefixes: set[str] = set(self._prefixesByNamespace.values())

    @property
    def namespaces(self) -> dict[str, str]:
        return {
            prefix: namespace
            for namespace, prefix in sorted(
                self._prefixesByNamespace.items(),
                key=operator.itemgetter(1)
            )
        }

    def __contains__(self, namespace: str) -> bool:
        return namespace in self._prefixesByNamespace

    def getPrefix(self, namespace: str) -> str | None:
        return self._prefixesByNamespace.get(namespace)

    def addNamespace(self, namespace: str, preferredPrefix: str) -> str:
        prefix = self._prefixesByNamespace.get(namespace)
        if prefix is not None:
            return prefix

        prefix = reservedUriAliases.get(namespace)
        if prefix is None:
            prefix = preferredPrefix
            i = 2
            while prefix in self._usedPrefixes:
                prefix = f"{preferredPrefix}{i}"
                i += 1
        self._prefixesByNamespace[namespace] = prefix
        self._usedPrefixes.add(prefix)
        return prefix


def saveLoadableOIM(
    modelXbrl: ModelXbrl,
    oimFile: str,
    outputZip: zipfile.ZipFile | None = None,
    # arguments to add extension features to OIM document
    extensionPrefixes: dict[str, str] | None = None,
    extensionReportObjects: dict[str, Any] | None = None,
    extensionFactPropertiesMethod: Callable[[ModelFact, OimFact], None] | None = None,
    extensionReportFinalizeMethod: Callable[[OimReport], None] | None = None,
    *args: Any,
    **kwargs: Any,
) -> None:
    isJSON = oimFile.endswith(".json")
    isCSV = oimFile.endswith(".csv")
    isXL = oimFile.endswith(".xlsx")
    isCSVorXL = isCSV or isXL
    if not isJSON and not isCSVorXL:
        oimFile = oimFile + ".json"
        isJSON = True

    namespacePrefixes = NamespacePrefixes({nsOim: "xbrl"})
    if extensionPrefixes:
        for extensionPrefix, extensionNamespace in extensionPrefixes.items():
            namespacePrefixes.addNamespace(extensionNamespace, extensionPrefix)
    linkTypeAliases = {}
    groupAliases = {}

    def compileQname(qname: QName) -> None:
        if qname.namespaceURI is not None and qname.namespaceURI not in namespacePrefixes:
            namespacePrefixes.addNamespace(qname.namespaceURI, qname.prefix or "")

    aspectsDefined = {qnOimConceptAspect, qnOimEntityAspect, qnOimPeriodAspect}

    def oimValue(obj: Any) -> Any:
        if isinstance(obj, list):
            # set-valued enumeration fact
            return " ".join([oimValue(o) for o in obj])
        if isinstance(obj, QName) and obj.namespaceURI is not None:
            if obj.namespaceURI not in namespacePrefixes:
                namespacePrefixes.addNamespace(obj.namespaceURI, obj.prefix or "_")
            return f"{namespacePrefixes.getPrefix(obj.namespaceURI)}:{obj.localName}"
        if isinstance(obj, (float, Decimal)):
            try:
                if isinf(obj):
                    return "-INF" if obj < 0 else "INF"
                elif isnan(obj):
                    return "NaN"
                elif isinstance(obj, Decimal):
                    # XML canonical representation of decimal requires a decimal point.
                    # https://www.w3.org/TR/xmlschema-2/#decimal-canonical-representation
                    if obj % 1 == 0:
                        return f"{obj:.1f}"
                    intPart, fracPart = f"{obj:f}".split(".")
                    canonicalFracPart = fracPart.rstrip("0") or "0"
                    return f"{intPart}.{canonicalFracPart}"
                else:
                    return f"{obj}"
            except Exception:
                return str(obj)
        if isinstance(obj, bool):
            return "true" if obj else "false"
        if isinstance(
            obj,
            (
                DateTime,
                YearMonthDuration,
                DayTimeDuration,
                Time,
                gYearMonth,
                gMonthDay,
                gYear,
                gMonth,
                gDay,
                IsoDuration,
                int,
            ),
        ):
            return str(obj)
        return obj

    def oimPeriodValue(cntx: ModelContext) -> dict[str, str]:
        if cntx.isStartEndPeriod:
            s = cntx.startDatetime
            e = cntx.endDatetime
        else:  # instant
            s = e = cntx.instantDatetime
        assert isinstance(s, datetime) and isinstance(e, datetime)
        return {
            str(qnOimPeriodAspect): (
                f"{s.year:04}-{s.month:02}-{s.day:02}T{s.hour:02}:{s.minute:02}:{s.second:02}{tzinfoStr(s)}/"
                if cntx.isStartEndPeriod
                else ""
            )
            + f"{e.year:04}-{e.month:02}-{e.day:02}T{e.hour:02}:{e.minute:02}:{e.second:02}{tzinfoStr(e)}"
        }

    hasTuple = False
    hasLang = False
    hasUnits = False
    hasNumeric = False

    footnotesRelationshipSet = ModelRelationshipSet(modelXbrl, "XBRL-footnotes")

    # compile QNames in instance for OIM
    for fact in modelXbrl.factsInInstance:
        concept = fact.concept
        if concept is not None:
            if concept.isNumeric:
                hasNumeric = True
            if concept.baseXbrliType in ("string", "normalizedString", "token") and fact.xmlLang:
                hasLang = True
        compileQname(fact.qname)
        if hasattr(fact, "xValue") and isinstance(fact.xValue, QName):
            compileQname(fact.xValue)
        unit = fact.unit
        if unit is not None:
            hasUnits = True
        if fact.modelTupleFacts:
            hasTuple = True
    if hasTuple:
        modelXbrl.error(
            "arelleOIMsaver:tuplesNotAllowed", "Tuples are not allowed in an OIM document", modelObject=modelXbrl
        )
        return

    entitySchemePrefixes = {}
    for cntx in modelXbrl.contexts.values():
        if cntx.entityIdentifierElement is not None:
            scheme = cntx.entityIdentifier[0]
            if scheme not in entitySchemePrefixes:
                if not entitySchemePrefixes:  # first one is just scheme
                    if scheme == "http://www.sec.gov/CIK":
                        _schemePrefix = "cik"
                    elif scheme == "http://standard.iso.org/iso/17442":
                        _schemePrefix = "lei"
                    else:
                        _schemePrefix = "scheme"
                else:
                    _schemePrefix = f"scheme{len(entitySchemePrefixes) + 1}"
                namespacePrefixes.addNamespace(scheme, _schemePrefix)
                entitySchemePrefixes[scheme] = namespacePrefixes.getPrefix(scheme)
        for dim in cntx.qnameDims.values():
            compileQname(dim.dimensionQname)
            aspectsDefined.add(dim.dimensionQname)
            if dim.isExplicit:
                compileQname(dim.memberQname)

    for unit in modelXbrl.units.values():
        if unit is not None:
            for measures in unit.measures:
                for measure in measures:
                    compileQname(measure)

    if hasLang:
        aspectsDefined.add(qnOimLangAspect)
    if hasUnits:
        aspectsDefined.add(qnOimUnitAspect)

    for footnoteRel in footnotesRelationshipSet.modelRelationships:
        if footnoteRel.arcrole not in linkTypeAliases:
            typePrefix = reservedUriAliases.get(footnoteRel.arcrole, f"ftTyp_{os.path.basename(footnoteRel.arcrole)}")
            linkTypeAliases[footnoteRel.arcrole] = typePrefix
        if footnoteRel.linkrole not in groupAliases:
            groupPrefix = reservedUriAliases.get(footnoteRel.linkrole, f"ftGrp_{os.path.basename(footnoteRel.linkrole)}")
            groupAliases[footnoteRel.linkrole] = groupPrefix

    dtsReferences = set()
    assert modelXbrl.modelDocument is not None
    baseUrl = modelXbrl.modelDocument.uri.partition("#")[0]
    for doc, ref in sorted(modelXbrl.modelDocument.referencesDocument.items(), key=lambda _item: _item[0].uri):
        if ref.referringModelObject.qname in SCHEMA_LB_REFS:
            dtsReferences.add(relativeUri(baseUrl, doc.uri))
    for refType in ("role", "arcrole"):
        for refElt in sorted(
            modelXbrl.modelDocument.xmlRootElement.iterchildren(f"{{http://www.xbrl.org/2003/linkbase}}{refType}Ref"),
            key=lambda elt: elt.get(refType + "URI"),
        ):
            dtsReferences.add(refElt.get("{http://www.w3.org/1999/xlink}href").partition("#")[0])
    dtsReferences = sorted(dtsReferences)  # turn into list
    footnoteFacts = set()

    def factFootnotes(
        fact: ModelFact,
        oimFact: dict[str, Any] | None = None,
        csvLinks: dict[str, Any] | None = None,
    ):
        footnotes = []
        oimLinks = {}
        if isCSVorXL and csvLinks is not None:
            oimLinks = csvLinks
        for footnoteRel in footnotesRelationshipSet.fromModelObject(fact):
            srcId = fact.id if fact.id else f"f{fact.objectIndex}"
            toObj = footnoteRel.toModelObject
            # json
            typePrefix = linkTypeAliases[footnoteRel.arcrole]
            groupPrefix = groupAliases[footnoteRel.linkrole]
            if typePrefix not in oimLinks:
                oimLinks[typePrefix] = {}
            _link = oimLinks[typePrefix]
            if groupPrefix not in _link:
                if isJSON:
                    _link[groupPrefix] = []
                elif isCSVorXL:
                    _link[groupPrefix] = {}
            tgtId = toObj.id if toObj.id else f"f{toObj.objectIndex}"
            if isJSON:
                tgtIdList = _link[groupPrefix]
                tgtIdList.append(tgtId)
            elif isCSVorXL:
                # Footnote links in xBRL-CSV include the CSV table identifier.
                tgtIdList = _link[groupPrefix].setdefault(f"facts.r_{srcId}.value", [])
                tgtIdList.append(f"footnotes.r_{tgtId}.footnote")
            footnote = {
                "group": footnoteRel.linkrole,
                "footnoteType": footnoteRel.arcrole,
            }
            if isinstance(toObj, ModelFact):
                footnote["factRef"] = tgtId
            else:  # text footnotes
                footnote["id"] = tgtId
                footnote["footnote"] = toObj.viewText()
                if toObj.xmlLang:
                    footnote["language"] = toObj.xmlLang
                footnoteFacts.add(toObj)
                footnotes.append(footnote)
        if oimLinks:
            _links = {
                typePrefix: {groupPrefix: idList for groupPrefix, idList in sorted(groups.items())}
                for typePrefix, groups in sorted(oimLinks.items())
            }
            if isJSON and oimFact is not None:
                oimFact["links"] = _links

        return footnotes

    def factAspects(fact: ModelFact) -> dict[str, Any]:
        oimFact = {}
        aspects = {}
        if isCSVorXL:
            oimFact["id"] = fact.id or f"f{fact.objectIndex}"
        concept = fact.concept
        if concept is not None:
            aspects[str(qnOimConceptAspect)] = oimValue(concept.qname)
            if concept.type is not None and concept.type.isOimTextFactType and fact.xmlLang:
                aspects[str(qnOimLangAspect)] = fact.xmlLang
        if fact.isItem:
            if fact.isNil:
                _value = None
            else:
                _inferredDecimals = inferredDecimals(fact)
                _value = oimValue(fact.xValue)
            oimFact["value"] = _value
            if fact.concept is not None and fact.concept.isNumeric:
                _numValue = fact.xValue
                if isinstance(_numValue, Decimal) and not isinf(_numValue) and not isnan(_numValue):
                    _numValue = int(_numValue) if _numValue == _numValue.to_integral() else float(_numValue)
                if not fact.isNil and not isinf(_inferredDecimals):  # accuracy omitted if infinite
                    oimFact["decimals"] = _inferredDecimals
        oimFact["dimensions"] = aspects
        cntx = fact.context
        if cntx is not None:
            if cntx.entityIdentifierElement is not None and cntx.entityIdentifier != ENTITY_NA_QNAME:
                aspects[str(qnOimEntityAspect)] = oimValue(qname(*cntx.entityIdentifier))
            if cntx.period is not None and not cntx.isForeverPeriod:
                aspects.update(oimPeriodValue(cntx))
            for _qn, dim in sorted(cntx.qnameDims.items(), key=lambda item: item[0]):
                if dim.isExplicit:
                    dimVal = oimValue(dim.memberQname)
                else:  # typed
                    if dim.typedMember.get("{http://www.w3.org/2001/XMLSchema-instance}nil") in ("true", "1"):
                        dimVal = None
                    else:
                        dimVal = dim.typedMember.stringValue
                aspects[str(dim.dimensionQname)] = dimVal
        unit = fact.unit
        if unit is not None:
            _mMul, _mDiv = unit.measures
            _sMul = "*".join(oimValue(m) for m in sorted(_mMul, key=lambda m: oimValue(m)))
            if _mDiv:
                _sDiv = "*".join(oimValue(m) for m in sorted(_mDiv, key=lambda m: oimValue(m)))
                if len(_mDiv) > 1:
                    _sUnit = f"({_sMul})/({_sDiv})" if len(_mMul) > 1 else f"{_sMul}/({_sDiv})"
                else:
                    _sUnit = f"({_sMul})/{_sDiv}" if len(_mMul) > 1 else f"{_sMul}/{_sDiv}"
            else:
                _sUnit = _sMul
            if _sUnit != "xbrli:pure":
                aspects[str(qnOimUnitAspect)] = _sUnit

        if isJSON:
            factFootnotes(fact, oimFact=oimFact)
        return oimFact

    # common metadata
    oimReport = {}  # top level of oim json output
    oimReport["documentInfo"] = oimDocInfo = {}
    oimDocInfo["documentType"] = nsOim + ("/xbrl-json" if isJSON else "/xbrl-csv")
    if isJSON:
        oimDocInfo["features"] = oimFeatures = {}
    oimDocInfo["namespaces"] = namespacePrefixes.namespaces
    if linkTypeAliases:
        oimDocInfo["linkTypes"] = {a: u for u, a in sorted(linkTypeAliases.items(), key=operator.itemgetter(1))}
    if groupAliases:
        oimDocInfo["linkGroups"] = {a: u for u, a in sorted(groupAliases.items(), key=operator.itemgetter(1))}
    oimDocInfo["taxonomy"] = dtsReferences
    if isJSON:
        oimFeatures["xbrl:canonicalValues"] = True

    factsToSave = modelXbrl.facts
    pluginData = modelXbrl.modelManager.cntlr.getPluginData(PLUGIN_NAME)
    if isinstance(pluginData, SaveLoadableOIMPluginData) and pluginData.deduplicateFactsType is not None:
        deduplicatedFacts = frozenset(ValidateDuplicateFacts.getDeduplicatedFacts(modelXbrl, pluginData.deduplicateFactsType))
        duplicateFacts = frozenset(f for f in modelXbrl.facts if f not in deduplicatedFacts)
        if duplicateFacts:
            for fact in duplicateFacts:
                ValidateDuplicateFacts.logDeduplicatedFact(modelXbrl, fact)
            factsToSave = [f for f in factsToSave if f not in duplicateFacts]

    if isJSON:
        # save JSON
        oimReport["facts"] = oimFacts = {}
        # add in report level extension objects
        if extensionReportObjects:
            for extObjQName, extObj in extensionReportObjects.items():
                oimReport[extObjQName] = extObj

        def saveJsonFacts(facts: list[ModelFact], oimFacts: dict[str, dict[str, Any]]) -> None:
            for fact in facts:
                oimFact = factAspects(fact)
                # add in fact level extension objects
                if extensionFactPropertiesMethod:
                    extensionFactPropertiesMethod(fact, oimFact)
                id = fact.id if fact.id else f"f{fact.objectIndex}"
                oimFacts[id] = oimFact
                if fact.modelTupleFacts:
                    saveJsonFacts(fact.modelTupleFacts, oimFacts)

        saveJsonFacts(factsToSave, oimFacts)

        # add footnotes as pseudo facts
        for ftObj in footnoteFacts:
            ftId = ftObj.id if ftObj.id else f"f{ftObj.objectIndex}"
            oimFacts[ftId] = oimFact = {}
            oimFact["value"] = ftObj.viewText()
            oimFact["dimensions"] = {
                "concept": "xbrl:note",
                "noteId": ftId,
            }
            if ftObj.xmlLang:
                oimFact["dimensions"]["language"] = ftObj.xmlLang.lower()

        # allow extension report final editing before writing json structure
        # (possible example, reorganize facts into array vs object)
        if extensionReportFinalizeMethod:
            extensionReportFinalizeMethod(oimReport)

        with io.StringIO() if outputZip else open(oimFile, "w", encoding="utf-8") as fh:
            fh.write(json.dumps(oimReport, indent=1))
            if outputZip:
                fh.seek(0)
                outputZip.writestr(os.path.basename(oimFile), fh.read())
        if not outputZip:
            modelXbrl.modelManager.cntlr.showStatus(_("Saved JSON OIM file {}").format(oimFile))

    elif isCSVorXL:
        # save CSV
        oimReport["tables"] = oimTables = {}
        oimReport["tableTemplates"] = csvTableTemplates = {}
        oimTables["facts"] = csvTable = {}
        csvTable["template"] = "facts"
        csvTable["url"] = "tbd"
        csvTableTemplates["facts"] = csvTableTemplate = {}
        csvTableTemplate["rowIdColumn"] = "id"
        csvTableTemplate["dimensions"] = csvTableDimensions = {}
        csvTableTemplate["columns"] = csvFactColumns = {}
        csvLinks = {}
        if footnotesRelationshipSet.modelRelationships:
            oimReport["links"] = csvLinks
        aspectQnCol = {}
        aspectsHeader = []

        def addAspectQnCol(aspectQn: QName | str) -> None:
            colQName = str(aspectQn).replace("xbrl:", "")
            colNCName = colQName.replace(":", "_")
            if colNCName not in aspectQnCol:
                aspectQnCol[colNCName] = len(aspectsHeader)
                aspectsHeader.append(colNCName)
                if colNCName == "value":
                    csvFactColumns[colNCName] = col = {}
                    col["dimensions"] = {}
                elif colNCName == "id":
                    csvFactColumns[colNCName] = {}  # empty object
                elif colNCName == "decimals":
                    csvFactColumns[colNCName] = {}  # empty object
                    csvTableTemplate[colQName] = "$" + colNCName
                else:
                    csvFactColumns[colNCName] = {}  # empty object
                    csvTableDimensions[colQName] = "$" + colNCName

        # pre-ordered aspect columns
        # if hasId:
        addAspectQnCol("id")
        addAspectQnCol(qnOimConceptAspect)
        if hasNumeric:
            addAspectQnCol("decimals")
        if qnOimEntityAspect in aspectsDefined:
            addAspectQnCol(qnOimEntityAspect)
        if qnOimPeriodAspect in aspectsDefined:
            addAspectQnCol(qnOimPeriodAspect)
        if qnOimUnitAspect in aspectsDefined:
            addAspectQnCol(qnOimUnitAspect)
        for aspectQn in sorted(aspectsDefined, key=lambda qn: str(qn)):
            if aspectQn.namespaceURI != nsOim:
                addAspectQnCol(aspectQn)
        addAspectQnCol("value")

        def aspectCols(fact: ModelFact) -> list[Any]:
            cols: list[Any] = [None] * len(aspectsHeader)

            def setColValues(aspects: dict[str, Any]) -> None:
                for aspectQn, aspectValue in aspects.items():
                    colQName = str(aspectQn).replace("xbrl:", "")
                    colNCName = colQName.replace(":", "_")
                    if isinstance(aspectValue, dict):
                        setColValues(aspectValue)
                    elif colNCName in aspectQnCol:
                        if aspectValue is None:
                            _aspectValue = "#nil"
                        elif aspectValue == "":
                            _aspectValue = "#empty"
                        elif isinstance(aspectValue, str) and aspectValue.startswith("#"):
                            _aspectValue = "#" + aspectValue
                        else:
                            _aspectValue = aspectValue
                        cols[aspectQnCol[colNCName]] = _aspectValue

            setColValues(factAspects(fact))
            return cols

        # metadata

        if isCSV:
            if oimFile.endswith("-facts.csv"):  # strip -facts.csv if a prior -facts.csv file was chosen
                _baseURL = oimFile[:-10]
            elif oimFile.endswith(".csv"):
                _baseURL = oimFile[:-4]
            else:
                _baseURL = oimFile
            _csvInfo = {}  # open file, writer

            def _open(filesuffix: str | None, tabname: str, csvTable: dict[str, Any] | None = None) -> None:
                _filename = _baseURL if filesuffix is None else _baseURL + filesuffix
                if csvTable is not None:
                    csvTable["url"] = os.path.basename(_filename)  # located in same directory with metadata
                _csvInfo["file"] = open(_filename, csvOpenMode, newline=csvOpenNewline, encoding="utf-8-sig")
                _csvInfo["writer"] = csv.writer(_csvInfo["file"], dialect="excel")

            def _writerow(row: Iterable[str], header: bool = False) -> None:
                _csvInfo["writer"].writerow(row)

            def _close() -> None:
                _csvInfo["file"].close()
                _csvInfo.clear()
        elif isXL:
            headerWidths = {
                "concept": 40,
                "decimals": 8,
                "language": 9,
                "value": 50,
                "entity": 20,
                "period": 20,
                "unit": 20,
                "metadata": 100,
            }

            # Excel's light orange fill color = 00FF990
            hdrCellFill = PatternFill(patternType=fills.FILL_SOLID, fgColor=Color("00FFBF5F"))
            workbook = Workbook()
            # remove pre-existing worksheets
            while len(workbook.worksheets) > 0:
                workbook.remove(workbook.worksheets[0])
            _xlInfo = {}  # open file, writer

            def _open(filesuffix: str | None, tabname: str, csvTable: dict[str, Any] | None = None) -> None:
                if csvTable is not None:
                    csvTable["url"] = tabname + "!"
                _xlInfo["ws"] = workbook.create_sheet(title=tabname)

            def _writerow(rowvalues: Iterable[str], header: bool = False) -> None:
                row = []
                _ws = _xlInfo["ws"]
                for i, v in enumerate(rowvalues):
                    cell = WriteOnlyCell(_ws, value=v)
                    if header:
                        cell.fill = hdrCellFill
                        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                        colLetter = chr(ord("A") + i)
                        _ws.column_dimensions[colLetter] = ColumnDimension(_ws, customWidth=True)
                        _ws.column_dimensions[colLetter].width = headerWidths.get(v, 40)

                    else:
                        cell.alignment = Alignment(
                            horizontal="right"
                            if isinstance(v, Number)
                            else "center"
                            if isinstance(v, bool)
                            else "left",
                            vertical="top",
                            wrap_text=isinstance(v, str),
                        )
                    row.append(cell)
                _ws.append(row)

            def _close() -> None:
                _xlInfo.clear()

        # save facts
        _open("-facts.csv", "facts", csvTable)
        _writerow(aspectsHeader, header=True)

        def saveCSVfacts(facts: list[ModelFact]) -> None:
            for fact in facts:
                _writerow(aspectCols(fact))
                saveCSVfacts(fact.modelTupleFacts)

        saveCSVfacts(factsToSave)
        _close()

        # save footnotes
        if footnotesRelationshipSet.modelRelationships:
            footnotesDeduplicatedById = {
                footnote["id"]: footnote
                for fact in modelXbrl.facts
                for footnote in factFootnotes(fact, csvLinks=csvLinks)
            }
            footnotes = sorted(footnotesDeduplicatedById.values(), key=operator.itemgetter("id"))
            if footnotes:  # text footnotes
                footnotesTable = {"template": "footnotes"}
                oimTables["footnotes"] = footnotesTable
                csvTableTemplates[footnotesTable["template"]] = {
                    "rowIdColumn" : "id",
                    "dimensions": {
                        "language": "$language"
                    },
                    "columns": {
                        "id": {},
                        "footnote": {
                            "dimensions": {
                                "concept": "xbrl:note",
                            },
                        },
                        "language": {},
                    }
                }
                _open("-footnotes.csv", "footnotes", footnotesTable)
                cols = ("id", "footnote", "language")
                _writerow(cols, header=True)
                for footnote in footnotes:
                    _writerow(tuple(footnote.get(col, "") for col in cols))
                _close()

        # save metadata
        if isCSV:
            assert isinstance(_baseURL, str)
            csvMetadataFile = _baseURL + "-metadata.json"
            with open(csvMetadataFile, "w", encoding="utf-8") as fh:
                fh.write(json.dumps(oimReport, ensure_ascii=False, indent=2, sort_keys=False))
            modelXbrl.modelManager.cntlr.showStatus(_("Saved CSV OIM metadata file {}").format(csvMetadataFile))
        elif isXL:
            _open(None, "metadata")
            _writerow(["metadata"], header=True)
            _writerow([json.dumps(oimReport, ensure_ascii=False, indent=1, sort_keys=False)])
            _close()
            workbook.save(oimFile)
            modelXbrl.modelManager.cntlr.showStatus(_("Saved Excel file {}").format(oimFile))


def saveLoadableOIMMenuCommand(cntlr: CntlrWinMain) -> None:
    # save DTS menu item has been invoked
    if (
        cntlr.config is None
        or cntlr.modelManager is None
        or cntlr.modelManager.modelXbrl is None
        or cntlr.modelManager.modelXbrl.modelDocument is None
        or cntlr.modelManager.modelXbrl.modelDocument.type
        not in (ModelDocument.Type.INSTANCE, ModelDocument.Type.INLINEXBRL, ModelDocument.Type.INLINEXBRLDOCUMENTSET)
    ):
        cntlr.addToLog(
            messageCode="arelleOIMsaver",
            message=_("No supported XBRL instance documents loaded that can be saved to OIM format."),
        )
        return
        # get file name into which to save log file while in foreground thread
    oimFile = cntlr.uiFileDialog(
        "save",
        title=_("arelle - Save Loadable OIM file"),
        initialdir=cntlr.config.setdefault("loadableOIMFileDir", "."),
        filetypes=[(_("JSON file .json"), "*.json"), (_("CSV file .csv"), "*.csv"), (_("XLSX file .xlsx"), "*.xlsx")],
        defaultextension=".json",
    )  # type: ignore[no-untyped-call]
    if not isinstance(oimFile, str):
        # User cancelled file dialog.
        return

    cntlr.config["loadableOIMFileDir"] = os.path.dirname(oimFile)
    cntlr.saveConfig()

    thread = threading.Thread(
        target=lambda _modelXbrl=cntlr.modelManager.modelXbrl, _oimFile=oimFile: saveLoadableOIM(
            _modelXbrl, _oimFile, None
        )
    )
    thread.daemon = True
    thread.start()


def saveOimFiles(
    cntlr: Cntlr,
    modelXbrl: ModelXbrl,
    oimFiles: list[str],
    responseZipStream: BinaryIO | None = None,
) -> None:
    try:
        if responseZipStream is None:
            for oimFile in oimFiles:
                saveLoadableOIM(modelXbrl, oimFile, None)
        else:
            with zipfile.ZipFile(responseZipStream, "a", zipfile.ZIP_DEFLATED, True) as _zip:
                for oimFile in oimFiles:
                    saveLoadableOIM(modelXbrl, oimFile, _zip)
            responseZipStream.seek(0)
    except Exception as ex:
        cntlr.addToLog(f"Exception saving OIM {ex}")


@dataclass
class SaveLoadableOIMPluginData(PluginData):
    deduplicateFactsType: ValidateDuplicateFacts.DeduplicationType | None
    saveTestcaseOimFileSuffix: str | None


class SaveLoadableOIMPlugin(PluginHooks):
    @staticmethod
    def cntlrCmdLineOptions(
        parser: OptionParser,
        *args: Any,
        **kwargs: Any,
    ) -> None:
        parser.add_option(
            "--saveLoadableOIM",
            action="store",
            dest="saveLoadableOIM",
            help=_("Save Loadable OIM file (JSON, CSV or XLSX)"),
        )
        parser.add_option(
            "--saveLoadableOIMDirectory",
            action="store",
            dest="saveLoadableOIMDirectory",
            help=_("Directory to export all supported OIM formats (JSON and CSV)"),
        )
        parser.add_option(
            "--saveTestcaseOIM",
            action="store",
            dest="saveTestcaseOimFileSuffix",
            help=_("Save Testcase Variation OIM file (argument file suffix and type, such as -savedOim.csv"),
        )
        parser.add_option(
            "--deduplicateOimFacts",
            action="store",
            choices=[a.value for a in ValidateDuplicateFacts.DeduplicationType],
            dest="deduplicateOimFacts",
            help=_("Remove duplicate facts when saving the OIM instance"))

    @staticmethod
    def cntlrCmdLineUtilityRun(
        cntlr: Cntlr,
        options: RuntimeOptions,
        *args: Any,
        **kwargs: Any,
    ) -> None:
        deduplicateOimFacts = cast(Optional[str], getattr(options, "deduplicateOimFacts", None))
        saveTestcaseOimFileSuffix = cast(Optional[str], getattr(options, "saveTestcaseOimFileSuffix", None))
        deduplicateFactsType = None
        if deduplicateOimFacts is not None:
            deduplicateFactsType = ValidateDuplicateFacts.DeduplicationType(deduplicateOimFacts)
        pluginData = SaveLoadableOIMPluginData(PLUGIN_NAME, deduplicateFactsType, saveTestcaseOimFileSuffix)
        cntlr.setPluginData(pluginData)

    @staticmethod
    def cntlrWinMainMenuTools(
        cntlr: CntlrWinMain,
        menu: Menu,
        *args: Any,
        **kwargs: Any,
    ) -> None:
        menu.add_command(label="Save Loadable OIM", underline=0, command=lambda: saveLoadableOIMMenuCommand(cntlr))

    @staticmethod
    def cntlrCmdLineXbrlRun(
        cntlr: CntlrCmdLine,
        options: RuntimeOptions,
        modelXbrl: ModelXbrl,
        entrypoint: dict[str, str] | None = None,
        sourceZipStream: BinaryIO | None = None,
        responseZipStream: BinaryIO | None = None,
        *args: Any,
        **kwargs: Any,
    ) -> None:
        # extend XBRL-loaded run processing for this option
        oimFile = cast(Optional[str], getattr(options, "saveLoadableOIM", None))
        allOimDirectory = cast(Optional[str], getattr(options, "saveLoadableOIMDirectory", None))
        if (oimFile or allOimDirectory) and (
            modelXbrl is None
            or modelXbrl.modelDocument is None
            or modelXbrl.modelDocument.type
            not in {
                ModelDocument.Type.INSTANCE,
                ModelDocument.Type.INLINEXBRL,
                ModelDocument.Type.INLINEXBRLDOCUMENTSET,
            }
        ):
            cntlr.addToLog("No XBRL instance has been loaded.")
            return
        if oimFile:
            saveOimFiles(cntlr, modelXbrl, [oimFile], responseZipStream)
        if allOimDirectory:
            oimDir = Path(allOimDirectory)
            try:
                oimDir.mkdir(parents=True, exist_ok=True)
            except OSError as err:
                cntlr.addToLog(
                    _("Unable to save OIM files into requested directory: {}, {}").format(allOimDirectory, err.strerror)
                )
                return
            assert modelXbrl.modelDocument is not None
            basefileStem = None
            if hasattr(modelXbrl, "ixdsTarget"):
                ixdsTarget = modelXbrl.ixdsTarget
                for doc in modelXbrl.modelDocument.referencesDocument:
                    if doc.type in {ModelDocument.Type.INLINEXBRL, ModelDocument.Type.INSTANCE}:
                        instanceFilename = Path(doc.uri)
                        if ixdsTarget is not None:
                            basefileStem = f"{instanceFilename.stem}.{ixdsTarget}"
                        else:
                            basefileStem = instanceFilename.stem
                        break
            if basefileStem is None:
                basefileStem = Path(modelXbrl.modelDocument.basename).stem
            oimFiles = [str(oimDir.joinpath(basefileStem + ext)) for ext in (".csv", ".json")]
            saveOimFiles(cntlr, modelXbrl, oimFiles, responseZipStream)

    @staticmethod
    def testcaseVariationValidated(
        testcaseDTS: ModelXbrl,
        testInstanceDTS: ModelXbrl,
        extraErrors: list[str],
        *args: Any,
        **kwargs: Any,
    ) -> None:
        pluginData = testcaseDTS.modelManager.cntlr.getPluginData(PLUGIN_NAME)
        oimFileSuffix = (
            pluginData.saveTestcaseOimFileSuffix if isinstance(pluginData, SaveLoadableOIMPluginData) else None
        )
        if (
            oimFileSuffix
            and testInstanceDTS is not None
            and testInstanceDTS.modelDocument is not None
            and testInstanceDTS.modelDocument.type
            in (ModelDocument.Type.INSTANCE, ModelDocument.Type.INLINEXBRL, ModelDocument.Type.INLINEXBRLDOCUMENTSET)
            and not any(oimErrorPattern.match(error) for error in testInstanceDTS.errors if error is not None)
        ):  # no OIM errors
            try:
                saveLoadableOIM(testInstanceDTS, testInstanceDTS.modelDocument.uri + oimFileSuffix)
            except Exception as ex:
                testcaseDTS.modelManager.cntlr.addToLog(f"Exception saving OIM {ex}")

    @staticmethod
    def saveLoadableOimSave(
        modelXbrl: ModelXbrl,
        oimFile: str,
        *args: Any,
        **kwargs: Any,
    ) -> None:
        return saveLoadableOIM(modelXbrl, oimFile, *args, **kwargs)


__pluginInfo__ = {
    "name": PLUGIN_NAME,
    "version": "1.3",
    "description": "This plug-in saves XBRL in OIM JSON, CSV or XLSX that can be re-loaded per se.",
    "license": "Apache-2",
    "author": authorLabel,
    "copyright": copyrightLabel,
    # classes of mount points (required)
    "CntlrWinMain.Menu.Tools": SaveLoadableOIMPlugin.cntlrWinMainMenuTools,
    "CntlrCmdLine.Options": SaveLoadableOIMPlugin.cntlrCmdLineOptions,
    "CntlrCmdLine.Utility.Run": SaveLoadableOIMPlugin.cntlrCmdLineUtilityRun,
    "CntlrCmdLine.Xbrl.Run": SaveLoadableOIMPlugin.cntlrCmdLineXbrlRun,
    "TestcaseVariation.Validated": SaveLoadableOIMPlugin.testcaseVariationValidated,
    "SaveLoadableOim.Save": SaveLoadableOIMPlugin.saveLoadableOimSave,
}
