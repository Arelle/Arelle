# Copyright Exbee Ltd 2026 
# License Apache 2
"""
tlbTemplateGen.py — XBRL Model Layout Object Feasibility Demonstrator

Purpose
-------
This module demonstrates that the XbrlModel ``layout`` objects defined in the
OIM Taxonomy specification (oim-taxonomy.md, §Layout object, §Data table object,
§Axis object etc.) can represent the same table-rendering semantics currently
expressed using XBRL Table Linkbase (TLB).

As its input it reads a regulator's DPM Annotated Templates Excel workbook (EIOPA
Solvency II, EBA COREP/FINREP/DORA, etc.) — the authoritative human-readable
source for those reporting templates — rather than the TLB itself.  The workbook
encodes table structure visually (multi-level headers, dimension/member labels,
Z-axis sheets, rollups, crossed-out gaps) using the cell conventions described
below.  The module decodes that structure and emits a single JSON file in the
``https://xbrl.org/2026/compiled`` document type, containing a compiled taxonomy
module (concepts, dimensions, members, cubes, labels) and a ``layout`` object
whose ``dataTables`` correspond one-to-one with the Excel worksheets.

Source Excel file
-----------------
``EIOPA_Solvency_II_DPM_Annotated_Templates_2.8.2_Hotfix.xlsx``
located at: oim/specifications/oim-taxonomy/examples/TLB/

Prior sample output
-------------------
``EIOPA_Solvency_II.json`` in the same directory.

Excel template layout conventions decoded
-----------------------------------------
Each worksheet represents one table. Cell A1 contains "TableName - Title".

* **Z axis**: rows labelled "Sheets (Z)" (EIOPA) or "Sheet per ..." (EBA) above
  the grid define an out-of-page (filter) dimension.  Member choices come from
  either in-sheet dropdown lists or a shared "Enumerations" worksheet.
* **X axis** (columns): the grey header band above the first data row contains
  multi-level column headers with optional rollup markers.  The row immediately
  above the first data row holds column IDs.  Rows below the header band supply
  per-column dimension/member assignments.
* **Y axis** (rows): leftmost columns hold multi-level row headers.  The
  rightmost header column holds row IDs.  Rows to the right of the ID column
  supply per-row dimension/member assignments.
* **Grey cells** (fill >= #CCCCCC): column/row header cells — not data entry.
* **Crossed-out cells** (diagonal border both ways): intentionally blank data
  cells (structural gaps in the table).
* **"Metrics" / "Main Property" dimension label**: treated as ``xbrl:concept``
  (the standard XBRL concept dimension).
* **Rollup markers**: a span with ``rollUpLocation`` = "start" or "end"
  indicates a subtotal column/row.

Output objects emitted (oim-taxonomy spec)
------------------------------------------
Each worksheet becomes a ``gridLayout`` data table in a single ``layout`` object.
The generator emits a *pure grid* layout — each axis uses ``gridHeaders`` +
``gridAxis.axisItems`` and never ``axisHeaders`` (the spec forbids ``gridHeaders``
and ``axisHeaders`` on the same axis, and the visual ``gridHeaders`` carry the
richer multi-level labels).

* ``XbrlLayout`` — top-level layout container (one per output file).
* ``XbrlDataTable`` — one ``gridLayout`` table per worksheet; holds ``xAxis``,
  ``yAxis``, optional ``zAxis``, and a ``cubeName`` (see below).
* ``gridHeaders`` — per-level header labels with optional ``span`` /
  ``rollUpLocation`` (the visual row/column headings).
* ``gridAxis.axisItems`` — maps each column/row ``axisId`` to its dimension
  member tuple; the intersection of x/y(/z) items assigns a fact to a cell.

Cubes emitted per table
-----------------------
* One ``xbrl:reportCube`` per table (``<table>_Cube``) enumerating the concept
  dimension, every dimension used on its axes (marked ``optional``), and the
  period/entity/unit core dimensions.  This defines the fact space of the table.
  NB: cube dimensions use the cube-object keys ``dimension`` / ``domainNetwork``,
  NOT the axis objects' ``dimensionName``.
* One ``xbrl:negativeCube`` per crossed-out (X) cell, referenced from the report
  cube's ``excludeCubes``, so a fact in a structural-gap cell has no fact space
  (oimte:noFactSpaceForFact).  Each pins the cell's members via deduplicated
  single-member "solo" domains.  (Follow-up: coalesce cells into rectangles.)

Format profiles
---------------
``PROFILES[FMT]`` collects the per-authority conventions (skip-sheets, header
column offset, Z-axis tokens, enumeration source, entry-comment prefix, gray
threshold, anchor strategy) so a new authority is mostly a new profile entry;
only ``anchorStrategy`` needs code (EIOPA gray-band vs EBA "Rows"/"Columns").

Configuration constants (top of module)
----------------------------------------
* ``EXCELWB`` / ``LAYOUTFILE`` — input and output file paths.
* ``FMT`` — profile key (``"EIOPA"`` or ``"EBA"``) selecting ``PROFILES[FMT]``.
* ``TABLEFILTER`` — set to a table name (e.g. ``"S.02.01.01.01"``) to process
  only that sheet; useful for debugging.  Set to ``None`` to process all tables.
"""

import os, csv, json
import regex as re
from zipfile import ZipFile
from lxml import etree
from openpyxl import load_workbook
from collections import OrderedDict, defaultdict
from ordered_set import OrderedSet

EXCELWB = "/Users/hermf/Documents/projects/Arelle/ArelleProject/hermfischer-xb/temp/EIOPA_Solvency_II_DPM_Annotated_Templates_2.8.2_Hotfix.xlsx"
LAYOUTFILE = "/Users/hermf/Documents/projects/Arelle/ArelleProject/hermfischer-xb/temp/eiopa.json" # "/Users/hermf/Documents/projects/XBRL.org/oim/specifications/oim-taxonomy/examples/TLB/EBA_Solvency_II_layout.json"
FMT = "EIOPA" # EBA or EIOPA
TABLEFILTER = "S.02.01.01.01" # None # None or table name to limit processing for debugging
#TABLEFILTER = "S.05.02.04.02"
#EXCELWB = "/Users/hermf/Documents/projects/EBA/templates/20260106 Annotated Table Layout  COREP 4.2 COREP_ALMCOREP 4.2.xlsx"
#LAYOUTFILE = "/Users/hermf/temp/corep.json"
#FMT = "EBA" # EBA or EIOPA
#TABLEFILTER = "C_66.01.a

# Format profiles.  Each authority/authoring-tool encodes the same layout
# semantics with different conventions.  Rather than scatter ``if FMT ==`` tests
# through the parser, the profile object collects every convention that differs
# between authorities so a new authority is (mostly) a new profile entry.  Only
# ``anchorStrategy`` requires per-profile code (the EIOPA gray-band detector vs
# the EBA "Rows"/"Columns" label detector); everything else is data.
PROFILES = {
    "EIOPA": {
        "skipSheets": ("Table of Contents", "CIC Tables"),
        "firstRowHdrCol": 1,
        "hasEnumerationsSheet": True,
        "zTokenCol1": "Sheets (Z)",       # token in column 1 marking the Z axis
        "zTokenCol2Prefix": None,          # EBA marks Z with "Sheet per" in column 2
        "entryCommentPrefix": "VariableID",
        "grayThreshold": "CCCCCC",         # fill lightness at/above which a cell is a header
        "anchorStrategy": "grayBand",
    },
    "EBA": {                               # profile stub; anchor detector not yet implemented (step 3)
        "skipSheets": ("TOC",),
        "firstRowHdrCol": 2,
        "hasEnumerationsSheet": False,
        "zTokenCol1": None,
        "zTokenCol2Prefix": "Sheet per",
        "entryCommentPrefix": "VariableID",
        "grayThreshold": "CCCCCC",
        "anchorStrategy": "rowColLabels",
    },
}
profile = PROFILES[FMT]
firstRowHdrCol = profile["firstRowHdrCol"]

# get enumeration references from raw Excel XML (coss-sheet Formula1 not accessible in python
sheetCellEnumRow = {}
if profile["hasEnumerationsSheet"]:
    with ZipFile(EXCELWB, 'r') as xl:
        for sf in xl.namelist():
            if "/sheet" in sf:
                with xl.open(sf, mode="r") as fp:
                    ws = etree.parse(fp)
                    title = None
                    for hlElt in ws.iter("{*}hyperlink"):
                        if hlElt.get("ref") == "A1":
                            title = hlElt.get("display","").partition(" ")[0]
                    if title:
                        for listEnum in ws.iter("{*}dataValidation"):
                            if listEnum.get("type") == "list":
                                cell = listEnum.findtext("{*}sqref")
                                m = re.match(r".*\$([0-9]+):", listEnum.findtext("{*}formula1/{*}f",""))
                                if cell and m:
                                    sheetCellEnumRow[(title,cell)] = m.group(1)

# Python excel processing
wb = load_workbook(EXCELWB) # need data_only False to get dropdown lists

def colLtr(colNum):
    if colNum is None:
        return ""
    return chr(ord('A') + colNum - 1)

numTables = 0
numZinSheet = 0
numZlists = 0

enums = []
if profile["hasEnumerationsSheet"]:
    wsEnum = wb["Enumerations"]
    for rowNum in range(1, wsEnum.max_row):
        row = []
        for colNum in range(1, wsEnum.max_column):
            v = wsEnum.cell(rowNum, colNum).value
            if v:
                row.append(v)
            else:
                break
        enums.append(row)
        
labels = {}
concepts = set()
enumConceptMembers = defaultdict(set)
dims = set()
dimMems = defaultdict(set)
dataTables = []
cubes = []
domains = []
classes = []
prefixes = set()

# Single-member domains used to pin one member of a dimension inside a negativeCube
# (exclusion cube).  Deduplicated across the whole model by (dimension, member).
# NOTE (follow-up): the domain `root` for xbrl:concept is a placeholder and the
# per-cell granularity should later be coalesced into rectangles; see review notes.
soloDomains = {}
def soloDomain(dim, mem):
    key = (dim, mem)
    if key not in soloDomains:
        name = f"exp:solo{len(soloDomains)+1}"
        root = "xbrl:concept" if dim == "xbrl:concept" else f"{dim}Cls"
        domains.append({"name": name, "root": root,
                        "relationships": [{"source": root, "target": mem}]})
        soloDomains[key] = name
    return soloDomains[key]

def qnFromLabel(qnWithLabel):
    m = re.match(r"([\w:_.]+) (?:[(](.+)[)])?", qnWithLabel)
    if m:
        q, lbl = m.groups()
        if q.endswith('.'): q = q[:-1]
    elif re.match(r"[\w_.]+:[\w:_.]+", qnWithLabel): # 
        q = qnWithLabel
    else:
        q = re.sub(r"[^\w:_.]", qnWithLabel, "_") # if not a proper QName clean it up
    if not q.startswith("xbrl:"):
        prefixes.add(q.rpartition(":")[0])
    if m and lbl:
        labels[q] = lbl.rpartition("Metric:")[2]
    return q

def qn(dim=None, mem=None):
    d = None
    if dim == "Metrics":
        q = d = "xbrl:concept"
    elif dim:
        q = d = qnFromLabel(dim)
        dims.add(q)
    if mem:
        q = qnFromLabel(mem)
        if d == "xbrl:concept":
            concepts.add(q)
        elif d:
            dimMems[d].add(q)
    return q

for ws in wb.worksheets:
    if ws.title in profile["skipSheets"]:
        continue # skip table of contents
    tableName, _, tableTitle = ws["A1"].value.partition(" - ")
    qnTable = f"exp:{tableName}"
    maxRow = ws.max_row
    maxCol = ws.max_column
    numTables += 1
    mergedRanges = [mr.bounds for mr in ws.merged_cells.ranges]

    def wsCell(rowNum, colNum, mergeSource=True): # row and col start at 1
        if rowNum < 1 or colNum < 1 or rowNum > maxRow or colNum > maxCol:
            return None # open tables set firstRowHdrCol=0, producing col 0 lookups openpyxl rejects
        if mergeSource:
            for (c1, r1, c2, r2) in mergedRanges:
                if r1 <= rowNum <= r2 and c1 <= colNum <= c2:
                    rowNum = r1 # use upper left value
                    colNum = c1
                    break
        return ws.cell(rowNum, colNum)

    def cellValue(rowNum, colNum, default=None, mergeSource=True):
        cell = wsCell(rowNum, colNum, mergeSource=mergeSource)
        if (cell is None or cell.value is None) and default is not None:
            return default
        return cell.value if cell is not None else None

    def isGrayCell(rowNum, colNum, crossedOut=False, bgIndexMatch=None):
        cell = wsCell(rowNum, colNum, mergeSource=False)
        bgColor = None
        if cell is not None and cell.fill:
            if cell.fill.start_color.type == "rgb":
                bgColor = cell.fill.start_color.rgb
            elif cell.fill.start_color.type == "indexed":
                return bgIndexMatch == cell.fill.start_color.indexed
            else:
                bgColor = "FFFFFF" # default white
        # NB: bgColor may be 8-char ARGB ("FFBBBBBB"); the comparison below is dominated
        # by the alpha byte, so in practice this tests "has an explicit fill" rather than a
        # true lightness threshold.  The parser's header-band detection is calibrated to
        # that behaviour, so it is deliberately preserved (do not strip the alpha here).
        return (cell is not None and bgColor is not None and bgColor >= profile["grayThreshold"] and
                crossedOut == (cell is not None and cell.border is not None and cell.border.diagonalDown and cell.border.diagonalUp))

    def isCrossedOut(rowNum, colNum):
        # A blocked cell is drawn with both diagonals (an "X"); its fill (~BBBBBB) is
        # darker than the header threshold, so it is identified by the borders, not the fill.
        cell = wsCell(rowNum, colNum, mergeSource=False)
        return (cell is not None and cell.border is not None and
                cell.border.diagonalDown and cell.border.diagonalUp)

    def isEntryCell(rowNum, colNum):
        cell = wsCell(rowNum, colNum, mergeSource=False)
        if isCrossedOut(rowNum, colNum):
            return False # a crossed-out (X) cell is a structural gap, not a data-entry cell
        return cell is None or cell.value is None or (cell.comment is not None and cell.comment.text.startswith(profile["entryCommentPrefix"]))

    if TABLEFILTER and TABLEFILTER != tableName:
        continue # limit for debugging

    # find Z and first row/col
    firstDataRow = firstDataCol = firstXhdrRow = firstXdimRow = firstYdimCol = ydimRow = maxDataCol = None  
    zDim = inZ = firstXhdrCol = pastXhdr = None
    zChoices = []
    zTokenCol1 = profile["zTokenCol1"]
    zTokenCol2Prefix = profile["zTokenCol2Prefix"]
    pastZ = not (zTokenCol1 and any(cellValue(rowNum, 1) == zTokenCol1 for rowNum in range(2,20)))
    isOpenTable = False # assuming closed table until we see data rows with no header cells above them
    for rowNum in range(2,20):
        if zTokenCol1 and cellValue(rowNum, 1) == zTokenCol1: #EIOPA style
            zDim = cellValue(rowNum, 2) # might be None
            inZ = True
            if any(cellValue(1, c) for c in range(2,20)):
                # print(f"{tableName} ====> Insheet zlist")
                numZinSheet += 1
            else:
                numZlists += 1
                enumRow = sheetCellEnumRow.get((tableName, f"B{rowNum}"))
                if enumRow:
                    zChoices.append(enums[int(enumRow)-1])  
            continue
        elif zTokenCol2Prefix and cellValue(rowNum, 2, default="").startswith(zTokenCol2Prefix): # EBA style
            inZ = True
            numZinSheet += 1
            rowNum += 1
            continue
        if not pastZ:
            if not inZ:
                continue
            if all(not cellValue(rowNum, c) for c in range(1,20)):
                pastZ = True
            else:
                continue
        if not pastXhdr:
            if not firstXhdrCol:
                for colNum in range(firstRowHdrCol+1,10):
                    if isGrayCell(rowNum, colNum):
                        firstXhdrCol = colNum
                        firstXhdrRow = rowNum
                        break
                continue
            elif not isGrayCell(rowNum, firstXhdrCol):
                pastXhdr = True
                # check if this looks like an open table or a row header of a grid table
                if all(isEntryCell(rowNum, colNum) and not isGrayCell(rowNum, colNum) for colNum in range(firstRowHdrCol, firstXhdrCol)):
                    firstDataRow = rowNum
                    firstRowHdrCol = 0 # no row headers
                    isOpenTable = True
            else:
                continue
        for colNum in range(firstRowHdrCol+1,maxCol):
            v = cellValue(rowNum, colNum, mergeSource=False)
            if isEntryCell(rowNum, colNum) and not isGrayCell(rowNum, colNum) and isGrayCell(rowNum-1, colNum):
                if not firstDataCol:
                    firstDataCol = colNum
                    firstDataRow = rowNum
                    for col2 in range(colNum, maxCol+1):
                        if isGrayCell(rowNum-1, col2):
                            maxDataCol = col2
                    if not maxDataCol and isOpenTable:
                        maxDataCol = maxCol
                    break
    
    if not firstDataRow:
        continue  # no table

    yKeyColsDims = [] # are the leftmost columns with dimension labels (not necessarily all dim cols)
    yKeyColIds = []
    # is this an open table with no row headers?
    if isOpenTable:
        # capture dimensions from col headers
        for colNum in range(firstDataCol, maxDataCol+1):
            dim = wsCell(firstDataRow-1, colNum).comment
            if dim:
                yKeyColsDims.append(qn(dim.text))
                yKeyColIds.append(None)
                if isGrayCell(firstDataRow-1, colNum):
                    m = re.match(r".*[(](\w+)[)]", cellValue(firstDataRow-1, colNum))
                    if m:
                        yKeyColIds[-1] = m.group(1)
            else:
                # does Xhdr start below this row?
                if cellValue(firstDataRow+1, colNum -1) == "Metrics" and not isGrayCell(firstDataRow+1, colNum):
                    firstXdimRow = firstDataRow + 1
                    maxDataRow = firstDataRow # open, unlimited entry rows
                break # no more key columns    
    # find y dimNameRows
    if not isOpenTable and not firstXdimRow:
        for rowNum in range(firstDataRow + 1, 100):
            if all(isEntryCell(rowNum, c) for c in range(1,20)):
                continue # typed dim entry row
            if (firstDataCol <= 2 or not cellValue(rowNum, 1)) and cellValue(rowNum, firstDataCol-1) and cellValue(rowNum, firstDataCol):
                firstXdimRow = rowNum
                break
        # x header dim cols for non-open rows
        for rowNum in range(firstDataRow, firstXdimRow or maxRow+1):
            if not cellValue(rowNum, firstDataCol):
                for colNum in range(firstDataCol + 1, 100):
                    if cellValue(rowNum, colNum): # end of data coluns
                        firstYdimCol = colNum
                        break
            if firstYdimCol:
                break
        if firstYdimCol:
            for rowNum in range(firstXhdrRow, firstDataRow+1):
                if not isEntryCell(rowNum, firstYdimCol):
                    ydimRow = rowNum
                    break
            
        if not firstXdimRow:
            for rowNum in range(firstDataRow, maxRow+1):
                if all(isEntryCell(rowNum, colNum) and not isGrayCell(rowNum, colNum) for colNum in range(firstDataCol, maxDataCol+1)):
                    maxDataRow = rowNum
                else:
                    break

        else:
            maxDataRow = firstXdimRow - 1

    print(f"ws {tableName} firstXHdr {colLtr(firstDataCol)}{firstXhdrRow} firstData {colLtr(firstDataCol)}{firstDataRow} xDim {colLtr(firstDataCol-1)}{firstXdimRow} yDim {colLtr(firstYdimCol)}{ydimRow}")
    
    # prepare dataTable object
    xAxis = {}
    yAxis = {}
    dt = {"name": qnTable, "tableType": "gridLayout", "xAxis": xAxis, "yAxis": yAxis}
    if zDim is not None or zChoices:
        zAxis = {}
        dt["zAxis"] = zAxis
    dataTables.append(dt)

   # extract axes and labels
    cubeXDomMems = defaultdict(list)
    xDims = []
    rollupCols = set()
    # x header labels, use header row if present otherwise data rows, but only for columns with data
    xHdrs = []
    for xhRow in range(firstXhdrRow, firstDataRow-1):
        xhCol = firstDataCol
        lvl = firstDataRow - xhRow - 1
        while xhCol < maxDataCol+1:
            # is span in a rollup
            if xhCol in rollupCols:
                xhCol += 1
                continue
            # is cell start of merged range?
            span = levelSpan = 1
            for (c1, r1, c2, r2) in mergedRanges:
                if r1 <= xhRow <= r2 and c1 <= xhCol <= c2:
                    span = c2 + 1 - xhCol
                    levelSpan = r2 - xhRow
                    break
            lbl = {"label": cellValue(xhRow, xhCol) or "", "labelLevel": lvl} # label is a required string
            if span > 1:
                lbl["span"] = span
            if xhRow < firstDataRow - 2: # check if there is a rollup?
                if levelSpan > 1 or not cellValue(xhRow + 1, xhCol):
                    lbl["rollUpLocation"] = "start"
                    rollupCols.add(xhCol)
                elif not cellValue(xhRow, xhCol + span - 1):
                    lbl["rollUpLocation"] = "end"
                    rollupCols.add(xhCol + span - 1)
                elif span > 1:
                    lbl["rollUpLocation"] = "none"
            xHdrs.append(lbl)
            xhCol += span
   # x IDs for mapping to axis items
    xIds = [cellValue(firstDataRow-1, xdCol) for xdCol in range(firstDataCol, maxDataCol+1) if cellValue(firstDataRow-1, xdCol)]
    if isOpenTable:
        for i, xId in enumerate(yKeyColIds):
            if xId:
                xIds[i] = xId
    fullFirstXdimRow = fullFirstYdimCol = False
    if isOpenTable:
        xDimCol = len(yKeyColsDims)
    else:
        xDimCol = firstDataCol - 1
    if firstXdimRow:
        firstXdim = cellValue(firstXdimRow, xDimCol)
        firstXdimRowMems = [cellValue(firstXdimRow, xdCol) for xdCol in range(firstDataCol, firstYdimCol or maxDataCol+1)]
        fullFirstXdimRow = all(firstXdimRowMems)
        if fullFirstXdimRow:
            firstXdimRowQns = OrderedSet(qn(firstXdim, q) for q in firstXdimRowMems)
            cubeXDomMems[qn(firstXdim)] = list(firstXdimRowQns)
        for xdCol in range(firstDataCol, maxDataCol+1):
            xDims.append({})
            for memRow in range(firstXdimRow, firstXdimRow + 20):
                if isOpenTable and xdCol <= len(yKeyColsDims):
                    dim = yKeyColsDims[xdCol-1]
                    mem = None # mem comes from the data row key column cell
                else:
                    dim = cellValue(memRow, xDimCol)
                    mem = cellValue(memRow, xdCol)
                if not dim or not mem:
                    break
                qnMem = qn(dim,mem)
                qnDim = qn(dim)
                xDims[-1][qnDim] = qnMem
                # open dims with enumeration
                if dim in ("Metrics", "Main Property"):
                    dim = "xbrl:concept"
                if memRow == firstXdimRow:
                    enumRow = sheetCellEnumRow.get((tableName, f"{colLtr(xdCol)}{memRow-1}"))
                    if enumRow:
                        memEnums = [qn(mem=e) for e in enums[int(enumRow)-1]]
                        enumConceptMembers[qnMem].update(memEnums)
                        cubeXDomMems[qnMem] = memEnums
    cubeYDomMems = defaultdict(list)
    yDims = []
    # y header labels, use header row if present otherwise data rows, but only for columns with data
    yHdrs = []
    rollupRows = set()
    for yhCol in range(1, firstDataCol-1):
        yhRow = firstDataRow
        lvl = firstDataCol - yhCol - 1
        while yhRow < maxDataRow+1:
            # is span in a rollup
            if yhRow in rollupRows:
                yhRow += 1
                continue
            # is cell start of merged range?
            span = levelSpan = 1
            for (c1, r1, c2, r2) in mergedRanges:
                if r1 <= yhRow <= r2 and c1 <= yhCol <= c2:
                    span = r2 + 1 - yhRow
                    levelSpan = c2 - yhCol
                    break
            lbl = {"label": cellValue(yhRow, yhCol) or "", "labelLevel": lvl} # label is a required string
            if span > 1:
                lbl["span"] = span
            if yhCol < firstDataCol - 2: # check if there is a rollup?
                if levelSpan > 1 or not cellValue(yhRow, yhCol + 1):
                    lbl["rollUpLocation"] = "start"
                    rollupRows.add(yhRow)
                elif not cellValue(yhRow + span - 1, yhCol):
                    lbl["rollUpLocation"] = "end"
                    rollupRows.add(yhRow + span - 1)
                elif span > 1:
                    lbl["rollUpLocation"] = "none"
            yHdrs.append(lbl)
            yhRow += span
    # y IDs for mapping to axis items
    yIds = []
    if firstYdimCol and ydimRow:
        yIds = [cellValue(ydRow, firstDataCol-1) for ydRow in range(firstDataRow, firstXdimRow or maxDataRow+1)if cellValue(ydRow, firstDataCol-1)]
        firstYdim = cellValue(ydimRow, firstYdimCol)
        firstYdimColMems = [cellValue(ydRow, firstYdimCol) for ydRow in range(firstDataRow, firstXdimRow or maxDataRow+1)]
        fullFirstYdimCol = all(firstYdimColMems)
        if fullFirstYdimCol:
            firstYdimColQns = OrderedSet(qn(firstYdim, q) for q in firstYdimColMems)
            cubeYDomMems[qn(firstYdim)] = list(firstYdimColQns)
        for ydRow in range(firstDataRow, firstXdimRow or maxDataRow+1):
            yDims.append({})
            for memCol in range(firstYdimCol, firstYdimCol + 20):
                dim = cellValue(ydimRow, memCol)
                mem = cellValue(ydRow, memCol)
                if not dim or not mem:
                    break
                yDims[-1][qn(dim)] = qn(dim,mem)
                # metrics with enumeration
                if dim in ("Metrics", "Main Property"):
                    dim = "xbrl:concept"
                if memCol == firstYdimCol + 1:
                    enumRow = sheetCellEnumRow.get((tableName, f"{colLtr(firstYdimCol)}{ydRow}"))
                    if enumRow:
                        memEnums =[qn(mem=e) for e in enums[int(enumRow)-1]]
                        enumConceptMembers[qnMem].update(memEnums)
                        cubeYDomMems[qnMem] = memEnums

    zDims = zIds = None
    if zChoices:
        zDims = []; zIds = []
        for zDimsChoiceList in zChoices:
            zDimsVal = {}
            zDims.append(zDimsVal)
            zIds.append(None)
            for zDimsChoice in zDimsChoiceList:
                for zDim in zDimsChoice.split(' | '):
                    m = re.match(r"Z Axis: [(](\w+)[)]: ([^)]+[)]) (.*)$", zDim)
                    if m:
                        id, dim, mem = m.groups()
                        if id: zIds[-1] = id
                        qnDim = qn(dim)
                        qnMem = qn(dim, mem)
                        zDimsVal[qnDim] = qnMem
        
    # ---- report cube (one per table): defines the fact space assigned to this table ----
    # cubeDimensions use the cube-object property names `dimension`/`domainNetwork`
    # (NOT the axis objects' `dimensionName`).  Non-core dimensions are optional so a
    # fact need not carry every dimension.  Left unconstrained (no domainNetwork) a
    # dimension admits any member, which is sufficient to scope facts to this table.
    cubeName = f"{qnTable}_Cube"
    dt["cubeName"] = cubeName
    cubeDims = [{"dimension": "xbrl:concept"}]
    seenCubeDims = {"xbrl:concept"}
    for dimset in xDims + yDims + (zDims or []):
        for d in dimset:
            if d not in seenCubeDims:
                seenCubeDims.add(d)
                cubeDims.append({"dimension": d, "optional": True})
    for coreDim in ("xbrl:period", "xbrl:entity", "xbrl:unit"):
        cubeDims.append({"dimension": coreDim})
    cube = {"name": cubeName, "cubeType": "xbrl:reportCube", "cubeDimensions": cubeDims}
    cubes.append(cube)

    # ---- exclusion cubes: crossed-out (X) cells are structural gaps with no fact space ----
    # A gridLayout renders the cross-product of the x/y axis items, so a blocked
    # intersection would otherwise render as an input cell.  Each blocked cell becomes a
    # negativeCube pinning that cell's dimension members; the report cube excludes them,
    # so a fact landing there raises oimte:noFactSpaceForFact.  (Follow-up: coalesce
    # blocked cells into rectangles rather than one cube per cell.)
    if not isOpenTable and xDims and yDims:
        exclNames = []
        for j in range(len(yDims)):
            for i in range(len(xDims)):
                if not isCrossedOut(firstDataRow + j, firstDataCol + i):
                    continue
                exclDims = {**yDims[j], **xDims[i]}
                if zDims and len(zDims) == 1:
                    exclDims = {**exclDims, **zDims[0]}
                if not exclDims:
                    continue
                exclCubeDims = []
                for d, mem in exclDims.items():
                    ed = {"dimension": d, "domainNetwork": soloDomain(d, mem)}
                    if d != "xbrl:concept":
                        ed["optional"] = True
                    exclCubeDims.append(ed)
                exclCubeDims.append({"dimension": "xbrl:period"})
                exclCubeDims.append({"dimension": "xbrl:entity"})
                exclName = f"{qnTable}_Excl{len(exclNames)+1}"
                cubes.append({"name": exclName, "cubeType": "xbrl:negativeCube", "cubeDimensions": exclCubeDims})
                exclNames.append(exclName)
        if exclNames:
            cube["excludeCubes"] = exclNames

    # ---- axes: pure gridLayout ----
    # gridHeaders carry the visual multi-level labels/spans/rollups; gridAxis.axisItems
    # map each column/row id to its dimension members (the fact-to-cell assignment).
    # gridHeaders and axisHeaders MUST NOT coexist on one axis (spec grid-header
    # constraint), so this generator uses gridHeaders throughout and no axisHeaders.
    xAxis["gridAxis"] = xga = {}
    if xHdrs:
        xAxis["gridHeaders"] = xHdrs
    if xDims:
        xga["axisItems"] = [{"axisId": xIds[i] if i < len(xIds) else None, "dimensions": dm}
                            for i, dm in enumerate(xDims)]
    elif xIds:
        xga["axisItems"] = [{"axisId": xId, "dimensions": {}} for xId in xIds]

    yAxis["gridAxis"] = yga = {}
    if yHdrs:
        yAxis["gridHeaders"] = yHdrs
    if yDims:
        yga["axisItems"] = [{"axisId": yIds[i] if i < len(yIds) else None, "dimensions": dm}
                            for i, dm in enumerate(yDims)]
    elif yIds:
        yga["axisItems"] = [{"axisId": yId, "dimensions": {}} for yId in yIds]

    if zDims:
        zAxis["gridAxis"] = zga = {}
        zga["axisItems"] = [{"axisId": zIds[i], "dimensions": dm} for i, dm in enumerate(zDims)]

print(f"Num tables {numTables}, Num of Z in sheet = {numZinSheet}, Num Z lists {numZlists}")
print(f"labels {labels}\nconcepts {concepts}\ndims {dims}\ndimMems {dimMems}")


extensionPrefix = "exp"
targetNamespace = "http://example.com/taxonomy"

oimTxmy = OrderedDict()
oimTxmy["documentInfo"] = docInfo = OrderedDict()
docInfo["documentType"] = "https://xbrl.org/2026/compiled" # modelForm is "compiled"
docInfo["namespaces"] = namespaces = {
    extensionPrefix: targetNamespace,
    "xbrl": "https://xbrl.org/2026",
    "xbrlr": "https://xbrl.org/2026/report",  # was xbrli:.../2025/instance
    "xbrlm": "https://xbrl.org/2026/model",
    "xs": "http://www.w3.org/2001/XMLSchema"
    }
for p in prefixes:
    namespaces[p] = f"{targetNamespace}/{p}"
oimTxmy["xbrlModel"] = xbrlMdl = OrderedDict()

# provide consistent order to taxonomy properties and objects
xbrlMdl["name"] = f"{extensionPrefix}:{FMT.lower()}Example"
xbrlMdl["version"] = "0.9"
xbrlMdl["modelForm"] = "compiled" # all namespaces crammed together
xbrlMdl["abstracts"] = abstracts = []
def dt(qn):
    ln = qn.rpartition(":")[2]
    return {"m":"xbrlr:monetary",
            "e":"xs:QName",
            "i": "xs:integer",
            "b": "xs:boolean",
            "d": "xs:date",
            }.get(ln[0] if ln else "","xs:string")
def pt(qn):
    ln = qn.rpartition(":")[2]
    return "instant" if len(ln) > 1 and ln[1] == "i" else "duration"
xbrlMdl["concepts"] = []
for c in sorted(concepts):
    o = {"name": c, "dataType": dt(c), "periodType": pt(c)} 
    if c in enumConceptMembers:
        o["enumerationDomain"] = f"{c}Domain"
    xbrlMdl["concepts"].append(o)
        
        
xbrlMdl["cubes"] = cubes
xbrlMdl["dimensions"] = [{"name":d, "domainClass":f"{d}Cls"} for d in sorted(dims)]
xbrlMdl["domainClasses"] = [{"name":f"{d}Cls"} for d in sorted(dims)]
mems = set(v for vls in dimMems.values() for v in vls)
for m in enumConceptMembers.values():
    mems.update(m)
xbrlMdl["members"] = [{"name":f"{m}"} for m in sorted(mems)]
doms = set(f"{c}Domain" for c in enumConceptMembers.keys())
xbrlMdl["domains"] = domains
for d in sorted(doms):
    dom = {"name": d, "root": "xbrl:traitDomain"}
    if d.endswith("Domain") and d[:-6] in enumConceptMembers:
        dom["relationships"] = [{"source": "xbrl:traitDomain","target":m} for m in enumConceptMembers[d[:-6]]]
    domains.append(dom)
xbrlMdl["networks"] = networks = []
xbrlMdl["labels"] = labels = [{"relatedName":r,"language":"en","value":l,"labelType":"xbrl:label"} for r,l in labels.items()]
xbrlMdl["labelTypes"] = labelTypes = []
sharedLabelRefs = {}
xbrlMdl["layouts"] = [{
    "name": f"exp:{FMT.lower()}Layout",
    "tableConstruction": "topDown",
    "dataTables": dataTables
    }]

with open(LAYOUTFILE, "w") as fh:
    fh.write(json.dumps(oimTxmy, indent=3))

