# Copyright Exbee Ltd 2026
# License Apache 2
"""
tlbTemplate-eba.py — XBRL Model Layout generator for the EBA "Annotated Table Layout" family

Purpose
-------
Companion to ``tlbTemplate-eiopa.py`` (EIOPA).  The EBA DPM annotated templates (DORA,
COREP, FINREP, ...) use a different authoring convention from EIOPA, so this is a
separate parser that emits the same kind of output: an ``https://xbrl.org/2026/module``
document importing ``xbrlm:base``, with a ``gridLayout`` data table per worksheet and a
report cube per table.  All generated objects fold into one authority namespace
(``dora:``), source codes kept in the local name (see tlbTemplate-eiopa for the rationale).

EBA annotated-template conventions decoded
------------------------------------------
Each worksheet is one (usually open) table:

* **Anchors** are explicit text labels, not gray fill: ``Columns`` marks the column
  band's top-left; ``Rows`` (with ``Open Rows`` for open tables) marks the row band.
* Below ``Columns``: row+1 = column header labels, row+2 = column ids (0010, 0020, …),
  row+3 = the VariableID data row (holds "Rows"/"Open Rows" in cols A/B).
* **Concept (metric)**: the ``Main Property`` row (col B).  Each column's cell is
  ``(propID) label[dim:mem]`` where propID is the data-point/metric code (``mi…``
  monetary, ``ei…`` enumerated, ``si…`` string, ``di…`` date, ``ii…`` integer).
* **Dimensions**: rows below Main Property, col B = ``(DIM:DOMAIN) label``.
  Each column's cell for that row = ``(DOMAIN:member) label`` (explicit member) or
  ``(DOMAIN:) <Key value>`` (typed/open — no explicit member).

QName folding (single ``dora:`` module namespace)
------------------------------------------------
* concept  ``mi1310``          -> ``dora:met_mi1310``
* dimension ``CRZ``            -> ``dora:dim_CRZ``
* member   ``CO:x2``           -> ``dora:CO_x2``
* typed/open members (``(CR:) <Key value>``) are represented as a typed dimension
  (no explicit member is emitted for that column).

Configuration constants (top of module)
----------------------------------------
* ``EXCELWB`` / ``LAYOUTFILE`` — input workbook and output JSON.
* ``FMT`` — authority prefix (``"dora"``) = the single module namespace prefix.
* ``TABLEFILTER`` — a sheet name to process only that table (debug), or ``None``.

Status / limitations
--------------------
First cut for DORA-standalone (15 regular open tables).  The row axis of open tables is
left minimal (rows are unbounded data entries); the ``[dim:mem]`` constraint suffix on a
concept and multi-level column headers are not yet decoded.
"""

import json
import regex as re
from openpyxl import load_workbook
from collections import OrderedDict, defaultdict
from ordered_set import OrderedSet

EXCELWB = "/Users/hermf/Documents/projects/XBRL.org/oim/specifications/oim-taxonomy/examples/TLB/DORA 4.0 Templates.xlsx"
FMT = "dora"                              # authority prefix = single module namespace
TABLEFILTER = "B_02.01"                   # a table name to limit processing, or None for all
TLBDIR = "/Users/hermf/Documents/projects/XBRL.org/oim/specifications/oim-taxonomy/examples/TLB"
LAYOUTFILE = f"{TLBDIR}/DORA_layout{('_' + TABLEFILTER) if TABLEFILTER else ''}.json"

SKIP_SHEETS = ("TOC",)
PFX = FMT
targetNamespace = f"http://example.com/taxonomy/{FMT}"

wb = load_workbook(EXCELWB)

# ---- model accumulators ----
concepts = set()
dims = set()
members = set()
dimMems = defaultdict(set)                # dim QName -> set of member QNames
labels = {}                              # object QName -> label text
dataTables = []
cubes = []
domainNetworks = []

def clean(s):
    return re.sub(r"\s+", " ", str(s)).strip()

def foldConcept(propID, label=None):
    q = f"{PFX}:met_{propID}"
    concepts.add(q)
    if label: labels.setdefault(q, clean(label))
    return q

def foldDim(dimCode, label=None):
    q = f"{PFX}:dim_{dimCode}"
    dims.add(q)
    if label: labels.setdefault(q, clean(label))
    return q

def foldMember(domCode, memCode, label=None):
    q = f"{PFX}:{domCode}_{memCode}"
    members.add(q)
    if label: labels.setdefault(q, clean(label))
    return q

# EBA property code -> XBRL data type (first letter of the code)
def conceptDataType(propID):
    return {"m": "xbrlr:monetary", "i": "xs:integer", "d": "xs:date",
            "p": "xbrlr:percent", "b": "xs:boolean"}.get(propID[0], "xs:string")

reConcept = re.compile(r"\((\w+)\)\s*(.*)")            # (propID) label...
reDim = re.compile(r"\((\w+):(\w+)\)\s*(.*)")          # (DIM:DOMAIN) label
reMember = re.compile(r"\((\w+):(\w*)\)\s*(.*)")       # (DOMAIN:member) label  (member may be empty = typed)

numTables = 0
for ws in wb.worksheets:
    if ws.title in SKIP_SHEETS:
        continue
    a1 = ws["A1"].value or ""
    tableName = a1.partition(" - ")[0].strip() or ws.title
    if TABLEFILTER and TABLEFILTER != tableName:
        continue
    qnTable = f"{PFX}:{tableName}"
    maxRow, maxCol = ws.max_row, ws.max_column

    def cv(r, c):
        if r < 1 or c < 1 or r > maxRow or c > maxCol:
            return None
        return ws.cell(r, c).value

    # locate anchors
    colAnchor = rowAnchor = mainPropRow = None
    isOpen = False
    for r in range(1, 25):
        for c in range(1, 15):
            v = cv(r, c)
            if not isinstance(v, str):
                continue
            s = v.strip()
            if s == "Columns" and colAnchor is None:
                colAnchor = (r, c)
            elif s == "Rows" and rowAnchor is None:
                rowAnchor = (r, c)
            elif s == "Open Rows":
                isOpen = True
            elif s == "Main Property" and mainPropRow is None:
                mainPropRow = r
    if colAnchor is None:
        continue  # not a recognised table
    numTables += 1

    colTop, firstDataCol = colAnchor            # "Columns" sits at the band's top-left
    labelRow = colTop + 1
    idRow = colTop + 2
    # data columns = those with an id in idRow
    dataCols = [c for c in range(firstDataCol, maxCol + 1) if cv(idRow, c) not in (None, "")]
    if not dataCols:
        continue

    # dimension rows: col B, from Main Property row + 1 downward while "(DIM:DOM)" present
    dimRows = []
    if mainPropRow:
        r = mainPropRow + 1
        while r <= maxRow and isinstance(cv(r, 2), str) and reDim.match(cv(r, 2).strip()):
            dimRows.append(r)
            r += 1

    # build one axis item (dimension tuple) per data column
    xDims = []      # list of {dimQN: memberQN}
    xIds = []
    xHdrs = []
    for c in dataCols:
        dimset = {}
        # concept (Main Property)
        if mainPropRow:
            mp = cv(mainPropRow, c)
            if isinstance(mp, str):
                m = reConcept.match(mp.strip())
                if m:
                    dimset["xbrl:concept"] = foldConcept(m.group(1), m.group(2))
        # explicit dimension members
        for dr in dimRows:
            dm = reDim.match(cv(dr, 2).strip())
            dimCode, domCode = dm.group(1), dm.group(2)
            cell = cv(dr, c)
            if isinstance(cell, str):
                mm = reMember.match(cell.strip())
                if mm and mm.group(2):                       # explicit member (not typed <Key value>)
                    qDim = foldDim(dimCode, dm.group(3))
                    dimset[qDim] = foldMember(mm.group(1), mm.group(2), mm.group(3))
        xDims.append(dimset)
        xIds.append(clean(cv(idRow, c)))
        lbl = cv(labelRow, c)
        xHdrs.append({"label": clean(lbl) if lbl is not None else "", "labelLevel": 1})

    # A "definition"/closed table has no per-column Main Property; its concept is the single
    # row property (col B at the Rows anchor).  Synthesise a concept from that label and put
    # it on a single y-axis row item so every report cube has a concept dimension.
    tableConcepts = OrderedSet(d["xbrl:concept"] for d in xDims if "xbrl:concept" in d)
    yAxis = {"gridAxis": {}}                  # open table: rows are unbounded entries
    if not tableConcepts and rowAnchor:
        rowProp = clean(cv(rowAnchor[0], rowAnchor[1] + 1)) or tableName
        rowId = clean(cv(rowAnchor[0], rowAnchor[1] + 2))
        synth = foldConcept("s" + re.sub(r"\W+", "", rowProp)[:40], rowProp)  # 's' ⇒ xs:string
        tableConcepts = OrderedSet([synth])
        yAxis = {"gridAxis": {"axisItems": [{"axisId": rowId or "0010",
                                             "dimensions": {"xbrl:concept": synth}}]}}

    # ---- report cube (concept + axis dims), with a per-table concept domain ----
    conceptDomName = None
    if tableConcepts:
        conceptDomName = f"{qnTable}_ConceptDom"
        domainNetworks.append({"name": conceptDomName, "root": "xbrl:conceptDomain",
                               "relationships": [{"source": "xbrl:conceptDomain", "target": c} for c in tableConcepts]})
    cubeName = f"{qnTable}_Cube"
    conceptDim = {"dimension": "xbrl:concept"}
    if conceptDomName:
        conceptDim["domainNetwork"] = conceptDomName
    cubeDims = [conceptDim]
    seen = {"xbrl:concept"}
    for d in xDims:
        for k in d:
            if k not in seen:
                seen.add(k); cubeDims.append({"dimension": k, "optional": True})
    for coreDim in ("xbrl:period", "xbrl:entity", "xbrl:unit"):
        cubeDims.append({"dimension": coreDim})
    cubes.append({"name": cubeName, "cubeType": "xbrl:reportCube", "cubeDimensions": cubeDims})

    # ---- data table: columns on the x-axis (gridHeaders + gridAxis.axisItems); open rows ----
    xAxis = {"gridHeaders": xHdrs,
             "gridAxis": {"axisItems": [{"axisId": xIds[i], "dimensions": dm} for i, dm in enumerate(xDims)]}}
    dataTables.append({"name": qnTable, "tableType": "gridLayout", "cubeName": cubeName,
                       "xAxis": xAxis, "yAxis": yAxis})

# ---- assemble the OIM module ----
oim = OrderedDict()
oim["documentInfo"] = docInfo = OrderedDict()
docInfo["documentType"] = "https://xbrl.org/2026/module"
docInfo["namespaces"] = {
    PFX: targetNamespace,
    "xbrl": "https://xbrl.org/2026",
    "xbrlr": "https://xbrl.org/2026/report",
    "xbrlm": "https://xbrl.org/2026/model",
    "xs": "http://www.w3.org/2001/XMLSchema",
}
docInfo["documentNamespacePrefix"] = PFX
oim["xbrlModel"] = m = OrderedDict()
m["name"] = f"{PFX}:{FMT}Example"
m["version"] = "0.9"
m["importedTaxonomies"] = [{"xbrlModelName": "xbrlm:base"}]
m["concepts"] = [{"name": c, "dataType": conceptDataType(c.rpartition('_')[2]), "periodType": "duration"}
                 for c in sorted(concepts)]
m["cubes"] = cubes
m["dimensions"] = [{"name": d, "domainClass": f"{d}Cls"} for d in sorted(dims)]
m["domainClasses"] = [{"name": f"{d}Cls", "allowedDomainItem": "xbrl:memberObject"} for d in sorted(dims)]
m["members"] = [{"name": x} for x in sorted(members)]
m["domainNetworks"] = domainNetworks
definedNames = concepts | members | dims
m["labels"] = [{"forObject": r, "language": "en", "value": l, "labelType": "xbrl:label"}
               for r, l in labels.items() if r in definedNames]
m["layouts"] = [{"name": f"{PFX}:{FMT}Layout", "tableConstruction": "topDown", "dataTables": dataTables}]

with open(LAYOUTFILE, "w") as fh:
    fh.write(json.dumps(oim, indent=3))
print(f"tables {numTables} | concepts {len(concepts)} dims {len(dims)} members {len(members)} | -> {LAYOUTFILE}")
