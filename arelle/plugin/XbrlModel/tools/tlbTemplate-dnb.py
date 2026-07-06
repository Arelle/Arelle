# Copyright Exbee Ltd 2026
# License Apache 2
"""
tlbTemplate-dnb.py — XBRL Model Layout generator for the DNB DPM dictionary + templates

Purpose
-------
Third sibling to ``tlbTemplate-eiopa.py`` and ``tlbTemplate-eba.py``.  Unlike those two
(where the dimensional semantics are embedded in the template cells), the DNB workbook is
**dictionary-driven**: the ``T…`` template sheets are presentation only (labels + ids),
and the meaning lives in the dictionary + a fully-resolved ``Data Points`` sheet.  So this
is a different ingest that emits the same kind of output — an ``https://xbrl.org/2026/module``
document importing ``xbrlm:base``, one ``gridLayout`` table per template, all objects folded
into a single ``dnb:`` namespace.

Sources used
------------
* ``Dict Metrics`` — Metric Code -> data type (concepts).
* ``Data Points`` — the resolved table: per cell -> ``Metric`` (concept QName), ``Key(s)``
  (dimension members), ``X/Y/Z Ordinate`` (column/row/sheet ids), ``Unit Reference``.
* ``T…`` template sheets — the visual column/row header labels for each ordinate.

First cut
---------
Emits, per table, a report cube (concept + core dims, concept domain = the table's metrics)
and a gridLayout whose columns/rows come from the X/Y ordinates (labels from the template);
the metric (concept) is placed on the column axis items.  The ``Key(s)`` dimensional
breakdown is not yet decoded (its ``D01 D02 …`` shorthand needs the dimension mapping) — so
non-concept dimensions are a TODO.
"""

import json
import regex as re
from openpyxl import load_workbook
from collections import OrderedDict, defaultdict
from ordered_set import OrderedSet

EXCELWB = "/Users/hermf/Documents/projects/XBRL.org/oim/specifications/oim-taxonomy/examples/TLB/dnb-mes-dpm-dictionary-and-annotated-templates-4-0-0.xlsx"
FMT = "dnb"
TABLEFILTER = "T01.01"                    # a table code to limit processing, or None for all
TLBDIR = "/Users/hermf/Documents/projects/XBRL.org/oim/specifications/oim-taxonomy/examples/TLB"
LAYOUTFILE = f"{TLBDIR}/DNB_layout{('_' + TABLEFILTER) if TABLEFILTER else ''}.json"
PFX = FMT
targetNamespace = f"http://example.com/taxonomy/{FMT}"

wb = load_workbook(EXCELWB, data_only=True)

def clean(s):
    return re.sub(r"\s+", " ", str(s)).strip()

def headerIndex(ws):
    return {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1) if ws.cell(1, c).value}

# ---- Dict Metrics: metric code -> XBRL data type ----
DTYPE = {"monetaryItemType": "xbrlr:monetary", "percentItemType": "xbrlr:percent",
         "integerItemType": "xs:integer", "dateItemType": "xs:date",
         "booleanItemType": "xs:boolean", "decimalItemType": "xs:decimal"}
def dataTypeOf(metricType):
    ln = (metricType or "").rpartition(":")[2] or (metricType or "")
    return DTYPE.get(ln, "xs:string")

metricDataType = {}          # metric code (mi004) -> dataType
dm = wb["Dict Metrics"]; h = headerIndex(dm)
for r in range(2, dm.max_row + 1):
    code = dm.cell(r, h["Metric Code"]).value
    if code:
        metricDataType[str(code)] = dataTypeOf(dm.cell(r, h.get("Metric Type", 0)).value if h.get("Metric Type") else None)

# ---- model accumulators ----
concepts = {}                # concept QName -> dataType
dataTables = []
cubes = []
domainNetworks = []
labels = {}

reMetric = re.compile(r"([\w]+):([\w]+)")     # dnb_met:mi004

def foldConcept(metricQName):
    m = reMetric.match(metricQName)
    if not m:
        return None
    prefix, code = m.group(1), m.group(2)
    q = f"{PFX}:{prefix}_{code}"
    concepts.setdefault(q, metricDataType.get(code, "xs:string"))
    return q

# ---- template label helpers: for a table, map column-id/row-id -> header label ----
def templateLabels(tableCode):
    """Return (colLabels, rowLabels) keyed by ordinate id, from the T… sheet."""
    sheet = tableCode.replace(" ", "")
    if sheet not in wb.sheetnames:
        return {}, {}
    ws = wb[sheet]
    colLabels, rowLabels = {}, {}
    # column id row: the row whose cells are the numeric column ids (010, 020, …)
    for r in range(1, min(ws.max_row, 30) + 1):
        ids = {c: str(ws.cell(r, c).value).strip() for c in range(1, ws.max_column + 1)
               if isinstance(ws.cell(r, c).value, (str, int)) and re.fullmatch(r"0*\d+", str(ws.cell(r, c).value).strip() or "x")}
        if len(ids) >= 2:                     # a row of column ids
            for c, cid in ids.items():
                lbl = ws.cell(r - 1, c).value  # nearest header above the id
                colLabels[cid.lstrip("0") or "0"] = clean(lbl) if lbl else cid
            break
    # row ids: a column (usually C) whose cells are numeric ids, label in the column to its left
    for c in range(1, min(ws.max_column, 6) + 1):
        ids = {r: str(ws.cell(r, c).value).strip() for r in range(1, min(ws.max_row, 400) + 1)
               if isinstance(ws.cell(r, c).value, (str, int)) and re.fullmatch(r"0*\d+", str(ws.cell(r, c).value).strip() or "x")}
        if len(ids) >= 2:
            for r, rid in ids.items():
                lbl = ws.cell(r, c - 1).value
                rowLabels[rid.lstrip("0") or "0"] = clean(lbl) if lbl else rid
            break
    return colLabels, rowLabels

def norm(ordinal):                            # "010" -> "10" for label lookup, keep id as-is
    return (str(ordinal).lstrip("0") or "0") if ordinal is not None else None

# ---- Data Points: gather cells per table ----
dp = wb["Data Points"]; hp = headerIndex(dp)
cTab, cCell, cX, cY, cMet = hp["Table Code"], hp["Cell"], hp["X Ordinate"], hp["Y Ordinate"], hp["Metric"]
tableCells = defaultdict(list)                # table code -> [(x, y, conceptQN)]
for r in range(2, dp.max_row + 1):
    tc = dp.cell(r, cTab).value
    if not tc or (TABLEFILTER and tc != TABLEFILTER):
        continue
    metRaw = dp.cell(r, cMet).value
    conceptQN = foldConcept(clean(metRaw).split(" ")[0]) if metRaw else None
    tableCells[tc].append((dp.cell(r, cX).value, dp.cell(r, cY).value, conceptQN))

# ---- build a data table + report cube per table ----
for tc, cells in tableCells.items():
    qnTable = f"{PFX}:{tc.replace(' ', '')}"
    colLabels, rowLabels = templateLabels(tc)

    # columns = distinct X ordinates (with a representative concept), rows = distinct Y ordinates
    colConcept = OrderedDict()                # x ordinate -> conceptQN
    rowOrds = OrderedSet()
    for x, y, cq in cells:
        if x is not None and x not in colConcept:
            colConcept[x] = cq
        elif x is not None and cq and colConcept[x] is None:
            colConcept[x] = cq
        if y is not None:
            rowOrds.add(y)

    tableConcepts = OrderedSet(c for c in colConcept.values() if c) | OrderedSet(c for _, _, c in cells if c)
    conceptDomName = None
    if tableConcepts:
        conceptDomName = f"{qnTable}_ConceptDom"
        domainNetworks.append({"name": conceptDomName, "root": "xbrl:conceptDomain",
                               "relationships": [{"source": "xbrl:conceptDomain", "target": c} for c in tableConcepts]})
    cubeName = f"{qnTable}_Cube"
    conceptDim = {"dimension": "xbrl:concept"}
    if conceptDomName:
        conceptDim["domainNetwork"] = conceptDomName
    cubeDims = [conceptDim, {"dimension": "xbrl:period"}, {"dimension": "xbrl:entity"}, {"dimension": "xbrl:unit"}]
    cubes.append({"name": cubeName, "cubeType": "xbrl:reportCube", "cubeDimensions": cubeDims})

    xItems, xHdrs = [], []
    for x, cq in colConcept.items():
        xHdrs.append({"label": colLabels.get(norm(x), str(x)), "labelLevel": 1})
        xItems.append({"axisId": str(x), "dimensions": ({"xbrl:concept": cq} if cq else {})})
    yItems, yHdrs = [], []
    for y in rowOrds:
        yHdrs.append({"label": rowLabels.get(norm(y), str(y)), "labelLevel": 1})
        yItems.append({"axisId": str(y), "dimensions": {}})

    dataTables.append({"name": qnTable, "tableType": "gridLayout", "cubeName": cubeName,
                       "xAxis": {"gridHeaders": xHdrs, "gridAxis": {"axisItems": xItems}},
                       "yAxis": {"gridHeaders": yHdrs, "gridAxis": {"axisItems": yItems}}})

# ---- assemble the OIM module ----
oim = OrderedDict()
oim["documentInfo"] = docInfo = OrderedDict()
docInfo["documentType"] = "https://xbrl.org/2026/module"
docInfo["namespaces"] = {PFX: targetNamespace, "xbrl": "https://xbrl.org/2026",
                         "xbrlr": "https://xbrl.org/2026/report", "xbrlm": "https://xbrl.org/2026/model",
                         "xs": "http://www.w3.org/2001/XMLSchema"}
docInfo["documentNamespacePrefix"] = PFX
oim["xbrlModel"] = m = OrderedDict()
m["name"] = f"{PFX}:{FMT}Example"
m["version"] = "0.9"
m["importedTaxonomies"] = [{"xbrlModelName": "xbrlm:base"}]
m["concepts"] = [{"name": c, "dataType": dt, "periodType": "instant"} for c, dt in sorted(concepts.items())]
m["cubes"] = cubes
m["domainNetworks"] = domainNetworks
m["layouts"] = [{"name": f"{PFX}:{FMT}Layout", "tableConstruction": "topDown", "dataTables": dataTables}]

with open(LAYOUTFILE, "w") as fh:
    fh.write(json.dumps(oim, indent=3))
print(f"tables {len(dataTables)} | concepts {len(concepts)} | -> {LAYOUTFILE}")
