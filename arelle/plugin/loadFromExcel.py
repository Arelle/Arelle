# -*- coding: utf-8 -*-

'''
loadFromExcel.py is an example of a plug-in that will load an extension taxonomy from Excel
input and optionally save an (extension) DTS.

See COPYRIGHT.md for copyright information.
'''
import os, io, sys, time, re, traceback, json, posixpath
from fnmatch import fnmatch
from collections import defaultdict, OrderedDict
from arelle import PythonUtil, XbrlConst, ModelDocument, UrlUtil
from arelle.PythonUtil import OrderedDefaultDict, OrderedSet
from arelle.ModelDocument import Type, create as createModelDocument
from arelle.ModelValue import qname, QName
from arelle.Version import authorLabel, copyrightLabel
from arelle.XbrlConst import (qnLinkLabel, standardLabelRoles, qnLinkReference, standardReferenceRoles,
                              qnLinkPart, gen, link, defaultLinkRole,
                              conceptLabel, elementLabel, conceptReference, summationItem
                              )
qnXbrldtClosed = qname("{http://xbrl.org/2005/xbrldt}xbrldt:closed")

importColHeaderMap = defaultdict(list)
resourceParsePattern = re.compile(r"(label[s]?|reference[s]?|relationship to),?\s*([\w][\w\s#+-:/]+[\w#+-/])(\s*[(]([^)]+)[)])?$")
roleNumberPattern = re.compile(r"(.*)[#]([0-9][0-9A-Za-z]*)")
xlUnicodePattern = re.compile("_x([0-9A-F]{4})_")
excludeDesignatedEnumerations = False
annotateEnumerationsDocumentation = False
annotateElementDocumentation = False
saveXmlLang = None

NULLENTRY = ({},)

facetSortOrder = {
    "fractionDigits" : "_00",
    "length": "_01",
    "minInclusive": "_02",
    "maxInclusive": "_03",
    "minExclusive": "_04",
    "maxExclusive": "_05",
    "minLength": "_06",
    "maxLength": "_07",
    "pattern": "_08",
    "totalDigits": "_09",
    "whiteSpace": "_10",
    "enumeration": "_11"}

def loadFromExcel(cntlr, modelXbrl, excelFile, mappedUri):
    from openpyxl import load_workbook
    from arelle import ModelDocument, ModelXbrl, XmlUtil
    from arelle.ModelDocument import ModelDocumentReference
    from arelle.ModelValue import qname

    def xlUnicodeChar(match):
        return chr(int(match.group(1), 16))

    def xlValue(cell): # excel values may have encoded unicode, such as _0000D_
        v = cell.value
        if isinstance(v, str):
            return xlUnicodePattern.sub(xlUnicodeChar, v).replace('\r\n','\n').replace('\r','\n')
        return v

    defaultLabelLang = saveXmlLang or "en"

    importColumnHeaders = {
        "名前空間プレフィックス": "prefix",
        "prefix": "prefix",
        "要素名": "name",
        "name": "name",
        "type": "type",
        "typePrefix": "typePrefix", # usually part of type but optionally separate column
        "substitutionGroup": "substitutionGroup",
        "periodType": "periodType",
        "balance": "balance",
        "abstract": "abstract", # contains true if abstract
        "nillable": "nillable",
        "depth": "depth",
        "minLength": "minLength",
        "maxLength": "maxLength",
        "minInclusive": "minInclusive",
        "maxInclusive": "maxInclusive",
        "length": "length",
        "fixed": "fixed",
        "pattern": "pattern",
        "enumeration": "enumeration",
        "excludedEnumeration": "excludedEnumeration",
        "preferred label": "preferredLabel",
        "preferredLabel": "preferredLabel",
        "presentation parent": "presentationParent", # qname -- instead of label hierarchy and depth
        "calculation parent": "calculationParent", # qname
        "calculation weight": "calculationWeight",
        # label col heading: ("label", role, lang [indented]),
        "標準ラベル（日本語）": ("label", XbrlConst.standardLabel, "ja", "indented"),
        "冗長ラベル（日本語）": ("label", XbrlConst.verboseLabel, "ja"),
        "標準ラベル（英語）": ("label", XbrlConst.standardLabel, "en"),
        "冗長ラベル（英語）": ("label", XbrlConst.verboseLabel, "en"),
        "用途区分、財務諸表区分及び業種区分のラベル（日本語）": ("labels", XbrlConst.standardLabel, "ja"),
        "用途区分、財務諸表区分及び業種区分のラベル（英語）": ("labels", XbrlConst.standardLabel, "en"),
        # label [, role [(lang)]] : ("label", http resource role, lang [indented|overridePreferred])
        "label": ("label", XbrlConst.standardLabel, defaultLabelLang, "indented"),
        "label, standard": ("label", XbrlConst.standardLabel, defaultLabelLang, "overridePreferred"),
        "label, terse": ("label", XbrlConst.terseLabel, defaultLabelLang),
        "label, verbose": ("label", XbrlConst.verboseLabel, defaultLabelLang),
        "label, documentation": ("label", XbrlConst.documentationLabel, defaultLabelLang),
        "label, periodStart": ("label", XbrlConst.periodStartLabel, defaultLabelLang),
        "label, periodEnd": ("label", XbrlConst.periodEndLabel, defaultLabelLang),
        "group": "linkrole",
        "linkrole": "linkrole",
        "ELR": "linkrole",
        "dimension default": "dimensionDefault"
        # reference ("reference", reference http resource role, reference part QName)
        # reference, required": ("reference", "http://treasury.gov/dataact/role/taxonomyImplementationNote", qname("{http://treasury.gov/dataact/parts-2015-12-31}dataact-part:Required"))
        # attribute, qname (attribute on element in xsd)
        }

    fatalLoadingErrors = []

    startedAt = time.time()

    if os.path.isabs(excelFile):
        # allow relative filenames to loading directory
        priorCWD = os.getcwd()
        os.chdir(os.path.dirname(excelFile))
    else:
        priorCWD = None
    importExcelBook = load_workbook(excelFile, data_only=True)
    sheetNames = importExcelBook.get_sheet_names()
    dtsSheet = None
    if "XBRL DTS" in sheetNames:
        dtsSheet = "XBRL DTS"
    elif "DTS" in sheetNames:
        dtsSheet = "DTS"
    elif "Sheet2" in sheetNames:
        dtsSheet = "Sheet2"
    if dtsSheet:
        dtsWs = importExcelBook[dtsSheet]
    else:
        dtsWs = None
    imports = {"xbrli": ( ("namespace", XbrlConst.xbrli),
                          ("schemaLocation", "http://www.xbrl.org/2003/xbrl-instance-2003-12-31.xsd") )} # xml of imports
    importXmlns = {}
    hasPreLB = hasCalLB = hasDefLB = hasRefLB = hasGenLB = False
    # xxxLB structure [ (elr1, def1, "_ELR_", [roots]), (elr2, def2, "_ELR_", [rootw]) ...]
    #   roots = (rootHref, None, "_root_", [children])
    #   children = (childPrefix, childName, arcrole, [grandChildren])
    preLB = []
    defLB = []
    calLB = []
    refLB = []
    genLB = []

    def lbDepthList(lbStruct, depth, parentList=None):
        if len(lbStruct) > 0:
            if depth == topDepth or not (hasDepthColumn or hasPresentationIndentColumnIndentations):
                return lbStruct[-1].childStruct
            return lbDepthList(lbStruct[-1].childStruct, depth-1, list)
        else:
            if (hasDepthColumn or hasPresentationIndentColumnIndentations):
                cntlr.addToLog("Depth error, Excel sheet: {excelSheet} row: {excelRow}"
                               .format(excelSheet=importSheetName, excelRow=iRow),
                                messageCode="importExcel:depth")
            return None

    splitString = None # to split repeating groups (order, depth)
    importFileName = None # for alternate import file
    importSheetNames = []
    skipRows = []  # [(from,to),(from,to)]  row number starting at 1

    genDocs = {} # generated documents (schema + referenced linkbases)
    genElementsDoc = None
    def newDoc(name):
        genDocs[name] = PythonUtil.attrdict(
            name = name,
            initialComment = None,
            schemaDocumentation = None,
            extensionSchemaPrefix = "",
            extensionSchemaFilename = "",
            extensionSchemaRelDirname = None, # only non-null for relative directory path
            extensionSchemaNamespaceURI = "",
            extensionSchemaVersion = None, # <schema @version>
            extensionRoles = {}, # key is roleURI, value is role definition
            extensionRoleLabels= defaultdict(set), # key is roleURI, value is set( (lang, label) )
            extensionElements = {},
            extensionTypes = {}, # attrs are name, base.  has facets in separate dict same as elements
            extensionLabels = {},  # key = (prefix, name, lang, role), value = label text
            extensionReferences = OrderedDefaultDict(OrderedSet), # key = (prefix, name, role) values = (partQn, text)
            hasEnumerationDocumentation = False,
            imports = {"xbrli": ( ("namespace", XbrlConst.xbrli),
                                  ("schemaLocation", "http://www.xbrl.org/2003/xbrl-instance-2003-12-31.xsd") )}, # xml of imports
            includes = [], # just schemaLocation
            importXmlns = {},
            importFilenames = {}, # file names relative to base
            childGenDocs = [],
            linkbaseRefs = [],
            labelLinkbases = [],
            referenceLinkbases = [],
            hasPreLB = False,
            hasCalLB = False,
            hasDefLB = False,
            hasRefLB = False,
            hasGenLB = False,
            generated = False
            )
        return genDocs[name]

    thisDoc = newDoc(None)

    excelDir = os.path.dirname(excelFile) + os.path.sep

    def docRelpath(filename, baseDir=None):
        if baseDir is None:
            baseDir = thisDoc.extensionSchemaRelDirname
        if (baseDir is not None and
            not (UrlUtil.isAbsolute(filename) or os.path.isabs(filename))):
            return posixpath.relpath(filename, baseDir)
        return filename

    isUSGAAP = False
    isGenerateAndImport = True
    extensionPrefixForCoreLabels = None
    dtsActionColIndex = 0
    dtsFiletypeColIndex = 1
    dtsPrefixColIndex = 2
    dtsFilenameColIndex = 3
    dtsNamespaceURIColIndex = 4
    for iRow, row in enumerate(dtsWs.rows if dtsWs else ()):
        try:
            if (len(row) < 1):  # skip if col 1 is non-existent
                continue
            _col0 = row[0].value
            if isinstance(_col0, str) and _col0.startswith("#"): # empty or "#"
                continue
            if iRow == 0:
                # title row may have columns differently laid out
                for i, col in enumerate(row):
                    v = xlValue(col)
                    if isinstance(v, str):
                        if v == "specification": dtsActionColIndex = i
                        if v.startswith("file type"): dtsFiletypeColIndex = i
                        if v.startswith("prefix"): dtsPrefixColIndex = i
                        if v.startswith("file, href or role definition"): dtsFilenameColIndex = i
                        if v.startswith("namespace URI"): dtsNamespaceURIColIndex = i
                continue
            action = filetype = prefix = filename = namespaceURI = None
            if len(row) > dtsActionColIndex: action = xlValue(row[dtsActionColIndex])
            if len(row) > dtsFiletypeColIndex: filetype = xlValue(row[dtsFiletypeColIndex])
            if len(row) > dtsPrefixColIndex: prefix = xlValue(row[dtsPrefixColIndex])
            if len(row) > dtsFilenameColIndex: filename = xlValue(row[dtsFilenameColIndex])
            if len(row) > dtsNamespaceURIColIndex: namespaceURI = xlValue(row[dtsNamespaceURIColIndex])
            lbType = lang = None
            if action == "import":
                if filetype in ("role", "arcrole"):
                    continue
                elif filetype == "schema":
                    thisDoc.imports[prefix] = ( ("namespace", namespaceURI), ("schemaLocation", docRelpath(filename)) )
                    thisDoc.importXmlns[prefix] = namespaceURI
                    thisDoc.importFilenames[prefix] = filename
                    if re.match(r"http://[^/]+/us-gaap/", namespaceURI):
                        isUSGAAP = True
                elif filetype == "linkbase":
                    typeLang = prefix.split()
                    if len(typeLang) > 0:
                        lbType = typeLang[0]
                    else:
                        lbType = "unknown"
                    thisDoc.linkbaseRefs.append( (lbType, filename, False) )
            elif action == "include" and filename:
                thisDoc.includes.append(docRelpath(filename))
            elif action == "xmlns" and prefix and namespaceURI:
                thisDoc.importXmlns[prefix] = namespaceURI
            elif action in ("extension", "generate"):
                if filetype == "schema":
                    if prefix:
                        # starts new document.
                        if not thisDoc.name:
                            del genDocs[thisDoc.name] # remove anonymous doc
                        thisDoc = newDoc(prefix) # new doc with prefix as its name
                    thisDoc.extensionSchemaPrefix = prefix
                    thisDoc.extensionSchemaFilename = filename
                    thisDoc.extensionSchemaNamespaceURI = namespaceURI
                    if not UrlUtil.isAbsolute(filename) and not os.path.isabs(filename):
                        thisDoc.extensionSchemaRelDirname = posixpath.dirname(filename)
                    else:
                        thisDoc.extensionSchemaRelDirname = None
                elif filetype == "linkbase":
                    typeLang = prefix.split()
                    if len(typeLang) > 0:
                        lbType = typeLang[0]
                    else:
                        lbType = "unknown"
                    if len(typeLang) > 1:
                        lang = referenceRole = typeLang[1]
                    else:
                        lang = None
                        referenceRole = XbrlConst.standardReference
                    if lbType in ("label", "generic-label"):
                        # lang, if provided, is a regex pattern
                        thisDoc.labelLinkbases.append((lbType, lang, filename))
                        if action == "extension" and not extensionPrefixForCoreLabels:
                            extensionPrefixForCoreLabels = thisDoc.extensionSchemaPrefix
                    elif lbType in ("reference", "generic-reference"):
                        hasRefLB = True
                        thisDoc.referenceLinkbases.append((lbType, referenceRole, filename))
                    elif lbType == "presentation":
                        thisDoc.hasPreLB = hasPreLB = True
                    elif lbType == "definition":
                        thisDoc.hasDefLB = hasDefLB = True
                    elif lbType == "calculation":
                        thisDoc.hasCalLB = hasCalLB = True
                    elif lbType == "generic":
                        thisDoc.hasGenLB = hasGenLB = True
                    thisDoc.linkbaseRefs.append( (lbType, filename, True) )
                elif filetype == "initialComment" and prefix:
                    thisDoc.initialComment = prefix
                elif filetype == "schemaDocumentation" and prefix:
                    thisDoc.schemaDocumentation = prefix
                elif filetype == "enumerationDocumentation":
                    thisDoc.hasEnumerationDocumentation = True
                elif filetype == "role" and namespaceURI: # filename is definition, prefix is optional used-on QNames
                    thisDoc.extensionRoles[namespaceURI] = (filename, prefix)
                elif filetype == "role label" and namespaceURI and prefix: # filename is label, prefix is language
                    thisDoc.extensionRoleLabels[namespaceURI].add( (filename, prefix) )
                elif filetype == "schema-version" and filename:
                    thisDoc.extensionSchemaVersion = filename
                elif filetype == "table-style" and filename == "xbrl-us":
                    isUSGAAP = True
                elif filetype == "elements":
                    genElementsDoc = thisDoc
            elif action == "meta" and filetype == "table-style" and filename == "xbrl-us":
                isUSGAAP = True
            elif action == "meta" and filetype == "generate-style" and filename == "import-separately":
                isGenerateAndImport = False
            elif action == "workbook" and filename:
                importFileName = filename
            elif action == "worksheet" and filename:
                importSheetNames.append(filename)
            elif action == "colheader" and filename and namespaceURI:
                if namespaceURI == "split":
                    splitString = filename
                else:
                    importColHeaderMap[filename].append(namespaceURI)
                    if namespaceURI not in importColumnHeaders:
                        fatalLoadingErrors.append("colheader {} definition {} not recognized.".format(filename, namespaceURI))
            elif action == "skip rows" and filename:
                fromRow, _sep, toRow = filename.partition("-")
                try:
                    skipRows.append((int(fromRow), int(toRow) if toRow else int(fromRow)))
                except (ValueError, TypeError):
                    fatalLoadingErrors.append("Exception (at skip rows): {error}, Excel sheet: {excelSheet} row: {excelRow}"
                                              .format(error=err, excelSheet=dtsSheet, excelRow=iRow))


        except Exception as err:
            fatalLoadingErrors.append("Exception: {error}, Excel sheet: {excelSheet} row: {excelRow}, Traceback: {traceback}"
                                      .format(error=err, excelSheet=dtsSheet, excelRow=iRow, traceback=traceback.format_tb(sys.exc_info()[2])))
    # remove any imported linkbaseRefs that are also generated
    for thisDoc in genDocs.values():
        linkbaseRefsToRemove = [i
                                for i, (lbType, filename, generate) in enumerate(thisDoc.linkbaseRefs)
                                if not generate and (lbType, filename, True) in thisDoc.linkbaseRefs]
        while len(linkbaseRefsToRemove):
            i = linkbaseRefsToRemove.pop()
            thisDoc.linkbaseRefs.pop(i)

    dtsWs = None # dereference

    genOrder = []

    for name, doc in genDocs.items():
        insertPos = len(genOrder)
        for i, otherDoc in enumerate(genOrder):
            if doc.name in otherDoc.imports:
                insertPos = i # put this doc before any firstr doc that imports it
                break
        genOrder.insert(insertPos, doc)

    if importFileName: # alternative workbook
        importExcelBook = load_workbook(importFileName, read_only=True, data_only=True)
        sheetNames = importExcelBook.get_sheet_names()

    if importSheetNames:
        for importSheetName in importSheetNames:
            if importSheetName not in sheetNames:
                fatalLoadingErrors.append("Worksheet {} specified for Excel importing, but not present in workbook.".format(importSheetName))
    else:
        for s in sheetNames:
            if s.endswith("Concepts"):
                importSheetNames.append(s)
        if not importSheetNames:
            for s in sheetNames:
                if "xbrl" in s.lower() and "dts" not in s:
                    importSheetNames.append(s)
        if not importSheetNames:
            fatalLoadingErrors.append("Worksheet {} specified for Excel importing, but not present in workbook.".format(importSheetName))


    if not isUSGAAP and genOrder: # need extra namespace declaration
        genOrder[0].importXmlns["iod"] = "http://disclosure.edinet-fsa.go.jp/taxonomy/common/2013-03-31/iod"

    # find column headers row
    headerCols = OrderedDict()
    headerColsAllElrs = set()
    hasLinkroleSeparateRow = True
    hasPreferredLabelTextColumn = False
    hasConceptAttributeColumn = False
    hasDepthColumn = False
    hasPresentationParentColumn = False
    hasRelationshipToCol = False
    hasrelationshipAttributeColumn = False
    headerRows = set()
    topDepth = 999999

    for importSheetName in importSheetNames:
        if importSheetName not in sheetNames:
            continue
        headerCols.clear()
        headerRows.clear()
        hasConceptAttributeColumn = False
        hasDepthColumn = False
        hasPresentationParentColumn = False
        presentationIndentColumn = None
        hasPresentationIndentColumnIndentations = False
        hasRelationshipToCol = False
        hasrelationshipAttributeColumn = False
        conceptsWs = importExcelBook[importSheetName]

        def setHeaderCols(row):
            headerCols.clear()
            for iCol, colCell in enumerate(row):
                v = xlValue(colCell)
                if isinstance(v,str):
                    v = v.strip()
                if v in importColHeaderMap:
                    for hdr in importColHeaderMap[v]:
                        if hdr in importColumnHeaders:
                            headerCols[importColumnHeaders[hdr]] = iCol
                elif v in importColumnHeaders:
                    headerCols[importColumnHeaders[v]] = iCol
                elif isinstance(v,str):
                    if any(v.startswith(r) for r in ("label,", "labels,", "reference,", "references,", "relationship to,")):
                        # custom/extension label/reference
                        m = resourceParsePattern.match(v)
                        if m:
                            _resourceType = m.group(1)
                            _resourceRole = "/" + m.group(2) # last path seg of role
                            _resourceLangOrPart = m.group(4) # lang or part
                            headerCols[(_resourceType, _resourceRole, _resourceLangOrPart)] = iCol
                    else:
                        # custom/extension non-label/reference value column
                        headerCols[v] = iCol

        # find out which rows are header rows
        for iRow, row in enumerate(conceptsWs.rows if conceptsWs else ()):
            if any(fromRow <= iRow+1 <= toRow for fromRow,toRow in skipRows):
                continue

            #for iCol, colCell in enumerate(row):
            setHeaderCols(row)
            # must have some of these to be a header col
            if (sum(1 for h in headerCols if h in ("name", "type", "depth", "periodType")) >= 3 or
                sum(1 for h in headerCols if h == "name" or (isinstance(h, tuple) and h[0] == "relationship to")) >= 2):
                # it's a header col
                headerRows.add(iRow+1)
                presentationIndentColumn = None
                for h,i in headerCols.items():
                    if isinstance(h, tuple) and "indented" in h:
                       presentationIndentColumn = i
            if 'linkrole' in headerCols:
                hasLinkroleSeparateRow = False
            if 'preferredLabel' in headerCols and any(isinstance(h, tuple) and h[0] == 'label' and h[1] == '/preferredLabel'
                                                      for h in headerCols):
                hasPreferredLabelTextColumn = True
            if 'depth' in headerCols:
                hasDepthColumn = True
            if 'presentationParent' in headerCols:
                hasPresentationParentColumn = True
            if not hasDepthColumn and hasPresentationParentColumn:
                topDepth = 0
            if not hasPresentationParentColumn and not hasDepthColumn and presentationIndentColumn is not None:
                if row[presentationIndentColumn].value:
                    depth = int(row[presentationIndentColumn].alignment.indent)
                    if depth < topDepth:
                        hasPresentationIndentColumnIndentations = True
                        if depth < topDepth:
                            topDepth = depth
            hasRelationshipToCol = any(h[0] == "relationship to" for h in headerCols if isinstance(h, tuple))
            headerCols.clear()

        def cellHasValue(row, header, _type):
            if header in headerCols:
                iCol = headerCols[header]
                return iCol < len(row) and isinstance(row[iCol].value, _type)
            return False

        def cellValue(row, header, strip=False, nameChars=False, default=None):
            if header in headerCols:
                iCol = headerCols[header]
                if iCol < len(row):
                    v = xlValue(row[iCol])
                    if strip and isinstance(v, str):
                        v = v.strip()
                    if nameChars and isinstance(v, str):
                        v = ''.join(c for c in v if c.isalnum() or c in ('.', '_', '-'))
                    if v is None:
                        return default
                    return v
            return default

        def valueNameChars(v):
            return ''.join(c for c in v if c.isalnum() or c in ('.', '_', '-'))

        def rowPrefixNameValues(row):
            prefix = cellValue(row, 'prefix', nameChars=True)
            if cellHasValue(row, 'name', str):
                if not prefix: # maybe name is a qname
                    prefix, _sep, _name = cellValue(row, 'name').partition(":")
                    if not _sep: # no prefix at all, whole string is name
                        prefix = ""
                    name = cellValue(row, 'name', nameChars=True)[len(prefix):]
                else:
                    name = cellValue(row, 'name', nameChars=True)
            else:
                name = None
            if not prefix and "prefix" not in headerCols and genElementsDoc is not None:
                prefix = genElementsDoc.extensionSchemaPrefix
            return prefix, name

        def checkImport(thisDoc, qname):
            prefix, sep, localName = qname.partition(":")
            if sep:
                if prefix not in thisDoc.imports:
                    if prefix == "xbrldt":
                        thisDoc.imports["xbrldt"] = ("namespace", XbrlConst.xbrldt), ("schemaLocation", "http://www.xbrl.org/2005/xbrldt-2005.xsd")
                    elif prefix == "nonnum":
                        thisDoc.imports["nonnum"] = ("namespace", "http://www.xbrl.org/dtr/type/non-numeric"), ("schemaLocation", "http://www.xbrl.org/dtr/type/nonNumeric-2009-12-16.xsd")
                    elif prefix != thisDoc.extensionSchemaPrefix and prefix != "xs":
                        cntlr.addToLog("Warning: prefix schema file is not imported for: {qname}"
                               .format(qname=qname),
                                messageCode="importExcel:warning", file=thisDoc.extensionSchemaFilename)

        # find top depth
        for iRow, row in enumerate(conceptsWs.rows if conceptsWs else ()):
            if (iRow + 1) in headerRows:
                setHeaderCols(row)
                hasConceptAttributeColumn = any(v.startswith("attribute, ") for v in headerCols if isinstance(v,str))
                hasRelationshipAttributeColumn = any(v.startswith("relationship attribute, ") for v in headerCols if isinstance(v,str))
            elif not (hasLinkroleSeparateRow and (iRow + 1) in headerRows) and 'depth' in headerCols:
                depth = cellValue(row, 'depth')
                if isinstance(depth, int) and depth < topDepth:
                    topDepth = depth

        # find header rows
        currentELR = currentELRdefinition = None
        for iRow, row in enumerate(conceptsWs.rows if conceptsWs else ()):
            useLabels = False
            eltEnumRefsParts = None
            if any(fromRow <= iRow+1 <= toRow for fromRow,toRow in skipRows):
                continue
            if (all(col.value is None for col in row) or
                all(isinstance(row[i].value, str) and row[i].value.strip() == "n/a"
                   for i in (headerCols.get("name"), headerCols.get("type"), headerCols.get("value"))
                   if i is not None)):
                continue # skip blank row
            try:
                isHeaderRow = (iRow + 1) in headerRows
                isELRrow = hasLinkroleSeparateRow and (iRow + 2) in headerRows
                if isHeaderRow:
                    setHeaderCols(row)
                    headerColsAllElrs |= headerCols.keys() # accumulate all header cols for role checks
                elif isELRrow:
                    currentELR = currentELRdefinition = None
                    for colCell in row:
                        v = str(xlValue(colCell) or '')
                        if v.startswith("http://"):
                            currentELR = v
                        elif not currentELRdefinition and v.endswith("　科目一覧"):
                            currentELRdefinition = v[0:-5]
                        elif not currentELRdefinition:
                            currentELRdefinition = v
                    if currentELR or currentELRdefinition:
                        if hasPreLB:
                            preLB.append( LBentry(role=currentELR, name=currentELRdefinition, isELR=True) )
                            if hasPresentationParentColumn:
                                preRels = set()
                        if hasDefLB:
                            defLB.append( LBentry(role=currentELR, name=currentELRdefinition, isELR=True) )
                        if hasCalLB:
                            calLB.append( LBentry(role=currentELR, name=currentELRdefinition, isELR=True) )
                            calRels = set() # prevent duplications when same rel in different parts of tree
                        if hasGenLB:
                            genLB.append( LBentry(role=currentELR, name=currentELRdefinition, isELR=True) )
                elif headerCols:
                    if "linkrole" in headerCols and cellHasValue(row, 'linkrole', str):
                        v = cellValue(row, 'linkrole', strip=True)
                        _trialELR = _trialELRdefinition = None
                        if v.startswith("http://"):
                            _trialELR = v
                        elif v.endswith("　科目一覧"):
                            _trialELRdefinition = v[0:-5]
                        else:
                            _trialELRdefinition = v
                        if (_trialELR and _trialELR != currentELR) or (_trialELRdefinition and _trialELRdefinition != currentELRdefinition):
                            currentELR = _trialELR
                            currentELRdefinition = _trialELRdefinition
                            if currentELR or currentELRdefinition:
                                if hasPreLB:
                                    preLB.append( LBentry(role=currentELR, name=currentELRdefinition, isELR=True) )
                                if hasDefLB:
                                    defLB.append( LBentry(role=currentELR, name=currentELRdefinition, isELR=True) )
                                if hasCalLB:
                                    calLB.append( LBentry(role=currentELR, name=currentELRdefinition, isELR=True) )
                                    calRels = set() # prevent duplications when same rel in different parts of tree
                                if hasGenLB:
                                    genLB.append( LBentry(role=currentELR, name=currentELRdefinition, isELR=True) )
                    prefix, name = rowPrefixNameValues(row)
                    if cellHasValue(row, 'depth', int):
                        depth = cellValue(row, 'depth')
                    elif hasPresentationIndentColumnIndentations:
                        if row[presentationIndentColumn].value is not None:
                            depth = int(row[presentationIndentColumn].alignment.indent)
                        else:
                            depth = None
                    elif hasDepthColumn:
                        depth = None # non-ELR section, no depth
                    else: # depth provided by parent reference
                        depth = 0
                    subsGrp = cellValue(row, 'substitutionGroup')
                    isConcept = subsGrp in ("xbrli:item", "xbrli:tuple",
                                            "xbrldt:hypercubeItem", "xbrldt:dimensionItem")
                    if (prefix in genDocs) and name not in genDocs[prefix].extensionElements and name:
                        thisDoc = genDocs[prefix]
                        # elements row
                        eltType = cellValue(row, 'type')
                        eltTypePrefix = cellValue(row, 'typePrefix')
                        if not eltType:
                            eltType = 'xbrli:stringItemType'
                        elif eltTypePrefix and ':' not in eltType:
                            eltType = eltTypePrefix + ':' + eltType
                        elif ':' not in eltType and eltType.endswith("ItemType"):
                            eltType = 'xbrli:' + eltType
                        abstract = cellValue(row, 'abstract')
                        nillable = cellValue(row, 'nillable')
                        balance = cellValue(row, 'balance')
                        periodType = cellValue(row, 'periodType')
                        eltAttrs = {"name": name, "id": (prefix or "") + "_" + name}
                        if eltType:
                            eltAttrs["type"] = eltType
                            checkImport(thisDoc, eltType)
                        if subsGrp:
                            eltAttrs["substitutionGroup"] = subsGrp
                            checkImport(thisDoc, subsGrp)
                        if abstract or subsGrp in ("xbrldt:hypercubeItem", "xbrldt:dimensionItem"):
                            eltAttrs["abstract"] = abstract or "true"
                        if nillable:
                            eltAttrs["nillable"] = nillable
                        if balance:
                            eltAttrs["{http://www.xbrl.org/2003/instance}balance"] = balance
                        if periodType:
                            eltAttrs["{http://www.xbrl.org/2003/instance}periodType"] = periodType
                        if hasConceptAttributeColumn:
                            # custom attributes (attribute, prefix:localName in header)
                            for header in headerCols:
                                if isinstance(header, str) and header.startswith("attribute, "):
                                    value = cellValue(row, header)
                                    if value not in (None, ""):
                                        eltAttrs[header[11:]] = value # fix QName later after schemaElt exists
                        eltFacets = None
                        eltEnumRefParts = None
                        if eltType not in ("nonnum:domainItemType", "xbrli:booleanItemType", "xbrli:positiveIntegerItemType", "xbrli:dateItemType",
                                           "xbrli:gYearItemType"):
                            for facet in ("minLength", "maxLength", "minInclusive", "maxInclusive",
                                          "length", "fixed", "pattern", "enumeration", "excludedEnumeration"):
                                v = cellValue(row, facet)
                                if v is not None:
                                    if facet == "enumeration" and v.startswith("See tab "): # check for local or tab-contained enumeration
                                        _match = re.match(r"See tab ([^!]+)([!]([0-9]+):([0-9]+))?", v)
                                        if _match:
                                            _tab, _dummy, _rowFrom, _rowTo = _match.groups()
                                            if _tab in sheetNames:
                                                enumWs = importExcelBook[_tab]
                                                if _rowFrom and _rowTo:
                                                    # take cols named "enumeration" and "reference parts"
                                                    colHdrs = [enumWs.cell(row=1,column=i).value for i in range(1,enumWs.max_column+1)]
                                                    eltEnumValues = []
                                                    eltEnumRefsParts = []
                                                    for i in range(int(_rowFrom), int(_rowTo)+1):
                                                        _parts = []
                                                        eltEnumRefsParts.append(_parts)
                                                        for j, h in enumerate(colHdrs):
                                                            c = enumWs.cell(row=i,column=j+1).value
                                                            if c is not None:
                                                                if h == "enumeration":
                                                                    eltEnumValues.append(str(c))
                                                                else:
                                                                    m = resourceParsePattern.match(h)
                                                                    if m:
                                                                        _resourceType = m.group(1)
                                                                        _resourceRole = "/" + m.group(2) # last path seg of role
                                                                        _resourceLangOrPart = m.group(4) # lang or part
                                                                        _parts.append(((_resourceType, _resourceRole, _resourceLangOrPart), c))
                                                    v = "\n".join(eltEnumValues) if eltEnumValues else None
                                                else: # cols 1 and 2 are enum and labels
                                                    v = "\n".join(" = ".join(xlValue(col) for col in row if xlValue(col))
                                                                  for i, row in enumerate(enumWs.rows)
                                                                  if i > 0) # skip heading row
                                    if v is not None:
                                        if eltFacets is None: eltFacets = {}
                                        eltFacets[facet] = v
                        # if extension type is this schema, add extensionType for facets
                        if eltType and ':' in eltType:
                            _typePrefix, _sep, _typeName = eltType.rpartition(":")
                            baseType = cellValue(row, 'baseType')
                            baseTypePrefix = cellValue(row, 'baseTypePrefix')
                            if baseType and baseTypePrefix:
                                _baseType = "{}:{}".format(baseTypePrefix, baseType)
                            elif baseType:
                                _baseType = baseType
                            elif _typeName.endswith("ItemType"):
                                _baseType = "xbrli:tokenItemType" # should be a column??
                            else:
                                _baseType = "xs:token"
                            if _typePrefix in genDocs:
                                _typeDoc = genDocs[_typePrefix]
                                if _typeName not in _typeDoc.extensionTypes:
                                    _typeDoc.extensionTypes[_typeName] = ({"name":_typeName, "base":_baseType},eltFacets)
                                thisDoc.extensionElements[name] = (eltAttrs, None)
                            else: # not declarable
                                thisDoc.extensionElements[name] = (eltAttrs, eltFacets)
                        else:
                            thisDoc.extensionElements[name] = (eltAttrs, eltFacets)
                        thisDoc = None # deref for debugging
                    useLabels = True
                    if depth is not None or hasPresentationParentColumn:
                        if name is None:
                            _label = None
                            for colCell in row:
                                if colCell.value is not None:
                                    _label = xlValue(colCell)
                                    break
                            print ("Sheet {} row {} has relationships and no \"name\" field, label: {}".format(importSheetName, iRow+1, _label))
                        if hasPreLB:
                            preferredLabel = cellValue(row, 'preferredLabel')
                            if hasDepthColumn or hasPresentationIndentColumnIndentations:
                                entryList = lbDepthList(preLB, depth)
                                if entryList is not None and isConcept:
                                    if not name or not prefix:
                                        _name = "none"
                                    if depth == topDepth:
                                        entryList.append( LBentry(prefix=prefix, name=name, isRoot=True) )
                                    else:
                                        entryList.append( LBentry(prefix=prefix, name=name, arcrole=XbrlConst.parentChild,
                                                                  role=preferredLabel) )
                            elif hasPresentationParentColumn:
                                preParent = cellValue(row, 'presentationParent', default='') # only one top parent makes sense
                                if preParent:
                                    preParentPrefix, _sep, preParentName = preParent.rpartition(":")
                                    preParentName = valueNameChars(preParentName)
                                    entryList = lbDepthList(preLB, topDepth)
                                    if entryList is not None:
                                        preRel = (preParentPrefix, preParentName, prefix, name, currentELR or currentELRdefinition)
                                        if preRel not in preRels:
                                            entryList.append( LBentry(prefix=preParentPrefix, name=preParentName, isRoot=True, childStruct=
                                                                      [LBentry(prefix=prefix, name=name, arcrole=XbrlConst.parentChild,
                                                                               preferredLabel=preferredLabel )]) )
                                            preRels.add(preRel)
                                        else:
                                            pass
                        if hasDefLB and topDepth != 999999:
                            entryList = lbDepthList(defLB, depth)
                            if entryList is not None:
                                if depth == topDepth:
                                    if isConcept:
                                        entryList.append( LBentry(prefix=prefix, name=name, isRoot=True) )
                                else:
                                    if (not preferredLabel or # prevent start/end labels from causing duplicate dim-mem relationships
                                        not any(lbEntry.prefix == prefix and lbEntry.name == name
                                                for lbEntry in entryList)):
                                        # check if entry is a typed dimension
                                        eltAttrs = {}
                                        parentLBentry = lbDepthList(defLB, depth - 1)[-1]
                                        parentName = parentLBentry.name
                                        parentEltAttrs = {}
                                        for doc in genDocs.values():
                                            if name in doc.extensionElements:
                                                eltAttrs = doc.extensionElements.get(name, NULLENTRY)[0]
                                            if parentName in doc.extensionElements:
                                                parentEltAttrs = doc.extensionElements.get(parentName, NULLENTRY)[0]
                                        if (isUSGAAP and # check for typed dimensions
                                            parentEltAttrs.get("substitutionGroup") == "xbrldt:dimensionItem"
                                            and eltAttrs.get("type") != "nonnum:domainItemType"):
                                            # typed dimension, no LBentry
                                            typedDomainRef = "#" + eltAttrs.get("id", "")
                                            parentEltAttrs["{http://xbrl.org/2005/xbrldt}typedDomainRef"] = typedDomainRef
                                        elif isConcept:
                                            # explicit dimension
                                            role = None # default for a default dimension
                                            if "dimensionDefault" in headerCols and cellHasValue(row, 'dimensionDefault', (str,bool)):
                                                v = cellValue(row, 'dimensionDefault', strip=True)
                                                if v:
                                                    role = "_dimensionDefault_"
                                            entryList.append( LBentry(prefix=prefix, name=name, arcrole="_dimensions_", role=role) )
                        if hasCalLB:
                            calcParents = cellValue(row, 'calculationParent', default='').split()
                            calcWeights = str(cellValue(row, 'calculationWeight', default='')).split() # may be float or string
                            if calcParents and calcWeights:
                                # may be multiple parents split by whitespace
                                for i, calcParent in enumerate(calcParents):
                                    calcWeight = calcWeights[i] if i < len(calcWeights) else calcWeights[-1]
                                    calcParentPrefix, _sep, calcParentName = calcParent.rpartition(":")
                                    calcParentName = valueNameChars(calcParentName)
                                    entryList = lbDepthList(calLB, topDepth)
                                    if entryList is not None:
                                        calRel = (calcParentPrefix, calcParentName, prefix, name)
                                        if calRel not in calRels:
                                            entryList.append( LBentry(prefix=calcParentPrefix, name=calcParentName, isRoot=True, childStruct=
                                                                      [LBentry(prefix=prefix, name=name, arcrole=XbrlConst.summationItem, weight=calcWeight )]) )
                                            calRels.add(calRel)
                                        else:
                                            pass
                    hasRelationshipToCol = any(h[0] == "relationship to" for h in headerCols if isinstance(h, tuple))

                # accumulate extension labels and any reference parts
                if useLabels or hasRelationshipToCol:
                    prefix, name = rowPrefixNameValues(row)
                    if name is not None and (prefix in genDocs or extensionPrefixForCoreLabels or hasRelationshipToCol):
                        thisDoc = genDocs.get(extensionPrefixForCoreLabels or prefix) # None for relationshipTo a imported concept
                        preferredLabel = cellValue(row, 'preferredLabel')
                        for colItem, iCol in headerCols.items():
                            if isinstance(colItem, tuple):
                                colItemType = colItem[0]
                                role = colItem[1]
                                lang = part = colItem[2] # lang for label, part for reference
                                cell = row[iCol]
                                v = xlValue(cell)
                                if v is None or (isinstance(v, str) and not v):
                                    values = ()
                                else:
                                    v = str(v) # may be an int or float instead of str
                                    if colItemType in ("label", "reference", "relationship to"):
                                        values = (v,)
                                    elif colItemType in ("labels", "references"):
                                        values = v.split('\n')

                                if preferredLabel and "indented" in colItem and not hasPreferredLabelTextColumn:  # indented column sets preferredLabel if any
                                    role = preferredLabel
                                for i, value in enumerate(values):
                                    if colItemType == "relationship to": # doesn't require thisDoc
                                        entryList = lbDepthList(genLB, topDepth)
                                        if entryList is not None:
                                            toName = value
                                            if ":" in toName:
                                                toPrefix, _sep, toName = value.partition(":")
                                            else:
                                                toPrefix = prefix
                                            if hasRelationshipAttributeColumn:
                                                # custom attributes (attribute, prefix:localName in header)
                                                relAttrs = None
                                                for header in headerCols:
                                                    if isinstance(header, str) and header.startswith("relationship attribute, "):
                                                        attrValue = cellValue(row, header)
                                                        if attrValue not in (None, ""):
                                                            if relAttrs is None: relAttrs = {}
                                                            relAttrs[header[24:]] = attrValue # fix QName later after schemaElt exists
                                            entryList.append( LBentry(prefix=prefix, name=name, isRoot=True, childStruct=
                                                                      [LBentry(prefix=toPrefix, name=toName, arcrole=role, relAttrs=relAttrs)]) )
                                    elif thisDoc is None:
                                        pass
                                    # following options only apply to linkbases of generated taxonomies
                                    elif colItemType in ("label", "labels"):
                                        if isConcept:
                                            if hasPreferredLabelTextColumn and role == "/preferredLabel":
                                                role = preferredLabel
                                        else:
                                            if role == XbrlConst.standardLabel:
                                                role = XbrlConst.genStandardLabel # must go in generic labels LB
                                            elif role == XbrlConst.documentationLabel:
                                                role = XbrlConst.genDocumentationLabel
                                            else:
                                                continue
                                        thisDoc.extensionLabels[prefix, name, lang, role] = value.strip()
                                    elif hasRefLB and colItemType == "reference":
                                        if isConcept:
                                            # keep parts in order and not duplicated
                                            thisDoc.extensionReferences[prefix, name, role].add((part, value.strip()))
                                    elif hasRefLB and colItemType == "references":
                                        if isConcept:
                                            # role ending in # is appended with the value ordinal
                                            if role.endswith("#"):
                                                _role = "{}{:05.0f}".format(role, i)
                                            else:
                                                _role = role
                                            _value = value.strip().replace("\\n", "\n")
                                            if part is None: # part space value
                                                _part, _sep, _value = _value.partition(" ")
                                            else:
                                                _part = part
                                            # keep parts in order and not duplicated
                                            thisDoc.extensionReferences[prefix, name, _role].add((_part, _value))
                        if isConcept and eltEnumRefsParts and thisDoc is not None:
                            for i, _enumRefParts in enumerate(eltEnumRefsParts):
                                for (colItemType, role, part), value in _enumRefParts:
                                    if colItemType == "reference":
                                        _role = "{}#{:05.0f}".format(role, i+1)
                                        thisDoc.extensionReferences[prefix, name, _role].add((part, value.strip()))
                        thisDoc = None # deref for debugging

            except Exception as err:
                fatalLoadingErrors.append("Excel sheet: {excelSheet}, row: {excelRow}, error: {error}, Traceback: {traceback}"
                                   .format(error=err, excelSheet=importSheetName, excelRow=iRow, traceback=traceback.format_tb(sys.exc_info()[2])))            # uncomment to debug raise

        if not headerCols:
            if not conceptsWs:
                fatalLoadingErrors.append("Neither control worksheet (XBRL DTS tab) nor standard columns found, no DTS imported.")
            elif not currentELR:
                fatalLoadingErrors.append("Extended link role not found, no DTS imported.")

    if fatalLoadingErrors:
        raise Exception(",\n ".join(fatalLoadingErrors))

    if isUSGAAP and hasDefLB:
        # move line items above table
        def fixUsggapTableDims(lvl1Struct, level=0):
            foundTable = False
            emptyLinks = []
            foundHeadingItems = []
            foundLineItems = []
            for lvl1Entry in lvl1Struct:
                for lvl2Entry in lvl1Entry.childStruct:
                    if any(lvl2Entry.name.endswith(suffix) for suffix in ("Table", "_table", "Cube", "_cube")):
                        for lvl3Entry in lvl2Entry.childStruct:
                            if any(lvl3Entry.name.endswith(suffix) for suffix in ("LineItems", "_line_items")):
                                foundLineItems.append((lvl1Entry, lvl2Entry, lvl3Entry))
                                foundTable = True
                                break
                    else:
                        foundHeadingItems.append((lvl1Entry, lvl2Entry))
                if not foundLineItems:
                    foundNestedTable = fixUsggapTableDims(lvl1Entry.childStruct, level+1)
                    if level == 0 and not foundNestedTable:
                        emptyLinks.append(lvl1Entry)
                    foundTable |= foundNestedTable
                    del foundHeadingItems[:]
            #if foundLineItems or foundHeadingItems:
            #    print("lvlentry {}\n headingITems {}\n emptyLinks {}\n\n".format(foundLineItems, foundHeadingItems, emptyLinks))
            for lvl1Entry, lvl2Entry, lvl3Entry in foundLineItems:
                i1 = lvl1Entry.childStruct.index(lvl2Entry)
                lvl1Entry.childStruct.insert(i1, lvl3Entry)  # must keep lvl1Rel if it is __root__
                lvl3Entry.childStruct.insert(0, lvl2Entry)
                if any(lvl1Entry.name.endswith(suffix)
                       for suffix in ("Abstract", "_abstract", "Root", "_root", "_package", "_heading")):
                    lvl1Entry.childStruct.remove(lvl2Entry)
                lvl2Entry.childStruct.remove(lvl3Entry)
            for lvl1Entry, lvl2Entry in foundHeadingItems:
                lvl1Entry.childStruct.remove(lvl2Entry)
            for emptyLink in emptyLinks:
                lvl1Struct.remove(emptyLink)

            return foundTable

        fixUsggapTableDims(defLB)

    modelDocuments = []
    modelXbrl.blockDpmDBrecursion = True

    def generateDoc(thisDoc, parentDoc, visitedDocNames):
        if thisDoc.name in visitedDocNames:
            modelXbrl.error("loadFromExcel:circularDependency",
                            "Generation order dependency is circular: %(circularDependency)s",
                            modelXbrl=modelXbrl, circularDependency=",".join(visitedDocNames) + ", " + thisDoc.name)
            return
        visitedDocNames.append(thisDoc.name)

        if XbrlConst.xsd not in thisDoc.importXmlns.values():
            eltName = 'schema xmlns="{}"'.format(XbrlConst.xsd)
        else:
            for k,v in thisDoc.importXmlns.items():
                if v == XbrlConst.xsd:
                    eltName = "{}:schema".format(k)
                    break
        doc = createModelDocument(
              modelXbrl,
              Type.SCHEMA,
              thisDoc.extensionSchemaFilename,
              isEntry=(parentDoc is None),
              # initialComment="extracted from OIM {}".format(mappedUri),
              documentEncoding="utf-8",
              base='', # block pathname from becomming absolute
              initialXml='''
        <{eltName}
            targetNamespace="{targetNamespace}"
            attributeFormDefault="unqualified"
            elementFormDefault="qualified"
            xmlns:xs="http://www.w3.org/2001/XMLSchema"
            xmlns:{extensionPrefix}="{targetNamespace}"
            {importXmlns}
            xmlns:nonnum="http://www.xbrl.org/dtr/type/non-numeric"
            xmlns:link="http://www.xbrl.org/2003/linkbase"
            xmlns:xbrli="http://www.xbrl.org/2003/instance"
            xmlns:xlink="http://www.w3.org/1999/xlink"
            xmlns:xbrldt="http://xbrl.org/2005/xbrldt"
            {schemaVersion}{xmlLang} />
        '''.format(eltName=eltName,
                   targetNamespace=thisDoc.extensionSchemaNamespaceURI,
                   extensionPrefix=thisDoc.extensionSchemaPrefix,
                   importXmlns=''.join('xmlns:{0}="{1}"\n'.format(prefix, namespaceURI)
                                       for prefix, namespaceURI in thisDoc.importXmlns.items()),
                   schemaVersion='version="{}" '.format(thisDoc.extensionSchemaVersion) if thisDoc.extensionSchemaVersion else '',
                   xmlLang='\n xml:lang="{}"'.format(saveXmlLang) if saveXmlLang else "",
                   ),
                initialComment=thisDoc.initialComment
                )
        if parentDoc is None:
            modelXbrl.modelDocument = doc
        thisDoc.generated = True # prevent recursion
        doc.loadedFromExcel = True # signal to save generated taoxnomy in saveToFile below

        doc.inDTS = True  # entry document always in DTS
        doc.targetNamespace = thisDoc.extensionSchemaNamespaceURI # not set until schemaDiscover too late otherwise
        schemaElt = doc.xmlRootElement

        #foreach linkbase
        annotationElt = XmlUtil.addChild(schemaElt, XbrlConst.xsd, "annotation")
        if thisDoc.schemaDocumentation:
            XmlUtil.addChild(annotationElt, XbrlConst.xsd, "documentation", text=thisDoc.schemaDocumentation)
        appinfoElt = XmlUtil.addChild(annotationElt, XbrlConst.xsd, "appinfo")

        # add linkbaseRefs
        appinfoElt = XmlUtil.descendant(schemaElt, XbrlConst.xsd, "appinfo")

        # don't yet add linkbase refs, want to process imports first to get roleType definitions

        # add includes
        for filename in thisDoc.includes:
            XmlUtil.addChild(schemaElt, XbrlConst.xsd, "include", attributes=( ("schemaLocation", filename), ) )

        # add imports
        for importPrefix, importAttributes in sorted(thisDoc.imports.items(),
                                                     key=lambda item:item[1]):
            XmlUtil.addChild(schemaElt, XbrlConst.xsd, "import", attributes=importAttributes)
            # is the import an xsd which we have to generate
            if importPrefix in genDocs and not genDocs[importPrefix].generated:
                generateDoc(genDocs[importPrefix], doc, visitedDocNames) # generate document

        # add imports for gen LB if any role definitions (for discovery) and generic labels
        if any(roleURI in thisDoc.extensionRoleLabels for roleURI in thisDoc.extensionRoles.keys()):
            for importAttributes in ((("namespace", XbrlConst.gen), ("schemaLocation", "http://www.xbrl.org/2008/generic-link.xsd")),
                                     (("namespace", XbrlConst.genLabel), ("schemaLocation", "http://www.xbrl.org/2008/generic-label.xsd"))):
                XmlUtil.addChild(schemaElt, XbrlConst.xsd, "import", attributes=importAttributes )

        _enumNum = [1] # must be inside an object to be referenced in a nested procedure

        def addFacets(thisDoc, restrElt, facets):
            if facets:
                excludedEnumeration = facets.get("excludedEnumeration")
                if ((annotateEnumerationsDocumentation and excludedEnumeration == "X")
                    or excludedEnumeration == "D"):
                    # if generateEnumerationsDocumentationOnly annotation must be first child element
                    for facet, facetValue in facets.items():
                        if facet == "enumeration":
                            enumerationsDocumentation = []
                            for valLbl in facetValue.split("\n"):
                                val, _sep, _label = valLbl.partition("=")
                                val = val.strip()
                                if len(val):
                                    if val == "(empty)":
                                        val = ""
                                    _label = _label.strip()
                                    enumerationsDocumentation.append("{}: {}".format(val, _label) if _label else val)
                            XmlUtil.addChild(XmlUtil.addChild(restrElt, XbrlConst.xsd, "annotation"),
                                             XbrlConst.xsd, "documentation", text=
                                            " \n".join(enumerationsDocumentation))
                for facet, facetValue in sorted(facets.items(), key=lambda i:facetSortOrder.get(i[0],i[0])):
                    if facet == "enumeration":
                        if not annotateEnumerationsDocumentation and not excludedEnumeration:
                            for valLbl in facetValue.split("\n"):
                                val, _sep, _label = valLbl.partition("=")
                                val = val.strip()
                                _label = _label.strip()
                                if len(val):
                                    if val == "(empty)":
                                        val = ""
                                    _attributes = {"value":val}
                                    if _label:
                                        _labelsByLang = None
                                        if _label.startswith("{") and _label.endswith("}"):
                                            try:
                                                # multi-lingual labels are json dict
                                                _labelsByLang = json.loads(_label)
                                            except json.decoder.JSONDecodeError:
                                                _labelsByLang = None
                                        _name = "enum{}".format(_enumNum[0])
                                        _attributes["id"] = thisDoc.extensionSchemaPrefix + "_" + _name
                                        _enumNum[0] += 1
                                        if _labelsByLang: #multilingual
                                            for _lang, _langLabel in _labelsByLang.items():
                                                thisDoc.extensionLabels[thisDoc.extensionSchemaPrefix, _name, _lang, XbrlConst.genStandardLabel] = _langLabel
                                        else: # non-multi-lingual labels
                                            thisDoc.extensionLabels[thisDoc.extensionSchemaPrefix, _name, defaultLabelLang, XbrlConst.genStandardLabel] = _label
                                    enumElt = XmlUtil.addChild(restrElt, XbrlConst.xsd, facet, attributes=_attributes)
                                    if thisDoc.hasEnumerationDocumentation and _label:
                                        if _labelsByLang: #multilingual
                                            annotationElt = XmlUtil.addChild(enumElt, XbrlConst.xsd, "annotation")
                                            for _lang, _langLabel in _labelsByLang.items():
                                                thisDoc.extensionLabels[thisDoc.extensionSchemaPrefix, _name, _lang, XbrlConst.genStandardLabel] = _langLabel
                                                XmlUtil.addChild(annotationElt, XbrlConst.xsd, "documentation", text=_langLabel,
                                                                 attributes={"{http://www.w3.org/XML/1998/namespace}lang": _lang})
                                        else: # non-multi-lingual labels
                                            XmlUtil.addChild(XmlUtil.addChild(enumElt, XbrlConst.xsd, "annotation"),
                                                             XbrlConst.xsd, "documentation", text=_label)
                    elif facet != "excludedEnumeration":
                        XmlUtil.addChild(restrElt, XbrlConst.xsd, facet, attributes={"value":str(facetValue)})

        # add elements
        for eltName, eltDef in sorted(thisDoc.extensionElements.items(), key=lambda item: item[0]):
            eltAttrs, eltFacets = eltDef
            if eltFacets and "type" in eltAttrs:
                eltType = eltAttrs["type"]
                del eltAttrs["type"]
            if any(':' in attrname for attrname in eltAttrs.keys()): # fix up any prefixed attr names to be clark notation
                for attrname, attrvalue in eltAttrs.copy().items():
                    if not attrname.startswith('{') and ':' in attrname:
                        del eltAttrs[attrname]
                        eltAttrs[schemaElt.prefixedNameQname(attrname).clarkNotation] = attrvalue
            isConcept = eltAttrs.get('substitutionGroup') in (
                "xbrli:item", "xbrli:tuple", "xbrldt:hypercubeItem", "xbrldt:dimensionItem")
            elt = XmlUtil.addChild(schemaElt,
                                   XbrlConst.xsd, "element",
                                   attributes=eltAttrs)
            if annotateElementDocumentation:
                for labelRole in (XbrlConst.documentationLabel, XbrlConst.genDocumentationLabel):
                    labelKey = (thisDoc.extensionSchemaPrefix, eltAttrs["name"], defaultLabelLang, labelRole)
                    if labelKey in thisDoc.extensionLabels:
                        XmlUtil.addChild(XmlUtil.addChild(elt, XbrlConst.xsd, "annotation"),
                                         XbrlConst.xsd, "documentation", text=thisDoc.extensionLabels[labelKey])
                        break # if std doc label found, don't continue to look for generic doc labe
            if elt is not None and eltFacets and isConcept:
                cmplxType = XmlUtil.addChild(elt, XbrlConst.xsd, "complexType")
                cmplxCont = XmlUtil.addChild(cmplxType, XbrlConst.xsd, "simpleContent")
                restrElt = XmlUtil.addChild(cmplxCont, XbrlConst.xsd, "restriction", attributes={"base": eltType})
                addFacets(thisDoc, restrElt, eltFacets)
                del eltType

        for roleURI, (roleDefinition, usedOnRoles) in sorted(thisDoc.extensionRoles.items(), key=lambda rd: rd[1]):
            roleElt = XmlUtil.addChild(appinfoElt, XbrlConst.link, "roleType",
                                       attributes=(("roleURI",  roleURI),
                                                   ("id", "roleType_" + roleURI.rpartition("/")[2])))
            if roleDefinition:
                XmlUtil.addChild(roleElt, XbrlConst.link, "definition", text=roleDefinition)
            if usedOnRoles:
                for usedOnRole in usedOnRoles.split():
                    XmlUtil.addChild(roleElt, XbrlConst.link, "usedOn", text=usedOnRole)
            else:
                if hasPreLB and any(e.childStruct and e.isELR and (e.role == roleURI or e.name == roleDefinition) for e in preLB):
                    XmlUtil.addChild(roleElt, XbrlConst.link, "usedOn", text="link:presentationLink")
                if hasDefLB and any(e.childStruct and e.isELR and (e.role == roleURI or e.name == roleDefinition) for e in defLB):
                    XmlUtil.addChild(roleElt, XbrlConst.link, "usedOn", text="link:definitionLink")
                if hasCalLB and any(e.childStruct and e.isELR and (e.role == roleURI or e.name == roleDefinition) for e in calLB):
                    XmlUtil.addChild(roleElt, XbrlConst.link, "usedOn", text="link:calculationLink")
                if hasGenLB and any(e.childStruct and e.isELR and (e.role == roleURI or e.name == roleDefinition) for e in genLB):
                    XmlUtil.addChild(roleElt, XbrlConst.link, "usedOn", text=qname("{http://xbrl.org/2008/generic}genlink:link"))

            if not any(True for e in roleElt.iterchildren("{http://www.xbrl.org/2003/linkbase}usedOn")):
                modelXbrl.error("loadFromExcel:roleUsedOn",
                    "Role does not appear to be used in any relationships and has no usedOn elements: %(roleURI)s",
                    modelXbrl=modelXbrl, roleURI=roleURI, filename=thisDoc.extensionSchemaNamespaceURI)

        # add role definitions (for discovery) and generic labels
        if any(roleURI in thisDoc.extensionRoleLabels for roleURI in thisDoc.extensionRoles.keys()):
            # add appinfo generic linkbase for gen labels
            genLabLB = XmlUtil.addChild(appinfoElt, XbrlConst.link, "linkbase")
            XmlUtil.addChild(genLabLB, XbrlConst.link, "roleRef",
                             attributes=(("roleURI",  XbrlConst.genStandardLabel),
                                         ("{http://www.w3.org/1999/xlink}href",  "http://www.xbrl.org/2008/generic-label.xsd#standard-label"),
                                         ("{http://www.w3.org/1999/xlink}type",  "simple")))
            XmlUtil.addChild(genLabLB, XbrlConst.link, "arcroleRef",
                             attributes=(("arcroleURI",  elementLabel),
                                         ("{http://www.w3.org/1999/xlink}href",  "http://www.xbrl.org/2008/generic-label.xsd#element-label"),
                                         ("{http://www.w3.org/1999/xlink}type",  "simple")))
            linkElt = XmlUtil.addChild(genLabLB, qname("{http://xbrl.org/2008/generic}genlink:link"),
                                       attributes=(("{http://www.w3.org/1999/xlink}type", "extended"),
                                                   ("{http://www.w3.org/1999/xlink}role", defaultLinkRole)))
            for roleURI, _defLabel in sorted(thisDoc.extensionRoles.items(), key=lambda rd: rd[0]):
                if roleURI in thisDoc.extensionRoleLabels:
                    xlLabel = roleURI.rpartition("/")[2]
                    XmlUtil.addChild(linkElt, XbrlConst.link, "loc",
                                     attributes=(("{http://www.w3.org/1999/xlink}type", "locator"),
                                                 ("{http://www.w3.org/1999/xlink}href", "#roleType_" + xlLabel),
                                                 ("{http://www.w3.org/1999/xlink}label", "loc_" + xlLabel)))
                    XmlUtil.addChild(linkElt, XbrlConst.qnGenArc,
                                     attributes=(("{http://www.w3.org/1999/xlink}type", "arc"),
                                                 ("{http://www.w3.org/1999/xlink}arcrole", elementLabel),
                                                 ("{http://www.w3.org/1999/xlink}from", "loc_" + xlLabel),
                                                 ("{http://www.w3.org/1999/xlink}to", "label_" + xlLabel)))
                    for (text, lang) in thisDoc.extensionRoleLabels[roleURI]:
                        XmlUtil.addChild(linkElt, qname("{http://xbrl.org/2008/label}genlabel:label"),
                                         attributes=(("{http://www.w3.org/1999/xlink}type", "resource"),
                                                     ("{http://www.w3.org/1999/xlink}label", "label_" + xlLabel),
                                                     ("{http://www.w3.org/1999/xlink}role", XbrlConst.genStandardLabel),
                                                     ("{http://www.w3.org/XML/1998/namespace}lang", lang)),
                                         text=text)

        def addLinkbaseRef(lbType, lbFilename, lbDoc):
            role = "http://www.xbrl.org/2003/role/{0}LinkbaseRef".format(lbType)
            lbRefElt = XmlUtil.addChild(appinfoElt, XbrlConst.link, "linkbaseRef",
                                        attributes=(("{http://www.w3.org/1999/xlink}type",  "simple"),
                                                    ("{http://www.w3.org/1999/xlink}href",
                                                     docRelpath(lbFilename, thisDoc.extensionSchemaRelDirname)),
                                                    ("{http://www.w3.org/1999/xlink}arcrole",  "http://www.w3.org/1999/xlink/properties/linkbase"),
                                                    # generic label ref has no role
                                                    ) + (() if lbType.startswith("generic") else
                                                         (("{http://www.w3.org/1999/xlink}role",  role),))
                                        )
            if lbDoc: # provided for generated linbase refs
                doc.referencesDocument[lbDoc] = ModelDocumentReference("href", lbRefElt)

        # add referenced (not generated) linkbases
        for lbRefType, filename, generate in thisDoc.linkbaseRefs:
            if not generate:
                # if linkbase is generated by another doc which isn't generated yet, generate it
                for otherGenDoc in genDocs.values():
                    if not otherGenDoc.generated and any(
                                _otherLbRefType == lbRefType and _otherFilename == filename and _otherGenerate
                                for _otherLbRefType, _otherFilename, _otherGenerate in otherGenDoc.linkbaseRefs):
                        generateDoc(otherGenDoc, doc, visitedDocNames) # generate document
                addLinkbaseRef(lbRefType, filename, None)

        doc.schemaDiscover(schemaElt, False, False, thisDoc.extensionSchemaNamespaceURI)

        # add types after include and import are discovered
        # block creating any type which was previously provided by an include of the same namespace
        for typeName, typeDef in sorted(thisDoc.extensionTypes.items(), key=lambda item: item[0]):
            if qname(thisDoc.extensionSchemaNamespaceURI, typeName) in modelXbrl.qnameTypes:
                continue # type already exists, don't duplicate
            typeAttrs, typeFacets = typeDef
            if typeName.endswith("ItemType") or typeAttrs.get("base", "").endswith("ItemType"):
                cmplxType = XmlUtil.addChild(schemaElt, XbrlConst.xsd, "complexType", attributes={"name": typeAttrs["name"]})
                contElt = XmlUtil.addChild(cmplxType, XbrlConst.xsd, "simpleContent")
            else:
                contElt = XmlUtil.addChild(schemaElt, XbrlConst.xsd, "simpleType", attributes={"name": typeAttrs["name"]})
            restrElt = XmlUtil.addChild(contElt, XbrlConst.xsd, "restriction", attributes={"base": typeAttrs["base"]})
            # remove duplicitous facets already in base type
            baseQn = qname(schemaElt, typeAttrs.get("base"))
            if typeFacets:
                if baseQn and baseQn.namespaceURI not in (XbrlConst.xsd, XbrlConst.xbrli) and baseQn in modelXbrl.qnameTypes:
                    # remove duplicated facets of underlying type
                    baseTypeFacets = modelXbrl.qnameTypes[baseQn].facets or () # allow iteration if None
                    typeFacets = dict((facet, value)
                                      for facet, value in typeFacets.items()
                                      if facet not in baseTypeFacets or str(baseTypeFacets[facet]) != value)
                addFacets(thisDoc, restrElt, typeFacets)

        # find extension label roles, reference roles and parts
        extLabelRoles = {}
        extReferenceRoles = {}
        extReferenceParts = {}
        extReferenceSchemaDocs = {}
        extUnrecognizedRoles = set()
        relationshipArcroles = {}
        relationshipArcqnames = {}
        def setExtRefPart(partLocalName):
            if partLocalName not in extReferenceParts:
                for partConcept in modelXbrl.nameConcepts.get(partLocalName, ()):
                    if partConcept is not None and partConcept.subGroupHeadQname == qnLinkPart:
                        extReferenceParts[partLocalName] = partConcept.qname
                        extReferenceSchemaDocs[partConcept.qname.namespaceURI] = (
                            partConcept.modelDocument.uri if partConcept.modelDocument.uri.startswith("http://") else
                            partConcept.modelDocument.basename)
                        break
        for _headerColKey in headerColsAllElrs:
            if isinstance(_headerColKey, tuple) and len(_headerColKey) >= 3 and not _headerColKey[1].startswith("http://"):
                _resourceType = _headerColKey[0]
                _resourceRole = _headerColKey[1]
                _resourceLangOrPart = _headerColKey[2]
            elif isinstance(_headerColKey, str) and "!reference" in _headerColKey:
                m = resourceParsePattern.match(_headerColKey.partition("!")[2])
                _resourceType = m.group(1)
                _resourceRole = "/" + m.group(2)
                _resourceLangOrPart = m.group(4)
            else:
                continue
            _resourceQName, _standardRoles = {
                    "label": (qnLinkLabel, standardLabelRoles),
                    "labels": (qnLinkLabel, standardLabelRoles),
                    "reference": (qnLinkReference, standardReferenceRoles),
                    "references": (qnLinkReference, standardReferenceRoles)
                    }.get(_resourceType, (None,()))
            _resourceRoleURI = None
            # find resource role
            for _roleURI in _standardRoles:
                if _roleURI.endswith(_resourceRole):
                    _resourceRoleURI = _roleURI
                    _resourceRoleMatchPart = _resourceRole
                    break
            if _resourceRoleURI is None: # try custom roles
                _resourceRoleMatchPart = _resourceRole.partition("#")[0] # remove # part
                for _roleURI in modelXbrl.roleTypes:
                    if _roleURI.endswith(_resourceRoleMatchPart):
                        for _roleType in modelXbrl.roleTypes[_roleURI]:
                            if _resourceQName in _roleType.usedOns:
                                _resourceRoleURI = _roleURI
                                break
            if _resourceType in ("label", "labels"):
                if _resourceRoleURI:
                    extLabelRoles[_resourceRoleMatchPart] = _resourceRoleURI
                elif any(_resourceRoleMatchPart == k[2] for k in thisDoc.extensionLabels.keys()):
                    modelXbrl.error("loadFromExcel:labelResourceRole",
                        "Label resource role not found: %(role)s",
                        modelXbrl=modelXbrl, role=_resourceRoleMatchPart, filename=thisDoc.extensionSchemaNamespaceURI)
            elif _resourceType in ("reference", "references"):
                if _resourceRoleURI:
                    extReferenceRoles[_resourceRoleMatchPart] = _resourceRoleURI
                    # find part QName
                    setExtRefPart(_resourceLangOrPart)
                elif any(_resourceRoleMatchPart == k[2] for k in thisDoc.extensionReferences.keys()):
                    modelXbrl.error("loadFromExcel:referenceResourceRole",
                        "Reference resource role not found: %(role)s",
                        modelXbrl=modelXbrl, role=_resourceRoleMatchPart, filename=thisDoc.extensionSchemaNamespaceURI)
            elif _resourceType == "relationship to":
                for _arcroleURI in modelXbrl.arcroleTypes:
                    if _arcroleURI.endswith(_resourceRoleMatchPart):
                        for _arcroleType in modelXbrl.arcroleTypes[_arcroleURI]:
                            for _resourceQName in _arcroleType.usedOns:
                                break
                        break
                if _resourceQName is None:
                    modelXbrl.error("loadFromExcel:relationshipArcrole",
                        "Relationship arcrole not found: %(arcrole)s",
                        modelXbrl=modelXbrl, arcrole=_resourceRoleMatchPart, filename=thisDoc.extensionSchemaNamespaceURI)
                else:
                    relationshipArcroles[_resourceRoleMatchPart] = _arcroleURI
                    relationshipArcqnames[_arcroleURI] = _resourceQName

        # label linkbase
        for lbType, lang, filename in thisDoc.labelLinkbases:
            thisDoc.thisLBdir = posixpath.dirname(filename)
            langPattern = re.compile(lang or ".*")
            _isGeneric = lbType.startswith("generic")
            if _isGeneric and "http://xbrl.org/2008/label" not in modelXbrl.namespaceDocs:
                # must pre-load generic linkbases in order to create properly typed elements (before discovery because we're creating elements by lxml)
                ModelDocument.load(modelXbrl, "http://www.xbrl.org/2008/generic-link.xsd", isDiscovered=True)
                ModelDocument.load(modelXbrl, "http://www.xbrl.org/2008/generic-label.xsd", isDiscovered=True)
            lbDoc = ModelDocument.create(modelXbrl, ModelDocument.Type.LINKBASE, filename, base="", initialXml="""
            <linkbase
                xmlns="http://www.xbrl.org/2003/linkbase"
                xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance"
                xmlns:xlink="http://www.w3.org/1999/xlink"
                xmlns:xbrli="http://www.xbrl.org/2003/instance"
                {}
                xsi:schemaLocation="http://www.xbrl.org/2003/linkbase
                http://www.xbrl.org/2003/xbrl-linkbase-2003-12-31.xsd{}"
                {}>{}</linkbase>
            """.format("""
                xmlns:genlink="http://xbrl.org/2008/generic"
                xmlns:genlabel="http://xbrl.org/2008/label"
                """ if _isGeneric else "",
                       """
                http://xbrl.org/2008/generic  http://www.xbrl.org/2008/generic-link.xsd
                http://xbrl.org/2008/label  http://www.xbrl.org/2008/generic-label.xsd
                """ if _isGeneric else "",
                '\n xml:lang="{}"'.format(saveXmlLang) if saveXmlLang else "",
                """
                <arcroleRef arcroleURI="http://xbrl.org/arcrole/2008/element-label" xlink:href="http://www.xbrl.org/2008/generic-label.xsd#element-label" xlink:type="simple"/>
                """ if _isGeneric else ""),
                initialComment=thisDoc.initialComment)
            lbDoc.inDTS = True
            lbDoc.loadedFromExcel = True
            if isGenerateAndImport:
                addLinkbaseRef(lbType, filename, lbDoc) # must be explicitly imported
            lbElt = lbDoc.xmlRootElement
            linkElt = XmlUtil.addChild(lbElt,
                                       gen if _isGeneric else link,
                                       "link" if _isGeneric else "labelLink",
                                       attributes=(("{http://www.w3.org/1999/xlink}type", "extended"),
                                                   ("{http://www.w3.org/1999/xlink}role", defaultLinkRole)))
            firstLinkElt = linkElt
            locs = set()
            roleRefs = set()
            for labelKey, text in thisDoc.extensionLabels.items():
                prefix, name, labelLang, role = labelKey
                labelLang = labelLang or defaultLabelLang
                role = role.partition("#")[0] # remove # part
                role = extLabelRoles.get(role, role) # get custom role, if any
                if langPattern.match(labelLang) and _isGeneric == (role in (XbrlConst.genStandardLabel, XbrlConst.genDocumentationLabel)):
                    locLabel = prefix + "_" + name
                    if locLabel not in locs:
                        locs.add(locLabel)
                        XmlUtil.addChild(linkElt,
                                         XbrlConst.link, "loc",
                                         attributes=(("{http://www.w3.org/1999/xlink}type", "locator"),
                                                     ("{http://www.w3.org/1999/xlink}href", LBHref(thisDoc, prefix, name)),
                                                     ("{http://www.w3.org/1999/xlink}label", locLabel)))
                        XmlUtil.addChild(linkElt,
                                         gen if _isGeneric else link,
                                         "arc" if _isGeneric else "labelArc",
                                         attributes=(("{http://www.w3.org/1999/xlink}type", "arc"),
                                                     ("{http://www.w3.org/1999/xlink}arcrole", elementLabel if _isGeneric else conceptLabel),
                                                     ("{http://www.w3.org/1999/xlink}from", locLabel),
                                                     ("{http://www.w3.org/1999/xlink}to", "label_" + locLabel),
                                                     ("order", 1.0)))
                    XmlUtil.addChild(linkElt,
                                     XbrlConst.genLabel if _isGeneric else XbrlConst.link,
                                     "label",
                                     attributes=(("{http://www.w3.org/1999/xlink}type", "resource"),
                                                 ("{http://www.w3.org/1999/xlink}label", "label_" + locLabel),
                                                 ("{http://www.w3.org/1999/xlink}role", role)) + (
                                                (("{http://www.w3.org/XML/1998/namespace}lang", labelLang),)
                                                if True or lang != saveXmlLang else ()),
                                     text=text)
                    if role:
                        if role in XbrlConst.standardLabelRoles:
                            pass # no roleRef
                        elif role in modelXbrl.roleTypes:
                            roleType =  modelXbrl.roleTypes[role][0]
                            roleRefs.add(("roleRef", role, roleType.modelDocument.uri + "#" + roleType.id))
                        elif role.startswith("http://www.xbrl.org/2009/role/negated"):
                            roleRefs.add(("roleRef", role, "http://www.xbrl.org/lrr/role/negated-2009-12-16.xsd#" + role.rpartition("/")[2]))
                        else:
                            extUnrecognizedRoles.add(role)
            # add arcrole references
            for roleref, roleURI, href in roleRefs:
                XmlUtil.addChild(lbElt,
                                 XbrlConst.link, roleref,
                                 attributes=(("arcroleURI" if roleref == "arcroleRef" else "roleURI", roleURI),
                                             ("{http://www.w3.org/1999/xlink}type", "simple"),
                                             ("{http://www.w3.org/1999/xlink}href", href)),
                                 beforeSibling=firstLinkElt)
            lbDoc.linkbaseDiscover(lbElt)
            if extUnrecognizedRoles:
                modelXbrl.error("loadFromExcel:undefinedLabelRole",
                                "Label roles not defined: %(undefinedRoles)s",
                                modelXbrl=modelXbrl, undefinedRoles=",".join(sorted(extUnrecognizedRoles)))
                extUnrecognizedRoles.clear()

        # reference linkbase
        for lbType, referenceRole, filename in thisDoc.referenceLinkbases:
            thisDoc.thisLBdir = posixpath.dirname(filename)
            _isGeneric = lbType.startswith("generic")
            lbDoc = ModelDocument.create(modelXbrl, ModelDocument.Type.LINKBASE, filename, base="", initialXml="""
            <linkbase
                xmlns="http://www.xbrl.org/2003/linkbase"
                xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance"
                xmlns:xlink="http://www.w3.org/1999/xlink"
                xmlns:xbrli="http://www.xbrl.org/2003/instance"
                {}
                xsi:schemaLocation="http://www.xbrl.org/2003/linkbase
                http://www.xbrl.org/2003/xbrl-linkbase-2003-12-31.xsd{}{}"
                {}>{}</linkbase>
            """.format("""
                xmlns:genlink="http://xbrl.org/2008/generic"
                xmlns:genreference="http://xbrl.org/2008/rerference"
                """ if _isGeneric else "",
                "".join([" {} {}".format(_ns, _uri) for _ns, _uri in extReferenceSchemaDocs.items()]),
                """
                http://xbrl.org/2008/generic  http://www.xbrl.org/2008/generic-link.xsd
                http://xbrl.org/2008/reference  http://www.xbrl.org/2008/generic-reference.xsd
                """ if _isGeneric else "",
                '\n xml:lang="{}"'.format(saveXmlLang) if saveXmlLang else "",
               """
                <roleRef roleURI="http://www.xbrl.org/2008/role/label" xlink:href="http://www.xbrl.org/2008/generic-label.xsd#standard-label" xlink:type="simple"/>
                <arcroleRef arcroleURI="http://xbrl.org/arcrole/2008/element-reference" xlink:href="http://xbrl.org/2008/generic-reference.xsd#element-reference" xlink:type="simple"/>
                """ if _isGeneric else ""),
                initialComment=thisDoc.initialComment)
            lbDoc.inDTS = True
            lbDoc.loadedFromExcel = True
            if isGenerateAndImport:
                addLinkbaseRef(lbType, filename, lbDoc) # must be explicitly imported
            lbElt = lbDoc.xmlRootElement
            linkElt = XmlUtil.addChild(lbElt,
                                       XbrlConst.gen if _isGeneric else XbrlConst.link,
                                       "link" if _isGeneric else "referenceLink",
                                       attributes=(("{http://www.w3.org/1999/xlink}type", "extended"),
                                                   ("{http://www.w3.org/1999/xlink}role", defaultLinkRole)))
            firstLinkElt = linkElt
            locs = set()
            roleRefs = set()
            undefinedReferenceParts = set()
            for referenceKey, references in thisDoc.extensionReferences.items():
                prefix, name, role = referenceKey
                role = role.partition("#")[0] # remove # part
                role = extReferenceRoles.get(role, role) # get custom role, if any
                if fnmatch(role, referenceRole):
                    locLabel = prefix + "_" + name
                    # must use separate arcs with order to force Altova to display parts in order
                    if locLabel not in locs:
                        locs.add(locLabel)
                        order = 1
                    else:
                        for order in range(2,1000):
                            _locLabel = "{}_{}".format(locLabel, order)
                            if _locLabel not in locs:
                                locLabel = _locLabel
                                locs.add(locLabel)
                                break
                        if order > 999:
                            print("resource order de-duplicate failure, too many reference parts")
                    XmlUtil.addChild(linkElt,
                                     XbrlConst.link, "loc",
                                     attributes=(("{http://www.w3.org/1999/xlink}type", "locator"),
                                                 ("{http://www.w3.org/1999/xlink}href", LBHref(thisDoc, prefix, name)),
                                                 ("{http://www.w3.org/1999/xlink}label", locLabel)))
                    XmlUtil.addChild(linkElt,
                                     XbrlConst.link, "referenceArc",
                                     attributes=(("{http://www.w3.org/1999/xlink}type", "arc"),
                                                 ("{http://www.w3.org/1999/xlink}arcrole", conceptReference),
                                                 ("{http://www.w3.org/1999/xlink}from", locLabel),
                                                 ("{http://www.w3.org/1999/xlink}to", "label_" + locLabel),
                                                 ("order", order)))
                    referenceResource = XmlUtil.addChild(linkElt,
                                     XbrlConst.genReference if _isGeneric else XbrlConst.link,
                                     "reference",
                                     attributes=(("{http://www.w3.org/1999/xlink}type", "resource"),
                                                 ("{http://www.w3.org/1999/xlink}label", "label_" + locLabel),
                                                 ("{http://www.w3.org/1999/xlink}role", role)))
                    for part, text in references: # list to preserve desired order
                        setExtRefPart(part)
                        if part in extReferenceParts:
                            partQn = extReferenceParts.get(part, part) # get part QName if any
                            XmlUtil.addChild(referenceResource, partQn, text=text)
                        else:
                            undefinedReferenceParts.add(part)
                    if role:
                        if role in XbrlConst.standardLabelRoles:
                            pass # no roleRef
                        elif role in modelXbrl.roleTypes:
                            roleType = modelXbrl.roleTypes[role][0]
                            roleRefs.add(("roleRef", role, roleType.modelDocument.uri + "#" + roleType.id))
                        elif role.startswith("http://www.xbrl.org/2009/role/negated"):
                            roleRefs.add(("roleRef", role, "http://www.xbrl.org/lrr/role/negated-2009-12-16.xsd#" + role.rpartition("/")[2]))
                        else:
                            extUnrecognizedRoles.add(role)
            for part in sorted(undefinedReferenceParts):
                print("reference part not defined: {}".format(part))
            # add arcrole references
            for roleref, roleURI, href in roleRefs:
                XmlUtil.addChild(lbElt,
                                 XbrlConst.link, roleref,
                                 attributes=(("arcroleURI" if roleref == "arcroleRef" else "roleURI", roleURI),
                                             ("{http://www.w3.org/1999/xlink}type", "simple"),
                                             ("{http://www.w3.org/1999/xlink}href", href)),
                                 beforeSibling=firstLinkElt)
            lbDoc.linkbaseDiscover(lbElt)
            if extUnrecognizedRoles:
                modelXbrl.error("loadFromExcel:undefinedReferenceRole",
                                "Reference roles not defined: %(undefinedRoles)s",
                                modelXbrl=modelXbrl, undefinedRoles=",".join(sorted(extUnrecognizedRoles)))
                extUnrecognizedRoles.clear()

        prefixedNamespaces = modelXbrl.prefixedNamespaces
        def hrefConcept(prefix, name):
            qn = qname(prefixedNamespaces[prefix], name)
            if qn in modelXbrl.qnameConcepts:
                return modelXbrl.qnameConcepts[qn]
            elif name in modelXbrl.nameConcepts: # prefix may be null or ambiguous to multiple documents, try concept local name
                return modelXbrl.nameConcepts[name][0]
            if prefix not in prefixedNamespaces:
                modelXbrl.error("loadFromExcel:undefinedRelationshipElementPrefix",
                                "Prefix not defined: %(prefix)s",
                                modelXbrl=modelXbrl, prefix=prefix)
                return None
            modelXbrl.error("loadFromExcel:undefinedRelationshipElement",
                            "QName not defined: %(prefix)s:%(localName)s",
                            modelXbrl=modelXbrl, prefix=prefix, localName=name)
            return None

        def prefixedNameQName(prefixedName):
            if ":" not in prefixedName:
                return prefixedName
            prefix, _sep, name = prefixedName.rpartition(":")
            if prefix not in prefixedNamespaces:
                modelXbrl.error("loadFromExcel:undefinedRelationshipAttributePrefix",
                                "Prefix not defined: %(prefix)s",
                                modelXbrl=modelXbrl, prefix=prefix)
                return prefixedName
            return QName(prefix, prefixedNamespaces[prefix], name)

        def lbTreeWalk(lbType, parentElt, lbStruct, roleRefs, dimDef=False, locs=None, arcsFromTo=None, fromPrefix=None, fromName=None):
            order = 1.0
            for lbEntry in lbStruct:
                if lbEntry.isELR:
                    if not lbEntry.childStruct: # skip empty ELRs
                        continue
                    role = "unspecified"
                    if lbEntry.role and lbEntry.role.startswith("http://"): # have a role specified
                        role = lbEntry.role
                    elif lbEntry.name: #may be a definition
                        for linkroleUri, modelRoleTypes in modelXbrl.roleTypes.items():
                            definition = modelRoleTypes[0].definition
                            if lbEntry.name == definition and linkroleUri in thisDoc.extensionRoles:
                                role = linkroleUri
                                break
                    if role == "unspecified":
                        # don't generate for roles not for this schema
                        continue
                        #
                        #modelXbrl.error("loadFromExcel:linkRoleDefinition",
                        #    "Link role has no definition: %(role)s",
                        #    modelXbrl=modelXbrl, role=lbEntry.name, filename=thisDoc.extensionSchemaNamespaceURI)
                    if role not in thisDoc.extensionRoles:
                        # don't generate for roles not for this schema
                        continue
                    if role == XbrlConst.defaultLinkRole:
                        pass
                    elif role in thisDoc.extensionRoles:
                        roleRefs.add(("roleRef", role, doc.uri + "#roleType_" + role.rpartition("/")[2]))
                    elif role in modelXbrl.roleTypes: # add roleRef
                        roleType = modelRoleTypes[0]
                        roleRefs.add(("roleRef", role, roleType.modelDocument.uri + "#" + roleType.id))
                    else:
                        extUnrecognizedRoles.add(role)
                    linkElt = XmlUtil.addChild(parentElt,
                                               XbrlConst.gen if lbType == "generic" else XbrlConst.link,
                                               "link" if lbType == "generic" else lbType + "Link",
                                               attributes=(("{http://www.w3.org/1999/xlink}type", "extended"),
                                                           ("{http://www.w3.org/1999/xlink}role", role)))
                    locs = set()
                    arcsFromTo = set()
                    lbTreeWalk(lbType, linkElt, lbEntry.childStruct, roleRefs, dimDef, locs, arcsFromTo)
                else:
                    toPrefix = lbEntry.prefix
                    toName = lbEntry.name
                    toHref = LBHref(thisDoc, toPrefix, toName)
                    if toHref is None:
                        modelXbrl.error("loadFromExcel:invalidQName",
                                        "%(linkbase)s relationship element with prefix '%(prefix)s' localName '%(localName)s' not found",
                                        modelXbrl=modelXbrl, linkbase=lbType, prefix=lbEntry.prefix, localName=lbEntry.name)
                        continue
                    if not toPrefix and toName in modelXbrl.nameConcepts:
                        toPrefix = modelXbrl.nameConcepts[toName][0].qname.prefix
                    toLabel = "{}_{}".format(toPrefix, toName)
                    toLabelAlt = None
                    if not lbEntry.isRoot:
                        if not fromPrefix and fromName in modelXbrl.nameConcepts:
                            fromPrefix = modelXbrl.nameConcepts[fromName][0].qname.prefix
                        fromLabel = "{}_{}".format(fromPrefix, fromName)
                        if (fromLabel, toLabel) in arcsFromTo:
                            # need extra loc to prevent arc from/to duplication in ELR
                            for i in range(1, 1000):
                                toLabelAlt = "{}_{}".format(toLabel, i)
                                if (fromLabel, toLabelAlt) not in arcsFromTo:
                                    toLabel = toLabelAlt
                                    break
                    if (toHref not in locs or toLabelAlt) and not dimDef:
                        XmlUtil.addChild(parentElt,
                                         XbrlConst.link, "loc",
                                         attributes=(("{http://www.w3.org/1999/xlink}type", "locator"),
                                                     ("{http://www.w3.org/1999/xlink}href", toHref),
                                                     ("{http://www.w3.org/1999/xlink}label", toLabel)))
                        locs.add(toHref)
                    if not lbEntry.isRoot:
                        arcsFromTo.add( (fromLabel, toLabel) )
                        if lbType == "calculation" and lbEntry.weight is not None:
                            otherAttrs = ( ("weight", lbEntry.weight), )
                        elif lbType == "presentation" and lbEntry.role:
                            if not lbEntry.role.startswith("http://"):
                                # check if any defined labels for this role
                                _labelRoleMatchPart = "/" + lbEntry.role
                                for _roleURI in modelXbrl.roleTypes:
                                    if _roleURI.endswith(_labelRoleMatchPart):
                                        for _roleType in modelXbrl.roleTypes[_roleURI]:
                                            if XbrlConst.qnLinkLabel in _roleType.usedOns:
                                                lbEntry.role = _roleURI
                                                break
                            if not lbEntry.role.startswith("http://"):
                                # default to built in label roles
                                lbEntry.role = "http://www.xbrl.org/2003/role/" + lbEntry.role
                            otherAttrs = ( ("preferredLabel", lbEntry.role), )
                            if lbEntry.role and lbEntry.role not in XbrlConst.standardLabelRoles:
                                if lbEntry.role in modelXbrl.roleTypes:
                                    roleType = modelXbrl.roleTypes[lbEntry.role][0]
                                    roleRefs.add(("roleRef", lbEntry.role, roleType.modelDocument.uri + "#" + roleType.id))
                                else:
                                    extUnrecognizedRoles.add(lbEntry.role)
                        elif lbType == "generic" and lbEntry.arcrole:
                            if not lbEntry.arcrole.startswith("http://"):
                                # check if any defined labels for this role
                                for _arcroleURI in modelXbrl.arcroleTypes:
                                    if _arcroleURI.endswith(lbEntry.arcrole):
                                        lbEntry.arcrole = _arcroleURI
                                        break
                            otherAttrs = tuple( (prefixedNameQName(_key), _value) # may need to process qname in key into clark name
                                                for _key, _value in (lbEntry.relAttrs.items() if lbEntry.relAttrs is not None else ()))
                        else:
                            otherAttrs = ( )
                        if lbEntry.arcrole == "_dimensions_":  # pick proper consecutive arcrole
                            fromConcept = hrefConcept(fromPrefix, fromName)
                            toConcept = hrefConcept(toPrefix, toName)
                            if dimDef: # special case for default dimension
                                if lbEntry.role != "_dimensionDefault_" and not lbTreeHasDimDefault(lbEntry.childStruct):
                                    continue # forget subtree, no default
                                if toConcept is not None and (toConcept.isDimensionItem or lbEntry.role == "_dimensionDefault_"):
                                    if (toHref not in locs or toLabelAlt):
                                        XmlUtil.addChild(parentElt,
                                                         XbrlConst.link, "loc",
                                                         attributes=(("{http://www.w3.org/1999/xlink}type", "locator"),
                                                                     ("{http://www.w3.org/1999/xlink}href", toHref),
                                                                     ("{http://www.w3.org/1999/xlink}label", toLabel)))
                                        locs.add(toHref)
                                    if lbEntry.role != "_dimensionDefault_":
                                        lbTreeWalk(lbType, parentElt, lbEntry.childStruct, roleRefs, dimDef, locs, arcsFromTo, toPrefix, toName)
                                    else:
                                        XmlUtil.addChild(parentElt, XbrlConst.link, "definitionArc",
                                                         attributes=(("{http://www.w3.org/1999/xlink}type", "arc"),
                                                                     ("{http://www.w3.org/1999/xlink}arcrole", XbrlConst.dimensionDefault),
                                                                     ("{http://www.w3.org/1999/xlink}from", fromLabel),
                                                                     ("{http://www.w3.org/1999/xlink}to", toLabel),
                                                                     ("order", order)) + otherAttrs )
                                        order += 1.0
                                else:
                                    lbTreeWalk(lbType, parentElt, lbEntry.childStruct, roleRefs, dimDef, locs, arcsFromTo, fromPrefix, fromName)
                                continue
                            elif toConcept is not None and toConcept.isHypercubeItem:
                                arcrole = XbrlConst.all
                                otherAttrs += ( (XbrlConst.qnXbrldtContextElement, "segment"),
                                                (qnXbrldtClosed, "true") )
                            elif toConcept is not None and toConcept.isDimensionItem:
                                arcrole = XbrlConst.hypercubeDimension
                            elif fromConcept is not None and fromConcept.isDimensionItem:
                                arcrole = XbrlConst.dimensionDomain
                            else:
                                arcrole = XbrlConst.domainMember
                        else:
                            arcrole = lbEntry.arcrole
                        if arcrole in relationshipArcqnames:
                            arcqname = relationshipArcqnames[arcrole]
                            arcNS = arcqname.namespaceURI
                            arcLocalname = arcqname.localName
                        elif lbType == "generic":
                            arcNS = XbrlConst.gen
                            arcLocalname = "arc"
                        else:
                            arcNS = XbrlConst.link
                            arcLocalname = lbType + "Arc"
                        XmlUtil.addChild(parentElt,
                                         arcNS, arcLocalname,
                                         attributes=(("{http://www.w3.org/1999/xlink}type", "arc"),
                                                     ("{http://www.w3.org/1999/xlink}arcrole", arcrole),
                                                     ("{http://www.w3.org/1999/xlink}from", fromLabel),
                                                     ("{http://www.w3.org/1999/xlink}to", toLabel),
                                                     ("order", order)) + otherAttrs )
                        order += 1.0
                    if lbType != "calculation" or lbEntry.isRoot:
                        lbTreeWalk(lbType, parentElt, lbEntry.childStruct, roleRefs, dimDef, locs, arcsFromTo, toPrefix, toName)

        def lbTreeHasDimDefault(lbStruct):
            for lbEntry in lbStruct:
                if lbEntry.isELR:
                    if not lbEntry.childStruct:
                        continue
                    if lbTreeHasDimDefault(lbEntry.childStruct):
                        return True
                else:
                    if not lbEntry.isRoot and (lbEntry.arcrole == "_dimensions_" and lbEntry.role == "_dimensionDefault_"):
                        return True
                    if lbTreeHasDimDefault(lbEntry.childStruct):
                        return True
            return False

        for hasLB, lbType, lbLB in ((hasPreLB and thisDoc.hasPreLB, "presentation", preLB),
                                    (hasDefLB and thisDoc.hasDefLB, "definition", defLB),
                                    (hasCalLB and thisDoc.hasCalLB, "calculation", calLB),
                                    (hasGenLB and thisDoc.hasGenLB, "generic", genLB)):
            if hasLB:
                for lbRefType, filename, generate in thisDoc.linkbaseRefs:
                    thisDoc.thisLBdir = posixpath.dirname(filename)
                    if generate and lbType == lbRefType:
                        # output presentation linkbase
                        lbDoc = ModelDocument.create(modelXbrl, ModelDocument.Type.LINKBASE, filename, base='', initialXml="""
                        <linkbase
                            xmlns="http://www.xbrl.org/2003/linkbase"
                            xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance"
                            xmlns:xlink="http://www.w3.org/1999/xlink"
                            xmlns:xbrli="http://www.xbrl.org/2003/instance"{}
                            xsi:schemaLocation="http://www.xbrl.org/2003/linkbase
                            http://www.xbrl.org/2003/xbrl-linkbase-2003-12-31.xsd{}"
                            />
                        """.format("""
                        xmlns:generic="http://xbrl.org/2008/generic"
                        """ if lbType == "generic" else "",
                        """
                        http://xbrl.org/2008/generic  http://www.xbrl.org/2008/generic-link.xsd
                        """ if lbType == "generic" else ""
                        ),
                            initialComment=thisDoc.initialComment)
                        lbDoc.inDTS = True
                        lbDoc.loadedFromExcel = True
                        addLinkbaseRef(lbRefType, filename, lbDoc)
                        lbElt = lbDoc.xmlRootElement
                        roleRefs = set()
                        if lbType == "definition":
                            roleRefs.update((("arcroleRef", XbrlConst.all, "http://www.xbrl.org/2005/xbrldt-2005.xsd#all"),
                                             ("arcroleRef", XbrlConst.dimensionDefault, "http://www.xbrl.org/2005/xbrldt-2005.xsd#dimension-default"),
                                             ("arcroleRef", XbrlConst.dimensionDomain, "http://www.xbrl.org/2005/xbrldt-2005.xsd#dimension-domain"),
                                             ("arcroleRef", XbrlConst.domainMember, "http://www.xbrl.org/2005/xbrldt-2005.xsd#domain-member"),
                                             ("arcroleRef", XbrlConst.hypercubeDimension, "http://www.xbrl.org/2005/xbrldt-2005.xsd#hypercube-dimension")))
                        elif lbType == "generic":
                            for _arcroleURI in relationshipArcroles.values():
                                for _arcroleType in modelXbrl.arcroleTypes[_arcroleURI]:
                                    roleRefs.add(("arcroleRef", _arcroleURI, _arcroleType.modelDocument.uri + "#" + _arcroleType.id))
                                    break
                        lbTreeWalk(lbType, lbElt, lbLB, roleRefs)
                        if lbType == "definition" and lbTreeHasDimDefault(lbLB):
                            lbTreeWalk(lbType, lbElt, lbLB, roleRefs, dimDef=True) # second tree walk for any dimension-defaults
                        firstLinkElt = None
                        for firstLinkElt in lbElt.iterchildren():
                            break
                        # add arcrole references
                        for roleref, roleURI, href in roleRefs:
                            XmlUtil.addChild(lbElt,
                                             link, roleref,
                                             attributes=(("arcroleURI" if roleref == "arcroleRef" else "roleURI", roleURI),
                                                         ("{http://www.w3.org/1999/xlink}type", "simple"),
                                                         ("{http://www.w3.org/1999/xlink}href",
                                                          docRelpath(href, thisDoc.thisLBdir))),
                                             beforeSibling=firstLinkElt)
                        lbDoc.linkbaseDiscover(lbElt)
                        break
                if extUnrecognizedRoles:
                    modelXbrl.error("loadFromExcel:undefinedRole",
                                    "%(lbType)s linkbase roles not defined: %(undefinedRoles)s",
                                    modelXbrl=modelXbrl, lbType=lbType, undefinedRoles=",".join(sorted(extUnrecognizedRoles)))
                    extUnrecognizedRoles.clear()
        visitedDocNames.pop()

    def LBHref(thisDoc, prefix, name):
        if not prefix and name in modelXbrl.nameConcepts:
            _concept = modelXbrl.nameConcepts[name][0]
            filename = _concept.modelDocument.uri
            prefix = _concept.qname.prefix
        elif prefix == thisDoc.extensionSchemaPrefix:
            filename = thisDoc.extensionSchemaFilename
        elif prefix in thisDoc.importFilenames:
            filename = thisDoc.importFilenames[prefix]
        elif prefix in genDocs:
            doc = genDocs[prefix]
            if not doc.generated:
                # try to load recursively
                generateDoc(doc, thisDoc)
            if doc.generated:
                filename = doc.extensionSchemaFilename
            else:
                return None
        elif name in modelXbrl.nameConcepts:
            filename = None
            for _concept in modelXbrl.nameConcepts[name]:
                if prefix == _concept.qname.prefix:
                    filename = _concept.modelDocument.uri
                    break
            if not filename:
                return None
        else:
            return None
        return "{0}#{1}_{2}".format(docRelpath(filename, thisDoc.thisLBdir), prefix, name)

    for thisDoc in genOrder:
        if not thisDoc.generated:
            generateDoc(thisDoc, None, [])

    #cntlr.addToLog("Completed in {0:.2} secs".format(time.time() - startedAt),
    #               messageCode="loadFromExcel:info")

    if priorCWD:
        os.chdir(priorCWD) # restore prior current working directory
    return modelXbrl.modelDocument

def isExcelPath(filepath):
    return os.path.splitext(filepath)[1] in (".xlsx", ".xls", ".xlsm")

def isExcelLoadable(modelXbrl, mappedUri, normalizedUri, filepath, **kwargs):
    return isExcelPath(filepath)

def excelLoaderFilingStart(cntlr, options, filesource, entrypointFiles, *args, **kwargs):
    global excludeDesignatedEnumerations, annotateEnumerationsDocumentation, annotateElementDocumentation, saveXmlLang
    excludeDesignatedEnumerations = options.ensure_value("excludeDesignatedEnumerations", False)
    annotateEnumerationsDocumentation = options.ensure_value("annotateEnumerationsDocumentation", False)
    annotateElementDocumentation = options.ensure_value("annotateElementDocumentation", False)
    saveXmlLang = options.ensure_value("saveLang", None)

def excelLoader(modelXbrl, mappedUri, filepath, *args, **kwargs):
    if not isExcelLoadable(modelXbrl, mappedUri, None, filepath):
        return None # not an OIM file

    cntlr = modelXbrl.modelManager.cntlr
    cntlr.showStatus(_("Loading Excel file: {0}").format(os.path.basename(filepath)))
    doc = loadFromExcel(cntlr, modelXbrl, filepath, mappedUri)
    if doc is None:
        return None # not an OIM file
    modelXbrl.loadedFromExcel = True
    return doc

def saveDts(cntlr, modelXbrl, outputDtsDir):
    from arelle import ModelDocument
    import shutil
    excelFileDir = os.path.dirname(modelXbrl.fileSource.url)
    def saveToFile(url):
        if os.path.isabs(url):
            return url
        filepath = os.path.join(outputDtsDir, url)
        os.makedirs(os.path.dirname(filepath), exist_ok=True)
        return filepath
    # save generated schema and their linkbases
    for doc in modelXbrl.urlDocs.values():
        if getattr(doc, "loadedFromExcel", False):
            doc.save(saveToFile(doc.uri), updateFileHistory=False)
            cntlr.showStatus(_("Saving XBRL DTS: {0}").format(os.path.basename(doc.uri)))
            for refDoc in doc.referencesDocument.keys():
                if refDoc.inDTS:
                    if refDoc.type == ModelDocument.Type.LINKBASE:
                        cntlr.showStatus(_("Saving XBRL DTS: {0}").format(os.path.basename(refDoc.uri)))
                        refDoc.save(saveToFile(refDoc.uri), updateFileHistory=False)
        elif not (UrlUtil.isAbsolute(doc.uri) or os.path.isabs(doc.uri) or outputDtsDir == excelFileDir):
            srcfile = os.path.join(excelFileDir, doc.uri)
            destfile = saveToFile(doc.uri)
            if os.path.exists(srcfile):
                if not os.path.exists(destfile):
                    shutil.copyfile(srcfile, destfile)
            else:
                modelXbrl.error("loadFromExcel:missingReference",
                                "Missing source file to copy to output DTS directory: %(missingFile)s",
                                modelXbrl=modelXbrl, missingFile=doc.uri)


def guiXbrlLoaded(cntlr, modelXbrl, attach, *args, **kwargs):
    if cntlr.hasGui and getattr(modelXbrl, "loadedFromExcel", False):
        from tkinter.filedialog import askdirectory
        outputDtsDir = askdirectory(parent=cntlr.parent,
                                    initialdir=cntlr.config.setdefault("outputDtsDir","."),
                                    title='Please select a directory for output DTS Contents')
        cntlr.config["outputDtsDir"] = outputDtsDir
        cntlr.saveConfig()
        if outputDtsDir:
            saveDts(cntlr, modelXbrl, outputDtsDir)
        cntlr.showStatus(_("Excel loading completed"), 3500)

def cmdLineXbrlLoaded(cntlr, options, modelXbrl, *args, **kwargs):
    if options.saveExcelDTSdirectory and getattr(modelXbrl, "loadedFromExcel", False):
        saveDts(cntlr, modelXbrl, options.saveExcelDTSdirectory)

def excelLoaderOptionExtender(parser, *args, **kwargs):
    parser.add_option("--save-Excel-DTS-directory",
                      action="store",
                      dest="saveExcelDTSdirectory",
                      help=_("Save a DTS loaded from Excel into this directory."))
    parser.add_option("--exclude-designated-enumerations",
                      action="store_true",
                      dest="excludeDesignatedEnumerations",
                      help=_("Save a DTS loaded from Excel into this directory."))
    parser.add_option("--annotate-enumerations-documentation",
                      action="store_true",
                      dest="annotateEnumerationsDocumentation",
                      help=_("Save a DTS loaded from Excel into this directory."))
    parser.add_option("--annotate-element-documentation",
                      action="store_true",
                      dest="annotateElementDocumentation",
                      help=_("Save a DTS loaded from Excel into this directory."))
    parser.add_option("--save-lang",
                      action="store",
                      dest="saveLang",
                      help=_("Save an xml:lang on top level elements (schema, linkbase)."))

class LBentry:
    __slots__ = ("prefix", "name", "arcrole", "role", "childStruct", "preferredLabel", "relAttrs")
    def __init__(self, prefix=None, name=None, arcrole=None, role=None, weight=None,
                 isELR=False, isRoot=False, childStruct=None, preferredLabel=None, relAttrs=None):
        if childStruct is not None:
            self.childStruct = childStruct
        else:
            self.childStruct = []
        self.prefix = prefix
        self.name = name
        if isELR:
            self.arcrole = "_ELR_"
        elif isRoot:
            self.arcrole = "_root_"
        else:
            self.arcrole = arcrole
        if weight is not None:  # summationItem
            self.role = weight
        else:
            self.role = role # resource role, or "default" if conept is a default dimension
        self.preferredLabel = preferredLabel
        self.relAttrs = relAttrs

    @property
    def isELR(self):
        return self.arcrole == "_ELR_"

    @property
    def isRoot(self):
        return self.arcrole == "_root_"

    @property
    def weight(self):
        if self.arcrole == summationItem:
            return self.role
        return None

    def __repr__(self):
        return "LBentry(prefix={},name={})".format(self.prefix,self.name)

__pluginInfo__ = {
    'name': 'Load From Excel',
    'version': '1.02',
    'description': "This plug-in loads XBRL from Excel and saves the resulting XBRL DTS.",
    'license': 'Apache-2',
    'author': authorLabel,
    'copyright': copyrightLabel,
    # classes of mount points (required)
    'ModelDocument.IsPullLoadable': isExcelLoadable,
    'ModelDocument.PullLoader': excelLoader,
    'CntlrWinMain.Xbrl.Loaded': guiXbrlLoaded,
    'CntlrCmdLine.Filing.Start': excelLoaderFilingStart,
    'CntlrCmdLine.Options': excelLoaderOptionExtender,
    'CntlrCmdLine.Xbrl.Loaded': cmdLineXbrlLoaded
}
